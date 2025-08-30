import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import copy

def modificar_horario_con_division_columna():
    """
    Modifica el archivo horarioUnificado_con_6t.xlsx:
    1. Divide cada columna de día en dos columnas
    2. El encabezado del día cubre ambas columnas
    3. Aplica reglas específicas de renombrado:
       - Turnos que se dividen en dos partes (ocupan ambas columnas):
         * 6TT → TLPT/NLPT
         * 6RT → MLPR/NLPR
         * 6T → TANT/NANT
         * 6R → MAST/NANR
         * 6N → MANR/TANR
         * 6S → MASR/TASR
         * 6MT → MLPR/TLPR
         * 3 → TAST/HXN4
         * 7 → BLPT/NLPR
       - Turnos que se renombran pero ocupan solo la primera columna:
         * 1T → BLPT
         * 1 → BANT
       - Otros turnos se mantienen iguales pero ocupan solo la primera columna
    """
    
    # Cargar el archivo original
    wb = openpyxl.load_workbook('horarioUnificado_con_6t.xlsx')
    ws = wb.active
    
    print(f"Procesando archivo: {wb.active.title}")
    print(f"Dimensiones originales: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Crear nuevo workbook
    nuevo_wb = openpyxl.Workbook()
    nuevo_ws = nuevo_wb.active
    nuevo_ws.title = ws.title
    
    # Definir colores
    azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    rojo_claro = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    
    def aplicar_color_seguro(celda, color):
        """Aplica color solo si existe"""
        if color:
            celda.fill = color
    
    # Copiar la primera columna (trabajadores) sin cambios y aplicar color azul claro
    for fila in range(1, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        celda_nueva = nuevo_ws.cell(row=fila, column=1, value=valor)
        celda_nueva.fill = azul_claro
    
    # Procesar columnas de días (desde la columna 2 en adelante)
    nueva_col = 2
    for col_original in range(2, ws.max_column + 1):
        dia = ws.cell(row=1, column=col_original).value
        
        # Crear dos columnas para cada día
        col_primera = nueva_col
        col_segunda = nueva_col + 1
        
        # Combinar las dos columnas para el encabezado del día
        nuevo_ws.merge_cells(start_row=1, start_column=col_primera, 
                           end_row=1, end_column=col_segunda)
        celda_encabezado = nuevo_ws.cell(row=1, column=col_primera, value=dia)
        
        # Centrar el encabezado y aplicar color
        celda_encabezado.alignment = Alignment(horizontal='center')
        
        # Aplicar color según el tipo de día
        if dia and str(dia).startswith("SUN"):
            # Domingo: rojo claro
            celda_encabezado.fill = rojo_claro
        else:
            # Otros días: azul claro
            celda_encabezado.fill = azul_claro
        
        # Procesar cada fila para este día
        for fila in range(2, ws.max_row + 1):
            turno_original = ws.cell(row=fila, column=col_original).value
            
            if turno_original:
                turno_str = str(turno_original).strip()
                
                # Obtener el color original del turno
                celda_original = ws.cell(row=fila, column=col_original)
                # Crear una copia del color original
                if celda_original.fill.start_color.rgb:
                    color_original = PatternFill(
                        start_color=celda_original.fill.start_color.rgb,
                        end_color=celda_original.fill.end_color.rgb,
                        fill_type=celda_original.fill.fill_type
                    )
                else:
                    color_original = None
                
                # Aplicar reglas específicas de renombrado
                if turno_str == "6TT":
                    # Dividir 6TT en TLPT y NLPT (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPT")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6RT":
                    # Dividir 6RT en MLPR y NLPR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MLPR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6T":
                    # Dividir 6T en TANT y NANT (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TANT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NANT")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6R":
                    # Dividir 6R en MAST y NANR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MAST")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NANR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6N":
                    # Dividir 6N en MANR y TANR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MANR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TANR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6S":
                    # Dividir 6S en MASR y TASR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MASR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TASR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "6MT":
                    # Dividir 6MT en MLPR y TLPR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="MLPR")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="TLPR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "3":
                    # Dividir 3 en TAST y HXN4 (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="TAST")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="HXN4")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "7":
                    # Dividir 7 en BLPT y NLPR (ocupa ambas columnas)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="NLPR")
                    aplicar_color_seguro(celda1, color_original)
                    aplicar_color_seguro(celda2, color_original)
                elif turno_str == "1T":
                    # Renombrar 1T a BLPT (ocupa solo la primera columna)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BLPT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")  # Segunda columna vacía
                    aplicar_color_seguro(celda1, color_original)
                elif turno_str == "1":
                    # Renombrar 1 a BANT (ocupa solo la primera columna)
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value="BANT")
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")  # Segunda columna vacía
                    aplicar_color_seguro(celda1, color_original)
                else:
                    # Otros turnos se mantienen iguales pero ocupan solo la primera columna
                    celda1 = nuevo_ws.cell(row=fila, column=col_primera, value=turno_original)
                    celda2 = nuevo_ws.cell(row=fila, column=col_segunda, value="")  # Segunda columna vacía
                    aplicar_color_seguro(celda1, color_original)
        
        # Aplicar color rojo claro a las dos columnas de domingos
        if dia and str(dia).startswith("SUN"):
            for fila in range(2, ws.max_row + 1):
                # Aplicar color rojo claro a las celdas vacías de domingo
                celda1 = nuevo_ws.cell(row=fila, column=col_primera)
                celda2 = nuevo_ws.cell(row=fila, column=col_segunda)
                
                # Solo aplicar color si la celda no tiene contenido (está vacía)
                if not celda1.value:
                    celda1.fill = rojo_claro
                if not celda2.value:
                    celda2.fill = rojo_claro
        
        nueva_col += 2  # Avanzar dos columnas para el siguiente día
    
    # Ajustar ancho de columnas
    for col in range(1, nuevo_ws.max_column + 1):
        nuevo_ws.column_dimensions[get_column_letter(col)].width = 8
    
    # Guardar el archivo modificado
    nombre_archivo_salida = "excel_con_division_de_columna.xlsx"
    nuevo_wb.save(nombre_archivo_salida)
    
    print(f"\nArchivo modificado guardado como: {nombre_archivo_salida}")
    print(f"Nuevas dimensiones: {nuevo_ws.max_row} filas x {nuevo_ws.max_column} columnas")
    
    # Mostrar resumen de los cambios
    print("\nResumen de cambios realizados:")
    print("- Cada columna de día se dividió en dos columnas")
    print("- El encabezado del día cubre ambas columnas")
    print("- Turnos que se dividen en dos partes (ocupan ambas columnas):")
    print("  * 6TT → TLPT/NLPT")
    print("  * 6RT → MLPR/NLPR")
    print("  * 6T → TANT/NANT")
    print("  * 6R → MAST/NANR")
    print("  * 6N → MANR/TANR")
    print("  * 6S → MASR/TASR")
    print("  * 6MT → MLPR/TLPR")
    print("  * 3 → TAST/HXN4")
    print("  * 7 → BLPT/NLPR")
    print("- Turnos que se renombran pero ocupan solo la primera columna:")
    print("  * 1T → BLPT")
    print("  * 1 → BANT")
    print("- Otros turnos se mantienen iguales pero ocupan solo la primera columna")

def mostrar_estructura_archivo():
    """
    Muestra la estructura del archivo original para referencia
    """
    try:
        wb = openpyxl.load_workbook('horarioUnificado_con_6t.xlsx')
        ws = wb.active
        
        print(f"Estructura del archivo original:")
        print(f"Archivo: horarioUnificado_con_6t.xlsx")
        print(f"Hoja: {ws.title}")
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
        
        # Mostrar encabezados
        print("\nEncabezados (primera fila):")
        for col in range(1, min(ws.max_column + 1, 10)):  # Primeras 10 columnas
            valor = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {repr(valor)}")
        
        # Mostrar algunos trabajadores
        print("\nPrimeros trabajadores:")
        for fila in range(2, min(ws.max_row + 1, 8)):  # Primeros 6 trabajadores
            trabajador = ws.cell(row=fila, column=1).value
            print(f"  Fila {fila}: {trabajador}")
        
        # Buscar turnos específicos
        print("\nBuscando turnos específicos:")
        turnos_encontrados = set()
        for fila in range(2, ws.max_row + 1):
            for col in range(2, ws.max_column + 1):
                valor = ws.cell(row=fila, column=col).value
                if valor and str(valor).strip() in ["6TT", "6T", "6RT", "6R", "1T", "1", "7"]:
                    turnos_encontrados.add(str(valor).strip())
        
        print(f"Turnos encontrados: {sorted(turnos_encontrados)}")
        
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

if __name__ == "__main__":
    # Mostrar estructura del archivo original
    mostrar_estructura_archivo()
    
    print("\n" + "="*50)
    print("INICIANDO MODIFICACIÓN DEL ARCHIVO")
    print("="*50)
    
    # Ejecutar la modificación
    modificar_horario_con_division_columna() 