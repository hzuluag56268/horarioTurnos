import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Tuple, Set
import os


class AsignadorTurnos6R:
    """
    Asigna turno "6R" con estas reglas:
    - Solo se asigna entre trabajadores elegibles: ['PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
      'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE']
    - No se considera el conteo de personal (se intenta asignar 1 por día)
    - Verificar que en ese día NO exista ya un "6R"
    - No asignar si el día siguiente tiene 1T/1/7/BLPTD/BANTD
    - Prioridades basadas en el día siguiente (mañana):
        1) DESC/TROP/SIND mañana y NO tener 1T/1/7/BLPTD/BANTD mañana
        2) NO tener 1T/T1/1/7 mañana
        3) Resto
    - Equidad: buscar paridad entre personas en el grupo (6R + 6RT + 7). Se usa re-balanceo moviendo 6R si es necesario.
    - Colorea las celdas "6R" de azul medio oscuro
    - Actualiza hoja "Estadísticas":
        * 1T = 1T + 7 + 1
        * 6RT = 6RT + 7 + 6R
        * 6T = 6TT
    - Archivo de entrada preferido: "horarioUnificado_con_6tt.xlsx"
    """

    TRABAJADORES_ELEGIBLES = [
        'PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
        'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE'
    ]

    COLOR_6R = "4169E1"  # Azul medio oscuro (RoyalBlue)

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_1.xlsx",
            "horarioUnificado_con_6tt.xlsx",
            "horarioUnificado_con_1t.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            elegido = "horarioUnificado_con_1.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()

        # Snapshot del estado original para respetar asignaciones preexistentes
        self.original_nonempty: Set[Tuple[int, int]] = set()
        self.original_6r: Set[Tuple[int, int]] = set()
        self._snapshot_estado_original()

        # Contador de equidad para el grupo 6R+6RT+7
        self.contador_grupo_6rt: Dict[str, int] = defaultdict(int)

        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _snapshot_estado_original(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None or str(valor).strip() == "":
                    continue
                val = str(valor).strip().upper()
                self.original_nonempty.add((fila, col))
                if val == "6R":
                    self.original_6r.add((fila, col))

    def _es_celda_originalmente_vacia(self, fila: int, col: int) -> bool:
        return (fila, col) not in self.original_nonempty

    def _es_celda_original_6r(self, fila: int, col: int) -> bool:
        return (fila, col) in self.original_6r

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _tiene_prioridad_manana(self, trabajador: str, col_dia: int) -> bool:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"DESC", "TROP", "SIND"}

    def _tiene_extra_manana(self, trabajador: str, col_dia: int) -> bool:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"1T", "T1", "1", "7"}

    def _tiene_restriccion_dura_manana(self, trabajador: str, col_dia: int) -> bool:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"1T", "1", "7", "BLPTD", "BANTD"}

    def _obtener_trabajadores_disponibles(self, col_dia: int) -> List[str]:
        disponibles: List[str] = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if (
                (celda.value is None or str(celda.value).strip() == "")
                and self._es_celda_originalmente_vacia(fila, col_dia)
            ):
                disponibles.append(trabajador)
        return disponibles

    def _existe_6r_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            if str(valor).strip().upper() == "6R":
                return True
        return False

    def _existe_nanrd_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            if str(valor).strip().upper() == "NANRD":
                return True
        return False

    def _seleccionar_equitativo(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        min_val = min(self.contador_grupo_6rt[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_grupo_6rt[c] == min_val]
        return random.choice(empatados)

    def _inicializar_contadores_desde_hoja(self) -> None:
        max_col = self.ws.max_column
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val in {"6RT", "7", "6R"}:
                    self.contador_grupo_6rt[trabajador] += 1

    def _actualizar_contadores(self, trabajador: str) -> None:
        self.contador_grupo_6rt[trabajador] += 1

    def asignar_6r_en_dia(self, col_dia: int) -> Optional[str]:
        # No duplicar 6R en el día
        if self._existe_6r_en_dia(col_dia):
            return None
        # No asignar si existe NANRD en el día
        if self._existe_nanrd_en_dia(col_dia):
            return None

        # Disponibilidad
        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        if not disponibles:
            return None

        # Excluir por restricción dura de mañana
        disponibles = [t for t in disponibles if not self._tiene_restriccion_dura_manana(t, col_dia)]
        if not disponibles:
            return None

        # Prioridades por mañana
        nivel1: List[str] = []
        nivel2: List[str] = []
        nivel3: List[str] = []
        for t in disponibles:
            prioridad = self._tiene_prioridad_manana(t, col_dia)
            tiene_extra = self._tiene_extra_manana(t, col_dia)
            tiene_dura = self._tiene_restriccion_dura_manana(t, col_dia)
            if prioridad and not tiene_dura and not tiene_extra:
                nivel1.append(t)
            elif not tiene_extra:
                nivel2.append(t)
            else:
                nivel3.append(t)

        for candidatos in (nivel1, nivel2, nivel3):
            elegido = self._seleccionar_equitativo(candidatos)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                # Confirmar que la celda fue originalmente vacía
                if not self._es_celda_originalmente_vacia(fila, col_dia):
                    return None
                self.ws.cell(row=fila, column=col_dia, value="6R")
                celda = self.ws.cell(row=fila, column=col_dia)
                celda.fill = PatternFill(start_color=self.COLOR_6R, end_color=self.COLOR_6R, fill_type="solid")
                self._actualizar_contadores(elegido)
                return elegido

        return None

    def _rebalancear_para_paridad(self) -> None:
        # Rebalancear hasta lograr diferencia <= 1 entre el máximo y el mínimo
        while True:
            conteos_actuales: Dict[str, int] = {}
            for t in self.TRABAJADORES_ELEGIBLES:
                if self._obtener_fila_trabajador(t):
                    conteos_actuales[t] = self.contador_grupo_6rt[t]
            if not conteos_actuales:
                break

            trabajador_max = max(conteos_actuales, key=conteos_actuales.get)
            trabajador_min = min(conteos_actuales, key=conteos_actuales.get)
            if conteos_actuales[trabajador_max] - conteos_actuales[trabajador_min] <= 1:
                break

            fila_max = self._obtener_fila_trabajador(trabajador_max)
            fila_min = self._obtener_fila_trabajador(trabajador_min)
            if not fila_max or not fila_min:
                break

            columna_candidata = None
            for col in range(2, self.ws.max_column + 1):
                valor_max = self.ws.cell(row=fila_max, column=col).value
                valor_min = self.ws.cell(row=fila_min, column=col).value
                if (
                    valor_max is not None
                    and str(valor_max).strip().upper() == "6R"
                    and not self._es_celda_original_6r(fila_max, col)
                    and (valor_min is None or str(valor_min).strip() == "")
                    and self._es_celda_originalmente_vacia(fila_min, col)
                    and not self._tiene_restriccion_dura_manana(trabajador_min, col)
                    and not self._existe_nanrd_en_dia(col)
                ):
                    columna_candidata = col
                    break

            if columna_candidata is None:
                break

            # Mover 6R de trabajador_max a trabajador_min en la misma columna
            self.ws.cell(row=fila_max, column=columna_candidata, value=None)
            celda_original = self.ws.cell(row=fila_max, column=columna_candidata)
            celda_original.fill = PatternFill(fill_type=None)

            self.ws.cell(row=fila_min, column=columna_candidata, value="6R")
            celda_nueva = self.ws.cell(row=fila_min, column=columna_candidata)
            celda_nueva.fill = PatternFill(start_color=self.COLOR_6R, end_color=self.COLOR_6R, fill_type="solid")

            self.contador_grupo_6rt[trabajador_max] -= 1
            self.contador_grupo_6rt[trabajador_min] += 1

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")   # 1T + 7 + 1
        ws_stats.cell(row=1, column=4, value="6RT")  # 6RT + 7 + 6R
        ws_stats.cell(row=1, column=5, value="6T")   # solo 6TT

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, 6):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7 + 1
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"1")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = 6RT + 7 + 6R
            formula_6rt = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6R")'
            )
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 6T = solo 6TT
            formula_6t = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6TT")'
            ws_stats.cell(row=fila_destino, column=5, value=formula_6t)

            fila_destino += 1

        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 8
        ws_stats.column_dimensions['C'].width = 8
        ws_stats.column_dimensions['D'].width = 8
        ws_stats.column_dimensions['E'].width = 8

    def procesar_todos_los_dias(self) -> None:
        max_col = self.ws.max_column
        for col in range(2, max_col + 1):
            self.asignar_6r_en_dia(col)

        # Re-balanceo para paridad del grupo 6R+6RT+7
        self._rebalancear_para_paridad()

        # Actualizar estadísticas y guardar
        self._actualizar_hoja_estadisticas()

        salida = "horarioUnificado_con_6r.xlsx"
        try:
            self.wb.save(salida)
            print(f"Archivo guardado como: {salida}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"Archivo por defecto en uso. Guardado como: {alternativo}")


if __name__ == "__main__":
    asignador = AsignadorTurnos6R()
    asignador.procesar_todos_los_dias() 