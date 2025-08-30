import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict
import os


class AsignadorTurnos6TT:
    """
    Asigna turno "6TT" con estas reglas:
    - Decisión por día según turnos operativos (fila de conteo):
      * >=16 → NO asignar
      * <=15 → asignar "6TT"
    - Solo se asigna entre trabajadores elegibles: ['YIS','MAQ','DJO','AFG','JLF','JMV']
      * Si ninguno tiene celda vacía, intentar con respaldo: ['FCE','JBV','HZG']
    - Verificar que en ese día NO exista ya un "6TT"
    - Prioridad basada en el día siguiente (mañana):
      * Preferir quienes NO tienen 1T/1/7 mañana
    - Equidad: balancear usando conteo de 6TT por persona y rebalanceo posterior 6RT+6TT (±1)
    - Actualiza hoja "Estadísticas" agregando la columna 6T (cuenta solo 6TT) y 6RT+6TT (suma de ambos)
    - Guarda como "horarioUnificado_con_6tt.xlsx"
    """

    TRABAJADORES_ELEGIBLES = [ 'YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']
    TRABAJADORES_RESPALDO = ['FCE', 'JBV', 'HZG']

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        # Elegir el archivo de entrada más reciente disponible
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_6rt.xlsx",
            "horarioUnificado_con_1t.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            elegido = "horarioUnificado_procesado.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()
        self.contador_6tt: Dict[str, int] = defaultdict(int)
        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _obtener_conteo_operativos(self, col_dia: int) -> Optional[int]:
        # Buscar etiqueta explícita de conteo
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TURNOS OPERATIVOS":
                try:
                    return int(self.ws.cell(row=fila, column=col_dia).value)
                except Exception:
                    return None
        # Fallback: última fila
        try:
            return int(self.ws.cell(row=self.ws.max_row, column=col_dia).value)
        except Exception:
            return None

    def _debe_asignar_en_dia(self, col_dia: int) -> bool:
        disponible = self._obtener_conteo_operativos(col_dia)
        if disponible is None:
            return False
        return disponible <= 15

    def _existe_6tt_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            val = self.ws.cell(row=fila, column=col_dia).value
            if val is None:
                continue
            if str(val).strip().upper() == "6TT":
                return True
        return False

    def _tiene_extra_manana(self, trabajador: str, col_dia: int) -> bool:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        val = self.ws.cell(row=fila, column=col_dia + 1).value
        if val is None:
            return False
        return str(val).strip().upper() in {"1T", "1", "7"}

    def _obtener_disponibles_lista(self, lista: List[str], col_dia: int) -> List[str]:
        disponibles: List[str] = []
        for trabajador in lista:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if celda.value is None or str(celda.value).strip() == "":
                disponibles.append(trabajador)
        return disponibles

    def _seleccionar_equitativo(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        min_val = min(self.contador_6tt[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_6tt[c] == min_val]
        return random.choice(empatados)

    def _inicializar_contadores_desde_hoja(self) -> None:
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            for col in range(2, self.ws.max_column + 1):
                val = self.ws.cell(row=fila, column=col).value
                if val is None:
                    continue
                if str(val).strip().upper() == "6TT":
                    self.contador_6tt[str(trabajador).strip().upper()] += 1

    # --- NUEVO: helpers y rebalanceo 6RT+6TT ---
    def _contar_total_6rt_6tt_por_trabajador(self, trabajador: str) -> int:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return 0
        total = 0
        for col in range(2, self.ws.max_column + 1):
            val = self.ws.cell(row=fila, column=col).value
            if val is None:
                continue
            if str(val).strip().upper() in {"6RT", "6TT"}:
                total += 1
        return total

    def _dias_con_6tt(self, trabajador: str) -> List[int]:
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return []
        dias: List[int] = []
        for col in range(2, self.ws.max_column + 1):
            val = self.ws.cell(row=fila, column=col).value
            if val is None:
                continue
            if str(val).strip().upper() == "6TT":
                dias.append(col)
        return dias

    def _celda_esta_vacia(self, fila: int, col: int) -> bool:
        val = self.ws.cell(row=fila, column=col).value
        return val is None or str(val).strip() == ""

    def _rebalancear_6rt_6tt(self) -> None:
        # Considerar solo elegibles presentes en la hoja
        elegibles_presentes = [t for t in self.TRABAJADORES_ELEGIBLES if self._obtener_fila_trabajador(t)]
        if not elegibles_presentes:
            return

        # Totales iniciales (6RT + 6TT)
        totales: Dict[str, int] = {t: self._contar_total_6rt_6tt_por_trabajador(t) for t in elegibles_presentes}

        # Límite de seguridad para evitar bucles
        limite = sum(totales.values()) + 50
        pasos = 0

        # Repetir hasta alcanzar ±1 o no haya más movimientos posibles
        while pasos < limite:
            pasos += 1
            max_t = max(totales.values())
            min_t = min(totales.values())
            if max_t - min_t <= 1:
                break

            donantes = [t for t in elegibles_presentes if totales[t] == max_t]
            receptores = [t for t in elegibles_presentes if totales[t] == min_t]

            movimiento_realizado = False

            for d in donantes:
                dias_d = self._dias_con_6tt(d)
                if not dias_d:
                    continue

                for col_dia in dias_d:
                    # Ordenar receptores priorizando quienes NO tienen 1T/1/7 mañana
                    receptores_ordenados = sorted(
                        receptores, key=lambda r: self._tiene_extra_manana(r, col_dia)
                    )
                    for r in receptores_ordenados:
                        fila_r = self._obtener_fila_trabajador(r)
                        if not fila_r:
                            continue
                        if not self._celda_esta_vacia(fila_r, col_dia):
                            continue

                        fila_d = self._obtener_fila_trabajador(d)
                        if not fila_d:
                            continue
                        val_d = self.ws.cell(row=fila_d, column=col_dia).value
                        if not (val_d and str(val_d).strip().upper() == "6TT"):
                            continue

                        # Mover 6TT dentro del mismo día: preserva "máx 1 por día" y la decisión diaria original
                        self.ws.cell(row=fila_d, column=col_dia, value=None)
                        self.ws.cell(row=fila_r, column=col_dia, value="6TT")

                        # Actualizar contadores
                        self.contador_6tt[d] -= 1
                        self.contador_6tt[r] += 1
                        totales[d] -= 1
                        totales[r] += 1

                        movimiento_realizado = True
                        break
                    if movimiento_realizado:
                        break
                if movimiento_realizado:
                    break

            if not movimiento_realizado:
                break

    # --- FIN NUEVO ---

    def asignar_6tt_en_dia(self, col_dia: int) -> Optional[str]:
        # Reglas de decisión por personal
        if not self._debe_asignar_en_dia(col_dia):
            return None

        # No duplicar 6TT en el día
        if self._existe_6tt_en_dia(col_dia):
            return None

        # Disponibilidad en elegibles; si no, usar respaldo
        disponibles = self._obtener_disponibles_lista(self.TRABAJADORES_ELEGIBLES, col_dia)
        if not disponibles:
            disponibles = self._obtener_disponibles_lista(self.TRABAJADORES_RESPALDO, col_dia)
        if not disponibles:
            return None

        # Prioridad: preferir quienes NO tienen 1T/1/7 mañana
        nivel1 = [t for t in disponibles if not self._tiene_extra_manana(t, col_dia)]
        nivel2 = [t for t in disponibles if t not in nivel1]

        for candidatos in (nivel1, nivel2):
            elegido = self._seleccionar_equitativo(candidatos)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                self.ws.cell(row=fila, column=col_dia, value="6TT")
                self.contador_6tt[elegido] += 1
                return elegido

        return None

    def procesar_todos_los_dias(self) -> None:
        for col in range(2, self.ws.max_column + 1):
            self.asignar_6tt_en_dia(col)

        # Rebalanceo: igualar 6RT+6TT entre elegibles (±1) moviendo 6TT en el mismo día
        self._rebalancear_6rt_6tt()

        self._actualizar_hoja_estadisticas()

        salida = "horarioUnificado_con_6tt.xlsx"
        try:
            self.wb.save(salida)
            print(f"Archivo guardado como: {salida}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"Archivo por defecto en uso. Guardado como: {alternativo}")

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")   # 1T + 7
        ws_stats.cell(row=1, column=4, value="6RT")  # 6RT + 7
        ws_stats.cell(row=1, column=5, value="6T")   # solo 6TT
        ws_stats.cell(row=1, column=6, value="6RT+6TT")  # NUEVO: suma 6RT y 6TT

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, 7):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = 6RT + 7
            formula_6rt = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 6T = solo 6TT
            formula_6t = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6TT")'
            ws_stats.cell(row=fila_destino, column=5, value=formula_6t)

            # NUEVO: 6RT+6TT (combinado)
            formula_mix = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6TT")'
            )
            ws_stats.cell(row=fila_destino, column=6, value=formula_mix)

            fila_destino += 1

        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 8
        ws_stats.column_dimensions['C'].width = 8
        ws_stats.column_dimensions['D'].width = 8
        ws_stats.column_dimensions['E'].width = 8
        ws_stats.column_dimensions['F'].width = 10


if __name__ == "__main__":
    asignador = AsignadorTurnos6TT()
    asignador.procesar_todos_los_dias() 