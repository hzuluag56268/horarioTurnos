import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Set
import os


class AsignadorTurnosMofis:
    """
    Asigna turnos específicos a controladores del grupo MOFIS con estas reglas:
    - Trabajadores elegibles: ['MEI', 'VCM', 'ROP', 'WEH']
    - Turnos a asignar según cantidad de elegibles disponibles:
      * 4 elegibles: [MS, TS, MN, TN]
      * 3 elegibles: [S, MN, TN]
      * 2 elegibles: [S, N]
      * 1 elegible: [N]
    - NO asignar si tienen turnos no operativos (DESC, TROP, LIBR, VACA, etc.)
    - SÍ sobreescribir turnos como X que no están en la lista de no operativos
    - Priorizar equidad en cantidad de turnos S+N por trabajador
    - Colorear celdas de amarillo claro
    - Verificar que no existan ya estos turnos en el día
    - Actualizar hoja "Estadísticas" con columna 6S (S+N)
    - Guardar como "horarioUnificado_con_mofis.xlsx"
    """

    TRABAJADORES_ELEGIBLES = ['MEI', 'VCM', 'ROP', 'WEH']
    
    # Turnos no operativos que impiden asignación
    TURNOS_NO_OPERATIVOS = {
        # Turnos básicos
        "DESC", "TROP", "LIBR",
        # Turnos completos
        "VACA", "COME", "COMT", "COMS",
        # Turnos adicionales originales
        "SIND", "CMED", "CERT",
        # Formación, instrucción y entrenamiento
        "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
        "MENT", "TENT", "NENT", "AENT",
        "MINS", "TINS", "NINS", "AINS",
        # Gestión, oficinas y grupos de trabajo
        "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
        "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
        "MGST", "TGST", "MOFI", "TOFI",
        # Operativos y asignaciones especiales
        "CET", "ATC", "KATC", "XATC", "YATC", "ZATC"
    }
    
    # Mapeo de cantidad de elegibles a turnos a asignar
    TURNOS_POR_CANTIDAD = {
        4: ["MS", "TS", "MN", "TN"],
        3: ["S", "MN", "TN"],
        2: ["S", "N"],
        1: ["N"]
    }

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        # Elegir el archivo de entrada más reciente disponible
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_diurnas.xlsx",
            "horarioUnificado_con_6rt.xlsx",
            "horarioUnificado_con_1t.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            elegido = "horarioUnificado_procesado.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()
        self.contador_sn: Dict[str, int] = defaultdict(int)  # Contador de turnos S+N
        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _es_turno_no_operativo(self, turno: str) -> bool:
        """Verifica si un turno está en la lista de no operativos"""
        if turno is None:
            return False
        return str(turno).strip().upper() in self.TURNOS_NO_OPERATIVOS

    def _es_turno_s_o_n(self, turno: str) -> bool:
        """Verifica si un turno es S o N (para conteo de equidad)"""
        if turno is None:
            return False
        return str(turno).strip().upper() in {"S", "N"}

    def _obtener_elegibles_disponibles(self, col_dia: int) -> List[str]:
        """Obtiene lista de trabajadores elegibles que no tienen turnos no operativos"""
        disponibles = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            valor = celda.value
            
            # Si la celda está vacía o tiene un turno que NO está en la lista de no operativos, es elegible
            if valor is None or str(valor).strip() == "" or not self._es_turno_no_operativo(valor):
                disponibles.append(trabajador)
        
        return disponibles

    def _existe_turno_en_dia(self, turno: str, col_dia: int) -> bool:
        """Verifica si ya existe un turno específico en el día"""
        for fila in range(2, 26):
            val = self.ws.cell(row=fila, column=col_dia).value
            if val is None:
                continue
            if str(val).strip().upper() == turno.upper():
                return True
        return False

    def _seleccionar_equitativo(self, candidatos: List[str]) -> Optional[str]:
        """Selecciona el trabajador con menos turnos S+N para mantener equidad"""
        if not candidatos:
            return None
        min_val = min(self.contador_sn[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_sn[c] == min_val]
        return random.choice(empatados)

    def _inicializar_contadores_desde_hoja(self) -> None:
        """Inicializa contadores de turnos S+N desde el archivo existente"""
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            for col in range(2, self.ws.max_column + 1):
                val = self.ws.cell(row=fila, column=col).value
                if val is None:
                    continue
                if self._es_turno_s_o_n(val):
                    self.contador_sn[str(trabajador).strip().upper()] += 1

    def asignar_turnos_en_dia(self, col_dia: int) -> List[str]:
        """Asigna turnos MOFIS en un día específico"""
        asignaciones = []
        
        # Obtener trabajadores elegibles disponibles
        elegibles = self._obtener_elegibles_disponibles(col_dia)
        cantidad_elegibles = len(elegibles)
        
        # Si no hay elegibles, no asignar nada
        if cantidad_elegibles == 0:
            return asignaciones
        
        # Obtener turnos a asignar según cantidad de elegibles
        turnos_a_asignar = self.TURNOS_POR_CANTIDAD.get(cantidad_elegibles, [])
        
        # Verificar que no existan ya estos turnos en el día
        turnos_disponibles = []
        for turno in turnos_a_asignar:
            if not self._existe_turno_en_dia(turno, col_dia):
                turnos_disponibles.append(turno)
        
        # Si no hay turnos disponibles, no asignar nada
        if not turnos_disponibles:
            return asignaciones
        
        # Asignar turnos priorizando equidad
        for turno in turnos_disponibles:
            # Seleccionar trabajador con menos turnos S+N
            trabajador = self._seleccionar_equitativo(elegibles)
            if not trabajador:
                break
            
            # Asignar turno
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            
            celda = self.ws.cell(row=fila, column=col_dia, value=turno)
            
            # Colorear celda de amarillo claro
            celda.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            # Actualizar contador si es turno S o N
            if self._es_turno_s_o_n(turno):
                self.contador_sn[trabajador] += 1
            
            asignaciones.append(f"{trabajador}: {turno}")
            
            # Remover trabajador de la lista de elegibles para este día
            elegibles.remove(trabajador)
        
        return asignaciones

    def procesar_todos_los_dias(self) -> None:
        """Procesa todos los días del mes asignando turnos MOFIS"""
        total_asignaciones = 0
        
        for col in range(2, self.ws.max_column + 1):
            asignaciones = self.asignar_turnos_en_dia(col)
            if asignaciones:
                total_asignaciones += len(asignaciones)
                print(f"Día {col}: {', '.join(asignaciones)}")
        
        print(f"\nTotal de asignaciones realizadas: {total_asignaciones}")
        
        # Mostrar estadísticas de equidad
        print("\nEstadísticas de equidad (turnos S+N por trabajador):")
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            print(f"  {trabajador}: {self.contador_sn[trabajador]}")
        
        self._actualizar_hoja_estadisticas()

        salida = "horarioUnificado_con_mofis.xlsx"
        try:
            self.wb.save(salida)
            print(f"\nArchivo guardado como: {salida}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"Archivo por defecto en uso. Guardado como: {alternativo}")

    def _actualizar_hoja_estadisticas(self) -> None:
        """Actualiza la hoja de estadísticas conservando todas las columnas del módulo de diurnas"""
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Verificar si ya existe la columna "3" (turnos 3)
        columna_3_existe = False
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header == "3":
                columna_3_existe = True
                break

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados (conservando todas las columnas del módulo de diurnas)
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")
        ws_stats.cell(row=1, column=4, value="6RT")
        ws_stats.cell(row=1, column=5, value="6T")
        
        # Mantener la columna "3" si existía y agregar las nuevas columnas
        if columna_3_existe:
            ws_stats.cell(row=1, column=6, value="3")
            ws_stats.cell(row=1, column=7, value="6S")
            ws_stats.cell(row=1, column=8, value="6N")
            ws_stats.cell(row=1, column=9, value="DIURNA")  # 6S + 6N
            ws_stats.cell(row=1, column=10, value="1D")     # Nueva columna 1D
            ws_stats.cell(row=1, column=11, value="3D")     # Nueva columna 3D
            ws_stats.cell(row=1, column=12, value="6D")     # Nueva columna 6D
            num_columnas = 12
        else:
            ws_stats.cell(row=1, column=6, value="6S")
            ws_stats.cell(row=1, column=7, value="6N")
            ws_stats.cell(row=1, column=8, value="DIURNA")  # 6S + 6N
            ws_stats.cell(row=1, column=9, value="1D")      # Nueva columna 1D
            ws_stats.cell(row=1, column=10, value="3D")     # Nueva columna 3D
            ws_stats.cell(row=1, column=11, value="6D")     # Nueva columna 6D
            num_columnas = 11

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, num_columnas + 1):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7 + 1
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"1")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = 6RT + 7 + 6R
            formula_6rt = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6R")'
            )
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 6T = 6TT + 6T
            formula_6t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6TT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6T")'
            )
            ws_stats.cell(row=fila_destino, column=5, value=formula_6t)

            # Columna "3" (turnos 3) si existía
            if columna_3_existe:
                formula_3 = f'=COUNTIF({hoja}!B{fila}:AE{fila},"3")'
                ws_stats.cell(row=fila_destino, column=6, value=formula_3)
                
                # 6S
                formula_6s = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                ws_stats.cell(row=fila_destino, column=7, value=formula_6s)

                # 6N = 6N + S + N + MCORTS + MCORTN (incluyendo turnos MOFIS)
                formula_6n = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"N")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"MCORTS")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"MCORTN")'
                )
                ws_stats.cell(row=fila_destino, column=8, value=formula_6n)

                # DIURNA = 6S + 6N
                formula_diurna = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                )
                ws_stats.cell(row=fila_destino, column=9, value=formula_diurna)
                
                # 1D = BANTD + BLPTD + 6*(6ND + 6SN + 6MTD) - con ponderación por horas
                formula_1d = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"BANTD")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"BLPTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6ND")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6SN")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6MTD")'
                )
                ws_stats.cell(row=fila_destino, column=10, value=formula_1d)

                # 3D = 3D
                formula_3d = f'=COUNTIF({hoja}!B{fila}:AE{fila},"3D")'
                ws_stats.cell(row=fila_destino, column=11, value=formula_3d)

                # 6D = 6*(NLPTD + NLPRD + NANTD + NANRD) - con ponderación por horas
                formula_6d = (
                    f'=6*COUNTIF({hoja}!B{fila}:AE{fila},"NLPTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NLPRD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NANTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NANRD")'
                )
                ws_stats.cell(row=fila_destino, column=12, value=formula_6d)
            else:
                # 6S
                formula_6s = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                ws_stats.cell(row=fila_destino, column=6, value=formula_6s)

                # 6N = 6N + S + N + MCORTS + MCORTN (incluyendo turnos MOFIS)
                formula_6n = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"N")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"MCORTS")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"MCORTN")'
                )
                ws_stats.cell(row=fila_destino, column=7, value=formula_6n)

                # DIURNA = 6S + 6N
                formula_diurna = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                )
                ws_stats.cell(row=fila_destino, column=8, value=formula_diurna)
                
                # 1D = BANTD + BLPTD + 6*(6ND + 6SN + 6MTD) - con ponderación por horas
                formula_1d = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"BANTD")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"BLPTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6ND")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6SN")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"6MTD")'
                )
                ws_stats.cell(row=fila_destino, column=9, value=formula_1d)

                # 3D = 3D
                formula_3d = f'=COUNTIF({hoja}!B{fila}:AE{fila},"3D")'
                ws_stats.cell(row=fila_destino, column=10, value=formula_3d)

                # 6D = 6*(NLPTD + NLPRD + NANTD + NANRD) - con ponderación por horas
                formula_6d = (
                    f'=6*COUNTIF({hoja}!B{fila}:AE{fila},"NLPTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NLPRD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NANTD")'
                    f'+6*COUNTIF({hoja}!B{fila}:AE{fila},"NANRD")'
                )
                ws_stats.cell(row=fila_destino, column=11, value=formula_6d)

            fila_destino += 1

        # Ajustar anchos de columna
        if columna_3_existe:
            widths = [('A', 10), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 8), ('I', 10), ('J', 8), ('K', 8), ('L', 8)]
        else:
            widths = [('A', 10), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 10), ('I', 8), ('J', 8), ('K', 8)]
        
        for col, width in widths:
            ws_stats.column_dimensions[col].width = width


if __name__ == "__main__":
    asignador = AsignadorTurnosMofis()
    asignador.procesar_todos_los_dias() 