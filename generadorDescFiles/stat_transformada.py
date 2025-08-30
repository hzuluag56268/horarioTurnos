import openpyxl
import os
from openpyxl.styles import PatternFill, Font
from typing import Optional, List, Dict


class StatTransformada:
    """
    Transforma la hoja de estadísticas en una nueva hoja llamada "stats".
    
    Características:
    - Primera columna: "SIGLA" con las siglas de cada trabajador
    - Columnas de turnos: Basadas en la columna 1T de la hoja de estadísticas
    - Relleno: Para cada trabajador, si tiene un valor N en 1T, se rellenan N celdas con "1"
    - Color: Celdas con "1" coloreadas de amarillo
    - Encabezado: "5AM" para el grupo de columnas de turnos
    - Número máximo de columnas: valor máximo en 1T + 2
    
    Archivo de entrada: horarioUnificado_con_mofis → horarioUnificado_con_diurnas → horarioUnificado_con_6t
    Archivo de salida: mismo nombre + "_stats"
    """

    COLOR_AMARILLO = "FFFF00"  # Amarillo por defecto de Excel

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        # Elegir el archivo de entrada según el orden de prioridad
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_mofis.xlsx",
            "horarioUnificado_con_diurnas.xlsx",
            "horarioUnificado_con_6t.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        
        if not elegido:
            raise FileNotFoundError("No se encontró ningún archivo de entrada válido")
        
        self.archivo_entrada = elegido
        print(f"📁 Archivo de entrada seleccionado: {self.archivo_entrada}")
        
        # Intentar cargar primero con data_only=True para obtener valores calculados
        print("🔄 Intentando cargar valores calculados de fórmulas dinámicas...")
        self.wb = openpyxl.load_workbook(self.archivo_entrada, data_only=True)
        
        # Verificar si los valores se cargaron correctamente
        valores_validos = self._verificar_carga_valores()
        
        if not valores_validos:
            print("⚠️  Los valores calculados no están disponibles. Intentando estrategia alternativa...")
            # Cargar sin data_only para trabajar con las fórmulas directamente
            self.wb = openpyxl.load_workbook(self.archivo_entrada, data_only=False)
            self._procesar_formulas_dinamicas()
        
        self._procesar_transformacion()

    def _verificar_carga_valores(self) -> bool:
        """Verifica que los valores se cargaron correctamente desde las fórmulas dinámicas"""
        if "Estadísticas" not in self.wb.sheetnames:
            print("⚠️  No se encontró la hoja 'Estadísticas'")
            return False
        
        ws_stats = self.wb["Estadísticas"]
        print("🔍 Verificando carga de valores desde fórmulas dinámicas...")
        
        # Buscar las columnas clave
        columnas_clave = ['1T', '6N', '6S', '3', '6T', '6RT']
        posiciones_columnas = {}
        
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header in columnas_clave:
                posiciones_columnas[header] = col
        
        # Verificar que los valores se cargaron correctamente
        valores_cargados = 0
        valores_faltantes = 0
        
        for columna, pos in posiciones_columnas.items():
            for fila in range(2, min(ws_stats.max_row + 1, 7)):  # Verificar primeras 5 filas
                valor = ws_stats.cell(row=fila, column=pos).value
                sigla = ws_stats.cell(row=fila, column=1).value
                if sigla:
                    if valor is not None and isinstance(valor, (int, float)) and valor != 0:
                        valores_cargados += 1
                    else:
                        valores_faltantes += 1
        
        print(f"📊 Resumen de verificación:")
        print(f"  ✅ Valores numéricos válidos: {valores_cargados}")
        print(f"  ⚠️  Valores faltantes o cero: {valores_faltantes}")
        
        # Considerar válido si tenemos al menos algunos valores numéricos
        es_valido = valores_cargados > 0
        
        if es_valido:
            print(f"✅ La carga con data_only=True fue exitosa. Los valores están disponibles.")
            print(f"✅ El archivo original '{self.archivo_entrada}' NO será modificado.")
            
            # Mostrar algunos ejemplos de valores cargados
            for columna, pos in list(posiciones_columnas.items())[:3]:  # Solo las primeras 3 columnas
                print(f"\n  📋 Ejemplos en columna {columna}:")
                ejemplos_mostrados = 0
                for fila in range(2, min(ws_stats.max_row + 1, 7)):
                    if ejemplos_mostrados >= 3:
                        break
                    valor = ws_stats.cell(row=fila, column=pos).value
                    sigla = ws_stats.cell(row=fila, column=1).value
                    if sigla and valor is not None:
                        print(f"    {sigla}: {valor}")
                        ejemplos_mostrados += 1
        else:
            print("⚠️  No se encontraron suficientes valores calculados válidos.")
        
        return es_valido

    def _procesar_formulas_dinamicas(self):
        """Procesa las fórmulas dinámicas cuando data_only=True no funciona"""
        if "Estadísticas" not in self.wb.sheetnames:
            print("⚠️  No se encontró la hoja 'Estadísticas'")
            return
        
        if "HorarioUnificado" not in self.wb.sheetnames:
            print("⚠️  No se encontró la hoja 'HorarioUnificado' necesaria para evaluar fórmulas")
            return
        
        ws_stats = self.wb["Estadísticas"]
        ws_horario = self.wb["HorarioUnificado"]
        
        print("🔄 Procesando fórmulas dinámicas manualmente...")
        
        # Buscar las columnas clave en la hoja de estadísticas
        columnas_clave = ['1T', '6N', '6S', '3', '6T', '6RT', '1D', '3D', '6D']
        posiciones_columnas = {}
        
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header in columnas_clave:
                posiciones_columnas[header] = col
        
        formulas_procesadas = 0
        
        # Procesar cada trabajador
        for fila in range(2, ws_stats.max_row + 1):
            sigla = ws_stats.cell(row=fila, column=1).value
            if not sigla:
                continue
            
            # Procesar cada columna clave
            for columna, pos_col in posiciones_columnas.items():
                celda = ws_stats.cell(row=fila, column=pos_col)
                
                if isinstance(celda.value, str) and celda.value.startswith('='):
                    # Es una fórmula, intentar evaluarla manualmente
                    try:
                        nuevo_valor = self._evaluar_formula_countif(celda.value, ws_horario, fila)
                        if nuevo_valor is not None:
                            celda.value = nuevo_valor
                            formulas_procesadas += 1
                            
                            if formulas_procesadas <= 10:  # Mostrar las primeras conversiones
                                print(f"  ✅ {sigla} - {columna}: {nuevo_valor}")
                    except Exception as e:
                        print(f"  ⚠️  Error procesando {sigla} - {columna}: {str(e)}")
        
        print(f"✅ Total de fórmulas procesadas manualmente: {formulas_procesadas}")
        print(f"✅ El archivo original '{self.archivo_entrada}' NO será modificado.")

    def _evaluar_formula_countif(self, formula: str, ws_horario, fila: int) -> Optional[int]:
        """Evalúa manualmente fórmulas COUNTIF simples"""
        try:
            # Las fórmulas típicas son como: =COUNTIF(HorarioUnificado!B2:AE2,"1T")+COUNTIF(HorarioUnificado!B2:AE2,"7")+COUNTIF(HorarioUnificado!B2:AE2,"1")
            
            # Extraer los patrones de búsqueda de la fórmula
            import re
            
            # Buscar todos los COUNTIF en la fórmula
            countif_matches = re.findall(r'COUNTIF\([^,]+,"([^"]+)"\)', formula)
            
            if not countif_matches:
                return None
            
            # Contar ocurrencias en la fila correspondiente del horario
            total = 0
            rango_inicio = 2  # Columna B
            rango_fin = 31    # Columna AE (aproximadamente)
            
            for patron in countif_matches:
                for col in range(rango_inicio, min(rango_fin + 1, ws_horario.max_column + 1)):
                    valor_celda = ws_horario.cell(row=fila, column=col).value
                    if valor_celda == patron:
                        total += 1
            
            return total
            
        except Exception as e:
            print(f"    ❌ Error evaluando fórmula: {str(e)}")
            return None

    def _mostrar_resumen_valores(self, ws_stats):
        """Muestra un resumen de los valores en las columnas clave de la hoja de Estadísticas"""
        print("\n📊 Resumen de valores en columnas clave:")
        
        # Buscar las columnas clave
        columnas_clave = ['1T', '6N', '6S', '3', '6T', '6RT']
        posiciones_columnas = {}
        
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header in columnas_clave:
                posiciones_columnas[header] = col
        
        # Mostrar valores de las primeras filas para cada columna clave
        for columna, pos in posiciones_columnas.items():
            print(f"\n  📋 Columna {columna} (columna {openpyxl.utils.get_column_letter(pos)}):")
            valores_mostrados = 0
            for fila in range(2, min(ws_stats.max_row + 1, 7)):  # Mostrar hasta 5 filas
                valor = ws_stats.cell(row=fila, column=pos).value
                sigla = ws_stats.cell(row=fila, column=1).value
                if sigla and valor is not None:
                    print(f"    {sigla}: {valor} (tipo: {type(valor).__name__})")
                    valores_mostrados += 1
                elif sigla:
                    print(f"    {sigla}: None o vacío")
                    valores_mostrados += 1
            
            if valores_mostrados == 0:
                print("    No se encontraron valores en las primeras filas")

    def _obtener_hoja_estadisticas(self):
        """Obtiene la hoja de estadísticas del archivo"""
        if "Estadísticas" in self.wb.sheetnames:
            return self.wb["Estadísticas"]
        else:
            raise ValueError("No se encontró la hoja 'Estadísticas' en el archivo")

    def _obtener_valores_maximos(self, ws_stats) -> tuple:
        """Obtiene los valores máximos en las columnas 1T, 6N, 6S, 3, 6T y 6RT para determinar el número de columnas"""
        # Buscar las columnas
        col_1t = None
        col_6n = None
        col_6s = None
        col_3 = None
        col_6t = None
        col_6rt = None
        col_1d = None
        col_3d = None
        col_6d = None
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header == "1T":
                col_1t = col
            elif header == "6N":
                col_6n = col
            elif header == "6S":
                col_6s = col
            elif header == "3":
                col_3 = col
            elif header == "6T":
                col_6t = col
            elif header == "6RT":
                col_6rt = col
            elif header == "1D":
                col_1d = col
            elif header == "3D":
                col_3d = col
            elif header == "6D":
                col_6d = col
        
        if col_1t is None:
            raise ValueError("No se encontró la columna '1T' en la hoja de estadísticas")
        # Advertir sobre columnas faltantes pero no fallar
        if col_6n is None:
            print("⚠️  Advertencia: No se encontró la columna '6N' en la hoja de estadísticas")
        if col_6s is None:
            print("⚠️  Advertencia: No se encontró la columna '6S' en la hoja de estadísticas")
        if col_3 is None:
            print("⚠️  Advertencia: No se encontró la columna '3' en la hoja de estadísticas")
        if col_6t is None:
            print("⚠️  Advertencia: No se encontró la columna '6T' en la hoja de estadísticas")
        if col_6rt is None:
            print("⚠️  Advertencia: No se encontró la columna '6RT' en la hoja de estadísticas")
        if col_1d is None:
            print("⚠️  Advertencia: No se encontró la columna '1D' en la hoja de estadísticas")
        if col_3d is None:
            print("⚠️  Advertencia: No se encontró la columna '3D' en la hoja de estadísticas")
        if col_6d is None:
            print("⚠️  Advertencia: No se encontró la columna '6D' en la hoja de estadísticas")
        
        # Encontrar los valores máximos
        max_valor_1t = 0
        max_valor_6n = 0
        max_valor_6s = 0
        max_valor_diurnas = 0
        max_valor_3 = 0
        max_valor_6t = 0
        max_valor_6rt = 0
        
        for fila in range(2, ws_stats.max_row + 1):
            # Valor 1T
            valor_1t = ws_stats.cell(row=fila, column=col_1t).value
            if valor_1t is not None:
                try:
                    valor_int = int(valor_1t)
                    max_valor_1t = max(max_valor_1t, valor_int)
                except (ValueError, TypeError):
                    continue
            
            # Valor 6N
            if col_6n is not None:
                valor_6n = ws_stats.cell(row=fila, column=col_6n).value
                if valor_6n is not None:
                    try:
                        valor_int = int(valor_6n)
                        max_valor_6n = max(max_valor_6n, valor_int)
                    except (ValueError, TypeError):
                        continue
            
            # Valor 6S
            if col_6s is not None:
                valor_6s = ws_stats.cell(row=fila, column=col_6s).value
                if valor_6s is not None:
                    try:
                        valor_int = int(valor_6s)
                        max_valor_6s = max(max_valor_6s, valor_int)
                    except (ValueError, TypeError):
                        continue
            
            # Valor 3
            if col_3 is not None:
                valor_3 = ws_stats.cell(row=fila, column=col_3).value
                if valor_3 is not None:
                    try:
                        valor_int = int(valor_3)
                        max_valor_3 = max(max_valor_3, valor_int)
                    except (ValueError, TypeError):
                        continue
            
            # Valor 6T
            if col_6t is not None:
                valor_6t = ws_stats.cell(row=fila, column=col_6t).value
                if valor_6t is not None:
                    try:
                        valor_int = int(valor_6t)
                        max_valor_6t = max(max_valor_6t, valor_int)
                    except (ValueError, TypeError):
                        continue
            
            # Valor 6RT
            if col_6rt is not None:
                valor_6rt = ws_stats.cell(row=fila, column=col_6rt).value
                if valor_6rt is not None:
                    try:
                        valor_int = int(valor_6rt)
                        max_valor_6rt = max(max_valor_6rt, valor_int)
                    except (ValueError, TypeError):
                        continue
        
        # Calcular el máximo total de DIURNAS (6N + 6S)
        max_valor_diurnas = max_valor_6n + max_valor_6s

        return max_valor_1t, max_valor_diurnas, max_valor_3, max_valor_6t, max_valor_6rt, col_1d, col_3d, col_6d

    def _crear_hoja_stats(self, ws_stats, max_1t: int, max_diurnas: int, max_3: int, max_6t: int, max_6rt: int, col_1d: int, col_3d: int, col_6d: int):
        """Crea la nueva hoja 'stats' con la transformación"""
        # Eliminar hoja 'stats' si ya existe
        if "stats" in self.wb.sheetnames:
            self.wb.remove(self.wb["stats"])
        
        # Crear nueva hoja
        ws_stats_nueva = self.wb.create_sheet("stats")
        
        # Número de columnas para cada grupo (máximo valor + 2, pero no más de 10)
        num_columnas_5am = min(max_1t + 2, 10)
        num_columnas_diurnas = min(max_diurnas + 2, 10)
        num_columnas_sln = min(max_3 + 2, 10)
        num_columnas_tant_nant = min(max_6t + 2, 10)
        num_columnas_mast_nanr = min(max_6rt + 2, 10)
        
        print(f"📊 Creando hoja 'stats' con {num_columnas_5am} columnas 5AM, {num_columnas_diurnas} columnas DIURNAS, {num_columnas_sln} columnas SLN, {num_columnas_tant_nant} columnas TANT/NANT y {num_columnas_mast_nanr} columnas MAST/NANR")
        if max_1t + 2 > 10:
            print(f"  ⚠️  Límite aplicado: 5AM reducido de {max_1t + 2} a 10 columnas")
        if max_diurnas + 2 > 10:
            print(f"  ⚠️  Límite aplicado: DIURNAS reducido de {max_diurnas + 2} a 10 columnas")
        if max_3 + 2 > 10:
            print(f"  ⚠️  Límite aplicado: SLN reducido de {max_3 + 2} a 10 columnas")
        if max_6t + 2 > 10:
            print(f"  ⚠️  Límite aplicado: TANT/NANT reducido de {max_6t + 2} a 10 columnas")
        if max_6rt + 2 > 10:
            print(f"  ⚠️  Límite aplicado: MAST/NANR reducido de {max_6rt + 2} a 10 columnas")
        
        # Configurar encabezados
        ws_stats_nueva.cell(row=1, column=1, value="SIGLA")
        
        # Combinar celdas para el encabezado 5AM
        if num_columnas_5am > 0:
            ws_stats_nueva.merge_cells(start_row=1, start_column=2, end_row=1, end_column=num_columnas_5am + 1)
            ws_stats_nueva.cell(row=1, column=2, value="5AM")
        
        # Combinar celdas para el encabezado DIURNAS (después de 5AM)
        if num_columnas_diurnas > 0:
            ws_stats_nueva.merge_cells(start_row=1, start_column=num_columnas_5am + 2, end_row=1, end_column=num_columnas_5am + num_columnas_diurnas + 1)
            ws_stats_nueva.cell(row=1, column=num_columnas_5am + 2, value="DIURNAS")
        
        # Agregar encabezado para la columna SumaD (después de DIURNAS)
        col_suma_d = num_columnas_5am + num_columnas_diurnas + 2
        ws_stats_nueva.cell(row=1, column=col_suma_d, value="SumaD")
        
        # Agregar encabezado para la columna SumaN (después de SumaD)
        col_suma_n = num_columnas_5am + num_columnas_diurnas + 3
        ws_stats_nueva.cell(row=1, column=col_suma_n, value="SumaN")
        
        # Agregar encabezado para la columna SumTot (después de SumaN)
        col_sumtot = num_columnas_5am + num_columnas_diurnas + 4
        ws_stats_nueva.cell(row=1, column=col_sumtot, value="SumTot")
        
        # Combinar celdas para el encabezado SLN (después de SumTot)
        if num_columnas_sln > 0:
            ws_stats_nueva.merge_cells(start_row=1, start_column=num_columnas_5am + num_columnas_diurnas + 5, end_row=1, end_column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + 5, value="SLN")
        
        # Combinar celdas para el encabezado TANT/NANT (después de SLN)
        if num_columnas_tant_nant > 0:
            ws_stats_nueva.merge_cells(start_row=1, start_column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5, end_row=1, end_column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5, value="TANT/NANT")
        
        # Combinar celdas para el encabezado MAST/NANR (después de TANT/NANT)
        if num_columnas_mast_nanr > 0:
            ws_stats_nueva.merge_cells(start_row=1, start_column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5, end_row=1, end_column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 4)
            ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5, value="MAST/NANR")
        
        # Agregar encabezados para las nuevas columnas (después de MAST/NANR)
        col_diurf = num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 5
        ws_stats_nueva.cell(row=1, column=col_diurf, value="DiurF")
        
        col_nocfes = num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 6
        ws_stats_nueva.cell(row=1, column=col_nocfes, value="NocFes")
        
        col_difho = num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 7
        ws_stats_nueva.cell(row=1, column=col_difho, value="DifHo")
        
        # Aplicar formato a encabezados
        from openpyxl.styles import Alignment
        
        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Formato para SIGLA
        celda_sigla = ws_stats_nueva.cell(row=1, column=1)
        celda_sigla.fill = header_fill
        celda_sigla.font = header_font
        celda_sigla.alignment = header_alignment
        
        # Formato para 5AM (celda combinada) - con color amarillo
        if num_columnas_5am > 0:
            celda_5am = ws_stats_nueva.cell(row=1, column=2)
            celda_5am.fill = PatternFill(start_color=self.COLOR_AMARILLO, 
                                       end_color=self.COLOR_AMARILLO, 
                                       fill_type="solid")
            celda_5am.font = header_font
            celda_5am.alignment = header_alignment
        
        # Formato para DIURNAS (celda combinada) - con color verde claro
        if num_columnas_diurnas > 0:
            celda_diurnas = ws_stats_nueva.cell(row=1, column=num_columnas_5am + 2)
            celda_diurnas.fill = PatternFill(start_color="90EE90", 
                                           end_color="90EE90", 
                                           fill_type="solid")  # Verde claro
            celda_diurnas.font = header_font
            celda_diurnas.alignment = header_alignment
        
        # Formato para SLN (celda combinada) - con color morado claro
        if num_columnas_sln > 0:
            celda_sln = ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + 5)
            celda_sln.fill = PatternFill(start_color="DDA0DD", 
                                       end_color="DDA0DD", 
                                       fill_type="solid")  # Morado claro
            celda_sln.font = header_font
            celda_sln.alignment = header_alignment
        
        # Formato para TANT/NANT (celda combinada) - con color naranja claro
        if num_columnas_tant_nant > 0:
            celda_tant_nant = ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5)
            celda_tant_nant.fill = PatternFill(start_color="FFB6C1", 
                                             end_color="FFB6C1", 
                                             fill_type="solid")  # Naranja claro
            celda_tant_nant.font = header_font
            celda_tant_nant.alignment = header_alignment
        
        # Formato para MAST/NANR (celda combinada) - con color gris claro
        if num_columnas_mast_nanr > 0:
            celda_mast_nanr = ws_stats_nueva.cell(row=1, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5)
            celda_mast_nanr.fill = PatternFill(start_color="D3D3D3", 
                                             end_color="D3D3D3", 
                                             fill_type="solid")  # Gris claro
            celda_mast_nanr.font = header_font
            celda_mast_nanr.alignment = header_alignment
        
        # Formato para SumaD - con color azul claro
        celda_suma_d = ws_stats_nueva.cell(row=1, column=col_suma_d)
        celda_suma_d.fill = PatternFill(start_color="87CEEB", 
                                       end_color="87CEEB", 
                                       fill_type="solid")  # Azul claro
        celda_suma_d.font = header_font
        celda_suma_d.alignment = header_alignment
        
        # Formato para SumaN - con color verde claro
        celda_suma_n = ws_stats_nueva.cell(row=1, column=col_suma_n)
        celda_suma_n.fill = PatternFill(start_color="98FB98", 
                                       end_color="98FB98", 
                                       fill_type="solid")  # Verde claro
        celda_suma_n.font = header_font
        celda_suma_n.alignment = header_alignment
        
        # Formato para SumTot - con color dorado
        celda_sumtot = ws_stats_nueva.cell(row=1, column=col_sumtot)
        celda_sumtot.fill = PatternFill(start_color="FFD700", 
                                       end_color="FFD700", 
                                       fill_type="solid")  # Dorado
        celda_sumtot.font = header_font
        celda_sumtot.alignment = header_alignment
        
        # Formato para DiurF - con color rosa claro
        celda_diurf = ws_stats_nueva.cell(row=1, column=col_diurf)
        celda_diurf.fill = PatternFill(start_color="FFC0CB", 
                                      end_color="FFC0CB", 
                                      fill_type="solid")  # Rosa claro
        celda_diurf.font = header_font
        celda_diurf.alignment = header_alignment
        
        # Formato para NocFes - con color violeta claro
        celda_nocfes = ws_stats_nueva.cell(row=1, column=col_nocfes)
        celda_nocfes.fill = PatternFill(start_color="E6E6FA", 
                                       end_color="E6E6FA", 
                                       fill_type="solid")  # Violeta claro
        celda_nocfes.font = header_font
        celda_nocfes.alignment = header_alignment
        
        # Formato para DifHo - con color marrón claro
        celda_difho = ws_stats_nueva.cell(row=1, column=col_difho)
        celda_difho.fill = PatternFill(start_color="DEB887", 
                                      end_color="DEB887", 
                                      fill_type="solid")  # Marrón claro
        celda_difho.font = header_font
        celda_difho.alignment = header_alignment
        
        # Buscar la columna 1T en la hoja de estadísticas original
        col_1t = None
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header == "1T":
                col_1t = col
                break
        
        if col_1t is None:
            raise ValueError("No se encontró la columna '1T' en la hoja de estadísticas")
        
        # Buscar las columnas 6N, 6S, 3, 6T y 6RT en la hoja de estadísticas original
        col_6n = None
        col_6s = None
        col_3 = None
        col_6t = None
        col_6rt = None
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header == "6N":
                col_6n = col
            elif header == "6S":
                col_6s = col
            elif header == "3":
                col_3 = col
            elif header == "6T":
                col_6t = col
            elif header == "6RT":
                col_6rt = col
        
        # Advertir sobre columnas faltantes pero no fallar
        if col_6n is None:
            print("⚠️  Advertencia: No se encontró la columna '6N' en la hoja de estadísticas")
        if col_6s is None:
            print("⚠️  Advertencia: No se encontró la columna '6S' en la hoja de estadísticas")
        if col_3 is None:
            print("⚠️  Advertencia: No se encontró la columna '3' en la hoja de estadísticas")
        if col_6t is None:
            print("⚠️  Advertencia: No se encontró la columna '6T' en la hoja de estadísticas")
        if col_6rt is None:
            print("⚠️  Advertencia: No se encontró la columna '6RT' en la hoja de estadísticas")
        if col_1d is None:
            print("⚠️  Advertencia: No se encontró la columna '1D' en la hoja de estadísticas")
        if col_3d is None:
            print("⚠️  Advertencia: No se encontró la columna '3D' en la hoja de estadísticas")
        if col_6d is None:
            print("⚠️  Advertencia: No se encontró la columna '6D' en la hoja de estadísticas")
        
        # Procesar cada trabajador
        fila_destino = 2
        trabajadores_procesados = 0
        fila_gce = None  # Para recordar la fila donde está GCE
        
        for fila in range(2, ws_stats.max_row + 1):
            sigla = ws_stats.cell(row=fila, column=1).value
            if not sigla:
                continue
            
            # Recordar la posición de GCE
            if sigla == "GCE":
                fila_gce = fila_destino
            
            # Obtener valores de 1T, 6N, 6S, 3, 6T y 6RT
            valor_1t = ws_stats.cell(row=fila, column=col_1t).value
            valor_6n = ws_stats.cell(row=fila, column=col_6n).value if col_6n is not None else None
            valor_6s = ws_stats.cell(row=fila, column=col_6s).value if col_6s is not None else None
            valor_3 = ws_stats.cell(row=fila, column=col_3).value if col_3 is not None else None
            valor_6t = ws_stats.cell(row=fila, column=col_6t).value if col_6t is not None else None
            valor_6rt = ws_stats.cell(row=fila, column=col_6rt).value if col_6rt is not None else None
            
            # Obtener valores de las nuevas columnas (copiar exactamente)
            valor_1d = ws_stats.cell(row=fila, column=col_1d).value if col_1d is not None else None
            valor_3d = ws_stats.cell(row=fila, column=col_3d).value if col_3d is not None else None
            valor_6d = ws_stats.cell(row=fila, column=col_6d).value if col_6d is not None else None
            
            # Escribir sigla
            ws_stats_nueva.cell(row=fila_destino, column=1, value=sigla)
            
            # Procesar valor de 1T (5AM)
            if valor_1t is not None and valor_1t != 0:
                try:
                    valor_int = int(valor_1t)
                    if valor_int > 0:
                        # Rellenar celdas de 5AM con color amarillo
                        for col in range(2, num_columnas_5am + 2):
                            celda = ws_stats_nueva.cell(row=fila_destino, column=col)
                            celda.fill = PatternFill(start_color=self.COLOR_AMARILLO, 
                                                   end_color=self.COLOR_AMARILLO, 
                                                   fill_type="solid")
                            # Solo poner "1" en las primeras celdas según el valor de 1T
                            if col < valor_int + 2:
                                celda.value = 1
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 1T ({valor_1t})")
            
            # Procesar valores de 6N y 6S (DIURNAS) - sumar ambos valores
            valor_total_diurnas = 0
            if valor_6n is not None and valor_6n != 0:
                try:
                    valor_int_6n = int(valor_6n)
                    valor_total_diurnas += valor_int_6n
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 6N ({valor_6n})")
            
            if valor_6s is not None and valor_6s != 0:
                try:
                    valor_int_6s = int(valor_6s)
                    valor_total_diurnas += valor_int_6s
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 6S ({valor_6s})")
            
            # Aplicar el valor total de DIURNAS (6N + 6S)
            if valor_total_diurnas > 0:
                # Rellenar celdas de DIURNAS con color verde claro
                for col in range(num_columnas_5am + 2, num_columnas_5am + num_columnas_diurnas + 2):
                    celda = ws_stats_nueva.cell(row=fila_destino, column=col)
                    celda.fill = PatternFill(start_color="90EE90", 
                                           end_color="90EE90", 
                                           fill_type="solid")  # Verde claro
                    # Solo poner "6" en las primeras celdas según el valor total de DIURNAS
                    if col < num_columnas_5am + 2 + valor_total_diurnas:
                        celda.value = 6
            
            # Procesar valor de 3 (SLN)
            if valor_3 is not None and valor_3 != 0:
                try:
                    valor_int_3 = int(valor_3)
                    if valor_int_3 > 0:
                        # Rellenar celdas de SLN con color morado claro
                        for col in range(num_columnas_5am + num_columnas_diurnas + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5):
                            celda = ws_stats_nueva.cell(row=fila_destino, column=col)
                            celda.fill = PatternFill(start_color="DDA0DD", 
                                                   end_color="DDA0DD", 
                                                   fill_type="solid")  # Morado claro
                            # Solo poner "3" en las primeras celdas según el valor de 3
                            if col < num_columnas_5am + num_columnas_diurnas + 5 + valor_int_3:
                                celda.value = 3
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 3 ({valor_3})")
            
            # Procesar valor de 6T (TANT/NANT)
            if valor_6t is not None and valor_6t != 0:
                try:
                    valor_int_6t = int(valor_6t)
                    if valor_int_6t > 0:
                        # Rellenar celdas de TANT/NANT con color naranja claro
                        for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5):
                            celda = ws_stats_nueva.cell(row=fila_destino, column=col)
                            celda.fill = PatternFill(start_color="FFB6C1", 
                                                   end_color="FFB6C1", 
                                                   fill_type="solid")  # Naranja claro
                            # Solo poner "6" en las primeras celdas según el valor de 6T
                            if col < num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5 + valor_int_6t:
                                celda.value = 6
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 6T ({valor_6t})")
            
            # Procesar valor de 6RT (MAST/NANR)
            if valor_6rt is not None and valor_6rt != 0:
                try:
                    valor_int_6rt = int(valor_6rt)
                    if valor_int_6rt > 0:
                        # Rellenar celdas de MAST/NANR con color gris claro
                        for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 5):
                            celda = ws_stats_nueva.cell(row=fila_destino, column=col)
                            celda.fill = PatternFill(start_color="D3D3D3", 
                                                   end_color="D3D3D3", 
                                                   fill_type="solid")  # Gris claro
                            # Solo poner "6" en las primeras celdas según el valor de 6RT
                            if col < num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5 + valor_int_6rt:
                                celda.value = 6
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sigla}: valor no numérico en 6RT ({valor_6rt})")
            
            # Copiar valores exactos de las nuevas columnas (sin modificación)
            # DiurF (columna 1D)
            if col_1d is not None:
                ws_stats_nueva.cell(row=fila_destino, column=col_diurf, value=valor_1d)
                # Aplicar color de fondo rosa claro
                celda_diurf = ws_stats_nueva.cell(row=fila_destino, column=col_diurf)
                celda_diurf.fill = PatternFill(start_color="FFC0CB", 
                                              end_color="FFC0CB", 
                                              fill_type="solid")  # Rosa claro
            
            # NocFes (columna 3D)
            if col_3d is not None:
                ws_stats_nueva.cell(row=fila_destino, column=col_nocfes, value=valor_3d)
                # Aplicar color de fondo violeta claro
                celda_nocfes = ws_stats_nueva.cell(row=fila_destino, column=col_nocfes)
                celda_nocfes.fill = PatternFill(start_color="E6E6FA", 
                                               end_color="E6E6FA", 
                                               fill_type="solid")  # Violeta claro
            
            # DifHo (columna 6D)
            if col_6d is not None:
                ws_stats_nueva.cell(row=fila_destino, column=col_difho, value=valor_6d)
                # Aplicar color de fondo marrón claro
                celda_difho = ws_stats_nueva.cell(row=fila_destino, column=col_difho)
                celda_difho.fill = PatternFill(start_color="DEB887", 
                                              end_color="DEB887", 
                                              fill_type="solid")  # Marrón claro
            
            # Crear fórmula dinámica para SumaD: sumar columnas 5AM + columnas DIURNAS
            # La fórmula será: =SUM(B{fila_destino}:H{fila_destino}) + SUM(I{fila_destino}:R{fila_destino})
            primera_col_5am = openpyxl.utils.get_column_letter(2)
            ultima_col_5am = openpyxl.utils.get_column_letter(num_columnas_5am + 1)
            primera_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + 2)
            ultima_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 1)
            
            formula_suma_d = f"=SUM({primera_col_5am}{fila_destino}:{ultima_col_5am}{fila_destino})+SUM({primera_col_diurnas}{fila_destino}:{ultima_col_diurnas}{fila_destino})"
            
            # Escribir la fórmula en la columna SumaD
            celda_suma_d = ws_stats_nueva.cell(row=fila_destino, column=col_suma_d)
            celda_suma_d.value = formula_suma_d
            celda_suma_d.fill = PatternFill(start_color="87CEEB", 
                                           end_color="87CEEB", 
                                           fill_type="solid")  # Azul claro
            celda_suma_d.font = Font(bold=True)
            
            # Crear fórmula dinámica para SumaN: sumar columnas SLN + TANT/NANT + MAST/NANR
            # La fórmula será: =SUM(U{fila_destino}:Y{fila_destino}) + SUM(Z{fila_destino}:AF{fila_destino}) + SUM(AG{fila_destino}:AM{fila_destino})
            primera_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 5)
            ultima_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            primera_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5)
            ultima_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            primera_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5)
            ultima_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 4)
            
            formula_suma_n = f"=SUM({primera_col_sln}{fila_destino}:{ultima_col_sln}{fila_destino})+SUM({primera_col_tant_nant}{fila_destino}:{ultima_col_tant_nant}{fila_destino})+SUM({primera_col_mast_nanr}{fila_destino}:{ultima_col_mast_nanr}{fila_destino})"
            
            # Escribir la fórmula en la columna SumaN
            celda_suma_n = ws_stats_nueva.cell(row=fila_destino, column=col_suma_n)
            celda_suma_n.value = formula_suma_n
            celda_suma_n.fill = PatternFill(start_color="98FB98", 
                                           end_color="98FB98", 
                                           fill_type="solid")  # Verde claro
            celda_suma_n.font = Font(bold=True)
            
            # Crear fórmula dinámica para SumTot: sumar SumaD + SumaN + DiurF + NocFes
            col_suma_d_letter = openpyxl.utils.get_column_letter(col_suma_d)
            col_suma_n_letter = openpyxl.utils.get_column_letter(col_suma_n)
            col_diurf_letter = openpyxl.utils.get_column_letter(col_diurf)
            col_nocfes_letter = openpyxl.utils.get_column_letter(col_nocfes)
            
            formula_sum_tot = f"={col_suma_d_letter}{fila_destino}+{col_suma_n_letter}{fila_destino}+{col_diurf_letter}{fila_destino}+{col_nocfes_letter}{fila_destino}"
            
            # Escribir la fórmula en la columna SumTot
            celda_sum_tot = ws_stats_nueva.cell(row=fila_destino, column=col_sumtot)
            celda_sum_tot.value = formula_sum_tot
            celda_sum_tot.fill = PatternFill(start_color="FFD700", 
                                           end_color="FFD700", 
                                           fill_type="solid")  # Dorado
            celda_sum_tot.font = Font(bold=True)
            
            # Contar trabajadores procesados (si tiene al menos un turno en 1T, 6N, 6S, 3, 6T o 6RT)
            tiene_turnos = False
            try:
                if (valor_1t is not None and int(valor_1t) > 0) or valor_total_diurnas > 0 or (valor_3 is not None and int(valor_3) > 0) or (valor_6t is not None and int(valor_6t) > 0) or (valor_6rt is not None and int(valor_6rt) > 0):
                    tiene_turnos = True
            except (ValueError, TypeError):
                pass
            
            if tiene_turnos:
                trabajadores_procesados += 1
                print(f"  ✅ {sigla}: 1T={valor_1t}, 6N={valor_6n}, 6S={valor_6s}, Total DIURNAS={valor_total_diurnas}, 3={valor_3}, 6T={valor_6t}, 6RT={valor_6rt}, DiurF={valor_1d}, NocFes={valor_3d}, DifHo={valor_6d}")
            
            fila_destino += 1
        
        # Añadir fila de sumatoria debajo de GCE si se encontró
        if fila_gce is not None:
            # Insertar nueva fila debajo de GCE
            ws_stats_nueva.insert_rows(fila_gce + 1)
            
            # Añadir nombre "PARCI" en la primera columna de la fila de sumatoria parcial (GCE)
            celda_parci_gce = ws_stats_nueva.cell(row=fila_gce + 1, column=1)
            celda_parci_gce.value = "PARCI"
            celda_parci_gce.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
            celda_parci_gce.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo 5AM
            for col in range(2, num_columnas_5am + 2):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde arriba hasta GCE
                # La fórmula será: =SUM(B2:B{fila_gce}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria = f"=SUM({col_letter}2:{col_letter}{fila_gce})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria = ws_stats_nueva.cell(row=fila_gce + 1, column=col)
                celda_sumatoria.value = formula_sumatoria
                
                # Aplicar formato especial a la celda de sumatoria
                celda_sumatoria.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
                celda_sumatoria.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo DIURNAS
            for col in range(num_columnas_5am + 2, num_columnas_5am + num_columnas_diurnas + 2):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde arriba hasta GCE
                # La fórmula será: =SUM(N2:N{fila_gce}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria = f"=SUM({col_letter}2:{col_letter}{fila_gce})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria = ws_stats_nueva.cell(row=fila_gce + 1, column=col)
                celda_sumatoria.value = formula_sumatoria
                
                # Aplicar formato especial a la celda de sumatoria
                celda_sumatoria.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
                celda_sumatoria.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo SLN
            for col in range(num_columnas_5am + num_columnas_diurnas + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde arriba hasta GCE
                # La fórmula será: =SUM(T2:T{fila_gce}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria = f"=SUM({col_letter}2:{col_letter}{fila_gce})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria = ws_stats_nueva.cell(row=fila_gce + 1, column=col)
                celda_sumatoria.value = formula_sumatoria
                
                # Aplicar formato especial a la celda de sumatoria
                celda_sumatoria.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
                celda_sumatoria.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo TANT/NANT
            for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde arriba hasta GCE
                # La fórmula será: =SUM(Y2:Y{fila_gce}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria = f"=SUM({col_letter}2:{col_letter}{fila_gce})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria = ws_stats_nueva.cell(row=fila_gce + 1, column=col)
                celda_sumatoria.value = formula_sumatoria
                
                # Aplicar formato especial a la celda de sumatoria
                celda_sumatoria.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
                celda_sumatoria.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo MAST/NANR
            for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 5):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde arriba hasta GCE
                # La fórmula será: =SUM(AF2:AF{fila_gce}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria = f"=SUM({col_letter}2:{col_letter}{fila_gce})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria = ws_stats_nueva.cell(row=fila_gce + 1, column=col)
                celda_sumatoria.value = formula_sumatoria
                
                # Aplicar formato especial a la celda de sumatoria
                celda_sumatoria.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
                celda_sumatoria.font = Font(bold=True)
            
            # Aplicar fórmula de sumatoria para la columna SumaD
            col_suma_d_letter = openpyxl.utils.get_column_letter(col_suma_d)
            formula_sumatoria_suma_d = f"=SUM({col_suma_d_letter}2:{col_suma_d_letter}{fila_gce})"
            
            # Escribir la fórmula en la nueva fila, columna SumaD
            celda_sumatoria_suma_d = ws_stats_nueva.cell(row=fila_gce + 1, column=col_suma_d)
            celda_sumatoria_suma_d.value = formula_sumatoria_suma_d
            
            # Aplicar formato especial a la celda de sumatoria SumaD
            celda_sumatoria_suma_d.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
            celda_sumatoria_suma_d.font = Font(bold=True)
            
            # Aplicar fórmula de sumatoria para la columna SumaN
            col_suma_n_letter = openpyxl.utils.get_column_letter(col_suma_n)
            formula_sumatoria_suma_n = f"=SUM({col_suma_n_letter}2:{col_suma_n_letter}{fila_gce})"
            
            # Escribir la fórmula en la nueva fila, columna SumaN
            celda_sumatoria_suma_n = ws_stats_nueva.cell(row=fila_gce + 1, column=col_suma_n)
            celda_sumatoria_suma_n.value = formula_sumatoria_suma_n
            
            # Aplicar formato especial a la celda de sumatoria SumaN
            celda_sumatoria_suma_n.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
            celda_sumatoria_suma_n.font = Font(bold=True)
            
            # Añadir segunda fila con sumatoria total de todas las columnas
            ws_stats_nueva.insert_rows(fila_gce + 2)
            
            # Añadir nombre "TOTAL" en la primera columna de la fila de sumatoria total (GCE)
            celda_total_gce = ws_stats_nueva.cell(row=fila_gce + 2, column=1)
            celda_total_gce.value = "TOTAL"
            celda_total_gce.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_total_gce.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Crear fórmula para sumar solo las columnas de 5AM (fila_gce + 1)
            # La fórmula será: =SUM(B{fila_gce+1}:{última_col_5am}{fila_gce+1})
            primera_col_5am = openpyxl.utils.get_column_letter(2)
            ultima_col_5am = openpyxl.utils.get_column_letter(num_columnas_5am + 1)
            formula_sumatoria_total_5am = f"=SUM({primera_col_5am}{fila_gce+1}:{ultima_col_5am}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total 5AM en la segunda columna (B) de la nueva fila
            celda_sumatoria_total_5am = ws_stats_nueva.cell(row=fila_gce + 2, column=2)
            celda_sumatoria_total_5am.value = formula_sumatoria_total_5am
            
            # Aplicar formato especial a la celda de sumatoria total 5AM
            celda_sumatoria_total_5am.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_5am.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Crear fórmula para sumar solo las columnas de DIURNAS (fila_gce + 1)
            # La fórmula será: =SUM({primera_col_diurnas}{fila_gce+1}:{última_col_diurnas}{fila_gce+1})
            primera_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + 2)
            ultima_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 1)
            formula_sumatoria_total_diurnas = f"=SUM({primera_col_diurnas}{fila_gce+1}:{ultima_col_diurnas}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total DIURNAS en la columna correspondiente
            celda_sumatoria_total_diurnas = ws_stats_nueva.cell(row=fila_gce + 2, column=num_columnas_5am + 2)
            celda_sumatoria_total_diurnas.value = formula_sumatoria_total_diurnas
            
            # Aplicar formato especial a la celda de sumatoria total DIURNAS
            celda_sumatoria_total_diurnas.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_diurnas.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Crear fórmula para sumar solo las columnas de SLN (fila_gce + 1)
            # La fórmula será: =SUM({primera_col_sln}{fila_gce+1}:{última_col_sln}{fila_gce+1})
            primera_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 4)
            ultima_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 3)
            formula_sumatoria_total_sln = f"=SUM({primera_col_sln}{fila_gce+1}:{ultima_col_sln}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total SLN en la columna correspondiente
            celda_sumatoria_total_sln = ws_stats_nueva.cell(row=fila_gce + 2, column=num_columnas_5am + num_columnas_diurnas + 4)
            celda_sumatoria_total_sln.value = formula_sumatoria_total_sln
            
            # Aplicar formato especial a la celda de sumatoria total SLN
            celda_sumatoria_total_sln.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_sln.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Crear fórmula para sumar solo las columnas de TANT/NANT (fila_gce + 1)
            # La fórmula será: =SUM({primera_col_tant_nant}{fila_gce+1}:{última_col_tant_nant}{fila_gce+1})
            primera_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            ultima_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 3)
            formula_sumatoria_total_tant_nant = f"=SUM({primera_col_tant_nant}{fila_gce+1}:{ultima_col_tant_nant}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total TANT/NANT en la columna correspondiente
            celda_sumatoria_total_tant_nant = ws_stats_nueva.cell(row=fila_gce + 2, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            celda_sumatoria_total_tant_nant.value = formula_sumatoria_total_tant_nant
            
            # Aplicar formato especial a la celda de sumatoria total TANT/NANT
            celda_sumatoria_total_tant_nant.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_tant_nant.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Crear fórmula para sumar solo las columnas de MAST/NANR (fila_gce + 1)
            # La fórmula será: =SUM({primera_col_mast_nanr}{fila_gce+1}:{última_col_mast_nanr}{fila_gce+1})
            primera_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            ultima_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 3)
            formula_sumatoria_total_mast_nanr = f"=SUM({primera_col_mast_nanr}{fila_gce+1}:{ultima_col_mast_nanr}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total MAST/NANR en la columna correspondiente
            celda_sumatoria_total_mast_nanr = ws_stats_nueva.cell(row=fila_gce + 2, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            celda_sumatoria_total_mast_nanr.value = formula_sumatoria_total_mast_nanr
            
            # Aplicar formato especial a la celda de sumatoria total MAST/NANR
            celda_sumatoria_total_mast_nanr.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_mast_nanr.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Aplicar fórmula de sumatoria total para la columna SumaD
            formula_sumatoria_total_suma_d = f"=SUM({col_suma_d_letter}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total SumaD en la columna correspondiente
            celda_sumatoria_total_suma_d = ws_stats_nueva.cell(row=fila_gce + 2, column=col_suma_d)
            celda_sumatoria_total_suma_d.value = formula_sumatoria_total_suma_d
            
            # Aplicar formato especial a la celda de sumatoria total SumaD
            celda_sumatoria_total_suma_d.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_suma_d.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Aplicar fórmula de sumatoria total para la columna SumaN
            formula_sumatoria_total_suma_n = f"=SUM({col_suma_n_letter}{fila_gce+1})"
            
            # Escribir la fórmula de sumatoria total SumaN en la columna correspondiente
            celda_sumatoria_total_suma_n = ws_stats_nueva.cell(row=fila_gce + 2, column=col_suma_n)
            celda_sumatoria_total_suma_n.value = formula_sumatoria_total_suma_n
            
            # Aplicar formato especial a la celda de sumatoria total SumaN
            celda_sumatoria_total_suma_n.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rojo claro
            celda_sumatoria_total_suma_n.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            
            # Calcular valor inicial de la primera columna para mostrar en el mensaje
            sumatoria_inicial_5am = 0
            sumatoria_inicial_diurnas = 0
            for fila in range(2, fila_gce + 1):
                valor_celda_5am = ws_stats_nueva.cell(row=fila, column=2).value
                valor_celda_diurnas = ws_stats_nueva.cell(row=fila, column=num_columnas_5am + 2).value
                if valor_celda_5am == 1:
                    sumatoria_inicial_5am += 1
                if valor_celda_diurnas == 6:
                    sumatoria_inicial_diurnas += 1
            
            print(f"  📊 Sumatorias dinámicas añadidas debajo de GCE para todos los grupos")
            print(f"     5AM - Primera columna: {sumatoria_inicial_5am} turnos totales")
            print(f"     DIURNAS - Primera columna: {sumatoria_inicial_diurnas} turnos totales")
            print(f"     Fórmulas aplicadas: SUM() para {num_columnas_5am} columnas 5AM, {num_columnas_diurnas} columnas DIURNAS, {num_columnas_sln} columnas SLN, {num_columnas_tant_nant} columnas TANT/NANT y 1 columna SumaD")
            print(f"     Sumatoria total 5AM: {formula_sumatoria_total_5am}")
            print(f"     Sumatoria total DIURNAS: {formula_sumatoria_total_diurnas}")
            print(f"     Sumatoria total SLN: {formula_sumatoria_total_sln}")
            print(f"     Sumatoria total TANT/NANT: {formula_sumatoria_total_tant_nant}")
            print(f"     Sumatoria total MAST/NANR: {formula_sumatoria_total_mast_nanr}")
            print(f"     Sumatoria total SumaD: {formula_sumatoria_total_suma_d}")
            print(f"     Sumatoria total SumaN: {formula_sumatoria_total_suma_n}")
            print(f"     Fórmula SumaD: SUM(5AM) + SUM(DIURNAS)")
            print(f"     Fórmula SumaN: SUM(SLN) + SUM(TANT/NANT) + SUM(MAST/NANR)")
        else:
            print("  ⚠️  No se encontró GCE en la lista de trabajadores")
        
        # Buscar la fila donde está YIS (después de GCE) y JMV (última sigla)
        fila_yis = None
        fila_jmv = None
        
        # Buscar YIS y JMV en la hoja original
        for fila in range(2, ws_stats.max_row + 1):
            sigla = ws_stats.cell(row=fila, column=1).value
            if sigla == "YIS":
                # Encontrar la fila correspondiente en la nueva hoja (después de insertar filas de GCE)
                fila_yis_original = fila
                # Calcular la nueva posición considerando las filas insertadas debajo de GCE
                if fila_gce is not None and fila_yis_original > fila_gce:
                    fila_yis = fila_yis_original + 2  # +2 porque se insertaron 2 filas debajo de GCE
                else:
                    fila_yis = fila_yis_original
                break
        
        # Buscar JMV en la hoja original y calcular su nueva posición
        for fila in range(2, ws_stats.max_row + 1):
            sigla = ws_stats.cell(row=fila, column=1).value
            if sigla == "JMV":
                fila_jmv_original = fila
                # Calcular la nueva posición considerando las filas insertadas debajo de GCE
                if fila_gce is not None and fila_jmv_original > fila_gce:
                    fila_jmv = fila_jmv_original + 2  # +2 porque se insertaron 2 filas debajo de GCE
                else:
                    fila_jmv = fila_jmv_original
                break
        
        if fila_yis is not None and fila_jmv is not None:
            # Insertar primera fila de sumatoria justo después de JMV
            ws_stats_nueva.insert_rows(fila_jmv + 1)
            
            # Añadir nombre "PARCI" en la primera columna de la fila de sumatoria parcial
            celda_parci = ws_stats_nueva.cell(row=fila_jmv + 1, column=1)
            celda_parci.value = "PARCI"
            celda_parci.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
            celda_parci.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo 5AM
            # Sumatoria desde YIS hasta JMV
            for col in range(2, num_columnas_5am + 2):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde YIS hasta JMV
                # La fórmula será: =SUM(B{fila_yis}:B{fila_jmv}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria_final = f"=SUM({col_letter}{fila_yis}:{col_letter}{fila_jmv})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria_final = ws_stats_nueva.cell(row=fila_jmv + 1, column=col)
                celda_sumatoria_final.value = formula_sumatoria_final
                
                # Aplicar formato especial a la celda de sumatoria (verde)
                celda_sumatoria_final.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                celda_sumatoria_final.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo DIURNAS
            # Sumatoria desde YIS hasta JMV
            for col in range(num_columnas_5am + 2, num_columnas_5am + num_columnas_diurnas + 2):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde YIS hasta JMV
                # La fórmula será: =SUM(N{fila_yis}:N{fila_jmv}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria_final = f"=SUM({col_letter}{fila_yis}:{col_letter}{fila_jmv})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria_final = ws_stats_nueva.cell(row=fila_jmv + 1, column=col)
                celda_sumatoria_final.value = formula_sumatoria_final
                
                # Aplicar formato especial a la celda de sumatoria (verde)
                celda_sumatoria_final.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                celda_sumatoria_final.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo SLN
            # Sumatoria desde YIS hasta JMV
            for col in range(num_columnas_5am + num_columnas_diurnas + 4, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde YIS hasta JMV
                # La fórmula será: =SUM(T{fila_yis}:T{fila_jmv}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria_final = f"=SUM({col_letter}{fila_yis}:{col_letter}{fila_jmv})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria_final = ws_stats_nueva.cell(row=fila_jmv + 1, column=col)
                celda_sumatoria_final.value = formula_sumatoria_final
                
                # Aplicar formato especial a la celda de sumatoria (verde)
                celda_sumatoria_final.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                celda_sumatoria_final.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo TANT/NANT
            # Sumatoria desde YIS hasta JMV
            for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde YIS hasta JMV
                # La fórmula será: =SUM(Y{fila_yis}:Y{fila_jmv}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria_final = f"=SUM({col_letter}{fila_yis}:{col_letter}{fila_jmv})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria_final = ws_stats_nueva.cell(row=fila_jmv + 1, column=col)
                celda_sumatoria_final.value = formula_sumatoria_final
                
                # Aplicar formato especial a la celda de sumatoria (verde)
                celda_sumatoria_final.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                celda_sumatoria_final.font = Font(bold=True)
            
            # Aplicar fórmulas de sumatoria para todas las columnas bajo MAST/NANR
            # Sumatoria desde YIS hasta JMV
            for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 4):
                # Crear fórmula dinámica para sumar todos los valores en cada columna desde YIS hasta JMV
                # La fórmula será: =SUM(AF{fila_yis}:AF{fila_jmv}) para cada columna
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_sumatoria_final = f"=SUM({col_letter}{fila_yis}:{col_letter}{fila_jmv})"
                
                # Escribir la fórmula en la nueva fila, columna correspondiente
                celda_sumatoria_final = ws_stats_nueva.cell(row=fila_jmv + 1, column=col)
                celda_sumatoria_final.value = formula_sumatoria_final
                
                # Aplicar formato especial a la celda de sumatoria (verde)
                celda_sumatoria_final.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
                celda_sumatoria_final.font = Font(bold=True)
            
            # Aplicar fórmula de sumatoria para la columna SumaD (desde YIS hasta JMV)
            formula_sumatoria_final_suma_d = f"=SUM({col_suma_d_letter}{fila_yis}:{col_suma_d_letter}{fila_jmv})"
            
            # Escribir la fórmula en la nueva fila, columna SumaD
            celda_sumatoria_final_suma_d = ws_stats_nueva.cell(row=fila_jmv + 1, column=col_suma_d)
            celda_sumatoria_final_suma_d.value = formula_sumatoria_final_suma_d
            
            # Aplicar formato especial a la celda de sumatoria SumaD (verde)
            celda_sumatoria_final_suma_d.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
            celda_sumatoria_final_suma_d.font = Font(bold=True)
            
            # Aplicar fórmula de sumatoria para la columna SumaN (desde YIS hasta JMV)
            formula_sumatoria_final_suma_n = f"=SUM({col_suma_n_letter}{fila_yis}:{col_suma_n_letter}{fila_jmv})"
            
            # Escribir la fórmula en la nueva fila, columna SumaN
            celda_sumatoria_final_suma_n = ws_stats_nueva.cell(row=fila_jmv + 1, column=col_suma_n)
            celda_sumatoria_final_suma_n.value = formula_sumatoria_final_suma_n
            
            # Aplicar formato especial a la celda de sumatoria SumaN (verde)
            celda_sumatoria_final_suma_n.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
            celda_sumatoria_final_suma_n.font = Font(bold=True)
            
            # Insertar segunda fila con sumatoria total justo después de PARCI
            ws_stats_nueva.insert_rows(fila_jmv + 2)
            
            # Añadir nombre "TOTAL" en la primera columna de la fila de sumatoria total
            celda_total = ws_stats_nueva.cell(row=fila_jmv + 2, column=1)
            celda_total.value = "TOTAL"
            celda_total.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_total.font = Font(bold=True, color="000000")  # Texto negro
            
            # Crear fórmula para sumar solo las columnas de 5AM (fila_jmv + 1)
            # La fórmula será: =SUM(B{fila_jmv+1}:{última_col_5am}{fila_jmv+1})
            primera_col_5am = openpyxl.utils.get_column_letter(2)
            ultima_col_5am = openpyxl.utils.get_column_letter(num_columnas_5am + 1)
            formula_sumatoria_total_final_5am = f"=SUM({primera_col_5am}{fila_jmv+1}:{ultima_col_5am}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total 5AM en la segunda columna (B) de la nueva fila
            celda_sumatoria_total_final_5am = ws_stats_nueva.cell(row=fila_jmv + 2, column=2)
            celda_sumatoria_total_final_5am.value = formula_sumatoria_total_final_5am
            
            # Aplicar formato especial a la celda de sumatoria total 5AM final (azul)
            celda_sumatoria_total_final_5am.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_5am.font = Font(bold=True, color="000000")  # Texto negro
            
            # Crear fórmula para sumar solo las columnas de DIURNAS (fila_jmv + 1)
            # La fórmula será: =SUM({primera_col_diurnas}{fila_jmv+1}:{última_col_diurnas}{fila_jmv+1})
            primera_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + 2)
            ultima_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 1)
            formula_sumatoria_total_final_diurnas = f"=SUM({primera_col_diurnas}{fila_jmv+1}:{ultima_col_diurnas}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total DIURNAS en la columna correspondiente
            celda_sumatoria_total_final_diurnas = ws_stats_nueva.cell(row=fila_jmv + 2, column=num_columnas_5am + 2)
            celda_sumatoria_total_final_diurnas.value = formula_sumatoria_total_final_diurnas
            
            # Aplicar formato especial a la celda de sumatoria total DIURNAS final (azul)
            celda_sumatoria_total_final_diurnas.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_diurnas.font = Font(bold=True, color="000000")  # Texto negro
            
            # Crear fórmula para sumar solo las columnas de SLN (fila_jmv + 1)
            # La fórmula será: =SUM({primera_col_sln}{fila_jmv+1}:{última_col_sln}{fila_jmv+1})
            primera_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 4)
            ultima_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 3)
            formula_sumatoria_total_final_sln = f"=SUM({primera_col_sln}{fila_jmv+1}:{ultima_col_sln}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total SLN final en la columna correspondiente
            celda_sumatoria_total_final_sln = ws_stats_nueva.cell(row=fila_jmv + 2, column=num_columnas_5am + num_columnas_diurnas + 4)
            celda_sumatoria_total_final_sln.value = formula_sumatoria_total_final_sln
            
            # Aplicar formato especial a la celda de sumatoria total SLN final (azul)
            celda_sumatoria_total_final_sln.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_sln.font = Font(bold=True, color="000000")  # Texto negro
            
            # Crear fórmula para sumar solo las columnas de TANT/NANT (fila_jmv + 1)
            # La fórmula será: =SUM({primera_col_tant_nant}{fila_jmv+1}:{última_col_tant_nant}{fila_jmv+1})
            primera_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            ultima_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 3)
            formula_sumatoria_total_final_tant_nant = f"=SUM({primera_col_tant_nant}{fila_jmv+1}:{ultima_col_tant_nant}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total TANT/NANT final en la columna correspondiente
            celda_sumatoria_total_final_tant_nant = ws_stats_nueva.cell(row=fila_jmv + 2, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
            celda_sumatoria_total_final_tant_nant.value = formula_sumatoria_total_final_tant_nant
            
            # Aplicar formato especial a la celda de sumatoria total TANT/NANT final (azul)
            celda_sumatoria_total_final_tant_nant.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_tant_nant.font = Font(bold=True, color="000000")  # Texto negro
            
            # Crear fórmula para sumar solo las columnas de MAST/NANR (fila_jmv + 1)
            # La fórmula será: =SUM({primera_col_mast_nanr}{fila_jmv+1}:{última_col_mast_nanr}{fila_jmv+1})
            primera_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            ultima_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 3)
            formula_sumatoria_total_final_mast_nanr = f"=SUM({primera_col_mast_nanr}{fila_jmv+1}:{ultima_col_mast_nanr}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total MAST/NANR final en la columna correspondiente
            celda_sumatoria_total_final_mast_nanr = ws_stats_nueva.cell(row=fila_jmv + 2, column=num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
            celda_sumatoria_total_final_mast_nanr.value = formula_sumatoria_total_final_mast_nanr
            
            # Aplicar formato especial a la celda de sumatoria total MAST/NANR final (azul)
            celda_sumatoria_total_final_mast_nanr.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_mast_nanr.font = Font(bold=True, color="000000")  # Texto negro
            
            # Aplicar fórmula de sumatoria total para la columna SumaD final
            formula_sumatoria_total_final_suma_d = f"=SUM({col_suma_d_letter}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total SumaD final en la columna correspondiente
            celda_sumatoria_total_final_suma_d = ws_stats_nueva.cell(row=fila_jmv + 2, column=col_suma_d)
            celda_sumatoria_total_final_suma_d.value = formula_sumatoria_total_final_suma_d
            
            # Aplicar formato especial a la celda de sumatoria total SumaD final (azul)
            celda_sumatoria_total_final_suma_d.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_suma_d.font = Font(bold=True, color="000000")  # Texto negro
            
            # Aplicar fórmula de sumatoria total para la columna SumaN final
            formula_sumatoria_total_final_suma_n = f"=SUM({col_suma_n_letter}{fila_jmv+1})"
            
            # Escribir la fórmula de sumatoria total SumaN final en la columna correspondiente
            celda_sumatoria_total_final_suma_n = ws_stats_nueva.cell(row=fila_jmv + 2, column=col_suma_n)
            celda_sumatoria_total_final_suma_n.value = formula_sumatoria_total_final_suma_n
            
            # Aplicar formato especial a la celda de sumatoria total SumaN final (azul)
            celda_sumatoria_total_final_suma_n.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul claro
            celda_sumatoria_total_final_suma_n.font = Font(bold=True, color="000000")  # Texto negro
            
            print(f"  📊 Sumatorias añadidas justo después de JMV para todos los grupos")
            print(f"     Rango: desde YIS (fila {fila_yis}) hasta JMV (fila {fila_jmv})")
            print(f"     Fórmulas aplicadas: SUM() para {num_columnas_5am} columnas 5AM, {num_columnas_diurnas} columnas DIURNAS, {num_columnas_sln} columnas SLN, {num_columnas_tant_nant} columnas TANT/NANT y 1 columna SumaD (verde)")
            print(f"     Sumatoria total 5AM: {formula_sumatoria_total_final_5am} (azul)")
            print(f"     Sumatoria total DIURNAS: {formula_sumatoria_total_final_diurnas} (azul)")
            print(f"     Sumatoria total SLN: {formula_sumatoria_total_final_sln} (azul)")
            print(f"     Sumatoria total TANT/NANT: {formula_sumatoria_total_final_tant_nant} (azul)")
            print(f"     Sumatoria total MAST/NANR: {formula_sumatoria_total_final_mast_nanr} (azul)")
            print(f"     Sumatoria total SumaD: {formula_sumatoria_total_final_suma_d} (azul)")
            print(f"     Sumatoria total SumaN: {formula_sumatoria_total_final_suma_n} (azul)")
        else:
            print("  ⚠️  No se encontró YIS o JMV en la lista de trabajadores")
        
        # Añadir fila de verificación al final
        # Calcular la última fila después de todas las inserciones
        ultima_fila_final = ws_stats_nueva.max_row + 1
        
        # Insertar fila de verificación al final
        ws_stats_nueva.insert_rows(ultima_fila_final)
        
        # Añadir nombre "VERIFICACIÓN" en la primera celda (A)
        celda_verificacion_nombre = ws_stats_nueva.cell(row=ultima_fila_final, column=1)
        celda_verificacion_nombre.value = "VERIFICACIÓN"
        celda_verificacion_nombre.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
        celda_verificacion_nombre.font = Font(bold=True, color="000000")  # Texto negro
        
        # Añadir las constantes en celdas separadas para poder modificarlas
        # Constante 1 en columna C
        celda_constante1 = ws_stats_nueva.cell(row=ultima_fila_final, column=3)
        celda_constante1.value = 1
        celda_constante1.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
        celda_constante1.font = Font(bold=True)
        
        # Constante 52 en columna D
        celda_constante52 = ws_stats_nueva.cell(row=ultima_fila_final, column=4)
        celda_constante52.value = 52
        celda_constante52.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
        celda_constante52.font = Font(bold=True)
        
        # Crear fórmula de verificación: (Total1 + Total2) / constante1 - constante52
        # Total1 = sumatoria total debajo de GCE (fila_gce + 2, columna B)
        # Total2 = sumatoria total después de JMV (fila_jmv + 2, columna B)
        if fila_gce is not None and fila_jmv is not None:
            formula_verificacion = f"=({openpyxl.utils.get_column_letter(2)}{fila_gce+2}+{openpyxl.utils.get_column_letter(2)}{fila_jmv+2})/{openpyxl.utils.get_column_letter(3)}{ultima_fila_final}-{openpyxl.utils.get_column_letter(4)}{ultima_fila_final}"
            
            # Escribir la fórmula en la segunda celda (B)
            celda_verificacion_formula = ws_stats_nueva.cell(row=ultima_fila_final, column=2)
            celda_verificacion_formula.value = formula_verificacion
            celda_verificacion_formula.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
            celda_verificacion_formula.font = Font(bold=True, color="000000")  # Texto negro
            
            print(f"  🔍 Fila de verificación añadida al final")
            print(f"     Fórmula: {formula_verificacion}")
            print(f"     Operación: (Total GCE + Total JMV) / constante1 - constante52")
            print(f"     Constantes: 1 (columna C), 52 (columna D) - modificables")
        else:
            print("  ⚠️  No se pudo crear la fórmula de verificación (faltan GCE o JMV)")
        
        # Añadir fórmula de verificación DIURNAS en la fila VERIFICACIÓN existente
        # Crear fórmula de verificación DIURNAS: (Total 5AM + Total DIURNAS) / constante6 - constante104
        # Total 5AM = sumatoria total debajo de GCE (fila_gce + 2, columna B)
        # Total DIURNAS = sumatoria total debajo de GCE (fila_gce + 2, columna I)
        if fila_gce is not None:
            formula_verificacion_diurnas = f"=({openpyxl.utils.get_column_letter(2)}{fila_gce+2}+{openpyxl.utils.get_column_letter(num_columnas_5am+2)}{fila_gce+2})/{openpyxl.utils.get_column_letter(3)}{ultima_fila_final}-{openpyxl.utils.get_column_letter(4)}{ultima_fila_final}"
            
            # Escribir la fórmula en la primera columna de DIURNAS (columna I)
            celda_verificacion_diurnas_formula = ws_stats_nueva.cell(row=ultima_fila_final, column=num_columnas_5am + 2)
            celda_verificacion_diurnas_formula.value = formula_verificacion_diurnas
            celda_verificacion_diurnas_formula.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja
            celda_verificacion_diurnas_formula.font = Font(bold=True, color="000000")  # Texto negro
            
            print(f"  🔍 Fórmula de verificación DIURNAS añadida en la fila VERIFICACIÓN existente")
            print(f"     Fórmula: {formula_verificacion_diurnas}")
            print(f"     Operación: (Total 5AM + Total DIURNAS) / constante6 - constante104")
            print(f"     Posición: Columna {openpyxl.utils.get_column_letter(num_columnas_5am + 2)} (primera columna DIURNAS)")
        else:
            print("  ⚠️  No se pudo crear la fórmula de verificación DIURNAS (falta GCE)")
        
        # Ajustar anchos de columna dinámicamente
        ws_stats_nueva.column_dimensions['A'].width = 6  # SIGLA - ancho reducido para 3-4 caracteres
        
        # Para las columnas de turnos 5AM, usar ancho basado en el valor máximo
        for col in range(2, num_columnas_5am + 2):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_stats_nueva.column_dimensions[col_letter].width = 3  # Ancho reducido para el valor "1"
        
        # Para las columnas de turnos DIURNAS, usar ancho basado en el valor máximo
        for col in range(num_columnas_5am + 2, num_columnas_5am + num_columnas_diurnas + 2):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_stats_nueva.column_dimensions[col_letter].width = 3  # Ancho reducido para el valor "6"
        
        # Para las columnas de turnos SLN, usar ancho basado en el valor máximo
        for col in range(num_columnas_5am + num_columnas_diurnas + 3, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 3):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_stats_nueva.column_dimensions[col_letter].width = 3  # Ancho reducido para el valor "3"
        
        # Para las columnas de turnos TANT/NANT, usar ancho basado en el valor máximo
        for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 3, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 3):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_stats_nueva.column_dimensions[col_letter].width = 3  # Ancho reducido para el valor "6"
        
        # Para las columnas de turnos MAST/NANR, usar ancho basado en el valor máximo
        for col in range(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 3, num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 3):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws_stats_nueva.column_dimensions[col_letter].width = 3  # Ancho reducido para el valor "6"
        
        # Para la columna SumaD, usar ancho apropiado para números
        ws_stats_nueva.column_dimensions[openpyxl.utils.get_column_letter(col_suma_d)].width = 5  # Ancho reducido para números de 2-3 dígitos
        
        # Para la columna SumaN, usar ancho apropiado para números
        ws_stats_nueva.column_dimensions[openpyxl.utils.get_column_letter(col_suma_n)].width = 5  # Ancho reducido para números de 2-3 dígitos
        
        # Para la columna SumTot, usar ancho apropiado para números
        ws_stats_nueva.column_dimensions[openpyxl.utils.get_column_letter(col_sumtot)].width = 6  # Ancho reducido para números de 2-3 dígitos
        
        # Actualizar todas las fórmulas SumaD, SumaN y SumTot después de insertar las filas de sumatoria
        self._actualizar_formulas_suma(ws_stats_nueva, col_suma_d, col_suma_n, col_sumtot, col_diurf, col_nocfes, num_columnas_5am, num_columnas_diurnas, num_columnas_sln, num_columnas_tant_nant, num_columnas_mast_nanr)
        
        print(f"📈 Total de trabajadores procesados: {trabajadores_procesados}")

    def _actualizar_formulas_suma(self, ws_stats_nueva, col_suma_d: int, col_suma_n: int, col_sumtot: int, col_diurf: int, col_nocfes: int, num_columnas_5am: int, num_columnas_diurnas: int, num_columnas_sln: int, num_columnas_tant_nant: int, num_columnas_mast_nanr: int):
        """Actualiza todas las fórmulas SumaD, SumaN y SumTot para que referencien correctamente sus propias filas"""
        print("🔄 Actualizando fórmulas SumaD, SumaN y SumTot para corregir referencias de filas...")
        
        formulas_actualizadas = 0
        
        # Recorrer todas las filas y actualizar las fórmulas
        for fila in range(2, ws_stats_nueva.max_row + 1):
            sigla = ws_stats_nueva.cell(row=fila, column=1).value
            
            # Saltar filas de sumatoria (PARCI, TOTAL, VERIFICACIÓN)
            if sigla in ["PARCI", "TOTAL", "VERIFICACIÓN"]:
                continue
            
            # Actualizar fórmula SumaD para esta fila específica
            celda_suma_d = ws_stats_nueva.cell(row=fila, column=col_suma_d)
            if celda_suma_d.value and isinstance(celda_suma_d.value, str) and celda_suma_d.value.startswith('='):
                # Crear nueva fórmula con la fila correcta
                primera_col_5am = openpyxl.utils.get_column_letter(2)
                ultima_col_5am = openpyxl.utils.get_column_letter(num_columnas_5am + 1)
                primera_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + 2)
                ultima_col_diurnas = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 1)
                
                nueva_formula_suma_d = f"=SUM({primera_col_5am}{fila}:{ultima_col_5am}{fila})+SUM({primera_col_diurnas}{fila}:{ultima_col_diurnas}{fila})"
                celda_suma_d.value = nueva_formula_suma_d
                formulas_actualizadas += 1
                
                if formulas_actualizadas <= 5:  # Mostrar solo las primeras actualizaciones
                    print(f"  ✅ {sigla} (fila {fila}): SumaD actualizada → {nueva_formula_suma_d}")
            
            # Actualizar fórmula SumaN para esta fila específica
            celda_suma_n = ws_stats_nueva.cell(row=fila, column=col_suma_n)
            if celda_suma_n.value and isinstance(celda_suma_n.value, str) and celda_suma_n.value.startswith('='):
                # Crear nueva fórmula con la fila correcta
                primera_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + 5)
                ultima_col_sln = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 4)
                primera_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + 5)
                ultima_col_tant_nant = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 4)
                primera_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + 5)
                ultima_col_mast_nanr = openpyxl.utils.get_column_letter(num_columnas_5am + num_columnas_diurnas + num_columnas_sln + num_columnas_tant_nant + num_columnas_mast_nanr + 4)
                
                nueva_formula_suma_n = f"=SUM({primera_col_sln}{fila}:{ultima_col_sln}{fila})+SUM({primera_col_tant_nant}{fila}:{ultima_col_tant_nant}{fila})+SUM({primera_col_mast_nanr}{fila}:{ultima_col_mast_nanr}{fila})"
                celda_suma_n.value = nueva_formula_suma_n
                
                if formulas_actualizadas <= 5:  # Mostrar solo las primeras actualizaciones
                    print(f"  ✅ {sigla} (fila {fila}): SumaN actualizada → {nueva_formula_suma_n}")
            
            # Actualizar fórmula SumTot para esta fila específica
            celda_sum_tot = ws_stats_nueva.cell(row=fila, column=col_sumtot)
            if celda_sum_tot.value and isinstance(celda_sum_tot.value, str) and celda_sum_tot.value.startswith('='):
                # Crear nueva fórmula con la fila correcta
                col_suma_d_letter = openpyxl.utils.get_column_letter(col_suma_d)
                col_suma_n_letter = openpyxl.utils.get_column_letter(col_suma_n)
                col_diurf_letter = openpyxl.utils.get_column_letter(col_diurf)
                col_nocfes_letter = openpyxl.utils.get_column_letter(col_nocfes)
                
                nueva_formula_sum_tot = f"={col_suma_d_letter}{fila}+{col_suma_n_letter}{fila}+{col_diurf_letter}{fila}+{col_nocfes_letter}{fila}"
                celda_sum_tot.value = nueva_formula_sum_tot
                
                if formulas_actualizadas <= 5:  # Mostrar solo las primeras actualizaciones
                    print(f"  ✅ {sigla} (fila {fila}): SumTot actualizada → {nueva_formula_sum_tot}")
        
        print(f"✅ Total de fórmulas SumaD, SumaN y SumTot actualizadas: {formulas_actualizadas}")

    def _procesar_transformacion(self):
        """Procesa la transformación completa"""
        print("🔄 Iniciando transformación de estadísticas...")
        
        # Obtener hoja de estadísticas
        ws_stats = self._obtener_hoja_estadisticas()
        
        # Obtener valores máximos de 1T, 6N, 6S, 3, 6T y 6RT
        max_1t, max_diurnas, max_3, max_6t, max_6rt, col_1d, col_3d, col_6d = self._obtener_valores_maximos(ws_stats)
        print(f"📊 Valor máximo en columna 1T: {max_1t}")
        print(f"📊 Valor máximo total DIURNAS (6N + 6S): {max_diurnas}")
        print(f"📊 Valor máximo en columna 3: {max_3}")
        print(f"📊 Valor máximo en columna 6T: {max_6t}")
        print(f"📊 Valor máximo en columna 6RT: {max_6rt}")
        print(f"📊 Columnas encontradas: 1D={col_1d}, 3D={col_3d}, 6D={col_6d}")
        
        # Crear nueva hoja stats
        self._crear_hoja_stats(ws_stats, max_1t, max_diurnas, max_3, max_6t, max_6rt, col_1d, col_3d, col_6d)
        
        # Generar nombre del archivo de salida
        base_name = os.path.splitext(self.archivo_entrada)[0]
        archivo_salida = f"{base_name}_stats.xlsx"
        
        # Guardar archivo de salida (sin modificar el original)
        try:
            self.wb.save(archivo_salida)
            print(f"✅ Archivo guardado como: {archivo_salida}")
            print(f"✅ Archivo original '{self.archivo_entrada}' NO fue modificado")
        except PermissionError:
            import random
            base, ext = os.path.splitext(archivo_salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"✅ Archivo por defecto en uso. Guardado como: {alternativo}")
            print(f"✅ Archivo original '{self.archivo_entrada}' NO fue modificado")

    def generar_reporte(self):
        """Genera un reporte de la transformación realizada"""
        print("\n" + "="*60)
        print("REPORTE DE TRANSFORMACIÓN DE ESTADÍSTICAS")
        print("="*60)
        
        ws_stats = self._obtener_hoja_estadisticas()
        ws_stats_nueva = self.wb["stats"]
        
        print(f"Archivo de entrada: {self.archivo_entrada}")
        print(f"Archivo de salida: {os.path.splitext(self.archivo_entrada)[0]}_stats.xlsx")
        print(f"Trabajadores en estadísticas original: {ws_stats.max_row - 1}")
        print(f"Trabajadores en hoja stats: {ws_stats_nueva.max_row - 1}")
        print(f"Columnas de turnos creadas: {ws_stats_nueva.max_column - 1}")
        
        # Contar trabajadores con turnos asignados
        trabajadores_con_turnos = 0
        for fila in range(2, ws_stats_nueva.max_row + 1):
            for col in range(2, ws_stats_nueva.max_column + 1):
                if ws_stats_nueva.cell(row=fila, column=col).value == 1:
                    trabajadores_con_turnos += 1
                    break
        
        print(f"Trabajadores con turnos asignados: {trabajadores_con_turnos}")
        print("="*60)


if __name__ == "__main__":
    try:
        transformador = StatTransformada()
        transformador.generar_reporte()
    except Exception as e:
        print(f"❌ Error: {e}") 