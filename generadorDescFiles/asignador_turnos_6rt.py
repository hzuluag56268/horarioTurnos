import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Tuple, Set
import os


class AsignadorTurnos6RT:
    """
    Asigna turno "6RT" (6 horas adicionales) con estas reglas:
    - Decisión por día según personal disponible/turnos operativos (fila de conteo):
      * ≤9   → asignar turno "6TT" (paridad propia de 6TT)
      * 10-15 → asignar turno "6RT"
      * ≥16  → NO asignar
    - Solo se asigna entre trabajadores elegibles: ['YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']
    - Verificar que en ese día NO exista ya un "6RT" ni "7" (no duplicar)
    - Prioridades basadas en el día siguiente (mañana):
      1) DESC/TROP/SIND mañana y NO tener 1T/T1/1/7 mañana
      2) NO tener 1T/T1/1/7 mañana
      3) Resto
    - Equidad para 6RT: contar "6RT" + "7" juntos
    - 6TT: paridad solo entre 6TT; no asignar si ya hay otro 6TT en el día; preferir quien NO tenga 1T/T1/1/7 mañana (restricción blanda)
    - No modificar celdas con turnos preexistentes (respeta asignaciones originales)
    - Actualiza hoja "Estadísticas" con columnas: SIGLA, DESC, 1T (1T+7), 6RT (6RT+7)
    """

    TRABAJADORES_ELEGIBLES = ['YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']
    TRABAJADORES_RESPALDO = ['FCE', 'JBV', 'HZG']

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        # Elegir el archivo de entrada más reciente disponible
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_1t.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            # fallback duro
            elegido = "horarioUnificado_procesado.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()

        # Snapshot del estado original para no tocar asignaciones preexistentes
        self.original_nonempty: Set[Tuple[int, int]] = set()
        self.original_6rt: Set[Tuple[int, int]] = set()
        self.original_7: Set[Tuple[int, int]] = set()
        self._snapshot_estado_original()

        # Contadores
        self.contador_grupo_6rt: Dict[str, int] = defaultdict(int)  # 6RT + 7
        self.contador_6tt: Dict[str, int] = defaultdict(int)        # solo 6TT
        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _snapshot_estado_original(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None or str(valor).strip() == "":
                    continue
                val = str(valor).strip().upper()
                self.original_nonempty.add((fila, col))
                if val == "6RT":
                    self.original_6rt.add((fila, col))
                    # Colorear celdas 6RT existentes de morado claro
                    celda = self.ws.cell(row=fila, column=col)
                    celda.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                elif val == "7":
                    self.original_7.add((fila, col))

    def _es_celda_originalmente_vacia(self, fila: int, col: int) -> bool:
        return (fila, col) not in self.original_nonempty

    def _es_celda_original_6rt(self, fila: int, col: int) -> bool:
        return (fila, col) in self.original_6rt

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _tiene_prioridad_manana(self, trabajador: str, col_dia: int) -> bool:
        """True si mañana tiene DESC, TROP o SIND."""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"DESC", "TROP", "SIND"}

    def _tiene_extra_manana(self, trabajador: str, col_dia: int) -> bool:
        """True si mañana tiene 1T/T1/1 o 7."""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        val = str(valor).strip().upper()
        return val in {"1T", "T1", "1", "7"}

    def _obtener_trabajadores_disponibles(self, col_dia: int, pool: Optional[List[str]] = None) -> List[str]:
        candidatos = pool if pool is not None else self.TRABAJADORES_ELEGIBLES
        disponibles = []
        for trabajador in candidatos:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if (
                (celda.value is None or str(celda.value).strip() == "")
                and self._es_celda_originalmente_vacia(fila, col_dia)
            ):
                disponibles.append(trabajador)
        return disponibles

    # Nuevo: obtener conteo de turnos operativos exclusivamente desde la fila con etiqueta
    def _obtener_conteo_operativos(self, col_dia: int) -> Optional[int]:
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TURNOS OPERATIVOS":
                valor = self.ws.cell(row=fila, column=col_dia).value
                try:
                    return int(valor)
                except Exception:
                    return None
        # Si no se encuentra la etiqueta, no devolver conteo
        return None

    def _determinar_asignacion_por_personal(self, col_dia: int) -> bool:
        """Devuelve True si se debe asignar 6RT (conteo operativos entre 10 y 15 inclusive)."""
        disponible = self._obtener_conteo_operativos(col_dia)
        if disponible is None:
            return False
        return 10 <= disponible <= 15

    def _existe_6rt_o_7_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            val = str(valor).strip().upper()
            if val in {"6RT", "7"}:
                return True
        return False

    # Nuevo: detectar si hay "7" en el día (columna)
    def _existe_7_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            if str(valor).strip().upper() == "7":
                return True
        return False

    # Nuevo: detectar si hay "6TT" en el día (columna)
    def _existe_6tt_en_dia(self, col_dia: int) -> bool:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            if str(valor).strip().upper() == "6TT":
                return True
        return False

    def _seleccionar_equitativo(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        min_val = min(self.contador_grupo_6rt[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_grupo_6rt[c] == min_val]
        return random.choice(empatados)

    def _seleccionar_equitativo_6tt(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        min_val = min(self.contador_6tt[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_6tt[c] == min_val]
        return random.choice(empatados)

    def _actualizar_contadores(self, trabajador: str, turno: str) -> None:
        if turno == "6RT":
            self.contador_grupo_6rt[trabajador] += 1
        elif turno == "7":
            self.contador_grupo_6rt[trabajador] += 1
        elif turno == "6TT":
            self.contador_6tt[trabajador] += 1

    def _inicializar_contadores_desde_hoja(self) -> None:
        max_col = self.ws.max_column
        for trabajador in self.TRABAJADORES_ELEGIBLES + self.TRABAJADORES_RESPALDO:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val in {"6RT", "7"}:
                    self.contador_grupo_6rt[trabajador] += 1
                elif val == "6TT":
                    self.contador_6tt[trabajador] += 1

    # Nuevo: re-balanceo para asegurar paridad ±1 en 6RT+7 entre elegibles
    def _rebalancear_para_paridad(self) -> None:
        while True:
            # Construir conteos actuales solo para quienes existen en la hoja
            conteos_actuales = {}
            for t in self.TRABAJADORES_ELEGIBLES:
                if self._obtener_fila_trabajador(t):
                    conteos_actuales[t] = self.contador_grupo_6rt[t]
            if not conteos_actuales:
                break

            trabajador_max = max(conteos_actuales, key=conteos_actuales.get)
            trabajador_min = min(conteos_actuales, key=conteos_actuales.get)
            if conteos_actuales[trabajador_max] - conteos_actuales[trabajador_min] <= 1:
                break

            fila_max = self._obtener_fila_trabajador(trabajador_max)
            fila_min = self._obtener_fila_trabajador(trabajador_min)
            if not fila_max or not fila_min:
                break

            columna_candidata = None
            for col in range(2, self.ws.max_column + 1):
                valor_max = self.ws.cell(row=fila_max, column=col).value
                valor_min = self.ws.cell(row=fila_min, column=col).value
                if (
                    valor_max is not None
                    and str(valor_max).strip().upper() == "6RT"
                    and (valor_min is None or str(valor_min).strip() == "")
                    and self._es_celda_originalmente_vacia(fila_min, col)
                    and not self._es_celda_original_6rt(fila_max, col)
                    and not self._existe_7_en_dia(col)
                    and self._determinar_asignacion_por_personal(col)
                    and not self._tiene_extra_manana(trabajador_min, col)
                ):
                    # Elegible para mover 6RT de max a min en este día
                    columna_candidata = col
                    break

            if columna_candidata is None:
                # No hay movimientos factibles; salir
                break

            # Reasignar 6RT
            self.ws.cell(row=fila_max, column=columna_candidata, value=None)
            # Limpiar color de la celda original
            celda_original = self.ws.cell(row=fila_max, column=columna_candidata)
            celda_original.fill = PatternFill(fill_type=None)
            
            self.ws.cell(row=fila_min, column=columna_candidata, value="6RT")
            # Colorear celda de morado claro
            celda_nueva = self.ws.cell(row=fila_min, column=columna_candidata)
            celda_nueva.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            self.contador_grupo_6rt[trabajador_max] -= 1
            self.contador_grupo_6rt[trabajador_min] += 1

    def asignar_6rt_en_dia(self, col_dia: int) -> Optional[str]:
        # Decisión por personal (10-15 operativos)
        if not self._determinar_asignacion_por_personal(col_dia):
            return None

        # No duplicar 6RT/7 en el día
        if self._existe_6rt_o_7_en_dia(col_dia):
            return None

        # Intento con elegibles principales
        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        # Si no hay, intento con respaldo
        if not disponibles:
            disponibles = self._obtener_trabajadores_disponibles(col_dia, self.TRABAJADORES_RESPALDO)
            if not disponibles:
                return None

        # Prioridades: 1) DESC/TROP/SIND mañana y NO extra mañana; 2) NO extra mañana; 3) resto
        nivel1 = []
        nivel2 = []
        nivel3 = []
        for t in disponibles:
            prioridad = self._tiene_prioridad_manana(t, col_dia)
            extra = self._tiene_extra_manana(t, col_dia)
            if prioridad and not extra:
                nivel1.append(t)
            elif not extra:
                nivel2.append(t)
            else:
                nivel3.append(t)

        for candidatos in (nivel1, nivel2, nivel3):
            elegido = self._seleccionar_equitativo(candidatos)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                # Confirmar que la celda de destino fue originalmente vacía
                if not self._es_celda_originalmente_vacia(fila, col_dia):
                    return None
                self.ws.cell(row=fila, column=col_dia, value="6RT")
                # Colorear celda de morado claro
                celda = self.ws.cell(row=fila, column=col_dia)
                celda.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                self._actualizar_contadores(elegido, "6RT")
                return elegido

        return None

    # Nuevo: asignación de 6TT cuando operativos ≤9
    def asignar_6tt_en_dia(self, col_dia: int) -> Optional[str]:
        conteo = self._obtener_conteo_operativos(col_dia)
        if conteo is None or conteo > 9:
            return None
        # No duplicar 6TT en el día
        if self._existe_6tt_en_dia(col_dia):
            return None

        # Intento con elegibles principales
        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        # Si no hay, intento con respaldo
        if not disponibles:
            disponibles = self._obtener_trabajadores_disponibles(col_dia, self.TRABAJADORES_RESPALDO)
            if not disponibles:
                return None

        # Restricción blanda: preferir sin 1T/T1/1/7 mañana
        preferidos = [t for t in disponibles if not self._tiene_extra_manana(t, col_dia)]
        resto = [t for t in disponibles if t not in preferidos]

        for candidatos in (preferidos, resto):
            elegido = self._seleccionar_equitativo_6tt(candidatos)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                if not self._es_celda_originalmente_vacia(fila, col_dia):
                    return None
                self.ws.cell(row=fila, column=col_dia, value="6TT")
                self._actualizar_contadores(elegido, "6TT")
                return elegido
        return None

    def procesar_todos_los_dias(self) -> None:
        max_col = self.ws.max_column
        for col in range(2, max_col + 1):
            conteo = self._obtener_conteo_operativos(col)
            if conteo is None:
                continue
            if conteo <= 9:
                self.asignar_6tt_en_dia(col)
            elif 10 <= conteo <= 15:
                self.asignar_6rt_en_dia(col)
            else:
                # ≥16 no asignar
                pass

        # Forzar paridad ±1 en 6RT+7 cuando sea posible moviendo 6RT
        self._rebalancear_para_paridad()

        self._actualizar_hoja_estadisticas()

        salida = "horarioUnificado_con_6rt.xlsx"
        try:
            self.wb.save(salida)
            print(f"Archivo guardado como: {salida}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"Archivo por defecto en uso. Guardado como: {alternativo}")

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")   # 1T + 7
        ws_stats.cell(row=1, column=4, value="6RT")  # 6RT + 7

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, 5):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = 6RT + 7
            formula_6rt = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            fila_destino += 1

        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 8
        ws_stats.column_dimensions['C'].width = 8
        ws_stats.column_dimensions['D'].width = 8


if __name__ == "__main__":
    asignador = AsignadorTurnos6RT()
    asignador.procesar_todos_los_dias() 