import openpyxl
from typing import Dict, List

def verificar_asignaciones_mofis():
    """Verifica que las asignaciones MOFIS se realizaron correctamente"""
    
    # Trabajadores elegibles
    TRABAJADORES_ELEGIBLES = ['MEI', 'VCM', 'ROP', 'WEH']
    
    # Turnos MOFIS
    TURNOS_MOFIS = ["MS", "TS", "MN", "TN", "S", "N"]
    
    # Cargar archivo
    wb = openpyxl.load_workbook("horarioUnificado_con_mofis.xlsx")
    
    # Obtener hoja principal
    ws = None
    for nombre in wb.sheetnames:
        if nombre != "Estadísticas":
            ws = wb[nombre]
            break
    
    if not ws:
        print("No se encontró la hoja principal")
        return
    
    print(f"Verificando asignaciones MOFIS en: {ws.title}")
    print("=" * 50)
    
    # Contadores por trabajador
    contadores: Dict[str, Dict[str, int]] = {}
    for trabajador in TRABAJADORES_ELEGIBLES:
        contadores[trabajador] = {turno: 0 for turno in TURNOS_MOFIS}
        contadores[trabajador]['total_sn'] = 0
    
    # Contar asignaciones por día
    dias_con_asignaciones = 0
    total_asignaciones = 0
    
    for col in range(2, ws.max_column + 1):
        asignaciones_dia = []
        
        for trabajador in TRABAJADORES_ELEGIBLES:
            fila = None
            for r in range(2, 26):
                valor = ws.cell(row=r, column=1).value
                if valor and str(valor).strip().upper() == trabajador.upper():
                    fila = r
                    break
            
            if fila:
                valor = ws.cell(row=fila, column=col).value
                if valor and str(valor).strip().upper() in [t.upper() for t in TURNOS_MOFIS]:
                    turno = str(valor).strip().upper()
                    contadores[trabajador][turno] += 1
                    if turno in ['S', 'N']:
                        contadores[trabajador]['total_sn'] += 1
                    asignaciones_dia.append(f"{trabajador}: {turno}")
                    total_asignaciones += 1
        
        if asignaciones_dia:
            dias_con_asignaciones += 1
            print(f"Día {col}: {', '.join(asignaciones_dia)}")
    
    print("\n" + "=" * 50)
    print("ESTADÍSTICAS POR TRABAJADOR:")
    print("=" * 50)
    
    for trabajador in TRABAJADORES_ELEGIBLES:
        print(f"\n{trabajador}:")
        for turno in TURNOS_MOFIS:
            if contadores[trabajador][turno] > 0:
                print(f"  {turno}: {contadores[trabajador][turno]}")
        print(f"  Total S+N: {contadores[trabajador]['total_sn']}")
    
    print("\n" + "=" * 50)
    print("RESUMEN:")
    print("=" * 50)
    print(f"Días con asignaciones: {dias_con_asignaciones}")
    print(f"Total de asignaciones: {total_asignaciones}")
    
    # Verificar equidad en turnos S+N
    totales_sn = [contadores[t]['total_sn'] for t in TRABAJADORES_ELEGIBLES]
    min_sn = min(totales_sn)
    max_sn = max(totales_sn)
    diferencia = max_sn - min_sn
    
    print(f"\nEquidad en turnos S+N:")
    print(f"  Mínimo: {min_sn}, Máximo: {max_sn}")
    print(f"  Diferencia: {diferencia}")
    if diferencia <= 1:
        print("  ✅ Equidad correcta (diferencia ≤ 1)")
    else:
        print("  ⚠️  Equidad no óptima (diferencia > 1)")
    
    # Verificar hoja de estadísticas
    if "Estadísticas" in wb.sheetnames:
        ws_stats = wb["Estadísticas"]
        print(f"\nHoja de Estadísticas:")
        print(f"  Columnas: {ws_stats.max_column}")
        print(f"  Filas: {ws_stats.max_row}")
        
        # Verificar columna 6S
        if ws_stats.max_column >= 6:
            header_6s = ws_stats.cell(row=1, column=6).value
            print(f"  Columna 6S: {header_6s}")
        else:
            print("  ⚠️  No se encontró la columna 6S")

if __name__ == "__main__":
    verificar_asignaciones_mofis() 