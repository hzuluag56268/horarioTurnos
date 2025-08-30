import openpyxl

def inspect_excel(filename="horarioUnificado_con_6t.xlsx"):
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        
        print(f"Archivo: {filename}")
        print("Hojas disponibles:")
        for sheet_name in wb.sheetnames:
            print(f"  - {sheet_name}")
        
        # Revisar la hoja de estadísticas
        if "Estadísticas" in wb.sheetnames:
            ws = wb["Estadísticas"]
            print(f"\nColumnas en hoja 'Estadísticas' (primeras 20 columnas):")
            for col in range(1, min(21, ws.max_column + 1)):
                header = ws.cell(row=1, column=col).value
                print(f"  Columna {col}: '{header}'")
            
            print(f"\nPrimeras 5 filas de datos:")
            for fila in range(1, min(6, ws.max_row + 1)):
                datos = []
                for col in range(1, min(11, ws.max_column + 1)):
                    valor = ws.cell(row=fila, column=col).value
                    datos.append(str(valor)[:10] if valor is not None else "None")
                print(f"  Fila {fila}: {' | '.join(datos)}")
                
            # Revisar si hay datos reales (no None) en las columnas de turnos
            print(f"\nRevisando datos reales en columnas de turnos:")
            columnas_turnos = ['1T', '6RT', '6T', '3', '6S', '6N']
            for col_name in columnas_turnos:
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header == col_name:
                        # Revisar valores en esta columna
                        valores_no_nulos = 0
                        for fila in range(2, min(10, ws.max_row + 1)):
                            valor = ws.cell(row=fila, column=col).value
                            if valor is not None and valor != 0:
                                valores_no_nulos += 1
                        print(f"  {col_name} (col {col}): {valores_no_nulos} valores no nulos")
                        break
        else:
            print("\n❌ No se encontró la hoja 'Estadísticas'")
    except Exception as e:
        print(f"❌ Error al abrir {filename}: {e}")

if __name__ == "__main__":
    # Revisar varios archivos
    archivos = [
        "horarioUnificado_con_mofis.xlsx",
        "horarioUnificado_con_6t.xlsx", 
        "horarioUnificado_con_diurnas.xlsx",
        "horarioUnificado_con_6rt.xlsx"
    ]
    
    for archivo in archivos:
        print("="*60)
        inspect_excel(archivo)
        print() 