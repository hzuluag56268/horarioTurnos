"""
Asignador de turnos de sábados y festivos desde JSON para 'horarioUnificado_procesado.xlsx'.

Descripción general:
- Entrada JSON agrupada por turno, con elementos: { "fecha": <str>, "trabajador": <str> }.
- Encabezados de columnas esperados en la fila 1 del Excel: "DOW-DD" (por ejemplo, "THU-07", "SUN-10").
- Mapeo de fecha → columna por coincidencia exacta del encabezado "DOW-DD" (día de semana y día del mes).
- Algoritmo de asignación: primero intenta asignación directa; si no es posible, realiza un matching 1:1
  entre trabajadores y fechas del mismo turno (intercambios) en dos pasadas: sin violaciones blandas, y
  luego permitiéndolas en caso necesario.
- Restricciones tenidas en cuenta:
  * Restricción dura: si se asigna un turno NLPR/NANR (incluye variantes terminadas en "D"), el día
    siguiente del mismo trabajador NO puede ser BLPTD ni BANTD.
  * Restricción blanda: si se asigna un turno NLPT/NANT (incluye variantes terminadas en "D"), se debe
    evitar que el día siguiente del mismo trabajador sea BLPTD o BANTD; solo se permite si no existe
    alternativa viable.
- La validación de restricciones se hace contra:
  * El valor real que ya existe en la hoja en la columna del día siguiente.
  * El propio plan del JSON para BLPTD/BANTD en la columna del día siguiente.
    - Salida: guarda el Excel resultante como 'horario_procesado_con_sabados_domingos.xlsx' y un reporte tabulado en
  'reporte_asignador_sabados_festivos.txt' con el detalle de asignaciones (directas, intercambios, blandas) y no asignados.
- Formatos de fecha aceptados en el JSON: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY.
"""

import json
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Set

import openpyxl
from openpyxl.styles import PatternFill, Font

# ------------------------------------------------------------
# Utilidades de fechas y encabezados DOW-DD
# ------------------------------------------------------------
DOW_NAMES = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]


def parse_iso_date(date_str: str) -> datetime:
    # Acepta YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except Exception:
            continue
    raise ValueError(f"Fecha no reconocida: {date_str}")


def date_to_header_tuple(dt: datetime) -> Tuple[str, str]:
    dow = DOW_NAMES[dt.weekday()]
    dd = f"{dt.day:02d}"
    return dow, dd


def parse_header_cell(value: Optional[str]) -> Optional[Tuple[str, str]]:
    if value is None:
        return None
    val = str(value).strip().upper()
    # Esperado: DOW-DD, p.ej. MON-07
    if "-" not in val:
        return None
    dow, dd = val.split("-", 1)
    dow = dow.strip()
    dd = dd.strip()
    if dow not in DOW_NAMES:
        return None
    if not (len(dd) == 2 and dd.isdigit()):
        return None
    return dow, dd


# ------------------------------------------------------------
# Estructuras de datos
# ------------------------------------------------------------
@dataclass
class PedidoAsignacion:
    turno: str
    trabajador: str
    fecha_str: str
    fecha_dt: datetime


@dataclass
class ResultadoAsignacion:
    turno: str
    trabajador: str
    fecha_original: str
    fecha_final: Optional[str]
    columna_final: Optional[int]
    tipo: str  # directa | intercambio | blanda | no_asignado
    motivo: Optional[str] = None


# ------------------------------------------------------------
# Asignador desde JSON
# ------------------------------------------------------------
class AsignadorSabadosFestivos:
    """
    Carga un plan de turnos de sábados y festivos desde un archivo JSON y lo aplica sobre 'horarioUnificado_procesado.xlsx'.

    Parámetros del constructor:
    - excel_in: ruta del Excel base (por defecto 'horarioUnificado_procesado.xlsx').
    - json_path: ruta del JSON de entrada (por defecto 'cuentas1y2sabadosDomingo_asignado.json').
    - excel_out: ruta del Excel de salida (por defecto 'horario_procesado_con_sabados_domingos.xlsx').
    - modo_simulacion: si es True, no escribe en el Excel (solo genera reporte en memoria).

    Encabezados y fechas:
    - La fila 1 contiene encabezados de tipo 'DOW-DD' (p. ej., 'THU-07').
    - Las fechas del JSON se mapean a columnas por coincidencia exacta de 'DOW-DD'.

    Proceso de asignación por turno:
    1) Cargar pedidos y precomputar el plan de BLPTD/BANTD del JSON para el día siguiente.
    2) Construir un grafo bipartito entre (trabajadores) y (fechas del JSON) marcando aristas:
       - 'fuertes' (no violan restricción blanda) y
       - 'blandas' (violan la restricción blanda, pero no la dura).
    3) Hallar un matching máximo en dos pasadas: primero con aristas fuertes; luego, si quedan pendientes,
       permitir aristas blandas.
    4) Escribir al Excel respetando las restricciones y priorizando evitar las blandas.
        5) Garantizar unicidad por día: un mismo turno no se asigna a más de un trabajador en la misma columna.

    Restricciones implementadas:
    - Dura: 
      * NLPR/NANR/NLPRD/NANRD/6R/6RT: no pueden tener BLPTD/BANTD al día siguiente
      * BLPTD/BANTD: no pueden tener BLPTD/BANTD al día anterior ni al día siguiente
    - Blanda: NLPT/NANT/NLPTD/NANTD/TASTD/6T/3/6TT deben evitar BLPTD/BANTD al día siguiente
    - Si no existe 'día siguiente' (última columna), no aplica la validación.
    - Si no existe 'día anterior' (primera columna), no aplica la validación.

    Formato visual:
    - Las celdas con violaciones blandas se colorean de azul clarito (#87CEEB) para facilitar su identificación.

    Notas:
    - La comprobación considera tanto lo que ya está escrito en la hoja como lo que planea
      el propio JSON (si en las columnas anterior/siguiente hay un pedido BLPTD/BANTD para el mismo trabajador).
    - El reporte 'reporte_asignador_sabados_festivos.txt' detalla para cada pedido si fue 'directa', 'intercambio', 'blanda'
      o 'no_asignado', junto con la columna destino y el encabezado 'DOW-DD'.
    """
    def __init__(
        self,
        excel_in: str = "horarioUnificado_procesado.xlsx",
        json_path: str = "cuentas1y2sabadosDomingo_asignado.json",
        excel_out: str = "horario_procesado_con_sabados_domingos.xlsx",
        modo_simulacion: bool = True,
    ) -> None:
        self.excel_in = excel_in
        self.json_path = json_path
        self.excel_out = excel_out
        self.modo_simulacion = modo_simulacion

        if not os.path.exists(self.excel_in):
            raise FileNotFoundError(f"No se encontró el archivo Excel: {self.excel_in}")
        if not os.path.exists(self.json_path):
            raise FileNotFoundError(f"No se encontró el archivo JSON: {self.json_path}")

        self.wb = openpyxl.load_workbook(self.excel_in)
        self.ws = self.wb.active

        # Mapeos clave
        self.sigla_to_row: Dict[str, int] = {}
        self.header_map: Dict[Tuple[str, str], List[int]] = {}
        self.col_to_header_tuple: Dict[int, Tuple[str, str]] = {}

        # Conjuntos de reglas
        self.hard_source_turns: Set[str] = {"NLPR", "NANR", "NLPRD", "NANRD", "6R", "6RT", "BLPTD", "BANTD"}
        self.soft_source_turns: Set[str] = {"NLPT", "NANT", "NLPTD", "NANTD", "TASTD", "6T", "3", "6TT"}
        self.blocked_next_day_turns: Set[str] = {"BLPTD", "BANTD"}

        # Plan del propio JSON para BLPTD/BANTD por (trabajador, col)
        self.plan_blpt_bant_por_celda: Set[Tuple[str, int]] = set()

        # Color para violaciones blandas
        self.color_violacion_blanda = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul clarito

        # Reporte
        self.resultados: List[ResultadoAsignacion] = []

        # Color para violaciones duras (fucsia)
        self.color_violacion_dura = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")

        self._mapear_trabajadores()
        self._mapear_encabezados()

    # --------------------------------------------------------
    # Mapeos de hoja
    # --------------------------------------------------------
    def _mapear_trabajadores(self) -> None:
        for fila in range(2, 26):
            val = self.ws.cell(row=fila, column=1).value
            if not val:
                continue
            sigla = str(val).strip().upper()
            if sigla:
                self.sigla_to_row[sigla] = fila

    def _mapear_encabezados(self) -> None:
        max_col = self.ws.max_column
        for col in range(2, max_col + 1):
            header = parse_header_cell(self.ws.cell(row=1, column=col).value)
            if not header:
                continue
            self.col_to_header_tuple[col] = header
            self.header_map.setdefault(header, []).append(col)

    # --------------------------------------------------------
    # Utilidades de mapeo fecha → columna
    # --------------------------------------------------------
    def _columna_para_fecha_preferida(self, dt: datetime) -> Optional[int]:
        dow, dd = date_to_header_tuple(dt)
        key = (dow, dd)
        cols = self.header_map.get(key, [])
        if not cols:
            return None
        # Si hay múltiples columnas con el mismo encabezado (caso raro), usar la primera
        return cols[0]

    def _todas_columnas_para_fecha(self, dt: datetime) -> List[int]:
        dow, dd = date_to_header_tuple(dt)
        return list(self.header_map.get((dow, dd), []))

    # --------------------------------------------------------
    # Carga de JSON y normalización
    # --------------------------------------------------------
    def _cargar_json(self) -> Dict[str, List[PedidoAsignacion]]:
        with open(self.json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        pedidos_por_turno: Dict[str, List[PedidoAsignacion]] = {}
        for turno, items in data.items():
            lst: List[PedidoAsignacion] = []
            for it in items:
                fecha_str = str(it.get("fecha", "")).strip()
                trabajador = str(it.get("trabajador", "")).strip().upper()
                if not fecha_str or not trabajador:
                    continue
                try:
                    dt = parse_iso_date(fecha_str)
                except Exception:
                    # Si fecha no parsea, saltamos pero reportamos luego
                    continue
                lst.append(PedidoAsignacion(turno=turno.strip(), trabajador=trabajador, fecha_str=fecha_str, fecha_dt=dt))
            pedidos_por_turno[turno] = lst
        return pedidos_por_turno

    # --------------------------------------------------------
    # Plan de BLPT/BANT para validar restricciones contra el "día siguiente"
    # --------------------------------------------------------
    def _precomputar_plan_blpt_bant(self, pedidos_por_turno: Dict[str, List[PedidoAsignacion]]) -> None:
        claves_objetivo = {"BLPTD", "BANTD"}
        # Construir set de (trabajador, columna) para el plan del propio JSON
        for turno, lst in pedidos_por_turno.items():
            if turno.strip().upper() not in claves_objetivo:
                continue
            for p in lst:
                col = self._columna_para_fecha_preferida(p.fecha_dt)
                if col is not None:
                    self.plan_blpt_bant_por_celda.add((p.trabajador, col))

    # --------------------------------------------------------
    # Validaciones de celda y restricciones
    # --------------------------------------------------------
    def _celda_vacia(self, fila: int, col: int) -> bool:
        val = self.ws.cell(row=fila, column=col).value
        return val is None or str(val).strip() == ""

    def _valor_en(self, fila: int, col: int) -> Optional[str]:
        val = self.ws.cell(row=fila, column=col).value
        if val is None:
            return None
        return str(val).strip().upper()

    def _existe_turno_en_columna(self, col_dia: int, turno: str) -> bool:
        """True si en ese día (columna) ya existe el turno indicado en cualquier trabajador (filas 2-25)."""
        turno_u = turno.strip().upper()
        for fila in range(2, 26):
            val = self.ws.cell(row=fila, column=col_dia).value
            if val is None:
                continue
            if str(val).strip().upper() == turno_u:
                return True
        return False

    def _chequear_restricciones(self, trabajador: str, col_actual: int, turno_actual: str) -> Tuple[bool, bool, Optional[str]]:
        """
        Valida restricciones referidas al día anterior y siguiente para una asignación tentativa.

        Retorna:
        - (violacion_dura, violacion_blanda, motivo)

        Reglas:
        - Dura: 
          * NLPR/NANR/NLPRD/NANRD/6R/6RT: no pueden tener BLPTD/BANTD al día siguiente
          * BLPTD/BANTD: no pueden tener BLPTD/BANTD al día anterior ni al día siguiente
        - Blanda: NLPT/NANT/NLPTD/NANTD/TASTD/6T/3/6TT deben evitar BLPTD/BANTD al día siguiente

        Si 'col_actual + 1' excede el número de columnas, no se valida restricción de día siguiente.
        Si 'col_actual - 1' es menor a 2, no se valida restricción de día anterior.
        """
        # Retorna (violacion_dura, violacion_blanda, motivo)
        turno_u = turno_actual.strip().upper()
        fila = self.sigla_to_row.get(trabajador)
        if not fila:
            return False, False, None

        # Día siguiente
        next_col = col_actual + 1
        max_col = self.ws.max_column
        next_val = None
        en_plan_blpt_bant_next = False
        if next_col <= max_col:
            next_val = self._valor_en(fila, next_col)
            en_plan_blpt_bant_next = (trabajador, next_col) in self.plan_blpt_bant_por_celda

        # Día anterior
        prev_col = col_actual - 1
        prev_val = None
        en_plan_blpt_bant_prev = False
        if prev_col >= 2:
            prev_val = self._valor_en(fila, prev_col)
            en_plan_blpt_bant_prev = (trabajador, prev_col) in self.plan_blpt_bant_por_celda

        # Restricciones duras
        if turno_u in self.hard_source_turns:
            # Para BLPTD/BANTD: verificar día anterior y siguiente
            if turno_u in {"BLPTD", "BANTD"}:
                if ((prev_val in self.blocked_next_day_turns) or en_plan_blpt_bant_prev or
                    (next_val in self.blocked_next_day_turns) or en_plan_blpt_bant_next):
                    return True, False, f"Restricción dura: {turno_u} con BLPTD/BANTD en día anterior o siguiente"
            # Para otros turnos duros: solo verificar día siguiente
            else:
                if (next_val in self.blocked_next_day_turns) or en_plan_blpt_bant_next:
                    return True, False, f"Restricción dura: {turno_u} con BLPTD/BANTD al día siguiente"

        # Restricciones blandas: solo verificar día siguiente
        if turno_u in self.soft_source_turns:
            if (next_val in self.blocked_next_day_turns) or en_plan_blpt_bant_next:
                return False, True, f"Restricción blanda: {turno_u} con BLPTD/BANTD al día siguiente"

        return False, False, None

    # --------------------------------------------------------
    # Matching por turno (1:1 entre pedidos y fechas del JSON)
    # --------------------------------------------------------
    def _resolver_turno(self, turno: str, pedidos: List[PedidoAsignacion]) -> None:
        # Normalizar en mayúsculas para comparaciones
        turno_u = turno.strip().upper()

        # Construir slots de fecha (uno por ítem del JSON para este turno)
        slots_fechas: List[datetime] = [p.fecha_dt for p in pedidos]

        # Construir grafo bipartito: izquierda = índices de pedidos, derecha = índices de slots
        # Edges duras (permitidas) y blandas (a evitar)
        n = len(pedidos)
        m = len(slots_fechas)
        hard_edges: Dict[int, List[int]] = {i: [] for i in range(n)}
        soft_edges: Dict[int, List[int]] = {i: [] for i in range(n)}

        # Precomputar columnas por slot
        slot_col_preferida: List[Optional[int]] = [self._columna_para_fecha_preferida(dt) for dt in slots_fechas]

        for i, pedido in enumerate(pedidos):
            fila = self.sigla_to_row.get(pedido.trabajador)
            if not fila:
                continue

            for j, dt in enumerate(slots_fechas):
                col = slot_col_preferida[j]
                if col is None:
                    # Si no hay mapeo directo, intentar cualquier columna de ese DOW-MM
                    posibles = self._todas_columnas_para_fecha(dt)
                else:
                    posibles = [col]

                # Este slot j es viable si existe alguna columna posible
                viable_fuerte = False
                viable_blando = False
                for c in posibles:
                    if not self._celda_vacia(fila, c):
                        continue
                    violacion_dura, violacion_blanda, _ = self._chequear_restricciones(pedido.trabajador, c, turno_u)
                    if violacion_dura:
                        continue
                    if violacion_blanda:
                        viable_blando = True
                    else:
                        viable_fuerte = True
                        break  # preferimos fuerte
                if viable_fuerte:
                    hard_edges[i].append(j)
                elif viable_blando:
                    soft_edges[i].append(j)

        # Matching máximo con dos pasadas: primero solo aristas duras, luego agregando blandas
        match_r: Dict[int, int] = {}  # slot j -> pedido i

        def try_kuhn(i: int, seen: Set[int], use_soft: bool) -> bool:
            # intenta asignar pedido i a algún slot disponible con DFS de aumento
            for j in hard_edges.get(i, []):
                if j in seen:
                    continue
                seen.add(j)
                if j not in match_r or try_kuhn(match_r[j], seen, use_soft):
                    match_r[j] = i
                    return True
            if use_soft:
                for j in soft_edges.get(i, []):
                    if j in seen:
                        continue
                    seen.add(j)
                    if j not in match_r or try_kuhn(match_r[j], seen, use_soft):
                        match_r[j] = i
                        return True
            return False

        # Pasada 1: solo duras
        for i in range(n):
            try_kuhn(i, set(), use_soft=False)
        # Pasada 2: permitir blandas
        for i in range(n):
            if i not in match_r.values():
                try_kuhn(i, set(), use_soft=True)

        # Escribir resultados
        # Invertimos match_r: pedido i -> slot j
        pedido_to_slot: Dict[int, int] = {i: j for j, i in match_r.items()}

        for i, pedido in enumerate(pedidos):
            fila = self.sigla_to_row.get(pedido.trabajador)
            if not fila:
                self.resultados.append(
                    ResultadoAsignacion(
                        turno=turno,
                        trabajador=pedido.trabajador,
                        fecha_original=pedido.fecha_str,
                        fecha_final=None,
                        columna_final=None,
                        tipo="no_asignado",
                        motivo="Trabajador no encontrado en la hoja",
                    )
                )
                continue

            j = pedido_to_slot.get(i)
            if j is None:
                # Intento forzado: buscar cualquier columna viable (ocupando aunque viole restricción dura)
                fila = self.sigla_to_row.get(pedido.trabajador)
                col_forzada = None
                motivo = "Sin fecha viable (ocupado o restricciones)"
                # Buscar primero por la columna preferida del slot original
                dt_target = pedido.fecha_dt
                candidatas = []
                col_pref = self._columna_para_fecha_preferida(dt_target)
                if col_pref is not None:
                    candidatas.append(col_pref)
                candidatas.extend([c for c in self._todas_columnas_para_fecha(dt_target) if c not in candidatas])
                for c in candidatas:
                    if not self._celda_vacia(fila, c):
                        continue
                    # Evitar duplicado del mismo turno en la columna
                    if self._existe_turno_en_columna(c, turno):
                        continue
                    # Chequear si esto viola una dura
                    viol_dura, viol_blanda, _ = self._chequear_restricciones(pedido.trabajador, c, turno)
                    if viol_dura:
                        col_forzada = c
                        motivo = "Asignado forzado con violación dura"
                        break
                if col_forzada is not None and not self.modo_simulacion:
                    celda = self.ws.cell(row=fila, column=col_forzada, value=turno)
                    celda.fill = self.color_violacion_dura
                    dow_mm = self.col_to_header_tuple.get(col_forzada)
                    fecha_final_str = f"{dow_mm[0]}-{dow_mm[1]}" if dow_mm else None
                    self.resultados.append(
                        ResultadoAsignacion(
                            turno=turno,
                            trabajador=pedido.trabajador,
                            fecha_original=pedido.fecha_str,
                            fecha_final=fecha_final_str,
                            columna_final=col_forzada,
                            tipo="dura",
                            motivo=motivo,
                        )
                    )
                    continue

                # Si no se pudo forzar, registrar como no_asignado
                self.resultados.append(
                    ResultadoAsignacion(
                        turno=turno,
                        trabajador=pedido.trabajador,
                        fecha_original=pedido.fecha_str,
                        fecha_final=None,
                        columna_final=None,
                        tipo="no_asignado",
                        motivo=motivo,
                    )
                )
                continue

            # Determinar columna final elegida para ese slot
            dt_final = slots_fechas[j]
            col_pref = self._columna_para_fecha_preferida(dt_final)
            col_candidatas = [col_pref] if col_pref is not None else self._todas_columnas_para_fecha(dt_final)
            col_final: Optional[int] = None
            tipo_asig = "intercambio"

            # Elegir columna específica minimizando violaciones blandas
            for c in col_candidatas:
                if c is None:
                    continue
                if not self._celda_vacia(fila, c):
                    continue
                # Evitar duplicar el mismo turno en la columna (un solo trabajador por día con ese turno)
                if self._existe_turno_en_columna(c, turno):
                    continue
                violacion_dura, violacion_blanda, _ = self._chequear_restricciones(pedido.trabajador, c, turno)
                if violacion_dura:
                    continue
                if not violacion_blanda:
                    col_final = c
                    tipo_asig = "directa" if pedido.fecha_dt == dt_final else "intercambio"
                    break
            if col_final is None:
                # Aceptar blanda si no hubo fuerte
                for c in col_candidatas:
                    if c is None:
                        continue
                    if not self._celda_vacia(fila, c):
                        continue
                    # Evitar duplicar el mismo turno en la columna
                    if self._existe_turno_en_columna(c, turno):
                        continue
                    violacion_dura, violacion_blanda, _ = self._chequear_restricciones(pedido.trabajador, c, turno)
                    if violacion_dura or not violacion_blanda:
                        continue
                    col_final = c
                    tipo_asig = "blanda"
                    break

            if col_final is None:
                self.resultados.append(
                    ResultadoAsignacion(
                        turno=turno,
                        trabajador=pedido.trabajador,
                        fecha_original=pedido.fecha_str,
                        fecha_final=None,
                        columna_final=None,
                        tipo="no_asignado",
                        motivo="No se encontró columna viable al escribir",
                    )
                )
                continue

            # Escribir en la hoja
            if not self.modo_simulacion:
                celda = self.ws.cell(row=fila, column=col_final, value=turno)
                # Colorear solo si hay violación; de lo contrario, limpiar relleno
                if tipo_asig == "blanda":
                    celda.fill = self.color_violacion_blanda
                else:
                    celda.fill = PatternFill(fill_type=None)

            # Fecha final para reporte: reconstruida desde encabezado
            dow_mm = self.col_to_header_tuple.get(col_final)
            fecha_final_str = None
            if dow_mm:
                fecha_final_str = f"{dow_mm[0]}-{dow_mm[1]}"

            # Motivo para blanda
            motivo_final = None
            if tipo_asig == "blanda":
                motivo_final = "Asignado aceptando violación blanda (BLPTD/BANTD al día siguiente)"

            self.resultados.append(
                ResultadoAsignacion(
                    turno=turno,
                    trabajador=pedido.trabajador,
                    fecha_original=pedido.fecha_str,
                    fecha_final=fecha_final_str,
                    columna_final=col_final,
                    tipo=tipo_asig,
                    motivo=motivo_final,
                )
            )

    # --------------------------------------------------------
    # Reporte
    # --------------------------------------------------------
    def _guardar_reporte(self, path: str = "reporte_asignador_sabados_festivos.txt") -> None:
        lineas: List[str] = []
        lineas.append("turno\ttrabajador\tfecha_original\tfecha_final(DOW-MM)\tcolumna\ttipo\tmotivo")
        # Orden: primero no_asignado, luego dura, blanda, intercambio, directa
        prioridad = {"no_asignado": 0, "dura": 1, "blanda": 2, "intercambio": 3, "directa": 4}
        for r in sorted(self.resultados, key=lambda x: (prioridad.get(x.tipo, 99), x.turno, x.trabajador)):
            lineas.append(
                f"{r.turno}\t{r.trabajador}\t{r.fecha_original}\t{r.fecha_final or ''}\t{r.columna_final or ''}\t{r.tipo}\t{r.motivo or ''}"
            )
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lineas))

    # --------------------------------------------------------
    # Actualizar hoja de estadísticas con 1D, 3D, 6D
    # --------------------------------------------------------
    def _actualizar_hoja_estadisticas_sd(self) -> None:
        """
        Agrega/actualiza columnas en la hoja 'Estadísticas':
        - 1D: BANTD + BLPTD
        - 3D: 3 (o 3D si existiera)
        - 6D: NLPTD + NLPRD + NANTD + NANRD
        Conserva el resto de columnas existentes y solo escribe las necesarias.
        """
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Crear mapa de encabezados existentes (mayúsculas)
        encabezados = {}
        max_col_stats = ws_stats.max_column or 0
        for c in range(1, max_col_stats + 1):
            v = ws_stats.cell(row=1, column=c).value
            if v is None:
                continue
            encabezados[str(v).strip().upper()] = c

        def asegurar_columna(nombre: str) -> int:
            nombre_u = nombre.strip().upper()
            if nombre_u in encabezados:
                return encabezados[nombre_u]
            col_nueva = (ws_stats.max_column or 0) + 1
            ws_stats.cell(row=1, column=col_nueva, value=nombre)
            # Estilo encabezado
            header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            header_font = Font(bold=True)
            cell = ws_stats.cell(row=1, column=col_nueva)
            cell.fill = header_fill
            cell.font = header_font
            encabezados[nombre_u] = col_nueva
            return col_nueva

        # Asegurar SIGLA y columnas pedidas
        col_sigla = asegurar_columna("SIGLA")
        col_1d = asegurar_columna("1D")
        col_3d = asegurar_columna("3D")
        col_6d = asegurar_columna("6D")

        # Anchos de columnas
        try:
            ws_stats.column_dimensions['A'].width = max(ws_stats.column_dimensions['A'].width or 0, 10)
        except Exception:
            pass

        # Nombre de hoja principal y rango B..AC
        hoja = self.ws.title

        # Escribir filas
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=col_sigla, value=trabajador)

            # Fórmulas dinámicas
            # 1D = BANTD + BLPTD
            formula_1d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"BANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"BLPTD")'
            )
            ws_stats.cell(row=fila_destino, column=col_1d, value=formula_1d)

            # 3D = 3D
            formula_3d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"3D")'
            )
            ws_stats.cell(row=fila_destino, column=col_3d, value=formula_3d)

            # 6D = NLPTD + NLPRD + NANTD + NANRD
            formula_6d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"NLPTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NLPRD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANRD")'
            )
            ws_stats.cell(row=fila_destino, column=col_6d, value=formula_6d)

            fila_destino += 1

    def _recalcular_estaticos_operativos_y_torre(self) -> None:
        """
        Recalcula las dos filas estáticas en la hoja principal:
        - 'TURNOS OPERATIVOS'
        - 'Torre'
        Usando la misma lógica de conteo que en procesador_horarios.py
        """
        # Definir turnos no operativos (copiado de procesador_horarios.py)
        turnos_no_operativos = {
            # Turnos básicos
            "DESC", "TROP",
            # Turnos completos
            "VACA", "COME", "COMT", "COMS",
            # Turnos adicionales originales
            "SIND", "CMED", "CERT",
            # Formación, instrucción y entrenamiento
            "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
            "MENT", "TENT", "NENT", "AENT",
            "MINS", "TINS", "NINS", "AINS",
            # Gestión, oficinas y grupos de trabajo
            "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
            "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
            "MGST", "TGST", "MOFI", "TOFI",
            # Operativos y asignaciones especiales
            "CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X"
        }

        ws = self.ws
        max_row = ws.max_row
        max_col = ws.max_column

        # Buscar/crear filas destino
        fila_conteo = None
        fila_torre = None
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=1).value
            if not v:
                continue
            val = str(v).strip().upper()
            if val == "TURNOS OPERATIVOS":
                fila_conteo = r
            elif val == "TORRE":
                fila_torre = r
        # Si no existen, crearlas al final conservando el orden (conteo antes que torre)
        cursor = max_row + 1
        if fila_conteo is None:
            fila_conteo = cursor
            ws.cell(row=fila_conteo, column=1, value="TURNOS OPERATIVOS")
            cursor += 1
        if fila_torre is None:
            fila_torre = cursor
            ws.cell(row=fila_torre, column=1, value="Torre")

        # Mapa sigla -> fila (2-25)
        sigla_a_fila = {}
        for r in range(2, min(26, max_row + 1)):
            sig = ws.cell(row=r, column=1).value
            if isinstance(sig, str):
                s = sig.strip().upper()
                if s:
                    sigla_a_fila[s] = r

        # Filas objetivo para Torre
        siglas_torre = {"YIS", "MAQ", "DJO", "AFG", "JLF", "JMV"}
        filas_objetivo = [sigla_a_fila[s] for s in siglas_torre if s in sigla_a_fila]

        # Recalcular conteos por columna
        for col in range(2, max_col + 1):
            # Operativos
            conteo_operativos = 0
            for r in range(2, 26):
                cell_value = ws.cell(row=r, column=col).value
                if cell_value is None or str(cell_value).strip() == "":
                    conteo_operativos += 1
                else:
                    valor_limpio = str(cell_value).strip().upper()
                    if valor_limpio not in turnos_no_operativos:
                        conteo_operativos += 1
            ws.cell(row=fila_conteo, column=col, value=conteo_operativos)

            # Torre
            conteo_torre = 0
            for r in filas_objetivo:
                v = ws.cell(row=r, column=col).value
                if v is None or str(v).strip() == "":
                    conteo_torre += 1
                else:
                    if str(v).strip().upper() not in turnos_no_operativos:
                        conteo_torre += 1
            ws.cell(row=fila_torre, column=col, value=conteo_torre)

        # Asegurar etiquetas
        ws.cell(row=fila_conteo, column=1, value="TURNOS OPERATIVOS")
        ws.cell(row=fila_torre, column=1, value="Torre")

    # --------------------------------------------------------
    # Orquestador
    # --------------------------------------------------------
    def asignar(self) -> None:
        """
        Ejecuta el flujo completo de asignación:
        - Carga y normaliza el JSON de pedidos.
        - Precomputa el plan de BLPTdom/BANTdom para el día siguiente.
        - Resuelve por turno el matching 1:1 con dos pasadas (evitando primero violaciones blandas).
        - Escribe en el Excel (salvo 'modo_simulacion=True') y genera el reporte en disco.
        """
        pedidos_por_turno = self._cargar_json()
        # Precompute plan de BLPT/BANT en el siguiente día
        self._precomputar_plan_blpt_bant(pedidos_por_turno)

        # Procesar cada turno de forma independiente, manteniendo consistencia de restricciones
        for turno, pedidos in pedidos_por_turno.items():
            # Ordenar por fecha original para estabilidad
            pedidos_ordenados = sorted(pedidos, key=lambda p: (p.fecha_dt, p.trabajador))
            self._resolver_turno(turno, pedidos_ordenados)

        # Actualizar hoja de estadísticas con 1D, 3D, 6D
        self._actualizar_hoja_estadisticas_sd()

        # Recalcular filas estáticas
        self._recalcular_estaticos_operativos_y_torre()

        # Guardar archivos
        if not self.modo_simulacion:
            try:
                self.wb.save(self.excel_out)
                print(f"Archivo guardado como: {self.excel_out}")
            except PermissionError:
                base, ext = os.path.splitext(self.excel_out)
                alternativo = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
                self.wb.save(alternativo)
                print(f"Archivo en uso. Guardado como: {alternativo}")

        self._guardar_reporte()
        print("Reporte escrito en 'reporte_asignador_sabados_festivos.txt'")


if __name__ == "__main__":
    asignador = AsignadorSabadosFestivos(modo_simulacion=False)
    asignador.asignar() 