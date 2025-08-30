import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict
import os
from openpyxl.comments import Comment


class AsignadorTurnos:
    """
    Asigna turnos "1T" (1 hora extra) o "7" (1 hora extra + 6 horas adicionales)
    sobre el archivo procesado, con las siguientes reglas:
    - Decisión por día según personal disponible (fila cuyo encabezado en la columna A es "TURNOS OPERATIVOS"):
      * ≤8  → no asignar
      * =9  → asignar turno "7"
      * ≥10 → asignar turno "1T"
    - Solo se asigna entre trabajadores elegibles: ['GCE', 'YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']
    - Prioridad para quienes tuvieron DESC/TROP/SIND el día anterior
    - Evitar, en lo posible, asignar si el día anterior tuvo "1T" o "7"
    - Restricción dura: NO asignar si el día anterior tuvo BANTD, BLPTD, NLPRD, NANRD, 1T o 7
    - Restricción blanda: Evitar si el día anterior tuvo NANTD o NLPTD
    - Verificar que en ese día NO exista ya un turno "1T" ni "7" ni "BLPTD" ni "BANTD" (no duplicar)
    - No asignar si el día siguiente tiene BANTD, BLPTD, 1T o 7
    - Restricción Torre: para "GCE", solo asignar 1T si el conteo de la fila "Torre" es ≤3
    - Equidad:
      * Grupo 1T: cuenta "1T" + "7" (toda persona con 1 hora extra)
      * Grupo 6RT: cuenta solo "7" (6 horas adicionales)
      * Para turno "1T": balancear por el grupo 1T
      * Para turno "7": balancear por el grupo 1T; si hay empate, usar grupo 6RT como desempate
    - Actualiza hoja "Estadísticas" con columnas: SIGLA, DESC, 1T (1T+7), 6RT (solo 7)
    """

    TRABAJADORES_ELEGIBLES = ['GCE', 'YIS', 'MAQ', 'DJO', 'AFG', 'JLF', 'JMV']

    def __init__(self, archivo_procesado: Optional[str] = None) -> None:
        self.archivo_procesado = self._resolver_archivo_entrada(archivo_procesado)
        self.wb = openpyxl.load_workbook(self.archivo_procesado)
        self.ws = self._obtener_hoja_horario()
        self.contador_grupo_1t: Dict[str, int] = defaultdict(int)
        self.contador_grupo_6rt: Dict[str, int] = defaultdict(int)
        # Inicializar contadores a partir de asignaciones ya existentes
        self._inicializar_contadores_desde_hoja()
        random.seed()

    def _resolver_archivo_entrada(self, preferido: Optional[str]) -> str:
        candidatos = [
            preferido,
            "horario_procesado_con_sabados_domingos.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        for c in [c for c in candidatos if c]:
            if os.path.exists(c):
                return c
        return "horario_procesado_con_sabados_domingos.xlsx"

    def _obtener_hoja_horario(self):
        """
        Obtiene la hoja principal de horario (no la de estadísticas).
        Si existe una hoja llamada "Estadísticas", se omite.
        """
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        # Por seguridad, si solo existe "Estadísticas", usarla igualmente
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        """Devuelve la fila (int) donde está el trabajador en la columna A, o None si no se encuentra."""
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _tiene_prioridad_dia_anterior(self, trabajador: str, col_dia: int) -> bool:
        """True si el día anterior el trabajador tiene DESC, TROP o SIND."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"DESC", "TROP", "SIND"}

    def _tuvo_extra_dia_anterior(self, trabajador: str, col_dia: int) -> bool:
        """True si el día anterior el trabajador tuvo 1T o 7."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"1T", "7"}

    def _tuvo_restriccion_dura_ayer(self, trabajador: str, col_dia: int) -> bool:
        """True si ayer tuvo BANTD, BLPTD, NLPRD, NANRD, 1T o 7."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"BANTD", "BLPTD", "NLPRD", "NANRD", "1T", "7"}

    def _tiene_restriccion_dura_manana(self, trabajador: str, col_dia: int) -> bool:
        """True si mañana tiene BANTD, BLPTD, 1T o 7."""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if col_dia + 1 > self.ws.max_column:
            return False
        valor = self.ws.cell(row=fila, column=col_dia + 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"BANTD", "BLPTD", "1T", "7"}

    def _tuvo_restriccion_blanda_ayer(self, trabajador: str, col_dia: int) -> bool:
        """True si ayer tuvo NANTD o NLPTD (evitar si es posible)."""
        if col_dia <= 2:
            return False
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        valor = self.ws.cell(row=fila, column=col_dia - 1).value
        if valor is None:
            return False
        return str(valor).strip().upper() in {"NANTD", "NLPTD"}

    def _obtener_trabajadores_disponibles(self, col_dia: int) -> List[str]:
        """Devuelve los trabajadores elegibles cuya celda del día está vacía."""
        disponibles = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if celda.value is None or str(celda.value).strip() == "":
                disponibles.append(trabajador)
        return disponibles

    def _obtener_conteo_operativos(self, col_dia: int) -> Optional[int]:
        """Busca la fila con etiqueta 'TURNOS OPERATIVOS' en la columna A y devuelve el entero de esa columna del día."""
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TURNOS OPERATIVOS":
                valor = self.ws.cell(row=fila, column=col_dia).value
                try:
                    return int(valor)
                except Exception:
                    return None
        return None

    def _obtener_conteo_torre(self, col_dia: int) -> Optional[int]:
        """Busca la fila con etiqueta 'Torre' (columna A) y devuelve el entero de esa columna del día."""
        for fila in range(1, self.ws.max_row + 1):
            etiqueta = self.ws.cell(row=fila, column=1).value
            if etiqueta and str(etiqueta).strip().upper() == "TORRE":
                valor = self.ws.cell(row=fila, column=col_dia).value
                try:
                    return int(valor)
                except Exception:
                    return None
        return None

    def _determinar_turno_por_personal(self, col_dia: int) -> Optional[str]:
        """
        - ≤8: None (no asignar)
        - =9: "7"
        - ≥10: "1T"
        """
        disponible = self._obtener_conteo_operativos(col_dia)
        if disponible is None:
            return None
        if disponible <= 8:
            return None
        if disponible == 9:
            return "7"
        return "1T"

    def _existe_turno_1t_o_7_en_dia(self, col_dia: int) -> bool:
        """True si en ese día ya existe un 1T, 7, BLPTD o BANTD en cualquier trabajador (filas 2-25)."""
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=col_dia).value
            if valor is None:
                continue
            val = str(valor).strip().upper()
            if val in {"1T", "7", "BLPTD", "BANTD"}:
                return True
        return False

    def _seleccionar_equitativo(self, candidatos: List[str], turno: str) -> Optional[str]:
        """
        Para turno "1T": usar contador_grupo_1t.
        Para turno "7": usar contador_grupo_1t y, como desempate, contador_grupo_6rt.
        """
        if not candidatos:
            return None

        # Contador principal siempre es el de 1T (porque 7 también suma al grupo 1T)
        principal = self.contador_grupo_1t
        min_principal = min(principal[c] for c in candidatos)
        candidatos_min = [c for c in candidatos if principal[c] == min_principal]

        if turno == "7" and len(candidatos_min) > 1:
            # Desempate con grupo 6RT
            secundaria = self.contador_grupo_6rt
            min_sec = min(secundaria[c] for c in candidatos_min)
            candidatos_min = [c for c in candidatos_min if secundaria[c] == min_sec]

        return random.choice(candidatos_min)

    def _actualizar_contadores(self, trabajador: str, turno: str) -> None:
        if turno == "1T":
            self.contador_grupo_1t[trabajador] += 1
        elif turno == "7":
            self.contador_grupo_1t[trabajador] += 1
            self.contador_grupo_6rt[trabajador] += 1

    def _inicializar_contadores_desde_hoja(self) -> None:
        """Inicializa los contadores 1T (1T+7) y 6RT (solo 7) leyendo asignaciones ya presentes."""
        max_col = self.ws.max_column
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val == "1T":
                    self.contador_grupo_1t[trabajador] += 1
                elif val == "7":
                    self.contador_grupo_1t[trabajador] += 1
                    self.contador_grupo_6rt[trabajador] += 1

    def _marcar_alerta_restriccion_dura(self, col_dia: int, mensaje: str = "Bloqueado por restricción dura (BANTD/BLPTD/NLPTD/NANRD)") -> None:
        """Agrega un comentario en el encabezado del día para alertar restricción dura."""
        header_cell = self.ws.cell(row=1, column=col_dia)
        texto_existente = header_cell.comment.text if header_cell.comment else ""
        nuevo_texto = (texto_existente + "\n" if texto_existente else "") + mensaje
        header_cell.comment = Comment(nuevo_texto, "Asignador")

    def asignar_turno_en_dia(self, col_dia: int) -> Optional[str]:
        """Intenta asignar "1T" o "7" en el día (columna) indicado, retornando el trabajador o None."""
        turno = self._determinar_turno_por_personal(col_dia)
        if turno is None:
            return None

        # Verificar que ese día NO haya ya un 1T ni un 7 ni BLPTD ni BANTD
        if self._existe_turno_1t_o_7_en_dia(col_dia):
            return None

        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        if not disponibles:
            return None

        tenia_disponibles = bool(disponibles)

        # Excluir por restricción dura del día anterior
        disponibles = [t for t in disponibles if not self._tuvo_restriccion_dura_ayer(t, col_dia)]
        if not disponibles:
            if tenia_disponibles:
                self._marcar_alerta_restriccion_dura(col_dia, "Bloqueado por restricción dura (ayer: BANTD/BLPTD/NLPRD/NANRD/1T/7)")
            return None

        # Excluir por restricción dura del día siguiente
        disponibles = [t for t in disponibles if not self._tiene_restriccion_dura_manana(t, col_dia)]
        if not disponibles:
            if tenia_disponibles:
                self._marcar_alerta_restriccion_dura(col_dia, "Bloqueado por restricción dura (mañana: BANTD/BLPTD/1T/7)")
            return None

        # Restricción Torre para GCE cuando turno objetivo es 1T
        if turno == "1T":
            torre = self._obtener_conteo_torre(col_dia)
            if torre is not None and torre > 3 and "GCE" in disponibles:
                disponibles = [t for t in disponibles if t != "GCE"]
                if not disponibles:
                    return None

        # Prioridades con restricciones (blandas y prioridad DESC/TROP/SIND de ayer)
        nivel1 = []
        nivel2 = []
        nivel3 = []
        nivel4 = []
        nivel5 = []
        for t in disponibles:
            prioridad = self._tiene_prioridad_dia_anterior(t, col_dia)
            tuvo_extra = self._tuvo_extra_dia_anterior(t, col_dia)
            blanda = self._tuvo_restriccion_blanda_ayer(t, col_dia)
            if prioridad and not tuvo_extra and not blanda:
                nivel1.append(t)
            elif not tuvo_extra and not blanda:
                nivel2.append(t)
            elif prioridad and blanda:
                nivel3.append(t)
            elif blanda:
                nivel4.append(t)
            else:
                nivel5.append(t)

        for candidatos in (nivel1, nivel2, nivel3, nivel4, nivel5):
            elegido = self._seleccionar_equitativo(candidatos, turno)
            if elegido:
                fila = self._obtener_fila_trabajador(elegido)
                if not fila:
                    return None
                self.ws.cell(row=fila, column=col_dia, value=turno)
                self._actualizar_contadores(elegido, turno)
                return elegido

        return None

    def procesar_todos_los_dias(self) -> None:
        max_col = self.ws.max_column
        for col in range(2, max_col + 1):
            self.asignar_turno_en_dia(col)

        self._actualizar_hoja_estadisticas()

        salida = "horarioUnificado_con_1t.xlsx"
        self.wb.save(salida)
        print(f"Archivo guardado como: {salida}")

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")   # 1T + 7
        ws_stats.cell(row=1, column=4, value="6RT")  # Solo 7
        ws_stats.cell(row=1, column=5, value="1D")   # BANTD + BLPTD
        ws_stats.cell(row=1, column=6, value="3D")   # 3 + 3D
        ws_stats.cell(row=1, column=7, value="6D")   # NLPTD + NLPRD + NANTD + NANRD

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, 8):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        # Filas de trabajadores desde la hoja principal
        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP (fórmula dinámica)
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7 (fórmula dinámica)
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = solo 7 (fórmula dinámica)
            formula_6rt = f'=COUNTIF({hoja}!B{fila}:AE{fila},"7")'
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 1D = BANTD + BLPTD
            formula_1d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"BANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"BLPTD")'
            )
            ws_stats.cell(row=fila_destino, column=5, value=formula_1d)

            # 3D = 3 + 3D
            formula_3d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"3")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"3D")'
            )
            ws_stats.cell(row=fila_destino, column=6, value=formula_3d)

            # 6D = NLPTD + NLPRD + NANTD + NANRD
            formula_6d = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"NLPTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NLPRD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANTD")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"NANRD")'
            )
            ws_stats.cell(row=fila_destino, column=7, value=formula_6d)

            fila_destino += 1

        # Anchos de columna
        ws_stats.column_dimensions['A'].width = 10
        ws_stats.column_dimensions['B'].width = 8
        ws_stats.column_dimensions['C'].width = 8
        ws_stats.column_dimensions['D'].width = 8
        ws_stats.column_dimensions['E'].width = 8
        ws_stats.column_dimensions['F'].width = 8
        ws_stats.column_dimensions['G'].width = 8


if __name__ == "__main__":
    asignador = AsignadorTurnos()
    asignador.procesar_todos_los_dias() 