import openpyxl
import random
from collections import defaultdict
from openpyxl.styles import PatternFill, Font
from typing import List, Optional, Dict, Tuple, Set
import os


class AsignadorTurnosDiurnas:
    """
    Asigna turnos "6S" y "6N" con estas reglas:
    - Decisión por día según personal disponible/turnos operativos:
      * 9-10 personal → asignar "6N" a un trabajador y "6S" a otro trabajador
      * 11 personal → asignar solo "6S" a un trabajador (no se asigna "6N")
      * 12+ personal → NO asignar turnos en ese día
    - Solo se asigna entre trabajadores elegibles para turno 3: ['PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
      'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE']
    - No tiene restricciones blandas o duras
    - No asignar si ese día ya existe un turno "6S", "6N", "BLPTD" o "NANRD"
    - Equidad: los trabajadores deben tener la misma cantidad de turnos de cada tipo en lo posible
    - Balanceo para que la columna DIURNA (suma de 6S y 6N) tenga diferencia ≤1
    - Colores: rojo oscuro para 6S, rojo medio para 6N
    - Archivo de entrada: "horarioUnificado_con_3.xlsx"
    - Archivo de salida: "horarioUnificado_con_diurnas.xlsx"
    
    IMPORTANTE: Este módulo actualiza la fila de conteo operativo estático usando la misma
    lógica que procesador_horarios.py antes de realizar las asignaciones, asegurando que
    las decisiones se basen en conteos consistentes y actualizados.
    """

    TRABAJADORES_ELEGIBLES = [
        'PHD', 'HLG', 'MEI', 'VCM', 'ROP', 'ECE', 'WEH', 'DFB', 'MLS', 'FCE',
        'JBV', 'GMT', 'BRS', 'HZG', 'JIS', 'CDT', 'WGG', 'GCE'
    ]

    COLOR_6S = "8B0000"  # Rojo oscuro (DarkRed)
    COLOR_6N = "DC143C"  # Rojo medio (Crimson)

    def __init__(self, archivo_entrada: Optional[str] = None) -> None:
        candidatos = [
            archivo_entrada,
            "horarioUnificado_con_3.xlsx",
            "horarioUnificado_con_6r.xlsx",
            "horarioUnificado_procesado.xlsx",
        ]
        candidatos = [c for c in candidatos if c]
        elegido = None
        for c in candidatos:
            if os.path.exists(c):
                elegido = c
                break
        if not elegido:
            elegido = "horarioUnificado_con_3.xlsx"
        self.archivo_entrada = elegido

        self.wb = openpyxl.load_workbook(self.archivo_entrada)
        self.ws = self._obtener_hoja_horario()

        # Snapshot del estado original
        self.original_nonempty: Set[Tuple[int, int]] = set()
        self.original_6s: Set[Tuple[int, int]] = set()
        self.original_6n: Set[Tuple[int, int]] = set()
        self._snapshot_estado_original()

        # Contadores de equidad
        self.contador_6s: Dict[str, int] = defaultdict(int)
        self.contador_6n: Dict[str, int] = defaultdict(int)
        self.contador_diurna: Dict[str, int] = defaultdict(int)  # 6S + 6N

        random.seed()
        self._inicializar_contadores_desde_hoja()

    def _obtener_hoja_horario(self):
        for nombre in self.wb.sheetnames:
            if nombre != "Estadísticas":
                return self.wb[nombre]
        return self.wb.active

    def _nombre_hoja_horario(self) -> str:
        return self.ws.title

    def _snapshot_estado_original(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None or str(valor).strip() == "":
                    continue
                val = str(valor).strip().upper()
                self.original_nonempty.add((fila, col))
                if val == "6S":
                    self.original_6s.add((fila, col))
                elif val == "6N":
                    self.original_6n.add((fila, col))

    def _es_celda_originalmente_vacia(self, fila: int, col: int) -> bool:
        return (fila, col) not in self.original_nonempty

    def _obtener_fila_trabajador(self, trabajador: str) -> Optional[int]:
        for fila in range(2, 26):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip().upper() == trabajador.upper():
                return fila
        return None

    def _contar_personal_operativo(self, col_dia: int) -> int:
        """Cuenta el personal operativo usando la misma lógica que procesador_horarios.py"""
        # Definir turnos no operativos (misma lista que procesador_horarios.py)
        turnos_no_operativos = {
            # Turnos básicos
            "DESC", "TROP",
            # Turnos completos
            "VACA", "COME", "COMT", "COMS",
            # Turnos adicionales originales
            "SIND", "CMED", "CERT",
            # Formación, instrucción y entrenamiento
            "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
            "MENT", "TENT", "NENT", "AENT",
            "MINS", "TINS", "NINS", "AINS",
            # Gestión, oficinas y grupos de trabajo
            "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
            "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
            "MGST", "TGST", "MOFI", "TOFI",
            # Operativos y asignaciones especiales
            "CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X"
        }
        
        count = 0
        for fila in range(2, 26):  # Filas 2-25
            cell_value = self.ws.cell(row=fila, column=col_dia).value
            if cell_value is None or str(cell_value).strip() == "":
                count += 1  # Celda vacía = operativo
            else:
                valor_limpio = str(cell_value).strip().upper()
                if valor_limpio not in turnos_no_operativos:
                    count += 1  # No está en lista de no operativos = operativo
        return count

    def _existe_turno_conflictivo_en_dia(self, col_dia: int) -> bool:
        """Verifica si ya existe 6S, 6N, BLPTD o NANRD en el día"""
        for fila in range(2, 26):
            v = self.ws.cell(row=fila, column=col_dia).value
            if v is None:
                continue
            val = str(v).strip().upper()
            if val in {"6S", "6N", "BLPTD", "NANRD"}:
                return True
        return False

    def _obtener_trabajadores_disponibles(self, col_dia: int) -> List[str]:
        disponibles: List[str] = []
        for trabajador in self.TRABAJADORES_ELEGIBLES:
            fila = self._obtener_fila_trabajador(trabajador)
            if not fila:
                continue
            celda = self.ws.cell(row=fila, column=col_dia)
            if (
                (celda.value is None or str(celda.value).strip() == "")
                and self._es_celda_originalmente_vacia(fila, col_dia)
            ):
                disponibles.append(trabajador)
        return disponibles

    def _inicializar_contadores_desde_hoja(self) -> None:
        max_col = self.ws.max_column
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            trabajador = str(trabajador).strip().upper()
            for col in range(2, max_col + 1):
                valor = self.ws.cell(row=fila, column=col).value
                if valor is None:
                    continue
                val = str(valor).strip().upper()
                if val == "6S":
                    self.contador_6s[trabajador] += 1
                    self.contador_diurna[trabajador] += 1
                elif val == "6N":
                    self.contador_6n[trabajador] += 1
                    self.contador_diurna[trabajador] += 1

    def _seleccionar_equitativo_6s(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        # Equidad por menor conteo de 6S
        min_val = min(self.contador_6s[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_6s[c] == min_val]
        return random.choice(empatados)

    def _seleccionar_equitativo_6n(self, candidatos: List[str]) -> Optional[str]:
        if not candidatos:
            return None
        # Equidad por menor conteo de 6N
        min_val = min(self.contador_6n[c] for c in candidatos)
        empatados = [c for c in candidatos if self.contador_6n[c] == min_val]
        return random.choice(empatados)

    def _actualizar_contadores(self, trabajador: str, tipo_turno: str, delta: int = 1) -> None:
        if tipo_turno == "6S":
            self.contador_6s[trabajador] += delta
            self.contador_diurna[trabajador] += delta
        elif tipo_turno == "6N":
            self.contador_6n[trabajador] += delta
            self.contador_diurna[trabajador] += delta

    def _actualizar_fila_conteo_operativo(self) -> None:
        """Actualiza la fila de conteo operativo estático usando la misma lógica que procesador_horarios.py"""
        # Definir turnos no operativos (misma lista que procesador_horarios.py)
        turnos_no_operativos = {
            # Turnos básicos
            "DESC", "TROP",
            # Turnos completos
            "VACA", "COME", "COMT", "COMS",
            # Turnos adicionales originales
            "SIND", "CMED", "CERT",
            # Formación, instrucción y entrenamiento
            "CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
            "MENT", "TENT", "NENT", "AENT",
            "MINS", "TINS", "NINS", "AINS",
            # Gestión, oficinas y grupos de trabajo
            "MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
            "MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
            "MGST", "TGST", "MOFI", "TOFI",
            # Operativos y asignaciones especiales
            "CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X"
        }
        
        # Buscar la fila de conteo operativo estático
        fila_conteo = None
        for fila in range(1, self.ws.max_row + 1):
            valor = self.ws.cell(row=fila, column=1).value
            if valor and str(valor).strip() == "TURNOS OPERATIVOS":
                fila_conteo = fila
                break
        
        if fila_conteo is None:
            print("⚠️  No se encontró la fila 'TURNOS OPERATIVOS' para actualizar")
            return
        
        # Actualizar conteos para cada columna
        for col in range(2, self.ws.max_column + 1):
            conteo_operativos = 0
            # Contar turnos operativos según la lógica correcta
            for row in range(2, 26):  # Filas 2-25
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value is None or str(cell_value).strip() == "":
                    conteo_operativos += 1
                else:
                    valor_limpio = str(cell_value).strip().upper()
                    if valor_limpio not in turnos_no_operativos:
                        conteo_operativos += 1
            
            # Escribir el conteo actualizado
            celda_conteo = self.ws.cell(row=fila_conteo, column=col)
            celda_conteo.value = conteo_operativos
            
            # Aplicar colores según el conteo (misma lógica que procesador_horarios.py)
            rojo_intenso = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            rojo_medio = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
            azul_clarito = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
            verde_clarito = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            verde_intenso = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
            sin_relleno = PatternFill(fill_type=None)
            fuente_blanca = Font(color="FFFFFF")
            
            if conteo_operativos <= 8:
                celda_conteo.fill = rojo_intenso
                celda_conteo.font = fuente_blanca
            elif conteo_operativos == 9:
                celda_conteo.fill = rojo_medio
            elif conteo_operativos == 10:
                celda_conteo.fill = azul_clarito
            elif conteo_operativos == 11:
                celda_conteo.fill = verde_clarito
            elif conteo_operativos == 12:
                celda_conteo.fill = verde_intenso
            else:
                celda_conteo.fill = sin_relleno
        
        print("✅ Fila de conteo operativo estático actualizada")

    def _asignar_turno(self, trabajador: str, col_dia: int, tipo_turno: str) -> bool:
        """Asigna un turno específico a un trabajador"""
        fila = self._obtener_fila_trabajador(trabajador)
        if not fila:
            return False
        if not self._es_celda_originalmente_vacia(fila, col_dia):
            return False
        
        self.ws.cell(row=fila, column=col_dia, value=tipo_turno)
        celda = self.ws.cell(row=fila, column=col_dia)
        
        if tipo_turno == "6S":
            celda.fill = PatternFill(start_color=self.COLOR_6S, end_color=self.COLOR_6S, fill_type="solid")
        elif tipo_turno == "6N":
            celda.fill = PatternFill(start_color=self.COLOR_6N, end_color=self.COLOR_6N, fill_type="solid")
        
        self._actualizar_contadores(trabajador, tipo_turno, 1)
        return True

    def _puede_asignar_turnos(self, col_dia: int) -> Tuple[bool, int, int, str]:
        """
        Verifica si se pueden asignar turnos en un día.
        Retorna (puede_asignar, personal_operativo, disponibles_count, razon)
        """
        if self._existe_turno_conflictivo_en_dia(col_dia):
            return False, 0, 0, "Turno conflictivo"
            
        personal_operativo = self._contar_personal_operativo(col_dia)
        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        disponibles_count = len(disponibles)
        
        if personal_operativo < 9:
            return False, personal_operativo, disponibles_count, "Poco personal (<9)"
        elif personal_operativo >= 12:
            return False, personal_operativo, disponibles_count, "Mucho personal (≥12)"
        elif personal_operativo in [9, 10] and disponibles_count < 2:
            return False, personal_operativo, disponibles_count, "Pocos disponibles para 6N+6S"
        elif personal_operativo == 11 and disponibles_count < 1:
            return False, personal_operativo, disponibles_count, "Sin disponibles para 6S"
        else:
            return True, personal_operativo, disponibles_count, "OK"

    def asignar_turnos_en_dia(self, col_dia: int) -> Tuple[Optional[str], Optional[str]]:
        """
        Asigna turnos en un día según las reglas.
        Retorna (trabajador_6s, trabajador_6n) o (None, None) si no se asigna
        """
        puede_asignar, personal_operativo, _, _ = self._puede_asignar_turnos(col_dia)
        if not puede_asignar:
            return None, None

        disponibles = self._obtener_trabajadores_disponibles(col_dia)
        trabajador_6s = None
        trabajador_6n = None

        if personal_operativo >= 9 and personal_operativo <= 10:
            # Asignar 6N y 6S
            if len(disponibles) >= 2:
                # Seleccionar para 6N
                elegido_6n = self._seleccionar_equitativo_6n(disponibles)
                if elegido_6n and self._asignar_turno(elegido_6n, col_dia, "6N"):
                    trabajador_6n = elegido_6n
                    disponibles.remove(elegido_6n)

                # Seleccionar para 6S
                elegido_6s = self._seleccionar_equitativo_6s(disponibles)
                if elegido_6s and self._asignar_turno(elegido_6s, col_dia, "6S"):
                    trabajador_6s = elegido_6s

        elif personal_operativo == 11:
            # Asignar solo 6S
            elegido_6s = self._seleccionar_equitativo_6s(disponibles)
            if elegido_6s and self._asignar_turno(elegido_6s, col_dia, "6S"):
                trabajador_6s = elegido_6s

        return trabajador_6s, trabajador_6n

    def _rebalancear_para_paridad(self) -> None:
        """Rebalanceo moviendo turnos 6S y 6N para que DIURNA tenga diferencia ≤1"""
        max_iteraciones = 50
        iteracion = 0
        
        while iteracion < max_iteraciones:
            iteracion += 1
            
            # Calcular conteos actuales de DIURNA para trabajadores elegibles presentes
            conteos_actuales: Dict[str, int] = {}
            for t in self.TRABAJADORES_ELEGIBLES:
                if self._obtener_fila_trabajador(t):
                    conteos_actuales[t] = self.contador_diurna[t]
            
            if not conteos_actuales:
                break

            trabajador_max = max(conteos_actuales, key=conteos_actuales.get)
            trabajador_min = min(conteos_actuales, key=conteos_actuales.get)
            
            if conteos_actuales[trabajador_max] - conteos_actuales[trabajador_min] <= 1:
                break

            fila_max = self._obtener_fila_trabajador(trabajador_max)
            fila_min = self._obtener_fila_trabajador(trabajador_min)
            if not fila_max or not fila_min:
                break

            # Buscar una columna donde mover un turno
            movimiento_realizado = False
            for col in range(2, self.ws.max_column + 1):
                valor_max = self.ws.cell(row=fila_max, column=col).value
                valor_min = self.ws.cell(row=fila_min, column=col).value
                
                if (
                    valor_max is not None
                    and str(valor_max).strip().upper() in {"6S", "6N"}
                    and (valor_min is None or str(valor_min).strip() == "")
                    and self._es_celda_originalmente_vacia(fila_min, col)
                    and not self._existe_turno_conflictivo_en_dia(col)
                ):
                    tipo_turno = str(valor_max).strip().upper()
                    
                    # Remover del trabajador_max
                    self.ws.cell(row=fila_max, column=col, value=None)
                    celda_original = self.ws.cell(row=fila_max, column=col)
                    celda_original.fill = PatternFill(fill_type=None)
                    self._actualizar_contadores(trabajador_max, tipo_turno, -1)

                    # Asignar al trabajador_min
                    self.ws.cell(row=fila_min, column=col, value=tipo_turno)
                    celda_nueva = self.ws.cell(row=fila_min, column=col)
                    color = self.COLOR_6S if tipo_turno == "6S" else self.COLOR_6N
                    celda_nueva.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    self._actualizar_contadores(trabajador_min, tipo_turno, 1)

                    movimiento_realizado = True
                    break

            if not movimiento_realizado:
                break

    def _actualizar_hoja_estadisticas(self) -> None:
        nombre_stats = "Estadísticas"
        if nombre_stats in self.wb.sheetnames:
            ws_stats = self.wb[nombre_stats]
        else:
            ws_stats = self.wb.create_sheet(nombre_stats)

        # Verificar si ya existe la columna "3" (turnos 3)
        columna_3_existe = False
        for col in range(1, ws_stats.max_column + 1):
            header = ws_stats.cell(row=1, column=col).value
            if header == "3":
                columna_3_existe = True
                break

        # Limpiar hoja
        for fila in ws_stats.iter_rows():
            for celda in fila:
                celda.value = None
                celda.fill = PatternFill(fill_type=None)

        # Encabezados
        ws_stats.cell(row=1, column=1, value="SIGLA")
        ws_stats.cell(row=1, column=2, value="DESC")
        ws_stats.cell(row=1, column=3, value="1T")
        ws_stats.cell(row=1, column=4, value="6RT")
        ws_stats.cell(row=1, column=5, value="6T")
        
        # Mantener la columna "3" si existía
        if columna_3_existe:
            ws_stats.cell(row=1, column=6, value="3")
            ws_stats.cell(row=1, column=7, value="6S")
            ws_stats.cell(row=1, column=8, value="6N")
            ws_stats.cell(row=1, column=9, value="DIURNA")  # 6S + 6N
            num_columnas = 9
        else:
            ws_stats.cell(row=1, column=6, value="6S")
            ws_stats.cell(row=1, column=7, value="6N")
            ws_stats.cell(row=1, column=8, value="DIURNA")  # 6S + 6N
            num_columnas = 8

        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, num_columnas + 1):
            c = ws_stats.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        hoja = self._nombre_hoja_horario()
        fila_destino = 2
        for fila in range(2, 26):
            trabajador = self.ws.cell(row=fila, column=1).value
            if not trabajador:
                continue
            ws_stats.cell(row=fila_destino, column=1, value=trabajador)

            # DESC + TROP
            formula_desc = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"DESC")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"TROP")'
            )
            ws_stats.cell(row=fila_destino, column=2, value=formula_desc)

            # 1T = 1T + 7 + 1
            formula_1t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"1T")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"1")'
            )
            ws_stats.cell(row=fila_destino, column=3, value=formula_1t)

            # 6RT = 6RT + 7 + 6R
            formula_6rt = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6RT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"7")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6R")'
            )
            ws_stats.cell(row=fila_destino, column=4, value=formula_6rt)

            # 6T = 6TT + 6T
            formula_6t = (
                f'=COUNTIF({hoja}!B{fila}:AE{fila},"6TT")'
                f'+COUNTIF({hoja}!B{fila}:AE{fila},"6T")'
            )
            ws_stats.cell(row=fila_destino, column=5, value=formula_6t)

            # Columna "3" (turnos 3) si existía
            if columna_3_existe:
                formula_3 = f'=COUNTIF({hoja}!B{fila}:AE{fila},"3")'
                ws_stats.cell(row=fila_destino, column=6, value=formula_3)
                
                # 6S
                formula_6s = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                ws_stats.cell(row=fila_destino, column=7, value=formula_6s)

                # 6N
                formula_6n = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                ws_stats.cell(row=fila_destino, column=8, value=formula_6n)

                # DIURNA = 6S + 6N
                formula_diurna = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                )
                ws_stats.cell(row=fila_destino, column=9, value=formula_diurna)
            else:
                # 6S
                formula_6s = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                ws_stats.cell(row=fila_destino, column=6, value=formula_6s)

                # 6N
                formula_6n = f'=COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                ws_stats.cell(row=fila_destino, column=7, value=formula_6n)

                # DIURNA = 6S + 6N
                formula_diurna = (
                    f'=COUNTIF({hoja}!B{fila}:AE{fila},"6S")'
                    f'+COUNTIF({hoja}!B{fila}:AE{fila},"6N")'
                )
                ws_stats.cell(row=fila_destino, column=8, value=formula_diurna)

            fila_destino += 1

        # Ajustar anchos de columna
        if columna_3_existe:
            widths = [('A', 10), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 8), ('I', 10)]
        else:
            widths = [('A', 10), ('B', 8), ('C', 8), ('D', 8), ('E', 8), ('F', 8), ('G', 8), ('H', 10)]
        
        for col, width in widths:
            ws_stats.column_dimensions[col].width = width

    def _generar_reporte_detallado(self) -> None:
        """Genera un reporte detallado de disponibilidad y asignaciones por día"""
        print("\n" + "="*80)
        print("REPORTE DETALLADO DE ASIGNACIÓN DE TURNOS DIURNOS (6S y 6N)")
        print("="*80)
        
        max_col = self.ws.max_column
        asignaciones_realizadas = []
        dias_con_9_10_personal = []
        dias_con_11_personal = []
        dias_sin_asignar_12_mas = []
        dias_con_conflictos = []
        
        for col in range(2, max_col + 1):
            header = self.ws.cell(row=1, column=col).value
            if not header or header == "SIGLA ATCO":
                continue
                
            # Verificar si se puede asignar
            puede_asignar, personal_operativo, disponibles_count, razon = self._puede_asignar_turnos(col)
            
            if not puede_asignar:
                if razon == "Turno conflictivo":
                    dias_con_conflictos.append((header, personal_operativo, disponibles_count))
                elif personal_operativo in [9, 10]:
                    dias_con_9_10_personal.append((header, personal_operativo, disponibles_count, razon))
                elif personal_operativo == 11:
                    dias_con_11_personal.append((header, personal_operativo, disponibles_count, razon))
                else:
                    dias_sin_asignar_12_mas.append((header, personal_operativo, disponibles_count, razon))
            else:
                # Intentar asignar (solo si puede asignar)
                trabajador_6s, trabajador_6n = self.asignar_turnos_en_dia(col)
                if trabajador_6s or trabajador_6n:
                    asignaciones_realizadas.append((header, trabajador_6s, trabajador_6n, personal_operativo, disponibles_count))
                else:
                    # Si debería poder asignar pero no lo hizo, es un error
                    if personal_operativo in [9, 10]:
                        dias_con_9_10_personal.append((header, personal_operativo, disponibles_count, "Error en asignación"))
                    elif personal_operativo == 11:
                        dias_con_11_personal.append((header, personal_operativo, disponibles_count, "Error en asignación"))
        
        # Mostrar asignaciones realizadas
        print(f"\n🎯 ASIGNACIONES REALIZADAS: {len(asignaciones_realizadas)}")
        if asignaciones_realizadas:
            print("-" * 60)
            for header, t_6s, t_6n, personal, disponibles in asignaciones_realizadas:
                turnos = []
                if t_6s:
                    turnos.append(f"6S→{t_6s}")
                if t_6n:
                    turnos.append(f"6N→{t_6n}")
                print(f"{header:>8}: {', '.join(turnos):20} (Personal: {personal:2d}, Disponibles: {disponibles})")
        
        # Mostrar días con 9-10 personal que no se asignaron
        if dias_con_9_10_personal:
            print(f"\n⚠️  DÍAS CON 9-10 PERSONAL SIN ASIGNAR: {len(dias_con_9_10_personal)}")
            print("-" * 60)
            for item in dias_con_9_10_personal:
                if len(item) == 4:
                    header, personal, disponibles, razon = item
                else:
                    header, personal, disponibles = item
                    razon = "Pocos disponibles" if disponibles < 2 else "Error desconocido"
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles} - {razon}")
        
        # Mostrar días con 11 personal que no se asignaron
        if dias_con_11_personal:
            print(f"\n⚠️  DÍAS CON 11 PERSONAL SIN ASIGNAR: {len(dias_con_11_personal)}")
            print("-" * 60)
            for item in dias_con_11_personal:
                if len(item) == 4:
                    header, personal, disponibles, razon = item
                else:
                    header, personal, disponibles = item
                    razon = "Sin disponibles" if disponibles < 1 else "Error desconocido"
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles} - {razon}")
        
        # Mostrar días con 12+ personal (no se asigna por reglas)
        if dias_sin_asignar_12_mas:
            print(f"\n📊 DÍAS SIN ASIGNAR (12+ PERSONAL): {len(dias_sin_asignar_12_mas)}")
            print("-" * 60)
            for item in dias_sin_asignar_12_mas:
                if len(item) == 4:
                    header, personal, disponibles, razon = item
                else:
                    header, personal, disponibles = item
                    razon = "No asignar por reglas"
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles} - {razon}")
        
        # Mostrar días con conflictos
        if dias_con_conflictos:
            print(f"\n🚫 DÍAS CON CONFLICTOS (6S/6N/BLPTD/NANRD): {len(dias_con_conflictos)}")
            print("-" * 60)
            for header, personal, disponibles in dias_con_conflictos:
                print(f"{header:>8}: Personal: {personal:2d}, Disponibles: {disponibles} - Turno conflictivo")
        
        # Resumen de personal por rangos
        print(f"\n📈 RESUMEN POR RANGOS DE PERSONAL:")
        print("-" * 60)
        rango_9_10 = len(dias_con_9_10_personal) + len([a for a in asignaciones_realizadas if 9 <= a[3] <= 10])
        rango_11 = len(dias_con_11_personal) + len([a for a in asignaciones_realizadas if a[3] == 11])
        rango_12_mas = len(dias_sin_asignar_12_mas)
        conflictos = len(dias_con_conflictos)
        
        print(f"Días con 9-10 personal: {rango_9_10:2d} (deberían asignar 6N+6S)")
        print(f"Días con 11 personal:   {rango_11:2d} (deberían asignar solo 6S)")
        print(f"Días con 12+ personal:  {rango_12_mas:2d} (no asignar por reglas)")
        print(f"Días con conflictos:    {conflictos:2d} (no asignar por conflictos)")
        
        print("="*80)
        return len(asignaciones_realizadas)

    def procesar_todos_los_dias(self) -> None:
        # Actualizar la fila de conteo operativo estático antes de asignar
        print("🔄 Actualizando fila de conteo operativo estático...")
        self._actualizar_fila_conteo_operativo()
        
        # Generar reporte detallado y obtener número de asignaciones
        num_asignaciones = self._generar_reporte_detallado()
        
        # Rebalancear para paridad
        self._rebalancear_para_paridad()

        # Actualizar estadísticas
        self._actualizar_hoja_estadisticas()

        # Guardar archivo
        salida = "horarioUnificado_con_diurnas.xlsx"
        try:
            self.wb.save(salida)
            print(f"\n✅ Archivo guardado como: {salida}")
        except PermissionError:
            base, ext = os.path.splitext(salida)
            alternativo = f"{base}_{random.randint(1000,9999)}{ext}"
            self.wb.save(alternativo)
            print(f"\n✅ Archivo por defecto en uso. Guardado como: {alternativo}")


if __name__ == "__main__":
    asignador = AsignadorTurnosDiurnas()
    asignador.procesar_todos_los_dias() 