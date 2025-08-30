import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import copy

def quitar_division_columna():
    """
    Convierte el archivo excel_con_division_de_columna.xlsx de vuelta al formato original:
    1. Combina las dos columnas de cada día en una sola
    2. Aplica las reglas inversas de renombrado:
       - TLPT/NLPT → 6TT
       - MLPR/NLPR → 6RT
       - TANT/NANT → 6T
       - MAST/NANR → 6R
       - MANR/TANR → 6N
       - MASR/TASR → 6S
       - MLPR/TLPR → 6MT
       - TAST/HXN4 → 3
       - BLPT/NLPR → 7
       - BLPT → 1T
       - BANT → 1
    3. Conserva los colores originales
    """
    
    # Cargar el archivo con división de columnas
    wb = openpyxl.load_workbook('excel_con_division_de_columna.xlsx')
    ws = wb.active
    
    print(f"Procesando archivo: {wb.active.title}")
    print(f"Dimensiones con división: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Crear nuevo workbook
    nuevo_wb = openpyxl.Workbook()
    nuevo_ws = nuevo_wb.active
    nuevo_ws.title = ws.title
    
    # Copiar la primera columna (trabajadores) sin cambios
    for fila in range(1, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        nuevo_ws.cell(row=fila, column=1, value=valor)
    
    # Procesar columnas de días (desde la columna 2 en adelante, de dos en dos)
    nueva_col = 2
    for col_primera in range(2, ws.max_column, 2):  # Avanzar de 2 en 2
        col_segunda = col_primera + 1
        
        # Obtener el día del encabezado (está en la primera columna del par)
        dia = ws.cell(row=1, column=col_primera).value
        
        # Crear una sola columna para el día
        nuevo_ws.cell(row=1, column=nueva_col, value=dia)
        
        # Procesar cada fila para este día
        for fila in range(2, ws.max_row + 1):
            valor_primera = ws.cell(row=fila, column=col_primera).value
            valor_segunda = ws.cell(row=fila, column=col_segunda).value
            
            # Determinar el turno original basado en los valores de ambas columnas
            turno_original = determinar_turno_original(valor_primera, valor_segunda)
            
            if turno_original:
                # Crear la celda con el turno original
                celda_nueva = nuevo_ws.cell(row=fila, column=nueva_col, value=turno_original)
                
                # Aplicar el color de la primera columna (si existe)
                color_primera = ws.cell(row=fila, column=col_primera).fill
                if color_primera.start_color.rgb:
                    celda_nueva.fill = PatternFill(
                        start_color=color_primera.start_color.rgb,
                        end_color=color_primera.end_color.rgb,
                        fill_type=color_primera.fill_type
                    )
        
        nueva_col += 1  # Avanzar una columna para el siguiente día
    
    # Ajustar ancho de columnas
    for col in range(1, nuevo_ws.max_column + 1):
        nuevo_ws.column_dimensions[get_column_letter(col)].width = 8
    
    # Guardar el archivo convertido
    nombre_archivo_salida = "conversion_inversa_a_una_sola_columna.xlsx"
    nuevo_wb.save(nombre_archivo_salida)
    
    print(f"\nArchivo convertido guardado como: {nombre_archivo_salida}")
    print(f"Nuevas dimensiones: {nuevo_ws.max_row} filas x {nuevo_ws.max_column} columnas")
    
    # Mostrar resumen de los cambios
    print("\nResumen de conversión inversa realizada:")
    print("- Se combinaron las dos columnas de cada día en una sola")
    print("- Se aplicaron las reglas inversas de renombrado:")
    print("  * TLPT/NLPT → 6TT")
    print("  * MLPR/NLPR → 6RT")
    print("  * TANT/NANT → 6T")
    print("  * MAST/NANR → 6R")
    print("  * MANR/TANR → 6N")
    print("  * MASR/TASR → 6S")
    print("  * MLPR/TLPR → 6MT")
    print("  * TAST/HXN4 → 3")
    print("  * BLPT/NLPR → 7")
    print("  * BLPT → 1T")
    print("  * BANT → 1")
    print("  * r → r (sin cambios)")
    print("- Se conservaron los colores originales de los turnos")
    
    # Verificar que la conversión fue exitosa
    verificar_conversion_exitosa()

def determinar_turno_original(valor_primera, valor_segunda):
    """
    Determina el turno original basado en los valores de las dos columnas divididas
    """
    if not valor_primera:
        return None
    
    primera = str(valor_primera).strip()
    segunda = str(valor_segunda).strip() if valor_segunda else ""
    
    # Reglas inversas de renombrado
    if primera == "TLPT" and segunda == "NLPT":
        return "6TT"
    elif primera == "MLPR" and segunda == "NLPR":
        return "6RT"
    elif primera == "TANT" and segunda == "NANT":
        return "6T"
    elif primera == "MAST" and segunda == "NANR":
        return "6R"
    elif primera == "MANR" and segunda == "TANR":
        return "6N"
    elif primera == "MASR" and segunda == "TASR":
        return "6S"
    elif primera == "MLPR" and segunda == "TLPR":
        return "6MT"
    elif primera == "TAST" and segunda == "HXN4":
        return "3"
    elif primera == "BLPT" and segunda == "NLPR":
        return "7"
    elif primera == "BLPT" and not segunda:
        return "1T"
    elif primera == "BANT" and not segunda:
        return "1"
    elif primera == "r" and not segunda:
        return "r"
    else:
        # Si no coincide con ninguna regla, devolver el valor de la primera columna
        return valor_primera

def mostrar_estructura_archivo_dividido():
    """
    Muestra la estructura del archivo con división de columnas para referencia
    """
    try:
        wb = openpyxl.load_workbook('excel_con_division_de_columna.xlsx')
        ws = wb.active
        
        print(f"Estructura del archivo con división:")
        print(f"Archivo: excel_con_division_de_columna.xlsx")
        print(f"Hoja: {ws.title}")
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
        
        # Mostrar encabezados
        print("\nEncabezados (primera fila):")
        for col in range(1, min(ws.max_column + 1, 15)):  # Primeras 15 columnas
            valor = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {repr(valor)}")
        
        # Mostrar algunos trabajadores
        print("\nPrimeros trabajadores:")
        for fila in range(2, min(ws.max_row + 1, 8)):  # Primeros 6 trabajadores
            trabajador = ws.cell(row=fila, column=1).value
            print(f"  Fila {fila}: {trabajador}")
        
        # Buscar turnos específicos en las primeras columnas
        print("\nBuscando turnos específicos en primeras columnas:")
        turnos_encontrados = set()
        for fila in range(2, ws.max_row + 1):
            for col in range(2, min(ws.max_column + 1, 20), 2):  # Solo primeras columnas de cada día
                valor = ws.cell(row=fila, column=col).value
                if valor and str(valor).strip() in ["TLPT", "MLPR", "TANT", "MAST", "MANR", "MASR", "TAST", "BLPT", "BANT", "r"]:
                    turnos_encontrados.add(str(valor).strip())
        
        print(f"Turnos encontrados: {sorted(turnos_encontrados)}")
        
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

def verificar_conversion_exitosa():
    """
    Compara el archivo original con el archivo convertido para verificar que la conversión fue exitosa
    """
    print("\n" + "="*50)
    print("VERIFICANDO CONVERSIÓN")
    print("="*50)
    
    try:
        # Cargar archivo original
        wb_original = openpyxl.load_workbook('horarioUnificado_con_6t.xlsx')
        ws_original = wb_original.active
        
        # Cargar archivo convertido
        wb_convertido = openpyxl.load_workbook('conversion_inversa_a_una_sola_columna.xlsx')
        ws_convertido = wb_convertido.active
        
        print(f"Archivo original: {ws_original.max_row} filas x {ws_original.max_column} columnas")
        print(f"Archivo convertido: {ws_convertido.max_row} filas x {ws_convertido.max_column} columnas")
        
        # Verificar dimensiones
        dimensiones_coinciden = (ws_original.max_row == ws_convertido.max_row and 
                                ws_original.max_column == ws_convertido.max_column)
        
        print(f"✅ Dimensiones coinciden: {dimensiones_coinciden}")
        
        # Verificar encabezados
        encabezados_coinciden = True
        diferencias_encabezados = []
        
        for col in range(1, ws_original.max_column + 1):
            valor_original = ws_original.cell(row=1, column=col).value
            valor_convertido = ws_convertido.cell(row=1, column=col).value
            
            if valor_original != valor_convertido:
                encabezados_coinciden = False
                diferencias_encabezados.append(f"Col {col}: '{valor_original}' vs '{valor_convertido}'")
        
        print(f"✅ Encabezados coinciden: {encabezados_coinciden}")
        if diferencias_encabezados:
            print("   Diferencias en encabezados:")
            for diff in diferencias_encabezados[:5]:  # Mostrar solo las primeras 5
                print(f"     {diff}")
        
        # Verificar trabajadores (primera columna)
        trabajadores_coinciden = True
        diferencias_trabajadores = []
        
        for fila in range(2, ws_original.max_row + 1):
            valor_original = ws_original.cell(row=fila, column=1).value
            valor_convertido = ws_convertido.cell(row=fila, column=1).value
            
            if valor_original != valor_convertido:
                trabajadores_coinciden = False
                diferencias_trabajadores.append(f"Fila {fila}: '{valor_original}' vs '{valor_convertido}'")
        
        print(f"✅ Trabajadores coinciden: {trabajadores_coinciden}")
        if diferencias_trabajadores:
            print("   Diferencias en trabajadores:")
            for diff in diferencias_trabajadores[:5]:
                print(f"     {diff}")
        
        # Verificar turnos
        turnos_coinciden = True
        diferencias_turnos = []
        total_celdas_comparadas = 0
        
        for fila in range(2, ws_original.max_row + 1):
            for col in range(2, ws_original.max_column + 1):
                valor_original = ws_original.cell(row=fila, column=col).value
                valor_convertido = ws_convertido.cell(row=fila, column=col).value
                
                total_celdas_comparadas += 1
                
                # Normalizar valores para comparación
                val_orig = str(valor_original).strip() if valor_original else ""
                val_conv = str(valor_convertido).strip() if valor_convertido else ""
                
                if val_orig != val_conv:
                    turnos_coinciden = False
                    trabajador = ws_original.cell(row=fila, column=1).value
                    dia = ws_original.cell(row=1, column=col).value
                    diferencias_turnos.append(f"{trabajador} en {dia}: '{val_orig}' vs '{val_conv}'")
        
        print(f"✅ Turnos coinciden: {turnos_coinciden}")
        print(f"   Total de celdas comparadas: {total_celdas_comparadas}")
        
        if diferencias_turnos:
            print(f"   Diferencias encontradas: {len(diferencias_turnos)}")
            print("   Primeras diferencias:")
            for diff in diferencias_turnos[:10]:  # Mostrar solo las primeras 10
                print(f"     {diff}")
        
        # Resumen final
        print("\n" + "-"*50)
        if dimensiones_coinciden and encabezados_coinciden and trabajadores_coinciden and turnos_coinciden:
            print("🎉 CONVERSIÓN EXITOSA: Los archivos son idénticos!")
        else:
            print("⚠️  CONVERSIÓN PARCIAL: Se encontraron algunas diferencias")
            print("   Revisa los detalles arriba para más información")
        
        # Estadísticas de turnos
        print("\nEstadísticas de turnos encontrados:")
        turnos_originales = set()
        turnos_convertidos = set()
        
        for fila in range(2, ws_original.max_row + 1):
            for col in range(2, ws_original.max_column + 1):
                valor_orig = ws_original.cell(row=fila, column=col).value
                valor_conv = ws_convertido.cell(row=fila, column=col).value
                
                if valor_orig:
                    turnos_originales.add(str(valor_orig).strip())
                if valor_conv:
                    turnos_convertidos.add(str(valor_conv).strip())
        
        print(f"Turnos en archivo original: {sorted(turnos_originales)}")
        print(f"Turnos en archivo convertido: {sorted(turnos_convertidos)}")
        
    except Exception as e:
        print(f"❌ Error durante la verificación: {e}")

if __name__ == "__main__":
    # Mostrar estructura del archivo con división
    mostrar_estructura_archivo_dividido()
    
    print("\n" + "="*50)
    print("INICIANDO CONVERSIÓN INVERSA")
    print("="*50)
    
    # Ejecutar la conversión inversa
    quitar_division_columna() 