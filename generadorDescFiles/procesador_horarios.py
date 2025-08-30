import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os

def procesar_horarios():
	"""
	Procesa el archivo horarioUnificado.xlsx para contar turnos operativos
	usando valores calculados y aplicar formato de colores según especificaciones.
	"""
	
	# Definir turnos no operativos
	turnos_no_operativos = {
		# Turnos básicos
		"DESC", "TROP",
		# Turnos completos
		"VACA", "COME", "COMT", "COMS",
		# Turnos adicionales originales
		"SIND", "CMED", "CERT",
		# Formación, instrucción y entrenamiento
		"CAPA", "MCAE", "TCAE", "MCHC", "TCHC", "NCHC", "ACHC",
		"MENT", "TENT", "NENT", "AENT",
		"MINS", "TINS", "NINS", "AINS",
		# Gestión, oficinas y grupos de trabajo
		"MCOR", "TCOR", "MSMS", "TSMS", "MDBM", "TDBM",
		"MDOC", "TDOC", "MPRO", "TPRO", "MATF", "TATF",
		"MGST", "TGST", "MOFI", "TOFI",
		# Operativos y asignaciones especiales
		"CET", "ATC", "KATC", "XATC", "YATC", "ZATC", "X"
	}
	
	# Cargar el archivo Excel
	try:
		wb = openpyxl.load_workbook('horioUnificado.xlsx')
		ws = wb.active
		print("Archivo cargado exitosamente")
	except FileNotFoundError:
		print("Error: No se encontró el archivo 'horioUnificado.xlsx'")
		return
	except Exception as e:
		print(f"Error al cargar el archivo: {e}")
		return
	
	# Obtener dimensiones de la hoja
	max_row = ws.max_row
	max_col = ws.max_column
	
	print(f"Dimensiones del archivo: {max_row} filas, {max_col} columnas")
	
	# Definir colores según especificaciones
	rojo_intenso = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")		# ≤8
	rojo_medio = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")		# =9 (también para 'Torre' >4)
	azul_clarito = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")		# =10
	verde_clarito = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")		# =11
	verde_intenso = PatternFill(start_color="008000", end_color="008000", fill_type="solid")		# =12
	sin_relleno = PatternFill(fill_type=None)														# ≥13
	rojo_claro_encabezado = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")	# Solo encabezado domingo
	amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")			# Turnos no operativos
	
	# Definir fuentes
	fuente_blanca = Font(color="FFFFFF")  # Para rojo intenso
	
	# Limpiar todo el formato existente antes de aplicar nuevos colores
	print("Limpiando formato existente...")
	for row in range(1, max_row + 1):
		for col in range(1, max_col + 1):
			cell = ws.cell(row=row, column=col)
			# Limpiar solo el relleno, preservar el valor
			cell.fill = PatternFill(fill_type=None)
	
	# Fijar filas para nuevos conteos (dinámicos y estáticos)
	fila_dinamico_torre = max_row + 1
	fila_dinamico_operativos = max_row + 2
	fila_conteo = max_row + 3		# estático operativos (penúltima)
	fila_torre = max_row + 4		# estático torre (última)
	
	# Etiquetas de filas
	ws.cell(row=fila_dinamico_torre, column=1, value="TORRE (DIN)")
	ws.cell(row=fila_dinamico_operativos, column=1, value="TURNOS OPERATIVOS (DIN)")
	ws.cell(row=fila_conteo, column=1, value="TURNOS OPERATIVOS")
	ws.cell(row=fila_torre, column=1, value="Torre")
	
	# Calcular y aplicar conteos estáticos para cada columna (operativos)
	for col in range(2, max_col + 1):
		conteo_operativos = 0
		# Contar turnos operativos según la lógica correcta
		for row in range(2, 26):  # Filas 2-25
			cell_value = ws.cell(row=row, column=col).value
			if cell_value is None or str(cell_value).strip() == "":
				conteo_operativos += 1
			else:
				valor_limpio = str(cell_value).strip().upper()
				if valor_limpio not in turnos_no_operativos:
					conteo_operativos += 1
		# Escribir estático
		celda_conteo_estatico = ws.cell(row=fila_conteo, column=col)
		celda_conteo_estatico.value = conteo_operativos
		# Color estático
		if conteo_operativos <= 8:
			celda_conteo_estatico.fill = rojo_intenso
			celda_conteo_estatico.font = fuente_blanca
		elif conteo_operativos == 9:
			celda_conteo_estatico.fill = rojo_medio
		elif conteo_operativos == 10:
			celda_conteo_estatico.fill = azul_clarito
		elif conteo_operativos == 11:
			celda_conteo_estatico.fill = verde_clarito
		elif conteo_operativos == 12:
			celda_conteo_estatico.fill = verde_intenso
		else:
			celda_conteo_estatico.fill = sin_relleno
	
	# Agregar fila 'Torre' estático (subconjunto de siglas)
	siglas_torre = {"YIS", "MAQ", "DJO", "AFG", "JLF", "JMV"}
	sigla_a_fila = {}
	for r in range(2, min(26, max_row + 1)):
		sigla = ws.cell(row=r, column=1).value
		if isinstance(sigla, str):
			sigla_limpia = sigla.strip().upper()
			if sigla_limpia:
				sigla_a_fila[sigla_limpia] = r
	filas_objetivo = [sigla_a_fila[s] for s in siglas_torre if s in sigla_a_fila]
	
	for col in range(2, max_col + 1):
		conteo_torre = 0
		for r in filas_objetivo:
			v = ws.cell(row=r, column=col).value
			if v is None or str(v).strip() == "":
				conteo_torre += 1
			else:
				if str(v).strip().upper() not in turnos_no_operativos:
					conteo_torre += 1
		celda_torre_estatico = ws.cell(row=fila_torre, column=col, value=conteo_torre)
		if conteo_torre > 4:
			celda_torre_estatico.fill = rojo_medio
		else:
			celda_torre_estatico.fill = sin_relleno
	
	# Añadir fórmulas dinámicas (Solución 1)
	turnos_list = sorted(list(turnos_no_operativos))
	for col in range(2, max_col + 1):
		col_letra = get_column_letter(col)
		rango = f"{col_letra}2:{col_letra}25"
		# Operativos dinámico (columna completa B2:B25)
		sustracciones_rango = "".join([f"-COUNTIF({rango},\"{t}\")" for t in turnos_list])
		formula_oper_din = f"=COUNTBLANK({rango})+COUNTIF({rango},\"<>\"){sustracciones_rango}"
		ws.cell(row=fila_dinamico_operativos, column=col, value=formula_oper_din)
		# Torre dinámico (suma de 6 celdas)
		terminos_torre = []
		for r in filas_objetivo:
			celda = f"{col_letra}{r}"
			sustracciones_celda = "".join([f"-COUNTIF({celda},\"{t}\")" for t in turnos_list])
			terminos_torre.append(f"COUNTBLANK({celda})+COUNTIF({celda},\"<>\"){sustracciones_celda}")
		formula_torre_din = "=" + "+".join(terminos_torre) if terminos_torre else "=0"
		ws.cell(row=fila_dinamico_torre, column=col, value=formula_torre_din)
	
	# Colorear celdas con turnos no operativos de amarillo (solo los que están en la lista)
	for col in range(2, max_col + 1):
		for row in range(2, min(26, max_row + 1)):
			cell = ws.cell(row=row, column=col)
			cell_value = cell.value
			# Solo colorear de amarillo si el valor está en la lista de turnos no operativos
			if cell_value is not None and str(cell_value).strip() != "":
				valor_limpio = str(cell_value).strip().upper()
				if valor_limpio in turnos_no_operativos:
					cell.fill = amarillo
	
	# Colorear SOLO el encabezado de domingos de rojo claro (no todas las celdas)
	for col in range(2, max_col + 1):
		header_cell = ws.cell(row=1, column=col)
		header_value = header_cell.value
		if header_value and "SUN" in str(header_value).upper():
			header_cell.fill = rojo_claro_encabezado
	
	# Crear nueva hoja de estadísticas
	print("Creando hoja de estadísticas...")
	if "Estadísticas" in wb.sheetnames:
		ws_stats = wb["Estadísticas"]
	else:
		ws_stats = wb.create_sheet("Estadísticas")
	# Limpiar la hoja de estadísticas
	for row in ws_stats.iter_rows():
		for cell in row:
			cell.value = None
			cell.fill = PatternFill(fill_type=None)
	# Crear encabezados
	ws_stats.cell(row=1, column=1, value="SIGLA")
	ws_stats.cell(row=1, column=2, value="DESC")
	# Nombres de trabajadores
	trabajadores = []
	for row in range(2, 26):
		nombre_trabajador = ws.cell(row=row, column=1).value
		if nombre_trabajador:
			trabajadores.append(nombre_trabajador)
	# Filas estadísticas
	for i, trabajador in enumerate(trabajadores, start=2):
		ws_stats.cell(row=i, column=1, value=trabajador)
		formula_desc_trop = f'=COUNTIF(HorarioUnificado!B{i}:AC{i},"DESC")+COUNTIF(HorarioUnificado!B{i}:AC{i},"TROP")'
		ws_stats.cell(row=i, column=2, value=formula_desc_trop)
	# Formato encabezados
	header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
	header_font = Font(bold=True)
	for col in range(1, 3):
		cell = ws_stats.cell(row=1, column=col)
		cell.fill = header_fill
		cell.font = header_font
	# Anchos
	ws_stats.column_dimensions['A'].width = 8
	ws_stats.column_dimensions['B'].width = 6
	
	# Guardar el archivo procesado
	nombre_archivo_salida = "horarioUnificado_procesado.xlsx"
	wb.save(nombre_archivo_salida)
	
	print(f"Archivo procesado guardado como: {nombre_archivo_salida}")
	print("Resumen del procesamiento:")
	print("- Se limpió todo el formato de color existente")
	print("- Se agregaron filas dinámicas: 'TORRE (DIN)' y 'TURNOS OPERATIVOS (DIN)'")
	print("- Se mantuvieron los conteos estáticos en las dos últimas filas: 'TURNOS OPERATIVOS' y 'Torre'")
	print("- Se aplicaron valores calculados y colores a los conteos estáticos")
	print("- Se colorearon de amarillo SOLO las celdas con turnos no operativos de la lista")
	print("- Se colorearon de rojo claro SOLO los encabezados de domingos")
	print(f"- Turnos no operativos reconocidos: {len(turnos_no_operativos)} tipos")
	print("- Se creó nueva hoja 'Estadísticas' con SIGLA y conteo unificado DESC+TROP")
	print("\nNota: Las filas dinámicas usan fórmulas de Excel y se actualizan automáticamente.")
	print("Los conteos estáticos siguen presentes en las dos últimas filas.")
	print("\n" + "="*60)
	print("*** VERIFICACIONES REQUERIDAS ***")
	print("="*60)
	print("• Verificar que NO haya más de 4 disponibles en Torre")
	print("• En días con 9 operativos, verificar que NO haya más de 3 disponibles en Torre")
	print("• El orden es procesar horarios, generar sábados")
	print("•, luego 1T/7, 6RT/6tt, 6TT.")
	print("="*60)

if __name__ == "__main__":
	procesar_horarios() 