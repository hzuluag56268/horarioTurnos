#!/usr/bin/env python3
"""
Test Manual del Excel
"""
import openpyxl
import pandas as pd

def test_excel_manual():
    """
    Prueba manual del archivo Excel
    """
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook('TURNOS_FECHAS_ESPECIFICAS.xlsx')
        print("✅ Archivo Excel cargado exitosamente")
        print(f"📊 Hojas disponibles: {wb.sheetnames}")
        
        # Verificar hoja principal
        ws = wb['Turnos Específicos']
        print(f"\n📋 Hoja principal: {ws.title}")
        print(f"   Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
        
        # Verificar encabezados
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
        print(f"   Encabezados: {headers}")
        
        # Verificar datos de ejemplo
        print("\n📝 Datos de ejemplo:")
        for row in range(2, 6):
            empleado = ws.cell(row=row, column=1).value
            turno = ws.cell(row=row, column=2).value
            fecha_inicio = ws.cell(row=row, column=3).value
            fecha_fin = ws.cell(row=row, column=4).value
            if empleado:
                print(f"   {empleado}: {turno} desde {fecha_inicio} hasta {fecha_fin}")
        
        # Verificar hojas auxiliares
        print("\n📋 Hojas auxiliares:")
        if 'Lista_Empleados' in wb.sheetnames:
            ws_emp = wb['Lista_Empleados']
            empleados = []
            for row in range(2, ws_emp.max_row + 1):
                emp = ws_emp.cell(row=row, column=1).value
                if emp:
                    empleados.append(emp)
            print(f"   Empleados ({len(empleados)}): {empleados[:5]}...")
            
        if 'Lista_Turnos' in wb.sheetnames:
            ws_turnos = wb['Lista_Turnos']
            turnos = []
            for row in range(2, ws_turnos.max_row + 1):
                turno = ws_turnos.cell(row=row, column=1).value
                if turno:
                    turnos.append(turno)
            print(f"   Turnos ({len(turnos)}): {turnos}")
        
        # Verificar validaciones (método simplificado)
        print(f"\n🔍 Validaciones de datos: {len(ws.data_validations)} encontradas")
        
        if len(ws.data_validations) > 0:
            print("✅ Las validaciones han sido aplicadas al archivo")
        else:
            print("❌ No se encontraron validaciones")
            
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_con_pandas():
    """
    Prueba usando pandas
    """
    try:
        print("\n🐼 Prueba con pandas:")
        df = pd.read_excel('TURNOS_FECHAS_ESPECIFICAS.xlsx', sheet_name='Turnos Específicos')
        print(f"   Dimensiones: {df.shape}")
        print(f"   Columnas: {df.columns.tolist()}")
        
        # Mostrar datos no vacíos
        df_no_vacio = df.dropna(subset=['Empleado'])
        print(f"   Filas con datos: {len(df_no_vacio)}")
        
        if len(df_no_vacio) > 0:
            print("   Primeras filas con datos:")
            for _, row in df_no_vacio.head().iterrows():
                print(f"     {row['Empleado']}: {row['Turno Requerido']} - {row['Fecha Inicio']}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error con pandas: {e}")
        return False

if __name__ == "__main__":
    print("🔧 PRUEBA MANUAL DEL ARCHIVO EXCEL")
    print("=" * 50)
    
    test_excel_manual()
    test_con_pandas()
    
    print("\n💡 INSTRUCCIONES PARA PROBAR MANUALMENTE:")
    print("1. Abra el archivo TURNOS_FECHAS_ESPECIFICAS.xlsx en Excel")
    print("2. Haga clic en cualquier celda de la columna 'Empleado'")
    print("3. Debe ver una flecha pequeña a la derecha de la celda")
    print("4. Haga clic en la flecha para ver el dropdown")
    print("5. Si no ve la flecha, las validaciones no están funcionando")
    print("6. Repita el proceso para la columna 'Turno Requerido'")
    print("\nSi los dropdowns no funcionan, puede que Excel no soporte")
    print("las validaciones creadas con openpyxl en su versión.") 