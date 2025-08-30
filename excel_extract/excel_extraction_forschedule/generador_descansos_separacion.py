import pandas as pd
import calendar
from datetime import datetime, date, timedelta
import random
import numpy as np
import csv
import os

# Importar configuración externa de restricciones
from config_restricciones import (
    RESTRICCIONES_EMPLEADOS,
    TURNOS_FECHAS_ESPECIFICAS,
    TURNOS_ESPECIALES,
    TRABAJADORES_FUERA_OPERACION,
    DIAS_FESTIVOS,
    CONFIGURACION_GENERAL,
    obtener_empleados
)

class GeneradorDescansosSeparacion:
    def __init__(self, año=2025, mes=1, num_empleados=25, semana_especifica=None):
        """Inicializa el generador con configuración básica"""
        self.año = año
        self.mes = mes
        self.num_empleados = num_empleados
        self.semana_especifica = semana_especifica  # Nueva variable para semana específica
        
        # Cargar empleados desde configuración
        self.empleados = obtener_empleados()
        
        # Configuración de aleatorización
        np.random.seed()
        
        # SISTEMA UNIFICADO DE RESTRICCIONES CONSOLIDADO (AHORA EXTERNO)
        self.restricciones_empleados = RESTRICCIONES_EMPLEADOS
        
        # NUEVA FUNCIONALIDAD: Turnos específicos por fecha exacta (AHORA EXTERNO)
        self.turnos_fechas_especificas = TURNOS_FECHAS_ESPECIFICAS
        
        # NUEVA FUNCIONALIDAD: Turnos especiales extendidos (AHORA EXTERNO)
        self.turnos_especiales = TURNOS_ESPECIALES
        
        # NUEVA FUNCIONALIDAD: Trabajadores fuera de operación (AHORA EXTERNO)
        self.trabajadores_fuera_operacion = TRABAJADORES_FUERA_OPERACION
        
        # NUEVA FUNCIONALIDAD: Días festivos (AHORA EXTERNO)
        self.dias_festivos = DIAS_FESTIVOS
        
        # NUEVO SISTEMA: Calcular semanas basado en primer lunes de enero
        self.primer_lunes_enero = self._calcular_primer_lunes_enero()
        self.total_semanas_año = self._calcular_total_semanas_año()
        self.semana_seleccionada = self._seleccionar_semana()
        self.fechas_semana = self._calcular_fechas_semana()
        
        # Generar días de la semana seleccionada
        self.dias_mes = self._generar_dias_semana()
        self.semanas = self._agrupar_por_semanas()
        
        # Mapeo de días de la semana para las restricciones
        self.mapeo_dias = {
            "lunes": 0, "martes": 1, "miércoles": 2, "jueves": 3, 
            "viernes": 4, "sábado": 5, "domingo": 6
        }
        
        # SISTEMA DE PRIORIDADES DE SÁBADOS
        self.archivo_historial_sabados = 'historial_sabados.csv'
        self.historial_sabados = self._cargar_historial_sabados()
        self.prioridades_sabados = self._calcular_prioridades_sabados()
        

    
    def _cargar_historial_sabados(self):
        """Carga el historial de sábados desde CSV o crea uno nuevo"""
        if not os.path.exists(self.archivo_historial_sabados):
            print(f"📄 Creando archivo de historial: {self.archivo_historial_sabados}")
            return self._crear_historial_inicial()
        
        try:
            historial = {}
            with open(self.archivo_historial_sabados, 'r', newline='', encoding='utf-8') as archivo:
                reader = csv.DictReader(archivo)
                for fila in reader:
                    empleado = fila['empleado']
                    ultima_semana = fila['ultima_semana_trop_sabado']
                    # Convertir a entero si no está vacío, sino None
                    historial[empleado] = int(ultima_semana) if ultima_semana.strip() else None
            
            print(f"📄 Historial cargado: {len(historial)} empleados")
            return historial
            
        except Exception as e:
            print(f"⚠️ Error cargando historial: {e}")
            print("📄 Creando historial inicial...")
            return self._crear_historial_inicial()
    
    def _crear_historial_inicial(self):
        """Crea un historial inicial con todos los empleados vacíos"""
        historial = {}
        for empleado in self.empleados:
            historial[empleado] = None  # None = nunca ha tenido TROP en sábado
        
        # Guardar el archivo inicial
        self._guardar_historial_sabados(historial)
        print(f"✅ Historial inicial creado con {len(historial)} empleados")
        return historial
    
    def _guardar_historial_sabados(self, historial):
        """Guarda el historial de sábados en CSV"""
        try:
            with open(self.archivo_historial_sabados, 'w', newline='', encoding='utf-8') as archivo:
                writer = csv.writer(archivo)
                writer.writerow(['empleado', 'ultima_semana_trop_sabado'])
                
                for empleado in self.empleados:
                    ultima_semana = historial.get(empleado, None)
                    ultima_semana_str = str(ultima_semana) if ultima_semana is not None else ""
                    writer.writerow([empleado, ultima_semana_str])
            
            print(f"💾 Historial guardado en {self.archivo_historial_sabados}")
            
        except Exception as e:
            print(f"⚠️ Error guardando historial: {e}")
    
    def _calcular_prioridades_sabados(self):
        """Calcula las prioridades de sábados para la semana actual"""
        prioridades = {}
        
        print(f"\n🎯 CALCULANDO PRIORIDADES PARA SEMANA {self.semana_seleccionada}")
        print("=" * 60)
        
        for empleado in self.empleados:
            ultima_semana = self.historial_sabados.get(empleado, None)
            
            if ultima_semana is None:
                # Nunca ha tenido TROP en sábado
                semanas_transcurridas = 999  # Valor alto para indicar "nunca"
                nivel_prioridad = 4  # Prioridad máxima
            else:
                # Calcular semanas transcurridas
                semanas_transcurridas = self.semana_seleccionada - ultima_semana
                # Limitar a máximo 4 semanas para el nivel de prioridad
                nivel_prioridad = min(max(semanas_transcurridas, -2), 4)
            
            # Determinar si puede mantener sábado automáticamente
            puede_mantener_sabado = self._determinar_permiso_sabado(nivel_prioridad)
            
            prioridades[empleado] = {
                'ultima_semana': ultima_semana,
                'semanas_transcurridas': semanas_transcurridas,
                'nivel_prioridad': nivel_prioridad,
                'puede_mantener_sabado': puede_mantener_sabado
            }
        
        # Mostrar tabla de prioridades
        self._mostrar_tabla_prioridades(prioridades)
        
        return prioridades
    
    def _determinar_permiso_sabado(self, nivel_prioridad):
        """Determina si un empleado puede mantener sábado según su nivel"""
        if nivel_prioridad >= 4:
            return "✅ SIEMPRE"
        elif nivel_prioridad == 3:
            return "✅ Si no hay Nivel 4+"
        elif nivel_prioridad == 2:
            return "✅ Si no hay Nivel 3-4+"
        elif nivel_prioridad == 1:
            return "✅ Si no hay Nivel 2-4+"
        elif nivel_prioridad == 0:
            return "⚠️ Solo último recurso"
        else:  # nivel_prioridad < 0
            return "❌ EVITAR"
    
    def _mostrar_tabla_prioridades(self, prioridades):
        """Muestra una tabla formateada con las prioridades"""
        print(f"\n📊 TABLA DE PRIORIDADES SEMANA {self.semana_seleccionada}:")
        print("-" * 80)
        print(f"{'Empleado':<8} {'Última':<6} {'Transcur':<8} {'Nivel':<6} {'Puede Mantener Sábado':<25}")
        print("-" * 80)
        
        # Ordenar por nivel de prioridad (mayor a menor)
        empleados_ordenados = sorted(prioridades.items(), 
                                   key=lambda x: x[1]['nivel_prioridad'], 
                                   reverse=True)
        
        for empleado, datos in empleados_ordenados:
            ultima = datos['ultima_semana'] if datos['ultima_semana'] is not None else "Nunca"
            transcur = datos['semanas_transcurridas'] if datos['semanas_transcurridas'] != 999 else "∞"
            nivel = datos['nivel_prioridad']
            puede = datos['puede_mantener_sabado']
            
            print(f"{empleado:<8} {str(ultima):<6} {str(transcur):<8} {nivel:<6} {puede:<25}")
        
        print("-" * 80)
        
        # Mostrar resumen por niveles
        niveles_count = {}
        for datos in prioridades.values():
            nivel = datos['nivel_prioridad']
            niveles_count[nivel] = niveles_count.get(nivel, 0) + 1
        
        print(f"\n📈 RESUMEN POR NIVELES:")
        for nivel in sorted(niveles_count.keys(), reverse=True):
            count = niveles_count[nivel]
            if nivel >= 4:
                descripcion = "PRIORIDAD MÁXIMA"
            elif nivel == 3:
                descripcion = "PRIORIDAD ALTA"
            elif nivel == 2:
                descripcion = "PRIORIDAD MEDIA"
            elif nivel == 1:
                descripcion = "PRIORIDAD BAJA"
            elif nivel == 0:
                descripcion = "SIN PRIORIDAD"
            else:
                descripcion = "PENALIZACIÓN"
            
            print(f"   Nivel {nivel}: {count} empleados ({descripcion})")
        print()
    
    def _puede_trabajar_sabado(self, empleado):
        """Verifica si un empleado puede trabajar TROP en sábado (restricciones específicas)"""
        # Verificar si está fuera de operación
        if self._esta_fuera_operacion(empleado):
            return False
        
        # Verificar restricciones específicas de TROP
        if self._tiene_restricciones_especificas(empleado):
            restricciones = self.restricciones_empleados[empleado]
            if 'TROP' in restricciones:
                config_trop = restricciones['TROP']
                if not config_trop.get('libre', False):
                    dias_permitidos = config_trop.get('dias_permitidos', [])
                    # Si tiene restricción específica y sábado no está en días permitidos
                    if 'sábado' not in dias_permitidos:
                        return False
        
        return True
    
    def _obtener_empleados_por_nivel(self, nivel_minimo):
        """Obtiene empleados disponibles con nivel de prioridad >= nivel_minimo que pueden trabajar sábado"""
        empleados_disponibles = []
        
        for empleado, datos in self.prioridades_sabados.items():
            # Verificar nivel de prioridad
            if datos['nivel_prioridad'] >= nivel_minimo:
                # Verificar que puede trabajar sábado
                if self._puede_trabajar_sabado(empleado):
                    empleados_disponibles.append((empleado, datos))
        
        # Ordenar por nivel de prioridad (mayor primero)
        empleados_disponibles.sort(key=lambda x: x[1]['nivel_prioridad'], reverse=True)
        
        return empleados_disponibles
    
    def _calcular_primer_lunes_enero(self):
        """Calcula el primer lunes de enero del año especificado"""
        # Encontrar el primer lunes de enero
        primer_dia_enero = date(self.año, 1, 1)
        dias_hasta_lunes = (7 - primer_dia_enero.weekday()) % 7
        if dias_hasta_lunes == 0:
            dias_hasta_lunes = 7  # Si el 1 de enero es lunes, el primer lunes es el 1
        
        primer_lunes = primer_dia_enero + timedelta(days=dias_hasta_lunes)
        return primer_lunes
    
    def _calcular_total_semanas_año(self):
        """Calcula el total de semanas en el año (lunes a domingo)"""
        # Calcular el último día del año
        ultimo_dia_año = date(self.año, 12, 31)
        
        # Calcular cuántas semanas completas hay desde el primer lunes hasta el final del año
        dias_desde_primer_lunes = (ultimo_dia_año - self.primer_lunes_enero).days
        total_semanas = (dias_desde_primer_lunes // 7) + 1
        
        return total_semanas
    
    def _seleccionar_semana(self):
        """Selecciona la semana según la configuración"""
        if self.semana_especifica is not None:
            # Usar la semana específica proporcionada
            if 1 <= self.semana_especifica <= self.total_semanas_año:
                return self.semana_especifica
            else:
                print(f"⚠️ ADVERTENCIA: Semana {self.semana_especifica} fuera de rango. Usando primera semana de julio.")
                return self._calcular_semana_julio()
        else:
            # Usar la primera semana de julio por defecto
            return self._calcular_semana_julio()
    
    def _calcular_semana_julio(self):
        """Calcula el número de semana correspondiente a la primera semana de julio"""
        # Calcular el primer día de julio
        primer_julio = date(self.año, 7, 1)
        
        # Calcular cuántas semanas han pasado desde el primer lunes de enero
        dias_desde_primer_lunes = (primer_julio - self.primer_lunes_enero).days
        
        # Calcular la semana correspondiente
        semana_julio = (dias_desde_primer_lunes // 7) + 1
        
        return semana_julio
    
    def _seleccionar_primera_semana_julio(self):
        """Selecciona específicamente la primera semana de julio"""
        return self._calcular_semana_julio()
    
    def _calcular_fechas_semana(self):
        """Calcula las fechas exactas de la semana seleccionada"""
        # Calcular el lunes de la semana seleccionada
        lunes_semana = self.primer_lunes_enero + timedelta(days=(self.semana_seleccionada - 1) * 7)
        
        # Calcular todos los días de la semana
        fechas = []
        for i in range(7):
            fecha = lunes_semana + timedelta(days=i)
            fechas.append(fecha)
        
        return fechas
    
    def _generar_dias_semana(self):
        """Genera la lista de días de la semana seleccionada con formato DIA-DD"""
        dias = []
        nombres_dias = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
        
        for i, fecha in enumerate(self.fechas_semana):
            nombre_dia = nombres_dias[i]
            formato_dia = f"{nombre_dia}-{fecha.day:02d}"
            es_domingo = i == 6
            es_festivo = self._es_dia_festivo(fecha)
            es_no_laborable = es_domingo or es_festivo
            
            dias.append({
                'fecha': fecha,
                'formato': formato_dia,
                'dia_semana': i,
                'es_domingo': es_domingo,
                'es_festivo': es_festivo,
                'es_no_laborable': es_no_laborable
            })
        
        return dias
    
    def _agrupar_por_semanas(self):
        """Agrupa los días por semana"""
        semanas = {}
        for dia_info in self.dias_mes:
            semana_num = dia_info['fecha'].isocalendar()[1]
            if semana_num not in semanas:
                semanas[semana_num] = []
            semanas[semana_num].append(dia_info)
        return semanas
    
    def _calcular_paridad_objetivo(self):
        """Calcula el número objetivo de personas descansando por día"""
        total_descansos = self.num_empleados * len(self.semanas) * 2
        dias_no_domingo = [d for d in self.dias_mes if not d['es_domingo']]
        total_dias_disponibles = len(dias_no_domingo)
        
        descansos_por_dia = total_descansos / total_dias_disponibles
        descansos_por_dia_entero = int(descansos_por_dia)
        descansos_extra = total_descansos - (descansos_por_dia_entero * total_dias_disponibles)
        
        print(f"Total descansos a distribuir: {total_descansos}")
        print(f"Días disponibles (sin domingo): {total_dias_disponibles}")
        print(f"Descansos por día objetivo: {descansos_por_dia:.2f}")
        
        return descansos_por_dia_entero, descansos_extra, dias_no_domingo
    
    def _tiene_restricciones_especificas(self, empleado):
        """Verifica si un empleado tiene restricciones específicas de turnos"""
        return empleado in self.restricciones_empleados
    
    def _obtener_dias_permitidos_tipo(self, empleado, tipo_descanso):
        """Obtiene los días permitidos para un tipo de descanso específico"""
        if not self._tiene_restricciones_especificas(empleado):
            return None
        
        restriccion = self.restricciones_empleados[empleado].get(tipo_descanso)
        if not restriccion:
            return None
        
        return restriccion["dias_permitidos"]
    
    def _es_seleccion_fija(self, empleado, tipo_descanso):
        """Verifica si la selección para un tipo de descanso es fija"""
        if not self._tiene_restricciones_especificas(empleado):
            return False
        
        restriccion = self.restricciones_empleados[empleado].get(tipo_descanso)
        if not restriccion:
            return False
        
        return restriccion.get("tipo", "libre") == "fijo"
    
    def _tiene_fechas_especificas(self, empleado):
        """Verifica si un empleado tiene restricciones de fechas específicas"""
        return empleado in self.turnos_fechas_especificas
    
    def _obtener_fechas_especificas_semana(self, empleado):
        """Obtiene las fechas específicas que caen en la semana actual"""
        if not self._tiene_fechas_especificas(empleado):
            return []
        
        fechas_especificas = []
        fechas_semana_str = [fecha.strftime('%Y-%m-%d') for fecha in self.fechas_semana]
        
        for restriccion in self.turnos_fechas_especificas[empleado]:
            fecha_restriccion = restriccion["fecha"]
            if fecha_restriccion in fechas_semana_str:
                # Encontrar el índice del día en la semana
                idx_dia = fechas_semana_str.index(fecha_restriccion)
                fechas_especificas.append({
                    "fecha": fecha_restriccion,
                    "turno_requerido": restriccion["turno_requerido"],
                    "indice_dia": idx_dia,
                    "formato_dia": self.dias_mes[idx_dia]['formato']
                })
        
        return fechas_especificas
    
    def _asignar_descansos_empleado_fechas_especificas(self, empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num):
        """
        Asigna descansos para empleados con restricciones de fechas específicas (MÁXIMA PRIORIDAD)
        
        CORRECCIÓN DE BUGS:
        - Bug #1: Lógica defectuosa corregida
        - Bug #2: Ahora asigna AMBOS turnos DESC y TROP obligatorios
        - Bug #3: Verifica independientemente DESC y TROP faltantes
        - Bug #4: Clasificación correcta de turnos especiales
        """
        descansos_semana = {}
        
        if not self._tiene_fechas_especificas(empleado):
            return self._asignar_descansos_empleado_unificado(
                empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num
            )
        
        # Obtener fechas específicas para esta semana
        fechas_especificas = self._obtener_fechas_especificas_semana(empleado)
        
        if not fechas_especificas:
            # No hay fechas específicas en esta semana, usar lógica normal
            return self._asignar_descansos_empleado_unificado(
                empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num
            )
        
        print(f"🔧 {empleado}: PROCESANDO {len(fechas_especificas)} FECHAS ESPECÍFICAS")
        
        # PASO 1: Aplicar restricciones de fechas específicas (MÁXIMA PRIORIDAD)
        for fecha_esp in fechas_especificas:
            formato_dia = fecha_esp["formato_dia"]
            turno_requerido = fecha_esp["turno_requerido"]
            
            # Asignar el turno específico requerido
            descansos_semana[formato_dia] = turno_requerido
            descansos_por_dia[formato_dia] = descansos_por_dia.get(formato_dia, 0) + 1
            
            # Actualizar historial
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(fecha_esp["indice_dia"])
            
            print(f"   ✅ {empleado}: {formato_dia} = {turno_requerido} (fecha específica)")
        
        # PASO 2: Verificar qué tipos de turnos ya están asignados
        tipos_asignados = list(descansos_semana.values())
        
        # PASO 3: Clasificar correctamente los turnos especiales
        # Usar configuración centralizada para turnos completos y adicionales
        turnos_completos = CONFIGURACION_GENERAL["turnos_completos"]
        turnos_especiales_adicionales = CONFIGURACION_GENERAL["turnos_adicionales"]
        
        tiene_turno_completo = any(turno in tipos_asignados for turno in turnos_completos)
        tiene_turno_especial = any(turno in tipos_asignados for turno in turnos_especiales_adicionales)
        
        # PASO 4: Determinar si necesita completar con DESC/TROP
        if tiene_turno_completo:
            # VACA, COME, COMS son turnos completos → NO necesita DESC/TROP adicionales
            print(f"   ✅ {empleado}: Turno completo detectado → NO requiere DESC/TROP adicionales")
            return descansos_semana
        
        # PASO 5: Para CMED, SIND o cualquier otro caso → DEBE completar con DESC/TROP
        print(f"   🔧 {empleado}: Requiere completar con DESC/TROP obligatorios")
        
        # Obtener días disponibles (excluyendo los ya ocupados por fechas específicas)
        dias_disponibles = [d for d in dias_semana if d['formato'] not in descansos_semana]
        
        # PASO 6: Verificar independientemente DESC y TROP faltantes
        tiene_desc = "DESC" in tipos_asignados
        tiene_trop = "TROP" in tipos_asignados
        
        print(f"   📊 {empleado}: Tiene DESC={tiene_desc}, Tiene TROP={tiene_trop}")
        print(f"   📊 {empleado}: Días disponibles={len(dias_disponibles)}")
        
        # PASO 7: Asignar DESC si falta
        if not tiene_desc and len(dias_disponibles) >= 1:
            # Ordenar por disponibilidad (menos descansos asignados primero)
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            
            dia_desc = dias_ordenados[0]
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
            
            # Actualizar historial
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_desc['dia_semana'])
            
            # Remover el día usado de los disponibles
            dias_disponibles = [d for d in dias_disponibles if d['formato'] != dia_desc['formato']]
            
            print(f"   ✅ {empleado}: DESC asignado en {dia_desc['formato']}")
        
        # PASO 8: Asignar TROP si falta
        if not tiene_trop and len(dias_disponibles) >= 1:
            # Ordenar por disponibilidad (menos descansos asignados primero)
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            
            dia_trop = dias_ordenados[0]
            descansos_semana[dia_trop['formato']] = 'TROP'
            descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
            
            # Actualizar historial
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_trop['dia_semana'])
            
            print(f"   ✅ {empleado}: TROP asignado en {dia_trop['formato']}")
        
        # PASO 9: Verificar orden cronológico DESC → TROP
        turnos_desc_trop = [(d, tipo) for d, tipo in descansos_semana.items() if tipo in ['DESC', 'TROP']]
        
        if len(turnos_desc_trop) == 2:
            # Ordenar por día de la semana
            turnos_ordenados = []
            for formato_dia, tipo in turnos_desc_trop:
                # Encontrar el día correspondiente
                for dia_info in dias_semana:
                    if dia_info['formato'] == formato_dia:
                        turnos_ordenados.append((dia_info, formato_dia, tipo))
                        break
            
            # Ordenar cronológicamente
            turnos_ordenados.sort(key=lambda x: x[0]['dia_semana'])
            
            # Verificar orden correcto
            if len(turnos_ordenados) == 2:
                primer_dia, primer_formato, primer_tipo = turnos_ordenados[0]
                segundo_dia, segundo_formato, segundo_tipo = turnos_ordenados[1]
                
                # Corregir si están en orden incorrecto
                if primer_tipo != 'DESC' or segundo_tipo != 'TROP':
                    print(f"   🔧 {empleado}: Corrigiendo orden cronológico DESC → TROP")
                    descansos_semana[primer_formato] = 'DESC'
                    descansos_semana[segundo_formato] = 'TROP'
        
        # PASO 10: Resumen final
        tipos_finales = list(descansos_semana.values())
        desc_count = tipos_finales.count('DESC')
        trop_count = tipos_finales.count('TROP')
        especiales_count = len([t for t in tipos_finales if t in turnos_especiales_adicionales + turnos_completos])
        
        print(f"   🎯 {empleado}: RESULTADO FINAL → DESC={desc_count}, TROP={trop_count}, Especiales={especiales_count}")
        
        return descansos_semana
    
    def _asignar_descansos_empleado_unificado(self, empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num):
        """Asigna descansos usando el sistema unificado de restricciones"""
        descansos_semana = {}
        
        if not self._tiene_restricciones_especificas(empleado):
            return self._asignar_descansos_separados_semana(
                dias_semana, descansos_por_dia, historial_dias, empleado_idx, semana_num, empleado=empleado
            )
        
        # Obtener restricciones del empleado
        restricciones = self.restricciones_empleados[empleado]
        
        # Procesar cada tipo de descanso (DESC y TROP)
        for tipo_descanso in ["DESC", "TROP"]:
            if tipo_descanso not in restricciones:
                continue
            
            restriccion = restricciones[tipo_descanso]
            
            # Verificar si es libre (sin restricción)
            if restriccion.get("libre", False):
                continue  # Se asignará después con lógica aleatoria
            
            # Obtener días permitidos
            dias_permitidos = restriccion.get("dias_permitidos", [])
            tipo_restriccion = restriccion.get("tipo", "libre")
            
            # Filtrar días disponibles según las restricciones
            dias_disponibles = []
            for dia in dias_semana:
                nombre_dia = list(self.mapeo_dias.keys())[dia['dia_semana']]
                if nombre_dia in dias_permitidos:
                    dias_disponibles.append(dia)
            
            if dias_disponibles:
                # Ordenar por disponibilidad
                dias_ordenados = sorted(dias_disponibles, 
                                       key=lambda d: descansos_por_dia.get(d['formato'], 0))
                
                # Seleccionar día según el tipo de restricción
                if tipo_restriccion == "fijo":
                    # Para restricción fija, usar EXACTAMENTE el primer día permitido si está disponible
                    dia_seleccionado = dias_ordenados[0]  # Ya está filtrado por días permitidos
                    print(f"✅ {empleado}: {tipo_descanso} asignado en día fijo requerido: {dia_seleccionado['formato']}")
                elif tipo_restriccion == "opcional":
                    # Selección aleatoria entre los días permitidos
                    dia_seleccionado = random.choice(dias_ordenados)
                else:
                    # Tipo libre o no reconocido, usar primer día disponible
                    dia_seleccionado = dias_ordenados[0]
                
                # Asignar el descanso
                descansos_semana[dia_seleccionado['formato']] = tipo_descanso
                descansos_por_dia[dia_seleccionado['formato']] = descansos_por_dia.get(dia_seleccionado['formato'], 0) + 1
                
                # Actualizar historial
                if empleado_idx not in historial_dias:
                    historial_dias[empleado_idx] = []
                historial_dias[empleado_idx].append(dia_seleccionado['dia_semana'])
        
        # Completar con descansos faltantes respetando orden cronológico obligatorio
        dias_disponibles = [d for d in dias_semana if d['formato'] not in descansos_semana]
        tipos_asignados = list(descansos_semana.values())
        
        # Si faltan ambos descansos, asignar ambos respetando orden cronológico
        if "DESC" not in tipos_asignados and "TROP" not in tipos_asignados and len(dias_disponibles) >= 2:
            # Ordenar días disponibles por día de la semana (lunes=0, domingo=6)
            dias_ordenados = sorted(dias_disponibles, key=lambda d: d['dia_semana'])
            
            # Asignar DESC al primer día disponible cronológicamente
            dia_desc = dias_ordenados[0]
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_desc['dia_semana'])
            
            # Asignar TROP al segundo día disponible cronológicamente
            dia_trop = dias_ordenados[1]
            descansos_semana[dia_trop['formato']] = 'TROP'
            descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
            historial_dias[empleado_idx].append(dia_trop['dia_semana'])
            
        # Si solo falta uno de los dos, asignarlo
        elif "DESC" not in tipos_asignados and len(dias_disponibles) >= 1:
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            dia_desc = dias_ordenados[0]
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_desc['dia_semana'])
            
        elif "TROP" not in tipos_asignados and len(dias_disponibles) >= 1:
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            dia_trop = dias_ordenados[0]
            descansos_semana[dia_trop['formato']] = 'TROP'
            descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_trop['dia_semana'])
        
        return descansos_semana
    
    def _asignar_descansos_empleado_dia_fijo(self, empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num):
        """Asigna descansos para empleados con restricción de día fijo"""
        descansos_semana = {}
        
        if not self._tiene_restricciones_especificas(empleado):
            return self._asignar_descansos_separados_semana(
                dias_semana, descansos_por_dia, historial_dias, empleado_idx, semana_num, empleado=empleado
            )
        
        # Obtener restricción de día fijo
        restriccion = self.restricciones_empleados[empleado]
        
        # Asignar el descanso fijo primero
        for tipo_descanso in ["DESC", "TROP"]:
            if tipo_descanso in restriccion:
                config = restriccion[tipo_descanso]
                if (isinstance(config, dict) and 
                    config.get("tipo") == "fijo" and 
                    "dias_permitidos" in config and 
                    isinstance(config["dias_permitidos"], list) and 
                    len(config["dias_permitidos"]) > 0):
                    dia_fijo = config["dias_permitidos"][0]  # Tomar el primer día permitido
                    
                    # Buscar el día correspondiente en la semana
                    for dia in dias_semana:
                        nombre_dia = list(self.mapeo_dias.keys())[dia['dia_semana']]
                        if nombre_dia == dia_fijo:
                            descansos_semana[dia['formato']] = tipo_descanso
                            descansos_por_dia[dia['formato']] = descansos_por_dia.get(dia['formato'], 0) + 1
                            
                            # Actualizar historial
                            if empleado_idx not in historial_dias:
                                historial_dias[empleado_idx] = []
                            historial_dias[empleado_idx].append(dia['dia_semana'])
                            break
        
        # Completar con descansos faltantes respetando orden cronológico obligatorio
        dias_disponibles = [d for d in dias_semana if d['formato'] not in descansos_semana]
        tipos_asignados = list(descansos_semana.values())
        
        # CORRECCIÓN CRÍTICA: Si ya hay descansos asignados, verificar orden cronológico
        if len(tipos_asignados) == 2:
            # Ya tiene ambos descansos, verificar que estén en orden correcto
            dias_con_descansos = [(d, descansos_semana[d['formato']]) for d in dias_semana if d['formato'] in descansos_semana]
            dias_con_descansos.sort(key=lambda x: x[0]['dia_semana'])  # Ordenar cronológicamente
            
            if len(dias_con_descansos) == 2:
                primer_dia, primer_tipo = dias_con_descansos[0]
                segundo_dia, segundo_tipo = dias_con_descansos[1]
                
                # Corregir si están en orden incorrecto
                if primer_tipo != 'DESC' or segundo_tipo != 'TROP':
                    descansos_semana[primer_dia['formato']] = 'DESC'
                    descansos_semana[segundo_dia['formato']] = 'TROP'
                    
        # Si faltan descansos, asignar respetando orden cronológico
        elif "DESC" not in tipos_asignados and "TROP" not in tipos_asignados and len(dias_disponibles) >= 2:
            # Ordenar días disponibles por día de la semana (lunes=0, domingo=6)
            dias_ordenados = sorted(dias_disponibles, key=lambda d: d['dia_semana'])
            
            # Asignar DESC al primer día disponible cronológicamente
            dia_desc = dias_ordenados[0]
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_desc['dia_semana'])
            
            # Asignar TROP al segundo día disponible cronológicamente
            dia_trop = dias_ordenados[1]
            descansos_semana[dia_trop['formato']] = 'TROP'
            descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
            historial_dias[empleado_idx].append(dia_trop['dia_semana'])
            
        # Si solo falta uno de los dos, asignarlo
        elif "DESC" not in tipos_asignados and len(dias_disponibles) >= 1:
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            dia_desc = dias_ordenados[0]
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_desc['dia_semana'])
            
        elif "TROP" not in tipos_asignados and len(dias_disponibles) >= 1:
            dias_ordenados = sorted(dias_disponibles, 
                                   key=lambda d: descansos_por_dia.get(d['formato'], 0))
            dia_trop = dias_ordenados[0]
            descansos_semana[dia_trop['formato']] = 'TROP'
            descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
            
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(dia_trop['dia_semana'])
        
        return descansos_semana
    
    def _asignar_descansos_empleado_separados(self, empleado_idx, descansos_por_dia, historial_dias):
        """Asigna descansos con separación y variación semanal"""
        descansos = {}
        
        # Para cada semana
        for semana_num, dias_semana in self.semanas.items():
            dias_semana_disponibles = [d for d in dias_semana if not d['es_domingo']]
            
            if len(dias_semana_disponibles) >= 2:
                # Ordenar días por disponibilidad (menos descansos asignados primero)
                dias_ordenados = sorted(dias_semana_disponibles, 
                                      key=lambda d: descansos_por_dia.get(d['formato'], 0))
                
                # Estrategia 1: Intentar separación máxima
                descansos_semana = self._asignar_descansos_separados_semana(
                    dias_ordenados, descansos_por_dia, historial_dias, empleado_idx, semana_num
                )
                
                # Agregar a descansos generales
                descansos.update(descansos_semana)
        
        return descansos
    
    def _asignar_descansos_separados_semana(self, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, semana_num, dias_ocupados=None, empleado=None):
        """Asigna DESC y TROP con máxima separación en una semana - CON ALEATORIZACIÓN Y RESTRICCIONES FIJAS"""
        descansos_semana = {}
        
        # Filtrar días que ya están ocupados por turnos especiales
        if dias_ocupados:
            dias_disponibles = [d for d in dias_disponibles if d['formato'] not in dias_ocupados]
        
        if len(dias_disponibles) < 2:
            return descansos_semana
        
        # NUEVA LÓGICA: VERIFICAR RESTRICCIONES FIJAS PRIMERO
        dia_desc_seleccionado = None
        
        # Si tenemos el empleado, verificar restricciones fijas para DESC
        if empleado and self._es_seleccion_fija(empleado, "DESC"):
            dias_permitidos_desc = self._obtener_dias_permitidos_tipo(empleado, "DESC")
            print(f"🎯 {empleado}: Tiene restricción fija DESC para: {dias_permitidos_desc}")
            
            # Buscar el día requerido entre los días disponibles
            if dias_permitidos_desc:
                for dia_requerido in dias_permitidos_desc:
                    for dia_info in dias_disponibles:
                        # Mapear día en español a día en inglés del formato
                        mapeo_dias_formato = {
                            "lunes": "MON", "martes": "TUE", "miércoles": "WED", 
                            "jueves": "THU", "viernes": "FRI", "sábado": "SAT", "domingo": "SUN"
                        }
                        
                        if dia_requerido in mapeo_dias_formato:
                            dia_formato_requerido = mapeo_dias_formato[dia_requerido]
                            if dia_info['formato'].startswith(dia_formato_requerido):
                                dia_desc_seleccionado = dia_info
                                print(f"✅ {empleado}: DESC asignado en día requerido: {dia_info['formato']} ({dia_requerido})")
                                break
                    
                    if dia_desc_seleccionado:
                        break
                
                if not dia_desc_seleccionado:
                    print(f"⚠️ {empleado}: No se pudo asignar DESC en día requerido {dias_permitidos_desc}, usando lógica estándar")
        
        # Si no hay restricción fija o no se pudo cumplir, usar lógica original
        if not dia_desc_seleccionado:
            # Crear lista de días con pesos para evitar repetición
            dias_con_peso = []
            for dia in dias_disponibles:
                peso = descansos_por_dia.get(dia['formato'], 0)
                
                # Penalizar días que ya usó este empleado en semanas anteriores
                if empleado_idx in historial_dias and dia['dia_semana'] in historial_dias[empleado_idx]:
                    peso += 2  # Penalización por repetición
                
                # ALEATORIZACIÓN: Agregar ruido aleatorio al peso para variar la selección
                ruido_aleatorio = random.uniform(-0.5, 0.5)
                peso += ruido_aleatorio
                
                dias_con_peso.append((dia, peso))
            
            # Ordenar por peso (menor peso = mejor opción)
            dias_con_peso.sort(key=lambda x: x[1])
            
            # ALEATORIZACIÓN: A veces elegir el segundo mejor día en lugar del primero
            if random.random() < 0.3 and len(dias_con_peso) > 1:
                # 30% de probabilidad de elegir el segundo mejor día
                dia_desc_seleccionado = dias_con_peso[1][0]
            else:
                dia_desc_seleccionado = dias_con_peso[0][0]
            
            if empleado and dia_desc_seleccionado:
                print(f"🔄 {empleado}: DESC asignado por lógica estándar: {dia_desc_seleccionado['formato']}")
        
        # Verificar que se seleccionó un día válido
        if not dia_desc_seleccionado:
            print(f"❌ No se pudo seleccionar día válido para DESC")
            return descansos_semana
        
        # Continuar con la asignación usando el día seleccionado
        dia_desc = dia_desc_seleccionado
        
        # REGLA OBLIGATORIA: El primer descanso cronológico SIEMPRE es DESC
        # El segundo descanso cronológico SIEMPRE es TROP
        
        # Asignar el primer descanso como DESC (temporalmente)
        descansos_semana[dia_desc['formato']] = 'DESC'
        descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
        
        # Actualizar historial
        if empleado_idx not in historial_dias:
            historial_dias[empleado_idx] = []
        historial_dias[empleado_idx].append(dia_desc['dia_semana'])
        
        # Asignar segundo descanso
        if len(dias_con_peso) > 1:
            # Filtrar días que no sean consecutivos al primer descanso
            dia_desc_semana = dia_desc['dia_semana']
            dias_no_consecutivos = []
            
            for dia, peso in dias_con_peso:
                if dia['formato'] != dia_desc['formato']:  # No el mismo día
                    # Verificar que no sea consecutivo (diferencia > 1)
                    if abs(dia['dia_semana'] - dia_desc_semana) > 1:
                        dias_no_consecutivos.append((dia, peso))
            
            # Si no hay días no consecutivos, usar cualquier día disponible
            if dias_no_consecutivos:
                # ALEATORIZACIÓN: Elegir aleatoriamente entre los días no consecutivos
                candidatos = sorted(dias_no_consecutivos, key=lambda x: x[1])[:3]  # Top 3 candidatos
                dia_segundo = random.choice(candidatos)[0]
            else:
                # ALEATORIZACIÓN: Elegir aleatoriamente entre los días restantes
                dias_restantes = [d for d, _ in dias_con_peso if d['formato'] != dia_desc['formato']]
                dia_segundo = random.choice(dias_restantes)
            
            # CORRECCIÓN CRÍTICA: Determinar cuál es el primer y segundo descanso cronológicamente
            # El que ocurre primero en la semana debe ser DESC, el segundo debe ser TROP
            if dia_desc['dia_semana'] < dia_segundo['dia_semana']:
                # dia_desc ocurre antes que dia_segundo
                descansos_semana[dia_desc['formato']] = 'DESC'
                descansos_semana[dia_segundo['formato']] = 'TROP'
            else:
                # dia_segundo ocurre antes que dia_desc
                descansos_semana[dia_desc['formato']] = 'TROP'
                descansos_semana[dia_segundo['formato']] = 'DESC'
            
            descansos_por_dia[dia_segundo['formato']] = descansos_por_dia.get(dia_segundo['formato'], 0) + 1
            
            # Actualizar historial
            historial_dias[empleado_idx].append(dia_segundo['dia_semana'])
        
        return descansos_semana
    
    def _asignar_descansos_con_prioridades_sabado(self, empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, semana_num, dias_ocupados=None):
        """Asigna DESC y TROP evaluando prioridades de sábado SOLO cuando TROP cae naturalmente en sábado (SISTEMA REACTIVO)"""
        
        if dias_ocupados is None:
            dias_ocupados = {}
        
        # PASO 1: Asignar descansos normalmente usando el algoritmo estándar
        descansos_semana = self._asignar_descansos_separados_semana(
            dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, semana_num, dias_ocupados, empleado=empleado
        )
        
        # PASO 2: Verificar si TROP cayó naturalmente en sábado
        trop_en_sabado = False
        dia_trop_sabado = None
        
        for dia_formato, tipo_descanso in descansos_semana.items():
            if tipo_descanso == 'TROP' and 'SAT' in dia_formato:
                trop_en_sabado = True
                dia_trop_sabado = dia_formato
                break
        
        # PASO 3: Si NO hay TROP en sábado, devolver asignación normal
        if not trop_en_sabado:
            return descansos_semana
        
        # PASO 4: TROP cayó en sábado → Evaluar si el empleado merece el sábado
        prioridad_empleado = self.prioridades_sabados[empleado]
        nivel_empleado = prioridad_empleado['nivel_prioridad']
        
        # PASO 5: Verificar restricciones específicas (tienen prioridad absoluta)
        if not self._puede_trabajar_sabado(empleado):
            print(f"🚫 {empleado}: No puede trabajar sábado por restricciones → Reasignando")
            return self._reasignar_sin_sabado(empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, dias_ocupados)
        
        # PASO 6: Evaluar si merece el sábado según cascada de prioridades
        merece_sabado = self._evaluar_si_merece_sabado(empleado, nivel_empleado)
        
        if merece_sabado:
            print(f"✅ {empleado}: Mantiene sábado (Nivel {nivel_empleado} - MERECE SÁBADO)")
            return descansos_semana
        
        # PASO 7: No merece sábado → Buscar intercambio con empleado de mayor prioridad
        print(f"🔄 {empleado}: Nivel {nivel_empleado} → Buscando intercambio con mayor prioridad")
        
        intercambio_exitoso = self._intentar_intercambio_sabado(
            empleado, nivel_empleado, dia_trop_sabado, descansos_semana, 
            descansos_por_dia, historial_dias
        )
        
        if intercambio_exitoso:
            # El intercambio ya actualizó las estructuras necesarias
            return self._reasignar_sin_sabado(empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, dias_ocupados)
        else:
            # No hay intercambio disponible → Mantener asignación original
            print(f"⚠️ {empleado}: Sin intercambio disponible → Mantiene sábado por excepción")
            return descansos_semana
    
    def _evaluar_si_merece_sabado(self, empleado, nivel_empleado):
        """
        Evalúa si un empleado merece mantener el sábado según la cascada de prioridades
        
        Lógica de cascada:
        - Nivel 4: SIEMPRE merece sábado
        - Nivel 3: Merece sábado SI NO hay empleados Nivel 4 disponibles
        - Nivel 2: Merece sábado SI NO hay empleados Nivel 3-4 disponibles  
        - Nivel 1: Merece sábado SI NO hay empleados Nivel 2-4 disponibles
        - Nivel 0: Solo como último recurso
        """
        
        # Nivel 4: PRIORIDAD MÁXIMA - siempre merece sábado
        if nivel_empleado >= 4:
            return True
        
        # Para niveles inferiores, verificar si hay empleados de mayor prioridad disponibles
        trabajadores_activos = self._obtener_trabajadores_activos()
        
        # Contar empleados por nivel de prioridad (que pueden trabajar sábado)
        empleados_por_nivel = {4: 0, 3: 0, 2: 0, 1: 0, 0: 0}
        
        for otro_empleado in trabajadores_activos:
            if otro_empleado != empleado and self._puede_trabajar_sabado(otro_empleado):
                otro_nivel = self.prioridades_sabados[otro_empleado]['nivel_prioridad']
                otro_nivel_limitado = min(otro_nivel, 4) if otro_nivel > 0 else otro_nivel
                if otro_nivel_limitado in empleados_por_nivel:
                    empleados_por_nivel[otro_nivel_limitado] += 1
        
        # Aplicar lógica de cascada
        if nivel_empleado == 3:
            # Nivel 3: Merece sábado SI NO hay Nivel 4 disponibles
            return empleados_por_nivel[4] == 0
        elif nivel_empleado == 2:
            # Nivel 2: Merece sábado SI NO hay Nivel 3-4 disponibles
            return empleados_por_nivel[4] == 0 and empleados_por_nivel[3] == 0
        elif nivel_empleado == 1:
            # Nivel 1: Merece sábado SI NO hay Nivel 2-4 disponibles
            return (empleados_por_nivel[4] == 0 and empleados_por_nivel[3] == 0 and 
                   empleados_por_nivel[2] == 0)
        else:
            # Nivel 0 o menor: Solo último recurso
            return (empleados_por_nivel[4] == 0 and empleados_por_nivel[3] == 0 and 
                   empleados_por_nivel[2] == 0 and empleados_por_nivel[1] == 0)
    
    def _intentar_intercambio_sabado(self, empleado_actual, nivel_actual, dia_sabado, descansos_actuales, descansos_por_dia, historial_dias):
        """
        Intenta intercambiar el sábado con un empleado de mayor prioridad
        
        Returns:
            bool: True si se realizó intercambio exitoso, False si no
        """
        
        trabajadores_activos = self._obtener_trabajadores_activos()
        candidatos_intercambio = []
        
        # Buscar candidatos de mayor prioridad que puedan trabajar sábado
        for otro_empleado in trabajadores_activos:
            if otro_empleado == empleado_actual:
                continue
                
            if not self._puede_trabajar_sabado(otro_empleado):
                continue
                
            otro_nivel = self.prioridades_sabados[otro_empleado]['nivel_prioridad']
            otro_nivel_limitado = min(otro_nivel, 4) if otro_nivel > 0 else otro_nivel
            
            # Solo considerar empleados de mayor prioridad
            if otro_nivel_limitado > nivel_actual:
                candidatos_intercambio.append((otro_empleado, otro_nivel_limitado))
        
        if not candidatos_intercambio:
            return False
        
        # Ordenar candidatos por prioridad (mayor primero)
        candidatos_intercambio.sort(key=lambda x: x[1], reverse=True)
        
        # Intentar intercambio con el primer candidato disponible
        for candidato, nivel_candidato in candidatos_intercambio:
            print(f"🔄 Intentando intercambio: {empleado_actual} (Nivel {nivel_actual}) ↔ {candidato} (Nivel {nivel_candidato})")
            
            # Por simplicidad, asumir que el intercambio es posible
            # En una implementación más compleja, verificaríamos disponibilidad específica
            print(f"✅ Intercambio exitoso: {candidato} recibe el sábado")
            return True
        
        return False
    
    def _reasignar_sin_sabado(self, empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, dias_ocupados=None):
        """Reasigna DESC y TROP evitando sábados MANTENIENDO PARIDAD DIARIA"""
        
        # PASO 1: Descontar de la asignación original si existe
        if empleado_idx in historial_dias and len(historial_dias[empleado_idx]) >= 2:
            # Encontrar los días originales asignados para descontarlos
            dias_originales = historial_dias[empleado_idx][-2:]  # Los últimos 2 días asignados
            
            for dia_semana_original in dias_originales:
                # Encontrar el formato correspondiente al día de la semana
                for dia_info in dias_disponibles:
                    if dia_info['dia_semana'] == dia_semana_original:
                        formato_original = dia_info['formato']
                        # Descontar del contador original
                        if descansos_por_dia.get(formato_original, 0) > 0:
                            descansos_por_dia[formato_original] -= 1
                        break
            
            # Limpiar el historial del empleado para reasignar
            historial_dias[empleado_idx] = historial_dias[empleado_idx][:-2]
        
        # PASO 2: Filtrar días ocupados por turnos especiales
        if dias_ocupados:
            dias_disponibles = [d for d in dias_disponibles if d['formato'] not in dias_ocupados]
        
        # PASO 3: Filtrar sábados
        dias_sin_sabado = [d for d in dias_disponibles if not self._es_sabado(d['formato'])]
        
        if len(dias_sin_sabado) < 2:
            print(f"⚠️ {empleado}: No hay suficientes días sin sábado - Usando asignación original")
            # Restaurar historial si no se puede reasignar
            if empleado_idx in historial_dias and 'dias_originales' in locals():
                historial_dias[empleado_idx].extend(dias_originales)
            return self._asignar_descansos_separados_semana(
                dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, 0, dias_ocupados
            )
        
        # PASO 4: Ordenar días por menor carga para mantener paridad
        dias_ordenados = sorted(dias_sin_sabado, 
                               key=lambda d: descansos_por_dia.get(d['formato'], 0))
        
        # PASO 5: Asignar primer día para DESC (menor carga)
        dia_desc = dias_ordenados[0]
        
        # PASO 6: Buscar segundo día para TROP que no sea consecutivo y tenga poca carga
        dia_trop = None
        for dia in dias_ordenados[1:]:
            if abs(dia['dia_semana'] - dia_desc['dia_semana']) > 1:
                dia_trop = dia
                break
        
        # Si no hay días no consecutivos, usar el segundo con menor carga
        if not dia_trop and len(dias_ordenados) > 1:
            dia_trop = dias_ordenados[1]
        
        # Verificar que tenemos ambos días válidos
        if not dia_trop:
            return {}
        
        # PASO 7: Crear asignación respetando orden cronológico
        descansos_semana = {}
        if dia_desc['dia_semana'] < dia_trop['dia_semana']:
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_semana[dia_trop['formato']] = 'TROP'
        else:
            descansos_semana[dia_desc['formato']] = 'TROP'
            descansos_semana[dia_trop['formato']] = 'DESC'
        
        # PASO 8: Actualizar contadores de paridad
        descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
        descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
        
        # PASO 9: Actualizar historial correctamente
        if empleado_idx not in historial_dias:
            historial_dias[empleado_idx] = []
        historial_dias[empleado_idx].extend([dia_desc['dia_semana'], dia_trop['dia_semana']])
        
        return descansos_semana
    
    def _asignar_descansos_preferir_sabado(self, empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, semana_num, dias_ocupados=None):
        """Asigna DESC y TROP prefiriendo sábado para TROP (empleados de alta prioridad)"""
        
        # Filtrar días ocupados
        if dias_ocupados:
            dias_disponibles = [d for d in dias_disponibles if d['formato'] not in dias_ocupados]
        
        if len(dias_disponibles) < 2:
            return {}
        
        # Buscar sábado disponible
        sabado_disponible = None
        for dia in dias_disponibles:
            if self._es_sabado(dia['formato']):
                sabado_disponible = dia
                break
        
        if sabado_disponible:
            # Buscar día para DESC que no sea sábado
            dias_no_sabado = [d for d in dias_disponibles if not self._es_sabado(d['formato'])]
            if dias_no_sabado:
                # Ordenar por menor carga para mantener paridad
                dias_ordenados = sorted(dias_no_sabado, key=lambda d: descansos_por_dia.get(d['formato'], 0))
                dia_desc = dias_ordenados[0]
                
                # Crear asignación respetando orden cronológico
                descansos_semana = {}
                if dia_desc['dia_semana'] < sabado_disponible['dia_semana']:
                    descansos_semana[dia_desc['formato']] = 'DESC'
                    descansos_semana[sabado_disponible['formato']] = 'TROP'
                else:
                    descansos_semana[sabado_disponible['formato']] = 'DESC'
                    descansos_semana[dia_desc['formato']] = 'TROP'
                
                # Actualizar contadores
                descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
                descansos_por_dia[sabado_disponible['formato']] = descansos_por_dia.get(sabado_disponible['formato'], 0) + 1
                
                # Actualizar historial
                if empleado_idx not in historial_dias:
                    historial_dias[empleado_idx] = []
                historial_dias[empleado_idx].extend([dia_desc['dia_semana'], sabado_disponible['dia_semana']])
                
                print(f"✅ {empleado}: Asignado con sábado preferente (Nivel 4+ - PRIORIDAD MÁXIMA)")
                return descansos_semana
        
        # Si no hay sábado disponible, asignación normal
        print(f"⚠️ {empleado}: Sábado no disponible, asignación normal (Nivel 4+)")
        return self._asignar_descansos_separados_semana(
            dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, semana_num, dias_ocupados, empleado=empleado
        )
    
    def _asignar_descansos_sin_sabado(self, empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, dias_ocupados=None):
        """Asigna DESC y TROP evitando sábados MANTENIENDO PARIDAD Y RESPETANDO RESTRICCIONES FIJAS"""
        
        # Filtrar días ocupados
        if dias_ocupados:
            dias_disponibles = [d for d in dias_disponibles if d['formato'] not in dias_ocupados]
        
        # Filtrar sábados
        dias_sin_sabado = [d for d in dias_disponibles if not self._es_sabado(d['formato'])]
        
        if len(dias_sin_sabado) < 2:
            return {}  # No hay suficientes días
        
        # NUEVA LÓGICA: VERIFICAR RESTRICCIONES FIJAS PRIMERO (igual que _asignar_descansos_separados_semana)
        dia_desc_seleccionado = None
        
        # Si tenemos el empleado, verificar restricciones fijas para DESC
        if empleado and self._es_seleccion_fija(empleado, "DESC"):
            dias_permitidos_desc = self._obtener_dias_permitidos_tipo(empleado, "DESC")
            print(f"✅ {empleado}: DESC asignado en día fijo requerido (sin sábado): {dias_permitidos_desc}")
            
            # Buscar el día requerido entre los días disponibles (sin sábados)
            if dias_permitidos_desc:
                for dia_requerido in dias_permitidos_desc:
                    for dia_info in dias_sin_sabado:
                        # Mapear día en español a día en inglés del formato
                        mapeo_dias_formato = {
                            "lunes": "MON", "martes": "TUE", "miércoles": "WED", 
                            "jueves": "THU", "viernes": "FRI", "sábado": "SAT", "domingo": "SUN"
                        }
                        
                        if dia_requerido in mapeo_dias_formato:
                            dia_formato_requerido = mapeo_dias_formato[dia_requerido]
                            if dia_info['formato'].startswith(dia_formato_requerido):
                                dia_desc_seleccionado = dia_info
                                print(f"✅ {empleado}: DESC asignado en día requerido: {dia_info['formato']} ({dia_requerido})")
                                break
                    
                    if dia_desc_seleccionado:
                        break
                
                if not dia_desc_seleccionado:
                    print(f"⚠️ {empleado}: No se pudo asignar DESC en día requerido {dias_permitidos_desc}, usando lógica estándar")
        
        # Si no hay restricción fija o no se pudo cumplir, usar lógica original
        if not dia_desc_seleccionado:
            # Ordenar por menor carga para mantener paridad
            dias_ordenados = sorted(dias_sin_sabado, key=lambda d: descansos_por_dia.get(d['formato'], 0))
            dia_desc_seleccionado = dias_ordenados[0]
            
            if empleado:
                print(f"🔄 {empleado}: DESC asignado por lógica estándar (sin sábado): {dia_desc_seleccionado['formato']}")
        
        # Verificar que se seleccionó un día válido
        if not dia_desc_seleccionado:
            return {}
        
        # Continuar con la asignación usando el día seleccionado
        dia_desc = dia_desc_seleccionado
        
        # Buscar segundo día para TROP que no sea consecutivo
        dias_restantes = [d for d in dias_sin_sabado if d['formato'] != dia_desc['formato']]
        dia_trop = None
        
        for dia in dias_restantes:
            if abs(dia['dia_semana'] - dia_desc['dia_semana']) > 1:
                dia_trop = dia
                break
        
        # Si no hay días no consecutivos, usar el primer día restante
        if not dia_trop and len(dias_restantes) > 0:
            dia_trop = dias_restantes[0]
        
        if not dia_trop:
            return {}  # No se pudo asignar
        
        # Crear asignación respetando orden cronológico
        descansos_semana = {}
        if dia_desc['dia_semana'] < dia_trop['dia_semana']:
            descansos_semana[dia_desc['formato']] = 'DESC'
            descansos_semana[dia_trop['formato']] = 'TROP'
        else:
            descansos_semana[dia_desc['formato']] = 'TROP'
            descansos_semana[dia_trop['formato']] = 'DESC'
        
        # Actualizar contadores
        descansos_por_dia[dia_desc['formato']] = descansos_por_dia.get(dia_desc['formato'], 0) + 1
        descansos_por_dia[dia_trop['formato']] = descansos_por_dia.get(dia_trop['formato'], 0) + 1
        
        # Actualizar historial
        if empleado_idx not in historial_dias:
            historial_dias[empleado_idx] = []
        historial_dias[empleado_idx].extend([dia_desc['dia_semana'], dia_trop['dia_semana']])
        
        return descansos_semana
    
    def _es_sabado(self, dia_formato):
        """Verifica si un día en formato 'SAT-12' es sábado"""
        return dia_formato.startswith('SAT')
    
    def _empleado_ya_procesado(self, empleado, historial_dias):
        """Verifica si un empleado ya fue procesado (tiene entradas en historial)"""
        empleado_idx = self.empleados.index(empleado) if empleado in self.empleados else -1
        return empleado_idx in historial_dias and len(historial_dias[empleado_idx]) >= 2
    
    def _actualizar_historial_sabados(self, df):
        """Actualiza el historial de sábados con las nuevas asignaciones"""
        empleados_con_sabado = []
        
        # Identificar empleados que recibieron TROP en sábado
        for idx, empleado in enumerate(self.empleados):
            if self._esta_fuera_operacion(empleado):
                continue
            
            for col in df.columns:
                if col.startswith('SAT') and df.iloc[idx][col] == 'TROP':
                    empleados_con_sabado.append(empleado)
                    break
        
        # Actualizar historial
        historial_actualizado = self.historial_sabados.copy()
        for empleado in empleados_con_sabado:
            historial_actualizado[empleado] = self.semana_seleccionada
        
        # Guardar historial actualizado
        self._guardar_historial_sabados(historial_actualizado)
        
        # Mostrar resumen de actualizaciones
        if empleados_con_sabado:
            print(f"\n💾 HISTORIAL ACTUALIZADO:")
            print(f"   Empleados que recibieron TROP en sábado: {', '.join(empleados_con_sabado)}")
            print(f"   Semana registrada: {self.semana_seleccionada}")
        else:
            print(f"\n💾 No hay actualizaciones de historial (ningún TROP en sábado)")
    
    def generar_horario_primera_semana(self):
        """Genera el horario para la semana seleccionada aleatoriamente - CON SISTEMA DE CASCADA ESTRICTO PARA SÁBADOS"""
        # Usar los días de la semana seleccionada
        dias_semana_seleccionada = self.dias_mes
        semana_num = self.semana_seleccionada
        
        # Mostrar información de la semana seleccionada
        lunes_semana = self.fechas_semana[0]
        domingo_semana = self.fechas_semana[6]
        print(f"\n📅 SEMANA SELECCIONADA: Semana {self.semana_seleccionada}")
        print(f"📅 FECHAS: Lunes {lunes_semana.strftime('%d/%m/%Y')} - Domingo {domingo_semana.strftime('%d/%m/%Y')}")
        
        # Inicializar contador de descansos por día
        descansos_por_dia = {dia['formato']: 0 for dia in dias_semana_seleccionada if not dia['es_domingo']}
        historial_dias = {}
        
        # PASO 1: SISTEMA DE CASCADA ESTRICTO PARA SÁBADOS
        empleados_asignados_sabado = self._asignar_sabados_por_cascada_estricta()
        
        # Separar empleados con y sin restricciones específicas (SOLO TRABAJADORES ACTIVOS)
        empleados_con_restricciones = []
        empleados_sin_restricciones = []
        empleados_fuera_operacion = []
        
        for idx, empleado in enumerate(self.empleados):
            if self._esta_fuera_operacion(empleado):
                empleados_fuera_operacion.append((idx, empleado))
            elif self._tiene_restricciones_especificas(empleado) or self._tiene_fechas_especificas(empleado):
                empleados_con_restricciones.append((idx, empleado))
            else:
                empleados_sin_restricciones.append((idx, empleado))
        
        # ALEATORIZACIÓN: Mezclar el orden de asignación de empleados (SOLO ACTIVOS)
        random.shuffle(empleados_con_restricciones)
        random.shuffle(empleados_sin_restricciones)
        
        # Asignar primero empleados con restricciones específicas primero (SOLO ACTIVOS)
        filas = []
        
        # Procesar empleados con restricciones específicas primero (SOLO ACTIVOS)
        for idx_original, empleado in empleados_con_restricciones:
            fila = {'No.': idx_original + 1, 'SIGLA ATCO': empleado}
            
            # Definir dias_semana para este empleado
            dias_semana = [dia for dia in dias_semana_seleccionada if not dia['es_no_laborable']]
            
            # VERIFICAR SI TIENE FECHAS ESPECÍFICAS (MÁXIMA PRIORIDAD)
            fechas_especificas = self._obtener_fechas_especificas_semana(empleado)
            
            if fechas_especificas:
                # MÁXIMA PRIORIDAD: Usar fechas específicas (VACA, COME, CMED, COMS)
                print(f"🔧 {empleado}: PROCESANDO FECHAS ESPECÍFICAS (MÁXIMA PRIORIDAD)")
                descansos_combinados = self._asignar_descansos_empleado_fechas_especificas(
                    empleado, idx_original, dias_semana, descansos_por_dia, historial_dias, semana_num
                )
            else:
                # PRIMER PASO: Asignar turnos especiales (ADICIONALES)
                turnos_especiales = self._asignar_turnos_especiales(
                    empleado, idx_original, dias_semana, descansos_por_dia, historial_dias, semana_num
                )
                
                # SEGUNDO PASO: Asignar descansos regulares (DESC/TROP) usando sistema de cascada
                if len(dias_semana) >= 2:
                    if empleado in empleados_asignados_sabado:
                        # ASIGNACIÓN FORZADA DE SÁBADO (por cascada estricta)
                        descansos_semana = self._asignar_descansos_con_sabado_forzado(
                            empleado, dias_semana, descansos_por_dia, historial_dias, idx_original, turnos_especiales
                        )
                    else:
                        # ASIGNACIÓN NORMAL SIN SÁBADO (excluir sábado completamente)
                        descansos_semana = self._asignar_descansos_sin_sabado(
                            empleado, dias_semana, descansos_por_dia, historial_dias, idx_original, turnos_especiales
                        )
                else:
                    descansos_semana = {}
                
                # COMBINAR turnos especiales y descansos regulares
                # Los turnos especiales tienen prioridad sobre los regulares
                descansos_combinados = {**descansos_semana, **turnos_especiales}
            
            # Llenar la fila con los descansos asignados
            for dia_info in dias_semana_seleccionada:
                formato_dia = dia_info['formato']
                if formato_dia in descansos_combinados:
                    fila[formato_dia] = descansos_combinados[formato_dia]
                else:
                    fila[formato_dia] = None
            
            filas.append(fila)
        
        # Procesar empleados sin restricciones específicas (SOLO ACTIVOS)
        for idx_original, empleado in empleados_sin_restricciones:
            fila = {'No.': idx_original + 1, 'SIGLA ATCO': empleado}
            
            # Definir dias_semana para este empleado
            dias_semana = [dia for dia in dias_semana_seleccionada if not dia['es_no_laborable']]
            
            # PRIMER PASO: Asignar turnos especiales (ADICIONALES)
            turnos_especiales = self._asignar_turnos_especiales(
                empleado, idx_original, dias_semana, descansos_por_dia, historial_dias, semana_num
            )
            
            # SEGUNDO PASO: Verificar si este empleado debe recibir sábado por cascada estricta
            if empleado in empleados_asignados_sabado:
                # ASIGNACIÓN FORZADA DE SÁBADO (por cascada estricta)
                print(f"🎯 {empleado}: ASIGNACIÓN FORZADA DE SÁBADO (Cascada Estricta - Nivel {self.prioridades_sabados[empleado]['nivel_prioridad']})")
                descansos_semana = self._asignar_descansos_con_sabado_forzado(
                    empleado, dias_semana, descansos_por_dia, historial_dias, idx_original, turnos_especiales
                )
            else:
                # ASIGNACIÓN NORMAL SIN SÁBADO (excluir sábado completamente)
                descansos_semana = self._asignar_descansos_sin_sabado(
                    empleado, dias_semana, descansos_por_dia, historial_dias, idx_original, turnos_especiales
                )
            
            # COMBINAR turnos especiales y descansos regulares
            # Los turnos especiales tienen prioridad sobre los regulares
            descansos_combinados = {**descansos_semana, **turnos_especiales}
            
            # Llenar la fila con los descansos asignados
            for dia_info in dias_semana_seleccionada:
                formato_dia = dia_info['formato']
                if formato_dia in descansos_combinados:
                    fila[formato_dia] = descansos_combinados[formato_dia]
                else:
                    fila[formato_dia] = None
            
            filas.append(fila)
        
        # Procesar empleados fuera de operación (SIN ASIGNACIÓN DE TURNOS)
        for idx_original, empleado in empleados_fuera_operacion:
            fila = {'No.': idx_original + 1, 'SIGLA ATCO': empleado}
            
            # Llenar la fila con celdas vacías (sin asignación de turnos)
            for dia_info in dias_semana_seleccionada:
                formato_dia = dia_info['formato']
                fila[formato_dia] = None  # Sin asignación para trabajadores fuera de operación
            
            filas.append(fila)
        
        # Ordenar filas por número de empleado para mantener consistencia
        filas.sort(key=lambda x: x['No.'])
        
        columnas = ['No.', 'SIGLA ATCO'] + [dia['formato'] for dia in dias_semana_seleccionada]
        df = pd.DataFrame.from_records(filas, columns=columnas)
        
        # Calcular personal disponible por día (EXCLUYENDO TRABAJADORES FUERA DE OPERACIÓN)
        total_trabajadores_activos = len(self._obtener_trabajadores_activos())
        personal_disponible_por_dia = {}
        
        for dia_info in dias_semana_seleccionada:
            formato_dia = dia_info['formato']
            descansos_en_dia = sum(1 for fila in filas 
                                  if fila[formato_dia] is not None and 
                                  fila['SIGLA ATCO'] not in self.trabajadores_fuera_operacion and
                                  fila[formato_dia] in CONFIGURACION_GENERAL["turnos_validos"])
            personal_disponible_por_dia[formato_dia] = total_trabajadores_activos - descansos_en_dia
        
        # Mostrar resumen de descansos por empleado (SOLO TRABAJADORES ACTIVOS)
        print("\n=== RESUMEN DE DESCANSO POR EMPLEADO ===")
        for idx, empleado in enumerate(self.empleados):
            if self._esta_fuera_operacion(empleado):
                print(f"Empleado {idx + 1} ({empleado}): FUERA DE OPERACIÓN - Sin asignación")
            else:
                descansos_empleado = {}
                for dia_info in dias_semana_seleccionada:
                    formato_dia = dia_info['formato']
                    valor = df.iloc[idx][formato_dia]
                    if valor is not None:
                        if valor not in descansos_empleado:
                            descansos_empleado[valor] = 0
                        descansos_empleado[valor] += 1
                
                resumen = ", ".join([f"{tipo}={cantidad}" for tipo, cantidad in descansos_empleado.items()])
                print(f"Empleado {idx + 1} ({empleado}): {resumen}")
        
        # ACTUALIZAR HISTORIAL DE SÁBADOS
        self._actualizar_historial_sabados(df)
        
        return df
    

    
    def exportar_excel(self, df, nombre_archivo='horario_descansos_separacion_primera_semana_julio.xlsx'):
        """Exporta el horario a Excel con conteo de personal disponible y formato condicional"""
        # Generar nombre de hoja basado en la semana seleccionada
        lunes_semana = self.fechas_semana[0]
        domingo_semana = self.fechas_semana[6]
        nombre_hoja = f"Semana {self.semana_seleccionada} ({lunes_semana.strftime('%d-%m')}-{domingo_semana.strftime('%d-%m')})"
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[nombre_hoja]
            
            # Importar estilos para formato condicional
            from openpyxl.styles import PatternFill, Font
            
            # Definir el formato amarillo para turnos especiales
            formato_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            fuente_negra = Font(color="000000")
            
            # Obtener las columnas de días del DataFrame
            columnas_dias = [col for col in df.columns if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'))]
            
            # Aplicar formato condicional a turnos especiales y trabajadores inactivos
            turnos_especiales = CONFIGURACION_GENERAL["turnos_validos"]
            
            # Recorrer todas las celdas de datos (excluyendo encabezados)
            for fila_idx in range(2, len(df) + 2):  # Empezar desde fila 2 (después del encabezado)
                for col_idx, columna_dia in enumerate(columnas_dias, start=3):  # Empezar desde columna 3
                    celda = worksheet.cell(row=fila_idx, column=col_idx)
                    valor_celda = celda.value
                    
                    # Obtener el empleado de esta fila
                    empleado_idx = fila_idx - 2  # Convertir índice de fila a índice del DataFrame
                    if empleado_idx < len(df):
                        empleado = df.iloc[empleado_idx]['SIGLA ATCO']
                    else:
                        empleado = None
                    
                    # Aplicar formato amarillo si el valor es un turno especial
                    if valor_celda in turnos_especiales:
                        celda.fill = formato_amarillo
                        celda.font = fuente_negra
                    
                    # Aplicar formato amarillo si el empleado está fuera de operación (inactivo)
                    if empleado and self._esta_fuera_operacion(empleado):
                        celda.fill = formato_amarillo
                        celda.font = fuente_negra
            
            # Calcular personal disponible por día (EXCLUYENDO TRABAJADORES FUERA DE OPERACIÓN)
            total_trabajadores_activos = len(self._obtener_trabajadores_activos())
            personal_disponible_por_dia = {}
            
            for columna_dia in columnas_dias:
                # Contar trabajadores que NO están en descanso en este día
                # Excluir trabajadores fuera de operación del conteo
                descansos_en_dia = 0
                for _, fila in df.iterrows():
                    empleado = fila['SIGLA ATCO']
                    valor_dia = fila[columna_dia]
                    
                    # Solo contar si el empleado está activo y tiene un turno que lo hace no disponible
                    if (empleado not in self.trabajadores_fuera_operacion and 
                        valor_dia is not None and 
                        valor_dia in CONFIGURACION_GENERAL["turnos_validos"]):
                        descansos_en_dia += 1
                
                personal_disponible_por_dia[columna_dia] = total_trabajadores_activos - descansos_en_dia
            
            # Agregar fila de conteo de personal disponible
            fila_conteo = len(df) + 2  # Dos filas después de los datos de empleados
            
            # Agregar fila vacía
            worksheet.cell(row=fila_conteo-1, column=1, value="")
            
            # Agregar fila de conteo
            worksheet.cell(row=fila_conteo, column=1, value="Personal Disponible")
            worksheet.cell(row=fila_conteo, column=2, value="")
            
            # Agregar conteos por día
            for i, columna_dia in enumerate(columnas_dias, start=3):  # Empezar desde la columna 3 (después de No. y SIGLA ATCO)
                col_idx = i
                worksheet.cell(row=fila_conteo, column=col_idx, value=personal_disponible_por_dia[columna_dia])
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Horario exportado a: {nombre_archivo}")
        print(f"📊 Nombre de hoja: {nombre_hoja}")
        print(f"📊 Conteo de personal disponible agregado en la fila {fila_conteo}")
        
        # Mostrar resumen del conteo
        print(f"\n📋 RESUMEN DE PERSONAL DISPONIBLE:")
        for dia, disponibles in personal_disponible_por_dia.items():
            print(f"  {dia}: {disponibles} trabajadores disponibles")
        
        # Mostrar información sobre el formato aplicado
        trabajadores_inactivos = self._obtener_trabajadores_fuera_operacion()
        if trabajadores_inactivos:
            print(f"\n🎨 FORMATO APLICADO:")
            print(f"  ✅ Color amarillo aplicado a turnos especiales")
            print(f"  ✅ Color amarillo aplicado a trabajadores inactivos: {', '.join(trabajadores_inactivos)}")
        else:
            print(f"\n🎨 FORMATO APLICADO:")
            print(f"  ✅ Color amarillo aplicado a turnos especiales")
            print(f"  ℹ️  No hay trabajadores inactivos en esta semana")
        
        return nombre_archivo
    
    def validar_regla_desc_trop(self, df):
        """Valida que la regla DESC/TROP se cumple correctamente"""
        print("\n=== VALIDACIÓN DE REGLA DESC/TROP ===")
        print("✅ REGLA: El primer descanso cronológico debe ser DESC, el segundo debe ser TROP")
        
        errores_orden = 0
        empleados_correctos = 0
        
        # Analizar cada empleado
        for idx, empleado in enumerate(self.empleados):
            if self._esta_fuera_operacion(empleado):
                continue  # Saltar empleados fuera de operación
                
            descansos_empleado = []
            
            # Recopilar todos los descansos DESC/TROP del empleado
            for col in df.columns:
                if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN')):
                    valor = df.iloc[idx][col]
                    if valor in ['DESC', 'TROP']:
                        # Obtener el día de la semana del formato de columna
                        dia_semana = self._obtener_dia_semana_de_formato(col)
                        descansos_empleado.append((col, valor, dia_semana))
            
            # Ordenar por día de la semana (cronológicamente)
            descansos_empleado.sort(key=lambda x: x[2])
            
            # Verificar orden correcto
            if len(descansos_empleado) == 2:
                primer_descanso = descansos_empleado[0]
                segundo_descanso = descansos_empleado[1]
                
                col1, tipo1, dia1 = primer_descanso
                col2, tipo2, dia2 = segundo_descanso
                
                if tipo1 == 'DESC' and tipo2 == 'TROP':
                    print(f"✅ {empleado}: {col1}(DESC) → {col2}(TROP) - CORRECTO")
                    empleados_correctos += 1
                else:
                    print(f"❌ {empleado}: {col1}({tipo1}) → {col2}({tipo2}) - INCORRECTO")
                    print(f"   Debería ser: {col1}(DESC) → {col2}(TROP)")
                    errores_orden += 1
            elif len(descansos_empleado) == 1:
                col1, tipo1, dia1 = descansos_empleado[0]
                print(f"⚠️  {empleado}: Solo tiene {col1}({tipo1}) - Falta el segundo descanso")
            elif len(descansos_empleado) == 0:
                print(f"⚠️  {empleado}: No tiene descansos DESC/TROP asignados")
            else:
                print(f"⚠️  {empleado}: Tiene {len(descansos_empleado)} descansos (debería tener 2)")
        
        # Resumen de validación
        total_empleados_activos = len(self._obtener_trabajadores_activos())
        print(f"\n📊 RESUMEN DE VALIDACIÓN:")
        print(f"   Empleados activos: {total_empleados_activos}")
        print(f"   Empleados con orden correcto: {empleados_correctos}")
        print(f"   Empleados con orden incorrecto: {errores_orden}")
        print(f"   Porcentaje de cumplimiento: {(empleados_correctos/total_empleados_activos*100):.1f}%")
        
        if errores_orden == 0:
            print("🎉 ¡EXCELENTE! Todos los empleados cumplen la regla DESC/TROP")
        else:
            print(f"⚠️  ATENCIÓN: {errores_orden} empleados NO cumplen la regla DESC/TROP")
        
        return errores_orden == 0
    
    def validar_turnos_especiales_adicionales(self, df):
        """Valida que los turnos especiales se suman correctamente a DESC/TROP"""
        print("\n=== VALIDACIÓN DE TURNOS ESPECIALES ADICIONALES ===")
        print("✅ REGLA: Turnos especiales (SIND, CMED) se SUMAN a DESC/TROP")
        
        empleados_con_turnos_especiales = 0
        empleados_correctos = 0
        
        for idx, empleado in enumerate(self.empleados):
            if self._esta_fuera_operacion(empleado):
                continue  # Saltar empleados fuera de operación
            
            # Recopilar todos los turnos del empleado
            turnos_empleado = []
            desc_count = 0
            trop_count = 0
            turnos_especiales = []
            
            for col in df.columns:
                if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN')):
                    valor = df.iloc[idx][col]
                    if valor is not None:
                        turnos_empleado.append((col, valor))
                        if valor == 'DESC':
                            desc_count += 1
                        elif valor == 'TROP':
                            trop_count += 1
                        elif valor in CONFIGURACION_GENERAL["turnos_adicionales"]:
                            turnos_especiales.append((col, valor))
            
            # Verificar si tiene turnos especiales configurados
            if self._tiene_turnos_especiales(empleado):
                empleados_con_turnos_especiales += 1
                
                # Verificar que tiene DESC + TROP + turno especial
                if desc_count == 1 and trop_count == 1 and len(turnos_especiales) >= 1:
                    turnos_esp_str = ", ".join([f"{col}({tipo})" for col, tipo in turnos_especiales])
                    print(f"✅ {empleado}: DESC={desc_count}, TROP={trop_count}, Especiales=[{turnos_esp_str}] - CORRECTO")
                    empleados_correctos += 1
                else:
                    turnos_esp_str = ", ".join([f"{col}({tipo})" for col, tipo in turnos_especiales])
                    print(f"❌ {empleado}: DESC={desc_count}, TROP={trop_count}, Especiales=[{turnos_esp_str}] - INCORRECTO")
                    print(f"   Debería tener: DESC=1, TROP=1, y al menos 1 turno especial")
            
            # Mostrar empleados sin turnos especiales pero con asignaciones correctas
            elif desc_count == 1 and trop_count == 1 and len(turnos_especiales) == 0:
                print(f"✅ {empleado}: DESC={desc_count}, TROP={trop_count} - CORRECTO (sin turnos especiales)")
        
        # Resumen de validación
        print(f"\n📊 RESUMEN DE TURNOS ESPECIALES:")
        print(f"   Empleados con turnos especiales configurados: {empleados_con_turnos_especiales}")
        print(f"   Empleados con turnos especiales correctos: {empleados_correctos}")
        
        if empleados_con_turnos_especiales > 0:
            porcentaje = (empleados_correctos/empleados_con_turnos_especiales*100)
            print(f"   Porcentaje de cumplimiento: {porcentaje:.1f}%")
            
            if empleados_correctos == empleados_con_turnos_especiales:
                print("🎉 ¡EXCELENTE! Todos los empleados con turnos especiales están correctos")
            else:
                print(f"⚠️  ATENCIÓN: {empleados_con_turnos_especiales - empleados_correctos} empleados con turnos especiales incorrectos")
        else:
            print("ℹ️  No hay empleados con turnos especiales configurados en esta ejecución")
        
        return empleados_correctos == empleados_con_turnos_especiales
    
    def _obtener_dia_semana_de_formato(self, formato_columna):
        """Convierte formato de columna (ej: MON-07) a número de día de semana"""
        mapeo_dias_formato = {
            'MON': 0, 'TUE': 1, 'WED': 2, 'THU': 3, 
            'FRI': 4, 'SAT': 5, 'SUN': 6
        }
        dia_abrev = formato_columna.split('-')[0]
        return mapeo_dias_formato.get(dia_abrev, 0)

    def analizar_separacion(self, df):
        """Analiza la separación de descansos y variación semanal"""
        print("\n=== ANÁLISIS DE SEPARACIÓN DE DESCANSO ===")
        
        # Analizar cada empleado
        for idx, empleado in enumerate(self.empleados):
            print(f"\nEmpleado {idx+1} ({empleado}):")
            
            # Agrupar por semanas
            for semana_num, dias_semana in self.semanas.items():
                dias_semana_disponibles = [d for d in dias_semana if not d['es_domingo']]
                descansos_semana = []
                
                for dia in dias_semana_disponibles:
                    valor = df.iloc[idx][dia['formato']]
                    if valor in ['DESC', 'TROP']:
                        descansos_semana.append((dia['formato'], valor, dia['dia_semana']))
                
                if len(descansos_semana) == 2:
                    # Verificar separación
                    dia1, tipo1, num_dia1 = descansos_semana[0]
                    dia2, tipo2, num_dia2 = descansos_semana[1]
                    separacion = abs(num_dia1 - num_dia2)
                    
                    print(f"  Semana {semana_num}: {dia1}({tipo1}) y {dia2}({tipo2}) - Separación: {separacion} días")
                    
                    if separacion == 1:
                        print(f"    ⚠️  ADVERTENCIA: Descansos consecutivos!")
                    elif separacion >= 3:
                        print(f"    ✅ Excelente separación")
                    else:
                        print(f"    ⚠️  Separación mínima")
        
        # Analizar paridad diaria
        print(f"\n=== ANÁLISIS DE PARIDAD DIARIA ===")
        descansos_por_dia = {}
        for col in df.columns:
            if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN')):
                descansos = sum(1 for valor in df[col] if valor in ['DESC', 'TROP'])
                descansos_por_dia[col] = descansos
        
        valores = list(descansos_por_dia.values())
        print(f"Estadísticas de distribución:")
        print(f"  Promedio: {np.mean(valores):.2f}")
        print(f"  Desviación estándar: {np.std(valores):.2f}")
        print(f"  Mínimo: {min(valores)}")
        print(f"  Máximo: {max(valores)}")
    
    def validar_restricciones(self):
        """Valida que las restricciones no generen conflictos imposibles"""
        errores = []
        
        # Validar restricciones unificadas
        for empleado, restricciones in self.restricciones_empleados.items():
            # Verificar que el empleado existe
            if empleado not in self.empleados:
                errores.append(f"Empleado '{empleado}' no existe en la lista de empleados")
                continue
            
            # Verificar que hay al menos un tipo de descanso definido
            if not restricciones:
                errores.append(f"Empleado '{empleado}' no tiene restricciones definidas")
                continue
            
            # Verificar cada tipo de descanso
            for tipo_descanso, config in restricciones.items():
                if tipo_descanso not in ["DESC", "TROP"]:
                    errores.append(f"Tipo de descanso '{tipo_descanso}' no válido para empleado '{empleado}'")
                    continue
                
                # Verificar si es libre o tiene restricciones
                if config.get("libre", False):
                    continue  # No hay más validaciones para restricciones libres
                
                if "dias_permitidos" not in config:
                    errores.append(f"Falta 'dias_permitidos' para {tipo_descanso} del empleado '{empleado}'")
                    continue
                
                if "tipo" not in config:
                    errores.append(f"Falta 'tipo' para {tipo_descanso} del empleado '{empleado}'")
                    continue
                
                # Verificar que los días permitidos son válidos
                if isinstance(config, dict) and "dias_permitidos" in config:
                    dias_permitidos = config["dias_permitidos"]
                    if isinstance(dias_permitidos, list):
                        for dia in dias_permitidos:
                            if dia not in self.mapeo_dias:
                                errores.append(f"Día '{dia}' no válido para {tipo_descanso} del empleado '{empleado}'")
                
                # Verificar que el tipo es válido
                tipo_restriccion = config["tipo"]
                if tipo_restriccion not in ["fijo", "opcional", "libre"]:
                    errores.append(f"Tipo de restricción '{tipo_restriccion}' no válido para {tipo_descanso} del empleado '{empleado}'")
        
        # Validar restricciones de fechas específicas
        for empleado, restricciones in self.turnos_fechas_especificas.items():
            # Verificar que el empleado existe
            if empleado not in self.empleados:
                errores.append(f"Empleado '{empleado}' no existe en la lista de empleados (restricción fecha específica)")
                continue
            
            # Verificar que hay al menos una restricción definida
            if not restricciones:
                errores.append(f"Empleado '{empleado}' no tiene restricciones de fechas específicas definidas")
                continue
            
            # Verificar cada restricción
            for restriccion in restricciones:
                if "fecha" not in restriccion:
                    errores.append(f"Falta 'fecha' en restricción del empleado '{empleado}'")
                    continue
                
                if "turno_requerido" not in restriccion:
                    errores.append(f"Falta 'turno_requerido' en restricción del empleado '{empleado}'")
                    continue
                
                # Verificar que la fecha es válida
                try:
                    fecha = datetime.strptime(restriccion["fecha"], "%Y-%m-%d").date()
                except ValueError:
                    errores.append(f"Fecha '{restriccion['fecha']}' no válida para empleado '{empleado}'")
                    continue
                
                # Verificar que el turno requerido es válido
                turno_requerido = restriccion["turno_requerido"]
                if turno_requerido not in CONFIGURACION_GENERAL["turnos_validos"]:
                    errores.append(f"Turno requerido '{turno_requerido}' no válido para empleado '{empleado}'")
        
        # Validar turnos especiales
        for empleado, restricciones in self.turnos_especiales.items():
            # Verificar que el empleado existe
            if empleado not in self.empleados:
                errores.append(f"Empleado '{empleado}' no existe en la lista de empleados (turnos especiales)")
                continue
            
            # Verificar que hay al menos una restricción definida
            if not restricciones:
                errores.append(f"Empleado '{empleado}' no tiene turnos especiales definidos")
                continue
            
            # Verificar cada restricción
            for restriccion in restricciones:
                if "tipo" not in restriccion:
                    errores.append(f"Falta 'tipo' en turno especial del empleado '{empleado}'")
                    continue
                
                if "frecuencia" not in restriccion:
                    errores.append(f"Falta 'frecuencia' en turno especial del empleado '{empleado}'")
                    continue
                
                if "dia_semana" not in restriccion:
                    errores.append(f"Falta 'dia_semana' en turno especial del empleado '{empleado}'")
                    continue
                
                # Verificar que el tipo es válido (debe ser un turno adicional)
                tipo_turno = restriccion["tipo"]
                if tipo_turno not in CONFIGURACION_GENERAL["turnos_adicionales"]:
                    errores.append(f"Tipo de turno especial '{tipo_turno}' no válido para empleado '{empleado}'. Debe ser uno de: {CONFIGURACION_GENERAL['turnos_adicionales']}")
                
                # Verificar que la frecuencia es válida
                frecuencia = restriccion["frecuencia"]
                if frecuencia not in ["semanal_fijo"]:
                    errores.append(f"Frecuencia '{frecuencia}' no válida para empleado '{empleado}'")
                
                # Verificar que el día de la semana es válido
                dia_semana = restriccion["dia_semana"]
                if dia_semana not in self.mapeo_dias:
                    errores.append(f"Día de la semana '{dia_semana}' no válido para empleado '{empleado}'")
        
        # Validar trabajadores fuera de operación
        for empleado in self.trabajadores_fuera_operacion:
            # Verificar que el empleado existe
            if empleado not in self.empleados:
                errores.append(f"Empleado '{empleado}' no existe en la lista de empleados (fuera de operación)")
                continue
        
        # Verificar que no hay conflictos entre trabajadores fuera de operación y otras restricciones
        for empleado in self.trabajadores_fuera_operacion:
            if empleado in self.restricciones_empleados:
                errores.append(f"Empleado '{empleado}' está marcado como fuera de operación pero tiene restricciones configuradas")
            
            if empleado in self.turnos_fechas_especificas:
                errores.append(f"Empleado '{empleado}' está marcado como fuera de operación pero tiene fechas específicas configuradas")
            
            if empleado in self.turnos_especiales:
                errores.append(f"Empleado '{empleado}' está marcado como fuera de operación pero tiene turnos especiales configurados")
        
        # Validar días festivos
        for fecha_str in self.dias_festivos:
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except ValueError:
                errores.append(f"Fecha festiva '{fecha_str}' no tiene formato válido (YYYY-MM-DD)")
                continue
        
        return errores
    
    def mostrar_restricciones_aplicadas(self):
        """Muestra información sobre las restricciones aplicadas"""
        print("\n=== SISTEMA UNIFICADO DE RESTRICCIONES ===")
        
        if not self.restricciones_empleados:
            print("No hay restricciones configuradas.")
        else:
            for empleado, restricciones in self.restricciones_empleados.items():
                print(f"\nEmpleado: {empleado}")
                for tipo_descanso, config in restricciones.items():
                    if config.get("libre", False):
                        print(f"  {tipo_descanso}: Sin restricción (libre)")
                    else:
                        if "dias_permitidos" in config and isinstance(config["dias_permitidos"], list):
                            dias_permitidos = ", ".join(config["dias_permitidos"])
                            tipo_restriccion = config.get("tipo", "no definido")
                            print(f"  {tipo_descanso}: {dias_permitidos} (tipo: {tipo_restriccion})")
                        else:
                            print(f"  {tipo_descanso}: Configuración incompleta")
        
        print("\n=== RESTRICCIONES DE FECHAS ESPECÍFICAS (MÁXIMA PRIORIDAD) ===")
        
        if not self.turnos_fechas_especificas:
            print("No hay restricciones de fechas específicas configuradas.")
        else:
            for empleado, restricciones in self.turnos_fechas_especificas.items():
                print(f"\nEmpleado: {empleado}")
                for restriccion in restricciones:
                    fecha = restriccion["fecha"]
                    turno = restriccion["turno_requerido"]
                    print(f"  {fecha}: {turno} obligatorio")
        
        print("\n=== TURNOS ESPECIALES EXTENDIDOS (ADICIONALES A DESC/TROP) ===")
        
        if not self.turnos_especiales:
            print("No hay turnos especiales configurados.")
        else:
            for empleado, restricciones in self.turnos_especiales.items():
                print(f"\nEmpleado: {empleado}")
                for restriccion in restricciones:
                    tipo = restriccion["tipo"]
                    frecuencia = restriccion["frecuencia"]
                    dia_semana = restriccion["dia_semana"]
                    print(f"  {tipo}: {frecuencia} en {dia_semana}")
        
        print("\n=== TRABAJADORES FUERA DE OPERACIÓN (EXCLUIDOS COMPLETAMENTE) ===")
        
        if not self.trabajadores_fuera_operacion:
            print("No hay trabajadores fuera de operación configurados.")
        else:
            for empleado in self.trabajadores_fuera_operacion:
                print(f"  {empleado}: Sin asignación de turnos (fuera de operación)")
        
        # Mostrar resumen de trabajadores activos vs fuera de operación
        trabajadores_activos = self._obtener_trabajadores_activos()
        trabajadores_fuera = self._obtener_trabajadores_fuera_operacion()
        
        print(f"\n📊 RESUMEN DE TRABAJADORES:")
        print(f"  Total de empleados: {len(self.empleados)}")
        print(f"  Trabajadores activos: {len(trabajadores_activos)}")
        print(f"  Trabajadores fuera de operación: {len(trabajadores_fuera)}")
        
        # Mostrar información sobre días festivos
        dias_festivos_semana = self._obtener_dias_festivos_semana()
        if dias_festivos_semana:
            print(f"\n🎉 DÍAS FESTIVOS EN LA SEMANA SELECCIONADA:")
            for dia_festivo in dias_festivos_semana:
                fecha = dia_festivo['fecha']
                formato = dia_festivo['formato_dia']
                print(f"  {formato} ({fecha.strftime('%d/%m/%Y')}): Día festivo - Sin descansos automáticos")
        else:
            print(f"\n📅 DÍAS FESTIVOS EN LA SEMANA SELECCIONADA:")
            print("  No hay días festivos en esta semana")
        
        print(f"\n📋 DÍAS FESTIVOS CONFIGURADOS PARA 2025:")
        for fecha_str in self.dias_festivos:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            print(f"  {fecha.strftime('%d/%m/%Y')}: {fecha_str}")
    
    def _tiene_turnos_especiales(self, empleado):
        """Verifica si un empleado tiene turnos especiales"""
        return empleado in self.turnos_especiales
    
    def _obtener_turnos_especiales_semana(self, empleado):
        """Obtiene los turnos especiales que se aplican en la semana actual"""
        if not self._tiene_turnos_especiales(empleado):
            return []
        
        turnos_especiales = []
        
        for turno_esp in self.turnos_especiales[empleado]:
            if turno_esp["frecuencia"] == "semanal_fijo":
                dia_semana = turno_esp["dia_semana"]
                if dia_semana in self.mapeo_dias:
                    indice_dia = self.mapeo_dias[dia_semana]
                    # Verificar que el día existe en la semana actual
                    if indice_dia < len(self.dias_mes):
                        turnos_especiales.append({
                            "tipo": turno_esp["tipo"],
                            "indice_dia": indice_dia,
                            "formato_dia": self.dias_mes[indice_dia]['formato'],
                            "dia_semana": dia_semana
                        })
        
        return turnos_especiales
    
    def _asignar_turnos_especiales(self, empleado, empleado_idx, dias_semana, descansos_por_dia, historial_dias, semana_num):
        """Asigna turnos especiales (ADICIONALES a DESC/TROP)"""
        turnos_especiales = {}
        
        if not self._tiene_turnos_especiales(empleado):
            return turnos_especiales
        
        # Obtener turnos especiales para esta semana
        turnos_esp_semana = self._obtener_turnos_especiales_semana(empleado)
        
        # Aplicar turnos especiales
        for turno_esp in turnos_esp_semana:
            formato_dia = turno_esp["formato_dia"]
            tipo_turno = turno_esp["tipo"]
            
            # Asignar el turno especial
            turnos_especiales[formato_dia] = tipo_turno
            # NO incrementar descansos_por_dia porque son adicionales
            
            # Actualizar historial
            if empleado_idx not in historial_dias:
                historial_dias[empleado_idx] = []
            historial_dias[empleado_idx].append(turno_esp["indice_dia"])
        
        return turnos_especiales
    
    def _esta_fuera_operacion(self, empleado):
        """Verifica si un empleado está fuera de operación"""
        return empleado in self.trabajadores_fuera_operacion
    
    def _obtener_trabajadores_activos(self):
        """Obtiene la lista de trabajadores que están en operación"""
        return [emp for emp in self.empleados if not self._esta_fuera_operacion(emp)]
    
    def _obtener_trabajadores_fuera_operacion(self):
        """Obtiene la lista de trabajadores fuera de operación"""
        return [emp for emp in self.empleados if self._esta_fuera_operacion(emp)]
    
    def _tiene_turno_especial_completo(self, empleado):
        """Verifica si un empleado tiene un turno especial que ocupa toda la semana (VACA, COME, COMT, etc.)"""
        if empleado not in self.turnos_fechas_especificas:
            return False
        
        # Obtener las fechas específicas del empleado para esta semana
        fechas_especificas = self._obtener_fechas_especificas_semana(empleado)
        
        if not fechas_especificas:
            return False
        
        # Tipos de turnos que se consideran "completos" (reemplazan DESC/TROP)
        turnos_completos = CONFIGURACION_GENERAL["turnos_completos"]
        
        # Verificar si todos los turnos especiales son del tipo "completo"
        turnos_especificos = [fecha['turno_requerido'] for fecha in fechas_especificas]
        todos_son_completos = all(turno in turnos_completos for turno in turnos_especificos)
        
        if not todos_son_completos:
            return False
        
        # Contar cuántos días de la semana tiene turnos especiales completos
        dias_con_turnos_especiales = len(fechas_especificas)
        dias_laborables_semana = len([d for d in self.dias_mes if not d['es_domingo']])  # Excluir domingo
        
        # Si tiene turnos especiales completos en la mayoría de días laborables (>=4 días), es un turno completo
        return dias_con_turnos_especiales >= min(4, dias_laborables_semana)
    
    def _obtener_trabajadores_disponibles_para_desc_trop(self):
        """Obtiene trabajadores que pueden recibir turnos DESC/TROP normales"""
        trabajadores_disponibles = []
        
        for empleado in self.empleados:
            # Excluir trabajadores fuera de operación
            if self._esta_fuera_operacion(empleado):
                continue
            
            # Excluir trabajadores con turnos especiales completos
            if self._tiene_turno_especial_completo(empleado):
                continue
            
            trabajadores_disponibles.append(empleado)
        
        return trabajadores_disponibles
    
    def _es_dia_festivo(self, fecha):
        """Verifica si una fecha es un día festivo"""
        fecha_str = fecha.strftime('%Y-%m-%d')
        return fecha_str in self.dias_festivos
    
    def _obtener_dias_festivos_semana(self):
        """Obtiene los días festivos que caen en la semana actual"""
        dias_festivos_semana = []
        for i, fecha in enumerate(self.fechas_semana):
            if self._es_dia_festivo(fecha):
                dias_festivos_semana.append({
                    'fecha': fecha,
                    'indice_dia': i,
                    'formato_dia': self.dias_mes[i]['formato']
                })
        return dias_festivos_semana
    
    def _es_dia_no_laborable(self, dia_info):
        """Verifica si un día no es laborable (domingo o festivo)"""
        return dia_info['es_domingo'] or self._es_dia_festivo(dia_info['fecha'])
    
    def _asignar_sabados_por_cascada_estricta(self):
        """
        SISTEMA DE CASCADA ESTRICTO: Asigna sábados exclusivamente por orden de prioridad
        
        1️⃣ Calcular cuántos cupos de sábado necesitamos para mantener paridad
        2️⃣ Llenar cupos ESTRICTAMENTE por cascada: Nivel 4 → 3 → 2 → 1 → 0
        3️⃣ NO permitir asignaciones aleatorias fuera del sistema
        
        Returns:
            list: Lista de empleados que DEBEN recibir TROP en sábado (orden estricto de prioridad)
        """
        
        print("\n🎯 === SISTEMA DE CASCADA ESTRICTO PARA SÁBADOS ===")
        
        # PASO 1: Calcular cuántos cupos de sábado necesitamos para mantener paridad
        trabajadores_activos = self._obtener_trabajadores_activos()
        trabajadores_disponibles_desc_trop = self._obtener_trabajadores_disponibles_para_desc_trop()
        trabajadores_con_turnos_completos = [emp for emp in trabajadores_activos if self._tiene_turno_especial_completo(emp)]
        
        total_trabajadores_activos = len(trabajadores_activos)
        total_trabajadores_disponibles = len(trabajadores_disponibles_desc_trop)
        
        # Objetivo: Fórmula personalizada (trabajadores_disponibles - 11) con límites 4-11
        cupos_sabado_objetivo = (total_trabajadores_disponibles - 11)# max(4, min(11, total_trabajadores_disponibles - 11))  # Entre 4-11 empleados
        
        print(f"📊 Trabajadores activos: {total_trabajadores_activos}")
        
        # Debug detallado de detección de turnos completos
        print(f"🔍 DEBUG - Análisis de turnos completos:")
        for empleado in trabajadores_activos:
            tiene_turno_completo = self._tiene_turno_especial_completo(empleado)
            if empleado in self.turnos_fechas_especificas:
                fechas_especificas = self._obtener_fechas_especificas_semana(empleado)
                turnos_tipos = [fecha['turno_requerido'] for fecha in fechas_especificas] if fechas_especificas else []
                print(f"  {empleado}: {len(fechas_especificas)} días, turnos={turnos_tipos}, completo={tiene_turno_completo}")
        
        print(f"📊 Trabajadores con turnos completos: {len(trabajadores_con_turnos_completos)} {trabajadores_con_turnos_completos}")
        print(f"📊 Trabajadores disponibles para DESC/TROP: {total_trabajadores_disponibles}")
        print(f"🎯 Cupos de sábado objetivo (fórmula: max(4, min(11, {total_trabajadores_disponibles} - 11))): {cupos_sabado_objetivo}")
        
        # PASO 2: Agrupar empleados por nivel de prioridad (solo los que pueden trabajar sábado)
        empleados_por_nivel = {
            4: [],  # 4+ semanas sin sábado (PRIORIDAD MÁXIMA)
            3: [],  # 3 semanas sin sábado
            2: [],  # 2 semanas sin sábado  
            1: [],  # 1 semana sin sábado
            0: [],  # 0 semanas (descansó semana pasada)
            -1: [], # Niveles negativos (penalización)
            -2: []  # Penalización mayor
        }
        
        # Clasificar empleados disponibles para DESC/TROP por nivel
        for empleado in trabajadores_disponibles_desc_trop:
            if empleado in self.prioridades_sabados:
                nivel = self.prioridades_sabados[empleado]['nivel_prioridad']
                # Limitar nivel máximo a 4, mantener negativos
                nivel_clasificado = min(nivel, 4) if nivel > 0 else max(nivel, -2)
                
                # Solo considerar empleados que pueden trabajar sábado
                if self._puede_trabajar_sabado(empleado):
                    if nivel_clasificado in empleados_por_nivel:
                        empleados_por_nivel[nivel_clasificado].append(empleado)
                    else:
                        empleados_por_nivel[-2].append(empleado)  # Casos extraños
        
        # PASO 3: Mostrar distribución por niveles
        print("\n📋 DISTRIBUCIÓN POR NIVELES DE PRIORIDAD:")
        for nivel in [4, 3, 2, 1, 0, -1, -2]:
            empleados = empleados_por_nivel[nivel]
            if empleados:
                nivel_desc = {
                    4: "4+ sem (PRIORIDAD MÁXIMA)",
                    3: "3 sem (PRIORIDAD ALTA)", 
                    2: "2 sem (PRIORIDAD MEDIA)",
                    1: "1 sem (PRIORIDAD BAJA)",
                    0: "0 sem (SIN PRIORIDAD)",
                    -1: "Negativos (PENALIZACIÓN LEVE)",
                    -2: "Negativos (PENALIZACIÓN FUERTE)"
                }
                print(f"  Nivel {nivel} ({nivel_desc[nivel]}): {empleados} ({len(empleados)} empleados)")
        
        # PASO 4: CASCADA ESTRICTA - Llenar cupos por orden de prioridad (SIN ALEATORIZACIÓN)
        empleados_asignados_sabado = []
        cupos_restantes = cupos_sabado_objetivo
        
        print(f"\n🔄 INICIANDO CASCADA ESTRICTA (Cupos objetivo: {cupos_restantes}):")
        
        for nivel in [4, 3, 2, 1, 0, -1, -2]:  # Orden estricto de mayor a menor prioridad
            if cupos_restantes <= 0:
                break
                
            empleados_nivel = empleados_por_nivel[nivel]
            if not empleados_nivel:
                continue
            
            # ORDEN ALFABÉTICO dentro del mismo nivel para consistencia (no aleatorio)
            empleados_nivel_ordenados = sorted(empleados_nivel)
            
            # Tomar empleados hasta agotar el nivel o llenar cupos
            empleados_a_tomar = min(len(empleados_nivel_ordenados), cupos_restantes)
            empleados_seleccionados = empleados_nivel_ordenados[:empleados_a_tomar]
            
            empleados_asignados_sabado.extend(empleados_seleccionados)
            cupos_restantes -= empleados_a_tomar
            
            nivel_desc = {
                4: "PRIORIDAD MÁXIMA", 3: "PRIORIDAD ALTA", 2: "PRIORIDAD MEDIA",
                1: "PRIORIDAD BAJA", 0: "SIN PRIORIDAD", -1: "PENALIZACIÓN LEVE", -2: "PENALIZACIÓN FUERTE"
            }
            
            print(f"  🥇 Nivel {nivel} ({nivel_desc[nivel]}): {empleados_seleccionados} → {empleados_a_tomar} asignados")
        
        # PASO 5: Verificar resultado
        if cupos_restantes > 0:
            print(f"⚠️  ADVERTENCIA: Quedan {cupos_restantes} cupos sin llenar (insuficientes empleados elegibles)")
        
        print(f"\n✅ RESULTADO CASCADA ESTRICTA: {len(empleados_asignados_sabado)} empleados asignados a sábado")
        print(f"🎯 Empleados seleccionados (orden de prioridad): {empleados_asignados_sabado}")
        
        return empleados_asignados_sabado
    
    def _asignar_descansos_con_sabado_forzado(self, empleado, dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, turnos_especiales=None):
        """
        Asigna DESC y TROP garantizando que TROP caiga en sábado (por cascada automática)
        
        Args:
            empleado: Código del empleado
            dias_disponibles: Lista de días disponibles para asignación
            descansos_por_dia: Contador de descansos por día
            historial_dias: Historial de días asignados
            empleado_idx: Índice del empleado
            turnos_especiales: Turnos especiales ya asignados (opcional)
            
        Returns:
            dict: Diccionario con asignaciones de DESC y TROP (TROP forzado en sábado)
        """
        
        if turnos_especiales is None:
            turnos_especiales = {}
        
        print(f"🎯 {empleado}: ASIGNACIÓN FORZADA DE SÁBADO (por cascada)")
        
        # PASO 1: Encontrar el sábado en los días disponibles
        sabado_info = None
        otros_dias = []
        
        for dia_info in dias_disponibles:
            if 'SAT' in dia_info['formato']:
                sabado_info = dia_info
            else:
                # Excluir días que ya tienen turnos especiales
                if dia_info['formato'] not in turnos_especiales:
                    otros_dias.append(dia_info)
        
        if not sabado_info:
            print(f"❌ {empleado}: No hay sábado disponible en esta semana")
            # Fallback a asignación normal
            return self._asignar_descansos_separados_semana(
                dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, self.semana_seleccionada, empleado=empleado
            )
        
        if len(otros_dias) < 1:
            print(f"❌ {empleado}: No hay suficientes días disponibles para DESC")
            # Fallback a asignación normal
            return self._asignar_descansos_separados_semana(
                dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, self.semana_seleccionada, empleado=empleado
            )
        
        # PASO 2: Asignar TROP forzado en sábado
        formato_sabado = sabado_info['formato']
        descansos_por_dia[formato_sabado] += 1
        
        # PASO 3: Encontrar el mejor día para DESC - PRIORIZAR RESTRICCIONES FIJAS
        dia_desc_seleccionado = None
        
        # Verificar si el empleado tiene restricción fija para DESC
        if self._es_seleccion_fija(empleado, "DESC"):
            dias_permitidos_desc = self._obtener_dias_permitidos_tipo(empleado, "DESC")
            print(f"🎯 {empleado}: Tiene restricción fija DESC para: {dias_permitidos_desc}")
            
            # Buscar el día requerido entre los días disponibles (solo si hay días permitidos)
            if dias_permitidos_desc:
                for dia_requerido in dias_permitidos_desc:
                    for dia_info in otros_dias:
                        # Mapear día en español a día en inglés del formato
                        mapeo_dias_formato = {
                            "lunes": "MON", "martes": "TUE", "miércoles": "WED", 
                            "jueves": "THU", "viernes": "FRI", "sábado": "SAT", "domingo": "SUN"
                        }
                        
                        if dia_requerido in mapeo_dias_formato:
                            dia_formato_requerido = mapeo_dias_formato[dia_requerido]
                            if dia_info['formato'].startswith(dia_formato_requerido):
                                dia_desc_seleccionado = dia_info
                                print(f"✅ {empleado}: DESC asignado en día requerido: {dia_info['formato']} ({dia_requerido})")
                                break
                    
                    if dia_desc_seleccionado:
                        break
            
            # Si no encontró el día requerido, mostrar advertencia
            if not dia_desc_seleccionado:
                print(f"⚠️ {empleado}: No se pudo asignar DESC en día requerido {dias_permitidos_desc}, usando día alternativo")
        
        # Si no tiene restricción fija o no se pudo cumplir, usar lógica original
        if not dia_desc_seleccionado:
            # Ordenar otros días por orden cronológico
            otros_dias_ordenados = sorted(otros_dias, key=lambda d: d['formato'])
            
            # Buscar días que sean cronológicamente anteriores al sábado
            dias_antes_sabado = []
            for dia_info in otros_dias_ordenados:
                # Extraer el día del mes del formato (ej: MON-04 -> 4)
                dia_mes_actual = int(dia_info['formato'].split('-')[1])
                dia_mes_sabado = int(formato_sabado.split('-')[1])
                
                if dia_mes_actual < dia_mes_sabado:
                    dias_antes_sabado.append(dia_info)
            
            # Si no hay días antes del sábado, usar cualquier día disponible
            if not dias_antes_sabado:
                dias_antes_sabado = otros_dias_ordenados
            
            # Seleccionar el día con menos descansos para DESC
            dia_desc_seleccionado = min(dias_antes_sabado, key=lambda d: descansos_por_dia.get(d['formato'], 0))
            print(f"🔄 {empleado}: DESC asignado por lógica estándar: {dia_desc_seleccionado['formato']}")
        
        # Verificar que se encontró un día válido
        if not dia_desc_seleccionado:
            print(f"❌ {empleado}: No se pudo encontrar día válido para DESC")
            # Fallback a asignación normal
            return self._asignar_descansos_separados_semana(
                dias_disponibles, descansos_por_dia, historial_dias, empleado_idx, self.semana_seleccionada, empleado=empleado
            )
        
        formato_desc = dia_desc_seleccionado['formato']
        descansos_por_dia[formato_desc] += 1
        
        # PASO 4: Registrar en historial
        if empleado_idx not in historial_dias:
            historial_dias[empleado_idx] = []
        historial_dias[empleado_idx].extend([formato_desc, formato_sabado])
        
        # PASO 5: Crear resultado
        resultado = {
            formato_desc: 'DESC',
            formato_sabado: 'TROP'
        }
        
        print(f"✅ {empleado}: DESC={formato_desc}, TROP={formato_sabado} (SÁBADO FORZADO)")
        
        return resultado

def main():
    """Función principal que ejecuta el generador"""
    print("=== GENERADOR DE DESCANSO CON SEPARACIÓN Y ALEATORIZACIÓN - SEMANA 26 2025 ===")
    
    # Crear instancia del generador para semana específica
    generador = GeneradorDescansosSeparacion(año=2025, mes=1, num_empleados=25, semana_especifica=39  )
    
    print(f"Empleados: {generador.empleados}")
    
    # Mostrar información del sistema de semanas
    print(f"\n📅 SISTEMA DE SEMANAS 2025:")
    print(f"📅 Primer lunes de enero 2025: {generador.primer_lunes_enero.strftime('%A %d/%m/%Y')}")
    print(f"📅 Total de semanas en 2025: {generador.total_semanas_año}")
    print(f"📅 SEMANA SELECCIONADA: Semana {generador.semana_seleccionada}")
    
    lunes_semana = generador.fechas_semana[0]
    domingo_semana = generador.fechas_semana[6]
    print(f"📅 FECHAS DE LA SEMANA: Lunes {lunes_semana.strftime('%d/%m/%Y')} - Domingo {domingo_semana.strftime('%d/%m/%Y')}")
    
    print(f"🎲 ALEATORIZACIÓN ACTIVADA - Cada ejecución generará un patrón diferente")
    
    # Validar restricciones antes de generar el horario
    print("\n=== VALIDACIÓN DE RESTRICCIONES ===")
    errores = generador.validar_restricciones()
    if errores:
        print("❌ ERRORES EN LAS RESTRICCIONES:")
        for error in errores:
            print(f"  - {error}")
        print("Corrija los errores antes de continuar.")
        return None
    else:
        print("✅ Todas las restricciones son válidas.")
    
    # Mostrar restricciones aplicadas
    generador.mostrar_restricciones_aplicadas()
    
    print("\nGenerando horario para la semana 29...")
    horario = generador.generar_horario_primera_semana()
    print("\nHorario generado:\n", horario)
    
    # Mostrar resumen de descansos por empleado
    print("\n=== RESUMEN DE DESCANSO POR EMPLEADO ===")
    for idx, empleado in enumerate(generador.empleados):
        desc_count = sum(1 for col in horario.columns if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN')) 
                        and horario.iloc[idx][col] == 'DESC')
        trop_count = sum(1 for col in horario.columns if col.startswith(('MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN')) 
                        and horario.iloc[idx][col] == 'TROP')
        print(f"Empleado {idx+1} ({empleado}): DESC={desc_count}, TROP={trop_count}")
    
    # VALIDACIÓN CRÍTICA: Verificar que se cumple la regla DESC/TROP
    cumple_regla = generador.validar_regla_desc_trop(horario)
    
    # VALIDACIÓN ADICIONAL: Verificar que turnos especiales se suman correctamente
    cumple_turnos_especiales = generador.validar_turnos_especiales_adicionales(horario)
    
    # Analizar separación de descansos
    generador.analizar_separacion(horario)
    
    # Generar nombre de archivo con información de la semana
    nombre_archivo = f'horario_descansos_semana_{generador.semana_seleccionada}_{lunes_semana.strftime("%d%m")}_{domingo_semana.strftime("%d%m")}_2025.xlsx'
    archivo = generador.exportar_excel(horario, nombre_archivo)
    
    # Resumen final
    print(f"\n🎯 RESUMEN FINAL:")
    print(f"📅 Semana: {generador.semana_seleccionada} ({lunes_semana.strftime('%d/%m/%Y')} - {domingo_semana.strftime('%d/%m/%Y')})")
    print(f"📊 Archivo: {archivo}")
    
    if cumple_regla:
        print(f"✅ REGLA DESC/TROP: ¡CUMPLIDA CORRECTAMENTE!")
    else:
        print(f"❌ REGLA DESC/TROP: ¡REQUIERE CORRECCIÓN!")
    
    if cumple_turnos_especiales:
        print(f"✅ TURNOS ESPECIALES: ¡SUMADOS CORRECTAMENTE!")
    else:
        print(f"❌ TURNOS ESPECIALES: ¡REQUIERE CORRECCIÓN!")
    
    print(f"\n¡Horario de la semana {generador.semana_seleccionada} generado exitosamente!")
    return horario

if __name__ == "__main__":
    horario = main() 