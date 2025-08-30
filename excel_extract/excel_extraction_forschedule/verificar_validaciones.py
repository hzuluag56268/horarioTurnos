#!/usr/bin/env python3
"""
Verificar Validaciones de Datos en Excel
"""
import openpyxl

def verificar_validaciones():
    """
    Verifica que las validaciones de datos estén correctamente aplicadas
    """
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook('TURNOS_FECHAS_ESPECIFICAS.xlsx')
        ws = wb['Turnos Específicos']
        
        print("✅ Archivo Excel cargado exitosamente")
        print(f"📊 Hojas disponibles: {wb.sheetnames}")
        
        # Verificar validaciones
        print(f"\n🔍 Validaciones aplicadas: {len(ws.data_validations)}")
        
        for i, dv in enumerate(ws.data_validations, 1):
            print(f"\n📋 Validación {i}:")
            print(f"   Tipo: {dv.type}")
            print(f"   Rangos: {dv.ranges}")
            print(f"   Fórmula: {dv.formula1}")
            if dv.showDropDown:
                print(f"   Dropdown: Sí")
            else:
                print(f"   Dropdown: No")
        
        # Verificar hojas auxiliares
        print(f"\n📝 Verificando hojas auxiliares:")
        if 'Lista_Empleados' in wb.sheetnames:
            ws_emp = wb['Lista_Empleados']
            print(f"   Lista_Empleados: {ws_emp.max_row} empleados")
            
        if 'Lista_Turnos' in wb.sheetnames:
            ws_turnos = wb['Lista_Turnos']
            print(f"   Lista_Turnos: {ws_turnos.max_row} turnos")
            
        return True
        
    except Exception as e:
        print(f"❌ Error al verificar validaciones: {e}")
        return False

if __name__ == "__main__":
    verificar_validaciones() 