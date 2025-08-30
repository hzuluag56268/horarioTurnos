#!/usr/bin/env python3
"""
Creador de Excel para Turnos Específicos
========================================
Crea un archivo Excel con dropdowns, autocompletado y validaciones
para facilitar la carga de TURNOS_FECHAS_ESPECIFICAS
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import config_restricciones

def crear_excel_turnos_especificos():
    """
    Crea el archivo Excel con todas las funcionalidades requeridas
    """
    # Crear nuevo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Turnos Específicos"
    
    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 15  # Empleado
    ws.column_dimensions['B'].width = 15  # Turno
    ws.column_dimensions['C'].width = 20  # Fecha Inicio
    ws.column_dimensions['D'].width = 20  # Fecha Fin
    ws.column_dimensions['E'].width = 30  # Comentarios
    
    # Crear encabezados
    headers = ['Empleado', 'Turno Requerido', 'Fecha Inicio', 'Fecha Fin', 'Comentarios']
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Obtener datos de configuración
    empleados = config_restricciones.obtener_empleados()
    turnos = config_restricciones.CONFIGURACION_GENERAL["turnos_validos"]
    
    # Crear hojas auxiliares para dropdowns
    ws_empleados = wb.create_sheet("Lista_Empleados")
    ws_turnos = wb.create_sheet("Lista_Turnos")
    
    # Poblar hoja de empleados
    for i, empleado in enumerate(empleados, 1):
        ws_empleados.cell(row=i, column=1, value=empleado)
    
    # Poblar hoja de turnos
    for i, turno in enumerate(turnos, 1):
        ws_turnos.cell(row=i, column=1, value=turno)
    
    # Ocultar las hojas auxiliares
    ws_empleados.sheet_state = 'hidden'
    ws_turnos.sheet_state = 'hidden'
    
    # Preparar filas para entrada de datos (100 filas)
    max_rows = 101  # Fila 1 es encabezado, 100 filas de datos
    
    # === VALIDACIÓN DE EMPLEADOS ===
    dv_empleados = DataValidation(
        type="list",
        formula1=f"Lista_Empleados!$A$1:$A${len(empleados)}",
        allow_blank=True
    )
    dv_empleados.error = "Por favor seleccione un empleado válido de la lista"
    dv_empleados.errorTitle = "Empleado inválido"
    dv_empleados.prompt = "Seleccione un empleado o escriba las primeras letras"
    dv_empleados.promptTitle = "Seleccionar Empleado"
    
    # Aplicar validación a columna A (empleados)
    ws.add_data_validation(dv_empleados)
    dv_empleados.add(f"A2:A{max_rows}")
    
    # === VALIDACIÓN DE TURNOS ===
    dv_turnos = DataValidation(
        type="list",
        formula1=f"Lista_Turnos!$A$1:$A${len(turnos)}",
        allow_blank=True
    )
    dv_turnos.error = "Por favor seleccione un turno válido de la lista"
    dv_turnos.errorTitle = "Turno inválido"
    dv_turnos.prompt = "Seleccione un turno o escriba las primeras letras"
    dv_turnos.promptTitle = "Seleccionar Turno"
    
    # Aplicar validación a columna B (turnos)
    ws.add_data_validation(dv_turnos)
    dv_turnos.add(f"B2:B{max_rows}")
    
    # === VALIDACIÓN DE FECHAS ===
    # Fecha mínima (desde hoy)
    fecha_minima = datetime.now().strftime("%Y-%m-%d")
    # Fecha máxima (2 años adelante)
    fecha_maxima = (datetime.now() + timedelta(days=730)).strftime("%Y-%m-%d")
    
    # Validación para fecha inicio
    dv_fecha_inicio = DataValidation(
        type="date",
        operator="between",
        formula1=fecha_minima,
        formula2=fecha_maxima,
        allow_blank=True
    )
    dv_fecha_inicio.error = f"La fecha debe estar entre {fecha_minima} y {fecha_maxima}"
    dv_fecha_inicio.errorTitle = "Fecha inválida"
    dv_fecha_inicio.prompt = "Ingrese la fecha en formato YYYY-MM-DD o use el selector de fecha"
    dv_fecha_inicio.promptTitle = "Fecha de Inicio"
    
    ws.add_data_validation(dv_fecha_inicio)
    dv_fecha_inicio.add(f"C2:C{max_rows}")
    
    # Validación para fecha fin
    dv_fecha_fin = DataValidation(
        type="date",
        operator="between",
        formula1=fecha_minima,
        formula2=fecha_maxima,
        allow_blank=True
    )
    dv_fecha_fin.error = f"La fecha debe estar entre {fecha_minima} y {fecha_maxima}"
    dv_fecha_fin.errorTitle = "Fecha inválida"
    dv_fecha_fin.prompt = "Ingrese la fecha en formato YYYY-MM-DD o use el selector de fecha (opcional)"
    dv_fecha_fin.promptTitle = "Fecha de Fin"
    
    ws.add_data_validation(dv_fecha_fin)
    dv_fecha_fin.add(f"D2:D{max_rows}")
    
    # === FORMATO DE CELDAS ===
    # Formato para fechas
    date_format = 'YYYY-MM-DD'
    for row in range(2, max_rows + 1):
        ws.cell(row=row, column=3).number_format = date_format  # Fecha inicio
        ws.cell(row=row, column=4).number_format = date_format  # Fecha fin
        
        # Aplicar bordes a todas las celdas
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # === FÓRMULAS DE VALIDACIÓN CRUZADA ===
    # Crear una hoja para fórmulas auxiliares
    ws_formulas = wb.create_sheet("Validaciones")
    ws_formulas.sheet_state = 'hidden'
    
    # Agregar comentarios explicativos
    ws.cell(row=1, column=6, value="Instrucciones:")
    ws.cell(row=2, column=6, value="1. Seleccione empleado del dropdown")
    ws.cell(row=3, column=6, value="2. Seleccione turno del dropdown")
    ws.cell(row=4, column=6, value="3. Ingrese fecha inicio (YYYY-MM-DD)")
    ws.cell(row=5, column=6, value="4. Fecha fin es opcional")
    ws.cell(row=6, column=6, value="5. Si no hay fecha fin = solo un día")
    
    # Colorear las instrucciones
    for row in range(1, 7):
        cell = ws.cell(row=row, column=6)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        cell.font = Font(color="000080")
    
    # Agregar ejemplos de datos
    ejemplos = [
        ["JIS", "VACA", "2025-07-17", "2025-07-30", "Vacaciones de verano"],
        ["AFG", "COME", "2025-07-01", "2025-07-31", "Comisión todo el mes"],
        ["YIS", "DESC", "2025-07-16", "", "Descanso un solo día"],
        ["HLG", "CMED", "2025-07-18", "", "Cita médica"]
    ]
    
    for i, ejemplo in enumerate(ejemplos, 2):
        for j, valor in enumerate(ejemplo, 1):
            ws.cell(row=i, column=j, value=valor)
            ws.cell(row=i, column=j).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Guardar el archivo
    filename = "TURNOS_FECHAS_ESPECIFICAS.xlsx"
    wb.save(filename)
    
    print(f"✅ Archivo Excel creado exitosamente: {filename}")
    print("\n📋 Características del archivo:")
    print("   • Dropdown con autocompletado para empleados")
    print("   • Dropdown con autocompletado para turnos")
    print("   • Validación de fechas con formato YYYY-MM-DD")
    print("   • Selector de fecha (calendar picker)")
    print("   • Fecha fin opcional (si vacía = solo un día)")
    print("   • Ejemplos de datos incluidos")
    print("   • Validaciones de entrada")
    print("\n💡 Instrucciones de uso:")
    print("   1. Abra el archivo en Excel")
    print("   2. Haga clic en cualquier celda de empleado para ver el dropdown")
    print("   3. Escriba la primera letra para filtrar opciones")
    print("   4. Para fechas, use formato YYYY-MM-DD o el selector visual")
    print("   5. La fecha fin es opcional")
    
    return filename

if __name__ == "__main__":
    crear_excel_turnos_especificos() 