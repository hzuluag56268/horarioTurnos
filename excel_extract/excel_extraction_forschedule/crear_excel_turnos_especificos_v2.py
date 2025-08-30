#!/usr/bin/env python3
"""
Creador de Excel para Turnos Específicos - Versión 2
====================================================
Versión mejorada con validaciones de datos que funcionan correctamente en Excel
"""

import pandas as pd
from datetime import datetime, timedelta
import config_restricciones

def crear_excel_turnos_especificos_v2():
    """
    Crea el archivo Excel con validaciones de datos mejoradas
    """
    # Obtener datos de configuración
    empleados = config_restricciones.obtener_empleados()
    turnos = config_restricciones.CONFIGURACION_GENERAL["turnos_validos"]
    
    # Crear DataFrame principal con ejemplos
    ejemplos = [
        ["JIS", "VACA", "2025-07-17", "2025-07-30", "Vacaciones de verano"],
        ["AFG", "COME", "2025-07-01", "2025-07-31", "Comisión todo el mes"],
        ["YIS", "DESC", "2025-07-16", None, "Descanso un solo día"],
        ["HLG", "CMED", "2025-07-18", None, "Cita médica"]
    ]
    
    # Crear más filas vacías para llenar
    for i in range(96):  # 96 filas adicionales + 4 ejemplos = 100 filas
        ejemplos.append([None, None, None, None, None])
    
    df_principal = pd.DataFrame(ejemplos, columns=[
        'Empleado', 'Turno Requerido', 'Fecha Inicio', 'Fecha Fin', 'Comentarios'
    ])
    
    # Crear DataFrame para lista de empleados
    df_empleados = pd.DataFrame(empleados, columns=['Empleados'])
    
    # Crear DataFrame para lista de turnos
    df_turnos = pd.DataFrame(turnos, columns=['Turnos'])
    
    # Crear DataFrame de instrucciones
    instrucciones = [
        ["INSTRUCCIONES DE USO:"],
        ["1. Seleccione empleado del dropdown"],
        ["2. Seleccione turno del dropdown"],
        ["3. Ingrese fecha inicio (YYYY-MM-DD)"],
        ["4. Fecha fin es opcional"],
        ["5. Si no hay fecha fin = solo un día"],
        ["6. Guarde el archivo después de llenar"],
        ["7. Ejecute: python cargar_excel_turnos.py"],
        [""],
        ["EMPLEADOS DISPONIBLES:"] + [[emp] for emp in empleados],
        [""],
        ["TURNOS DISPONIBLES:"] + [[turno] for turno in turnos]
    ]
    
    # Aplanar la lista de instrucciones
    instrucciones_flat = []
    for item in instrucciones:
        if isinstance(item, list) and len(item) == 1:
            instrucciones_flat.append(item[0])
        elif isinstance(item, list):
            instrucciones_flat.extend(item)
        else:
            instrucciones_flat.append(item)
    
    df_instrucciones = pd.DataFrame(instrucciones_flat, columns=['Instrucciones'])
    
    # Crear archivo Excel con múltiples hojas
    filename = "TURNOS_FECHAS_ESPECIFICAS.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Hoja principal
        df_principal.to_excel(writer, sheet_name='Turnos Específicos', index=False)
        
        # Hojas auxiliares
        df_empleados.to_excel(writer, sheet_name='Lista_Empleados', index=False)
        df_turnos.to_excel(writer, sheet_name='Lista_Turnos', index=False)
        df_instrucciones.to_excel(writer, sheet_name='Instrucciones', index=False)
        
        # Obtener el workbook y worksheets
        workbook = writer.book
        ws_principal = writer.sheets['Turnos Específicos']
        ws_empleados = writer.sheets['Lista_Empleados']
        ws_turnos = writer.sheets['Lista_Turnos']
        ws_instrucciones = writer.sheets['Instrucciones']
        
        # Aplicar estilos a la hoja principal
        aplicar_estilos_principal(ws_principal)
        
        # Aplicar validaciones de datos
        aplicar_validaciones_datos(ws_principal, len(empleados), len(turnos))
        
        # Ocultar hojas auxiliares
        ws_empleados.sheet_state = 'hidden'
        ws_turnos.sheet_state = 'hidden'
        
        # Aplicar estilos a instrucciones
        aplicar_estilos_instrucciones(ws_instrucciones)
    
    print(f"✅ Archivo Excel creado exitosamente: {filename}")
    print("\n📋 Características del archivo:")
    print("   • Dropdown con lista de empleados")
    print("   • Dropdown con lista de turnos")
    print("   • Validación de fechas")
    print("   • Ejemplos de datos incluidos")
    print("   • Instrucciones completas")
    print("   • Formato profesional")
    
    return filename

def aplicar_estilos_principal(ws):
    """
    Aplica estilos a la hoja principal
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 15  # Empleado
    ws.column_dimensions['B'].width = 18  # Turno
    ws.column_dimensions['C'].width = 15  # Fecha Inicio
    ws.column_dimensions['D'].width = 15  # Fecha Fin
    ws.column_dimensions['E'].width = 30  # Comentarios
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Estilo para ejemplos
    ejemplo_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Aplicar estilos a encabezados
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Aplicar estilos a ejemplos (filas 2-5)
    for row in range(2, 6):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = ejemplo_fill
    
    # Formato de fecha para columnas C y D
    for row in range(2, 102):
        ws.cell(row=row, column=3).number_format = 'YYYY-MM-DD'
        ws.cell(row=row, column=4).number_format = 'YYYY-MM-DD'

def aplicar_validaciones_datos(ws, num_empleados, num_turnos):
    """
    Aplica validaciones de datos usando el método correcto
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Validación para empleados (columna A)
    dv_empleados = DataValidation(
        type="list",
        formula1='Lista_Empleados!$A$2:$A$' + str(num_empleados + 1),
        allow_blank=True,
        showDropDown=True
    )
    dv_empleados.error = "Seleccione un empleado válido de la lista"
    dv_empleados.errorTitle = "Empleado inválido"
    dv_empleados.prompt = "Seleccione un empleado de la lista desplegable"
    dv_empleados.promptTitle = "Seleccionar Empleado"
    
    # Validación para turnos (columna B)
    dv_turnos = DataValidation(
        type="list",
        formula1='Lista_Turnos!$A$2:$A$' + str(num_turnos + 1),
        allow_blank=True,
        showDropDown=True
    )
    dv_turnos.error = "Seleccione un turno válido de la lista"
    dv_turnos.errorTitle = "Turno inválido"
    dv_turnos.prompt = "Seleccione un turno de la lista desplegable"
    dv_turnos.promptTitle = "Seleccionar Turno"
    
    # Validación para fechas
    fecha_minima = datetime.now().strftime("%Y-%m-%d")
    fecha_maxima = (datetime.now() + timedelta(days=730)).strftime("%Y-%m-%d")
    
    dv_fecha = DataValidation(
        type="date",
        operator="between",
        formula1=fecha_minima,
        formula2=fecha_maxima,
        allow_blank=True
    )
    dv_fecha.error = f"La fecha debe estar entre {fecha_minima} y {fecha_maxima}"
    dv_fecha.errorTitle = "Fecha inválida"
    dv_fecha.prompt = "Ingrese una fecha válida en formato YYYY-MM-DD"
    dv_fecha.promptTitle = "Fecha"
    
    # Aplicar validaciones
    ws.add_data_validation(dv_empleados)
    ws.add_data_validation(dv_turnos)
    ws.add_data_validation(dv_fecha)
    
    # Agregar rangos
    dv_empleados.add('A2:A101')
    dv_turnos.add('B2:B101')
    dv_fecha.add('C2:D101')

def aplicar_estilos_instrucciones(ws):
    """
    Aplica estilos a la hoja de instrucciones
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    
    # Configurar ancho de columna
    ws.column_dimensions['A'].width = 50
    
    # Estilo para títulos
    titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    titulo_font = Font(color="FFFFFF", bold=True)
    
    # Estilo para contenido
    contenido_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Aplicar estilos
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if "INSTRUCCIONES" in str(cell.value) or "EMPLEADOS" in str(cell.value) or "TURNOS" in str(cell.value):
            cell.fill = titulo_fill
            cell.font = titulo_font
        else:
            cell.fill = contenido_fill

def main():
    """
    Función principal
    """
    crear_excel_turnos_especificos_v2()
    print("\n💡 Para probar las validaciones:")
    print("1. Abra el archivo en Excel")
    print("2. Haga clic en cualquier celda de la columna 'Empleado'")
    print("3. Debe aparecer una flecha de dropdown")
    print("4. Haga clic en la flecha para ver la lista")
    print("5. Lo mismo para la columna 'Turno Requerido'")

if __name__ == "__main__":
    main() 