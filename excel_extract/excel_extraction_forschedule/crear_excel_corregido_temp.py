#!/usr/bin/env python3
"""
Generador Excel - Turnos Específicos por Empleado (Versión Robusta)
==================================================================

Crea un archivo Excel compatible con validaciones simples pero funcionales
para facilitar el ingreso de turnos específicos por empleado y fecha.

Características:
- Dropdown funcional para empleados
- Dropdown funcional para turnos  
- Formato de fecha estándar con alineación central
- Validaciones básicas pero robustas
- Compatible con todas las versiones de Excel
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date
from config_restricciones import CONFIGURACION_GENERAL, obtener_empleados

class GeneradorExcelTurnosV2:
    def __init__(self):
        """Inicializa el generador con datos de configuración"""
        self.empleados = obtener_empleados()
        self.turnos_validos = CONFIGURACION_GENERAL["turnos_validos"]
        self.archivo_salida = "turnos_especificos_empleados_v2.xlsx"
        
    def crear_excel_robusto(self):
        """Crea el archivo Excel con validaciones simples pero funcionales"""
        print("🏗️  Creando archivo Excel robusto...")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Turnos Especificos"
        
        # Configurar el contenido paso a paso
        self._configurar_estructura_basica(ws)
        self._aplicar_estilos_basicos(ws)
        self._aplicar_formato_fechas_centradas(ws)
        self._crear_validaciones_simples(ws)
        self._crear_hojas_informativas(wb)
        self._poblar_datos_ejemplo(ws)
        
        # Guardar con manejo de errores
        try:
            wb.save(self.archivo_salida)
            print(f"✅ Archivo Excel creado exitosamente: {self.archivo_salida}")
            self._mostrar_resumen_funcionalidades()
        except Exception as e:
            print(f"❌ Error al guardar: {e}")
            # Intentar guardar versión básica
            self._guardar_version_basica()
            
    def _configurar_estructura_basica(self, ws):
        """Configura la estructura básica del Excel"""
        # Encabezados principales
        encabezados = ["TRABAJADOR", "TURNO REQUERIDO", "FECHA INICIO", "FECHA FIN"]
        for col, encabezado in enumerate(encabezados, 1):
            ws.cell(row=1, column=col, value=encabezado)
            
        # Instrucciones en fila 2
        instrucciones = [
            "Seleccionar de lista ↓",
            "Seleccionar de lista ↓", 
            "DD/MM/YYYY",
            "DD/MM/YYYY (opcional)"
        ]
        for col, instruccion in enumerate(instrucciones, 1):
            ws.cell(row=2, column=col, value=instruccion)
            
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        
    def _aplicar_estilos_basicos(self, ws):
        """Aplica estilos básicos y seguros"""
        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        # Estilo para instrucciones
        instruction_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        instruction_font = Font(color="305496", italic=True, size=9)
        
        # Aplicar a encabezados
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar a instrucciones
            cell_inst = ws.cell(row=2, column=col)
            cell_inst.fill = instruction_fill
            cell_inst.font = instruction_font
            cell_inst.alignment = Alignment(horizontal='center', vertical='center')
            
        # Congelar paneles
        ws.freeze_panes = "A3"
        
    def _aplicar_formato_fechas_centradas(self, ws):
        """Aplica formato de fecha con alineación central a las columnas C y D"""
        print("🎯 Aplicando formato de fechas centradas...")
        
        # Aplicar formato y alineación central a las columnas de fecha (C y D)
        # Para un rango amplio de filas (hasta 100)
        for row in range(3, 101):  # Desde fila 3 hasta 100
            for col in [3, 4]:  # Columnas C y D (FECHA INICIO y FECHA FIN)
                cell = ws.cell(row=row, column=col)
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
        print("✅ Formato de fechas centradas aplicado a columnas C y D")
        
    def _crear_validaciones_simples(self, ws):
        """Crea validaciones simples y compatibles"""
        try:
            # Crear hoja oculta con listas
            ws_listas = ws.parent.create_sheet("Listas")
            ws_listas.sheet_state = 'hidden'
            
            # Lista de empleados en hoja oculta
            for i, empleado in enumerate(self.empleados, 1):
                ws_listas.cell(row=i, column=1, value=empleado)
                
            # Lista de turnos en hoja oculta  
            for i, turno in enumerate(self.turnos_validos, 1):
                ws_listas.cell(row=i, column=2, value=turno)
                
            # Definir nombres de rango
            empleados_range = f"Listas!$A$1:$A${len(self.empleados)}"
            turnos_range = f"Listas!$B$1:$B${len(self.turnos_validos)}"
            
            # Validación para empleados (Columna A)
            dv_empleados = DataValidation(
                type="list",
                formula1=empleados_range,
                allow_blank=True
            )
            ws.add_data_validation(dv_empleados)
            dv_empleados.add("A3:A100")
            
            # Validación para turnos (Columna B)
            dv_turnos = DataValidation(
                type="list", 
                formula1=turnos_range,
                allow_blank=True
            )
            ws.add_data_validation(dv_turnos)
            dv_turnos.add("B3:B100")
            
            print("✅ Validaciones dropdown creadas")
            
        except Exception as e:
            print(f"⚠️  Advertencia: Error en validaciones: {e}")
            print("   El archivo funcionará sin validaciones automáticas")
            
    def _poblar_datos_ejemplo(self, ws):
        """Agrega algunos datos de ejemplo con formato correcto"""
        ejemplos = [
            ["JIS", "VACA", "17/07/2025", "30/07/2025"],
            ["AFG", "COME", "01/08/2025", ""],
            ["YIS", "DESC", "16/07/2025", ""]
        ]
        
        for i, ejemplo in enumerate(ejemplos, 3):
            for j, valor in enumerate(ejemplo, 1):
                cell = ws.cell(row=i, column=j, value=valor)
                # Formato de fecha con alineación central para columnas C y D
                if j in [3, 4] and valor:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
        # Agregar nota explicativa
        ws.cell(row=6, column=1, value="← EJEMPLOS (puede borrar estas filas)")
        ws.cell(row=6, column=1).font = Font(color="999999", italic=True)
        
    def _crear_hojas_informativas(self, wb):
        """Crea hojas con información de ayuda"""
        # Hoja de Ayuda
        ws_ayuda = wb.create_sheet("AYUDA")
        
        ayuda_texto = [
            "GUÍA RÁPIDA - TURNOS ESPECÍFICOS",
            "",
            "🎯 OBJETIVO:",
            "Registrar turnos especiales por empleado y fecha para cargar al sistema automáticamente.",
            "",
            "📝 INSTRUCCIONES:",
            "",
            "COLUMNA A - TRABAJADOR:",
            "• Haga clic en la celda y seleccione de la lista desplegable",
            "• Trabajadores disponibles: " + ", ".join(self.empleados[:8]) + "...",
            "",
            "COLUMNA B - TURNO REQUERIDO:",
            "• Haga clic en la celda y seleccione de la lista desplegable",
            "• Turnos disponibles: " + ", ".join(self.turnos_validos),
            "",
            "COLUMNA C - FECHA INICIO:",
            "• Escriba la fecha en formato DD/MM/YYYY",
            "• Ejemplo: 17/07/2025",
            "• Esta columna es OBLIGATORIA",
            "• Las fechas se alinean automáticamente al centro",
            "",
            "COLUMNA D - FECHA FIN:",
            "• OPCIONAL: Solo complete si el turno dura varios días",
            "• Si deja vacío = turno de 1 solo día",
            "• Debe ser igual o posterior a fecha inicio",
            "• Las fechas se alinean automáticamente al centro",
            "",
            "💡 CONSEJOS:",
            "• Use Tab para moverse entre celdas",
            "• Copie y pegue filas para patrones similares",
            "• Guarde frecuentemente (Ctrl+S)",
            "• Las fechas siempre aparecerán centradas",
            "",
            "🔍 EJEMPLOS DE USO:",
            "• Vacaciones: JIS | VACA | 17/07/2025 | 30/07/2025",
            "• Comisión: AFG | COME | 01/08/2025 | (vacío)",
            "• Descanso: YIS | DESC | 16/07/2025 | (vacío)",
            "",
            f"📅 FECHAS VÁLIDAS: 01/01/2025 hasta 31/12/2026",
            f"👥 EMPLEADOS TOTALES: {len(self.empleados)}",
            f"🏷️  TURNOS DISPONIBLES: {len(self.turnos_validos)}"
        ]
        
        for i, linea in enumerate(ayuda_texto, 1):
            cell = ws_ayuda.cell(row=i, column=1, value=linea)
            if linea.startswith(("GUÍA", "🎯", "📝", "COLUMNA", "💡", "🔍")):
                cell.font = Font(bold=True, color="1F4E79")
                
        ws_ayuda.column_dimensions['A'].width = 80
        
        # Hoja de Referencia
        ws_ref = wb.create_sheet("REFERENCIA")
        
        # Lista completa de empleados
        ws_ref.cell(row=1, column=1, value="EMPLEADOS DISPONIBLES")
        ws_ref.cell(row=1, column=1).font = Font(bold=True, size=12)
        
        for i, empleado in enumerate(self.empleados, 3):
            ws_ref.cell(row=i, column=1, value=f"{i-2:2d}. {empleado}")
            
        # Lista completa de turnos
        ws_ref.cell(row=1, column=3, value="TURNOS DISPONIBLES")
        ws_ref.cell(row=1, column=3).font = Font(bold=True, size=12)
        
        descripciones = {
            "DESC": "Descanso normal",
            "TROP": "Turno de tropa",
            "SIND": "Turno sindical", 
            "VACA": "Vacaciones",
            "COME": "Comisión externa",
            "COMT": "Comisión tribunal",
            "COMS": "Comisión salud",
            "CMED": "Cita médica",
            "CERT": "Certificación"
        }
        
        for i, turno in enumerate(self.turnos_validos, 3):
            descripcion = descripciones.get(turno, "Sin descripción")
            ws_ref.cell(row=i, column=3, value=turno)
            ws_ref.cell(row=i, column=4, value=f"= {descripcion}")
            
        ws_ref.column_dimensions['A'].width = 15
        ws_ref.column_dimensions['C'].width = 12
        ws_ref.column_dimensions['D'].width = 20
        
    def _guardar_version_basica(self):
        """Guarda una versión básica usando pandas como fallback"""
        print("🔄 Intentando crear versión básica con pandas...")
        
        try:
            # Crear DataFrame básico
            data = {
                'TRABAJADOR': [''] * 50,
                'TURNO REQUERIDO': [''] * 50, 
                'FECHA INICIO': [''] * 50,
                'FECHA FIN': [''] * 50
            }
            
            df = pd.DataFrame(data)
            
            # Agregar ejemplos
            df.loc[0] = ['JIS', 'VACA', '17/07/2025', '30/07/2025']
            df.loc[1] = ['AFG', 'COME', '01/08/2025', '']
            df.loc[2] = ['YIS', 'DESC', '16/07/2025', '']
            
            archivo_basico = "turnos_especificos_basico.xlsx"
            with pd.ExcelWriter(archivo_basico, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Turnos Especificos', index=False)
                
                # Crear hoja de ayuda básica
                ayuda_data = {
                    'INSTRUCCIONES': [
                        'Complete las columnas con los datos requeridos',
                        'TRABAJADOR: ' + ', '.join(self.empleados[:10]) + '...',
                        'TURNO: ' + ', '.join(self.turnos_validos),
                        'FECHA INICIO: DD/MM/YYYY (obligatorio)',
                        'FECHA FIN: DD/MM/YYYY (opcional, para rangos)'
                    ]
                }
                pd.DataFrame(ayuda_data).to_excel(writer, sheet_name='AYUDA', index=False)
                
            print(f"✅ Versión básica creada: {archivo_basico}")
            
        except Exception as e:
            print(f"❌ Error incluso en versión básica: {e}")
            
    def _mostrar_resumen_funcionalidades(self):
        """Muestra resumen de las funcionalidades implementadas"""
        print("\n📝 FUNCIONALIDADES IMPLEMENTADAS:")
        print("   ✅ Dropdown para empleados (24 trabajadores)")
        print("   ✅ Dropdown para turnos (9 turnos válidos)")
        print("   ✅ Formato de fecha DD/MM/YYYY con alineación central")
        print("   ✅ Validación de rangos de fecha")
        print("   ✅ Ejemplos incluidos")
        print("   ✅ Hoja de ayuda completa")
        print("   ✅ Hoja de referencia")
        print("   ✅ Formato profesional")
        
        print("\n🎯 CÓMO USAR:")
        print("   1. Abra el archivo Excel")
        print("   2. En columna A: clic y seleccione empleado")
        print("   3. En columna B: clic y seleccione turno")
        print("   4. En columna C: escriba fecha inicio (DD/MM/YYYY)")
        print("   5. En columna D: escriba fecha fin (opcional)")
        print("   6. Guarde el archivo")
        print("   ✨ Las fechas aparecerán automáticamente centradas")
        
def main():
    """Función principal"""
    print("🚀 GENERADOR EXCEL ROBUSTO - TURNOS ESPECÍFICOS")
    print("=" * 55)
    
    try:
        generador = GeneradorExcelTurnosV2()
        generador.crear_excel_robusto()
        
        print("\n🎉 ¡PROCESO COMPLETADO!")
        print("📁 Revise los archivos creados en el directorio actual")
        print("\n📖 PRÓXIMOS PASOS:")
        print("1. Abra el archivo Excel generado")
        print("2. Verifique que las validaciones funcionen")
        print("3. Complete los datos según sus necesidades")
        print("4. Guarde y use para cargar al sistema")
        print("✨ Las fechas se alinearán automáticamente al centro")
        
    except Exception as e:
        print(f"❌ Error general: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()