"""
Sistema de Asignación Automática de Turnos para Cuentas 1 y 2 (Sábados y Domingos)

Este script automatiza la asignación de trabajadores a turnos específicos basándose en:
- Configuración leída desde la hoja 'config' del archivo Excel
- Historial de fechas trabajadas por cada empleado
- Restricciones de trabajadores a omitir por turno

RESTRICCIONES Y REGLAS DE ASIGNACIÓN:
1. Una sigla por fecha: No se puede asignar el mismo trabajador a múltiples fechas
2. Prioridad por antigüedad: Se asigna primero a quien no ha trabajado por más tiempo
3. Desempate por carga: En caso de empate, se asigna al que tiene menos fechas totales
4. Desempate alfabético: Si persiste el empate, se usa orden alfabético de siglas
5. Trabajadores omitidos: Se excluyen los trabajadores especificados en la configuración
6. Validación de elegibles: Si no hay trabajadores elegibles, se omite el turno
7. No duplicación: Al escribir resultados, se evitan fechas duplicadas en la misma fila

ESTRUCTURA DE LA HOJA 'config':
- Columna A: Nombre del turno (debe coincidir con una hoja del Excel)
- Columna B: Fechas a solicitar (formato JSON array: ["08-07", "08-10", ...])
- Columna C: Trabajadores a omitir (formato JSON array: ["MEI", "VCM", ...])

SALIDAS GENERADAS:
1. Archivo Excel: cuentas1y2sabadosDomingo_asignado.xlsx
   - Se añaden las nuevas asignaciones a cada hoja de turno
   - Se incluye una tabla de resumen al final de cada hoja
   - Las nuevas asignaciones se colorean con un color aleatorio para facilitar identificación
2. Archivo JSON: cuentas1y2sabadosDomingo_asignado.json
   - Estructura agrupada por turno con fecha y trabajador asignado

USO:
    python seleccion_sabados_festivos.py

DEPENDENCIAS:
    openpyxl, datetime, heapq, json, ast, collections, pathlib, typing, random
"""

from datetime import datetime, date
import heapq
import json
import ast
import random
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import PatternFill

# Configuración
WORKBOOK_PATH = Path(__file__).with_name("cuentas1y2sabadosDomingo.xlsx")
SHEET_NAME = "BANTdom"

# Configuración por defecto (se sobrescribirá desde la hoja config)
INPUT_DATES_RAW = ["08-07","08-10", "08-17", "08-18"]
OMITIR = {"MEI", "VCM", "ROP", "WEH", "PHD"}


def norm_str(s: Optional[str]) -> str:
    if s is None:
        return ""
    return str(s).strip().upper()


def infer_default_year(all_dates: List[date]) -> int:
    years = [d.year for d in all_dates if isinstance(d, date)]
    return max(years) if years else date.today().year


def parse_cell_date(value, default_year: int) -> Optional[date]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip()
        fmts = (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%m-%d",
            "%m/%d",
            "%d-%m",
            "%d/%m",
        )
        for fmt in fmts:
            try:
                dt = datetime.strptime(s, fmt)
                if "%Y" in fmt:
                    return dt.date()
                return dt.replace(year=default_year).date()
            except Exception:
                continue
    return None


def parse_input_dates(raw_dates: List[str], default_year: int) -> List[date]:
    out: List[date] = []
    for s in raw_dates:
        s2 = s.strip()
        parsed = False
        for fmt in (
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%m-%d-%Y",
            "%m-%d",
            "%m/%d",
            "%d-%m",
            "%d/%m",
        ):
            try:
                dt = datetime.strptime(s2, fmt)
                out.append(dt.date() if "%Y" in fmt else dt.replace(year=default_year).date())
                parsed = True
                break
            except Exception:
                continue
        if not parsed:
            raise ValueError(f"No puedo interpretar la fecha de entrada: {s}")
    return sorted(out)


def read_config_from_sheet(wb) -> Dict[str, Dict]:
    """
    Lee la configuración desde la hoja 'config'.
    Espera una estructura como:
    - Columna A: Nombre del turno
    - Columna B: Fechas (en formato JSON array)
    - Columna C: Trabajadores a omitir (en formato JSON set)
    """
    config = {}
    
    if "config" not in wb.sheetnames:
        print("Advertencia: No se encontró la hoja 'config'. Usando configuración por defecto.")
        return {}
    
    ws_config = wb["config"]
    
    for row in range(2, ws_config.max_row + 1):  # Empezar desde la fila 2 (asumiendo headers en fila 1)
        turno = norm_str(ws_config.cell(row=row, column=1).value)
        if not turno:
            continue
            
        # Leer fechas (formato JSON array)
        fechas_raw = ws_config.cell(row=row, column=2).value
        fechas = []
        if fechas_raw:
            try:
                if isinstance(fechas_raw, str):
                    # Limpiar caracteres especiales y parsear JSON
                    fechas_raw_clean = fechas_raw.replace('\xa0', '').strip()
                    fechas = json.loads(fechas_raw_clean)
                else:
                    fechas = [str(fechas_raw).strip()]
            except json.JSONDecodeError as e:
                print(f"Error parseando fechas para {turno}: {e}")
                continue
        
        # Leer trabajadores a omitir (formato set literal de Python)
        omitir_raw = ws_config.cell(row=row, column=3).value
        omitir = set()
        if omitir_raw:
            try:
                if isinstance(omitir_raw, str):
                    # Limpiar caracteres especiales
                    omitir_raw_clean = omitir_raw.replace('\xa0', '').strip()
                    
                    # Intentar parsear como JSON primero
                    try:
                        omitir_list = json.loads(omitir_raw_clean)
                        omitir = {norm_str(w) for w in omitir_list}
                    except json.JSONDecodeError:
                        # Si falla JSON, intentar como set literal de Python
                        try:
                            omitir_set = ast.literal_eval(omitir_raw_clean)
                            if isinstance(omitir_set, set):
                                omitir = {norm_str(w) for w in omitir_set}
                            else:
                                omitir = {norm_str(str(omitir_set))}
                        except (ValueError, SyntaxError):
                            # Fallback: separar por comas
                            omitir = {norm_str(w.strip()) for w in omitir_raw_clean.split(',') if w.strip()}
                else:
                    omitir = {norm_str(str(omitir_raw))}
            except Exception as e:
                print(f"Error parseando omitir para {turno}: {e}")
                continue
        
        config[turno] = {
            'fechas': fechas,
            'omitir': omitir
        }
    
    return config


def extract_worker_dates_from_ws(ws):
    # Recolectar datos crudos
    all_seen_values = []
    raw_rows = []
    for r in range(2, ws.max_row + 1):
        sigla = norm_str(ws.cell(row=r, column=1).value)
        if not sigla:
            continue
        row_values = [ws.cell(row=r, column=c).value for c in range(2, ws.max_column + 1)]
        raw_rows.append((sigla, row_values))
        for v in row_values:
            if v not in (None, ""):
                all_seen_values.append(v)

    # Inferir año por muestras
    parsed_samples: List[date] = []
    for v in all_seen_values:
        if isinstance(v, (date, datetime)):
            parsed_samples.append(v if isinstance(v, date) else v.date())
        elif isinstance(v, (int, float)):
            try:
                parsed_samples.append(from_excel(v).date())
            except Exception:
                pass
        elif isinstance(v, str):
            s = v.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
                try:
                    parsed_samples.append(datetime.strptime(s, fmt).date())
                    break
                except Exception:
                    continue
    default_year = infer_default_year(parsed_samples)

    # Parseo definitivo
    worker_dates: Dict[str, List[date]] = {}
    for sigla, row_values in raw_rows:
        dates_parsed: List[date] = []
        for v in row_values:
            d = parse_cell_date(v, default_year)
            if d:
                dates_parsed.append(d)
        worker_dates[sigla] = sorted(set(dates_parsed))

    return worker_dates, default_year


def load_workers_and_dates(path: Path, sheet_name: str):
    wb = load_workbook(path)
    print("Hojas en el libro:", wb.sheetnames)
    if sheet_name not in wb.sheetnames:
        fallback = wb.sheetnames[0]
        print(f"Hoja '{sheet_name}' no encontrada; usando '{fallback}'")
        ws = wb[fallback]
    else:
        ws = wb[sheet_name]

    all_seen_dates = []
    raw_rows = []
    for r in range(2, ws.max_row + 1):
        sigla = norm_str(ws.cell(row=r, column=1).value)
        if not sigla:
            continue
        row_values = [ws.cell(row=r, column=c).value for c in range(2, ws.max_column + 1)]
        raw_rows.append((sigla, row_values))
        for v in row_values:
            if v not in (None, ""):
                all_seen_dates.append(v)

    print("Muestra de siglas:", [sr[0] for sr in raw_rows[:10]])

    parsed_samples: List[date] = []
    for v in all_seen_dates:
        if isinstance(v, (date, datetime)):
            parsed_samples.append(v if isinstance(v, date) else v.date())
        elif isinstance(v, (int, float)):
            try:
                parsed_samples.append(from_excel(v).date())
            except Exception:
                pass
        elif isinstance(v, str):
            s = v.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
                try:
                    parsed_samples.append(datetime.strptime(s, fmt).date())
                    break
                except Exception:
                    continue

    default_year = infer_default_year(parsed_samples)
    print("Año por defecto inferido:", default_year)

    worker_dates: Dict[str, List[date]] = {}
    for sigla, row_values in raw_rows:
        dates_parsed: List[date] = []
        for v in row_values:
            d = parse_cell_date(v, default_year)
            if d:
                dates_parsed.append(d)
        worker_dates[sigla] = sorted(set(dates_parsed))

    preview = {s: (max(ds).strftime("%Y-%m-%d") if ds else "") for s, ds in list(worker_dates.items())[:10]}
    print("Última fecha por sigla (muestra):", preview)

    return wb, ws, worker_dates, default_year


def build_priority(worker_dates: Dict[str, List[date]], omit: set):
    heap = []
    for sigla, dates_list in worker_dates.items():
        if sigla in omit:
            continue
        if not dates_list:
            last_d, total = date.min, 0
        else:
            last_d, total = max(dates_list), len(dates_list)
        heap.append([last_d, total, sigla])
    heap.sort()
    return heap


def assign_dates(priority_heap: List, input_dates: List[date]):
    heapq.heapify(priority_heap)
    if not priority_heap:
        raise RuntimeError("No hay trabajadores elegibles.")
    assignments: Dict[date, str] = {}
    by_worker: Dict[str, List[date]] = defaultdict(list)
    for d in input_dates:
        last_d, total, sigla = heapq.heappop(priority_heap)
        assignments[d] = sigla
        by_worker[sigla].append(d)
        heapq.heappush(priority_heap, [d, total + 1, sigla])
    return assignments, by_worker


def write_results(wb, ws, by_worker: Dict[str, List[date]], output_assignments: Dict[date, str], fill_color: PatternFill):
    # Mapear sigla -> fila
    sigla_to_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        sigla = norm_str(ws.cell(row=r, column=1).value)
        if sigla:
            sigla_to_row[sigla] = r

    # Escribir fechas asignadas en la primera(s) columna(s) vacía(s) de cada trabajador
    for sigla, fechas in by_worker.items():
        r = sigla_to_row.get(sigla)
        if not r:
            continue
        # Detectar la última columna no vacía en la fila (2..max_col) y escribir a continuación.
        last_filled = 1
        existing_dates = set()
        for c_scan in range(2, ws.max_column + 1):
            val = ws.cell(row=r, column=c_scan).value
            if val not in (None, ""):
                last_filled = c_scan
                # Normalizar a date
                if hasattr(val, 'date'):
                    existing_dates.add(val.date())
                else:
                    existing_dates.add(val)
        c = last_filled + 1
        for d in sorted(fechas):
            # Evitar duplicar si ya existe esa fecha en la fila
            if d in existing_dates:
                continue
            cell = ws.cell(row=r, column=c, value=d)
            cell.fill = fill_color  # Aplicar color de fondo
            existing_dates.add(d)
            c += 1

    # Agregar tabla de resumen al final de la misma hoja, dejando una fila en blanco
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="Fecha")
    ws.cell(row=start_row, column=2, value="Trabajador")
    row = start_row + 1
    for d in sorted(output_assignments.keys()):
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=2, value=output_assignments[d])
        row += 1


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"No encuentro el archivo: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH)
    print("Hojas en el libro:", wb.sheetnames)

    # Generar color aleatorio para esta ejecución
    colors = [
        "FFB6C1",  # Rosa claro
        "90EE90",  # Verde claro
        "87CEEB",  # Azul cielo
        "DDA0DD",  # Ciruela claro
        "F0E68C",  # Amarillo kaki
        "98FB98",  # Verde pálido
        "FFA07A",  # Salmón claro
        "B0C4DE",  # Azul acero claro
        "FFE4B5",  # Melocotón
        "E6E6FA",  # Lavanda
        "F5DEB3",  # Trigo
        "D8BFD8",  # Cardo
        "FFDAB9",  # Melocotón claro
        "B0E0E6",  # Azul polvo
        "F0FFF0",  # Miel
    ]
    selected_color = random.choice(colors)
    fill_color = PatternFill(start_color=selected_color, end_color=selected_color, fill_type="solid")
    
    print(f"Color seleccionado para esta ejecución: #{selected_color}")

    json_by_turno: Dict[str, List[Dict[str, str]]] = {}

    config = read_config_from_sheet(wb)

    for turno, data in config.items():
        print(f"\nProcesando turno: {turno}")
        
        # Verificar que la hoja del turno existe (comparación insensible a mayúsculas)
        sheet_found = None
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == turno.upper():
                sheet_found = sheet_name
                break
        
        if not sheet_found:
            print(f"Advertencia: La hoja '{turno}' no existe. Saltando este turno.")
            continue
            
        input_dates = parse_input_dates(data['fechas'], date.today().year) # Usar año actual para parsear
        print("Fechas de entrada normalizadas:", [d.strftime("%Y-%m-%d") for d in input_dates])
        print("Omitidos:", sorted(data['omitir']))

        worker_dates, default_year = extract_worker_dates_from_ws(wb[sheet_found]) # Cargar fechas del turno
        priority = build_priority(worker_dates, data['omitir'])
        print("Total elegibles:", len(priority))
        if not priority:
            print("No hay elegibles; se omite el turno.")
            continue

        assignments, by_worker = assign_dates(priority, input_dates)
        print("Asignaciones:")
        for d in sorted(assignments.keys()):
            print(f"{d.strftime('%Y-%m-%d')} -> {assignments[d]}")

        # Para JSON agrupado por turno (nombre de la hoja)
        json_by_turno[turno] = [
            {"fecha": d.strftime("%Y-%m-%d"), "trabajador": assignments[d]}
            for d in sorted(assignments.keys())
        ]

        write_results(wb, wb[sheet_found], by_worker, assignments, fill_color) # Escribir en la hoja del turno

    out_path = WORKBOOK_PATH.with_name(WORKBOOK_PATH.stem + "_asignado.xlsx")
    wb.save(out_path)
    print(f"\nGuardado en: {out_path}")

    # Guardar JSON agrupado por turno
    json_path = WORKBOOK_PATH.with_name(WORKBOOK_PATH.stem + "_asignado.json")
    json_path.write_text(json.dumps(json_by_turno, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"JSON guardado en: {json_path}")


if __name__ == "__main__":
    main() 