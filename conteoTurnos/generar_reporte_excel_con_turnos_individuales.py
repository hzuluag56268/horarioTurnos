import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import defaultdict

def generar_reporte_excel_con_turnos_individuales(archivo_excel, archivo_salida="reporte_parejas_turnos_con_individuales.xlsx"):
    """
    Genera un reporte Excel con formato profesional del conteo de parejas de turnos,
    incluyendo cuatro columnas de sumatoria y conteo de turnos individuales BANT y BLPT:
    1. MAST/NANR + MLPR/NLPR + BLPT/NLPR
    2. TAST/SLN4 + TAST/SLN3
    3. TANT/NANT + TLPT/NLPT
    4. Resto de parejas de turnos
    5. Conteo individual de BANT
    6. Conteo individual de BLPT
    
    Args:
        archivo_excel (str): Ruta del archivo Excel
        archivo_salida (str): Nombre del archivo Excel de salida
    """
    try:
        # Leer el archivo Excel directamente con openpyxl
        wb_original = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws_original = wb_original.active
        
        # Encontrar la fila de encabezados (días)
        fila_encabezados = None
        for row in range(1, ws_original.max_row + 1):
            for col in range(1, ws_original.max_column + 1):
                cell_value = ws_original.cell(row=row, column=col).value
                if cell_value and str(cell_value).startswith(('FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED', 'THU')):
                    fila_encabezados = row
                    break
            if fila_encabezados:
                break
        
        if fila_encabezados is None:
            print("No se encontraron encabezados de días")
            return
        
        # Encontrar las filas de trabajadores
        filas_trabajadores = []
        for row in range(fila_encabezados + 1, ws_original.max_row + 1):
            cell_no = ws_original.cell(row=row, column=1).value
            cell_codigo = ws_original.cell(row=row, column=2).value
            if cell_no and cell_codigo and str(cell_no).isdigit():
                filas_trabajadores.append(row)
        
        # Recolectar todas las parejas únicas y turnos individuales
        todas_parejas = set()
        todos_turnos_individuales = set()
        datos_trabajadores = []
        
        for fila_idx in filas_trabajadores:
            numero_trabajador = ws_original.cell(row=fila_idx, column=1).value
            codigo_trabajador = ws_original.cell(row=fila_idx, column=2).value
            
            # Contar parejas de turnos y turnos individuales
            parejas_encontradas = defaultdict(int)
            turnos_individuales = defaultdict(int)
            
            # Recorrer las columnas de días (cada día tiene 2 columnas)
            for col in range(3, ws_original.max_column + 1, 2):
                if col + 1 <= ws_original.max_column:
                    turno1 = ws_original.cell(row=fila_idx, column=col).value
                    turno2 = ws_original.cell(row=fila_idx, column=col + 1).value
                    
                    # Contar turnos individuales
                    if turno1:
                        turnos_individuales[str(turno1)] += 1
                        todos_turnos_individuales.add(str(turno1))
                    if turno2:
                        turnos_individuales[str(turno2)] += 1
                        todos_turnos_individuales.add(str(turno2))
                    
                    # Crear la pareja de turnos
                    if turno1 and turno2:
                        pareja = f"{turno1}/{turno2}"
                        parejas_encontradas[pareja] += 1
                        todas_parejas.add(pareja)
            
            # Guardar datos del trabajador
            datos_trabajadores.append({
                'No.': numero_trabajador,
                'Codigo': codigo_trabajador,
                'Parejas': dict(parejas_encontradas),
                'Turnos_Individuales': dict(turnos_individuales)
            })
        
        # Crear lista ordenada de parejas
        parejas_ordenadas = sorted(todas_parejas)
        
        # Definir las parejas para las tres primeras sumatorias
        parejas_sumatoria1 = ['MAST/NANR', 'MLPR/NLPR', 'BLPT/NLPR']
        parejas_sumatoria2 = ['TAST/SLN4', 'TAST/SLN3']
        parejas_sumatoria3 = ['TANT/NANT', 'TLPT/NLPT']
        
        # Obtener todas las parejas incluidas en las tres primeras sumatorias
        parejas_incluidas = set(parejas_sumatoria1 + parejas_sumatoria2 + parejas_sumatoria3)
        
        # Obtener las parejas restantes (para la cuarta sumatoria)
        parejas_resto = [pareja for pareja in parejas_ordenadas if pareja not in parejas_incluidas]
        
        # Crear archivo Excel con formato
        crear_excel_formateado_con_turnos_individuales(datos_trabajadores, parejas_ordenadas, archivo_salida, 
                                                     parejas_sumatoria1, parejas_sumatoria2, parejas_sumatoria3, parejas_resto)
        
        print(f"Reporte Excel generado exitosamente: {archivo_salida}")
        print(f"Total de trabajadores: {len(datos_trabajadores)}")
        print(f"Total de parejas diferentes: {len(parejas_ordenadas)}")
        print(f"Total de turnos individuales diferentes: {len(todos_turnos_individuales)}")
        
        # Calcular y mostrar resumen de las cuatro sumatorias y turnos individuales
        print(f"\nResumen de las cuatro sumatorias y turnos individuales:")
        
        # Calcular sumatorias
        total_sumatoria1 = 0
        total_sumatoria2 = 0
        total_sumatoria3 = 0
        total_sumatoria4 = 0
        trabajadores_con_sumatoria1 = 0
        trabajadores_con_sumatoria2 = 0
        trabajadores_con_sumatoria3 = 0
        trabajadores_con_sumatoria4 = 0
        
        # Calcular turnos individuales
        total_bant = 0
        total_blpt = 0
        trabajadores_con_bant = 0
        trabajadores_con_blpt = 0
        
        for trabajador in datos_trabajadores:
            sumatoria1 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria1)
            sumatoria2 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria2)
            sumatoria3 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria3)
            sumatoria4 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_resto)
            
            # Conteo de turnos individuales
            bant_count = trabajador['Turnos_Individuales'].get('BANT', 0)
            blpt_count = trabajador['Turnos_Individuales'].get('BLPT', 0)
            
            total_sumatoria1 += sumatoria1
            total_sumatoria2 += sumatoria2
            total_sumatoria3 += sumatoria3
            total_sumatoria4 += sumatoria4
            total_bant += bant_count
            total_blpt += blpt_count
            
            if sumatoria1 > 0:
                trabajadores_con_sumatoria1 += 1
            if sumatoria2 > 0:
                trabajadores_con_sumatoria2 += 1
            if sumatoria3 > 0:
                trabajadores_con_sumatoria3 += 1
            if sumatoria4 > 0:
                trabajadores_con_sumatoria4 += 1
            if bant_count > 0:
                trabajadores_con_bant += 1
            if blpt_count > 0:
                trabajadores_con_blpt += 1
        
        print(f"1. MAST/NANR + MLPR/NLPR + BLPT/NLPR: {total_sumatoria1} veces en {trabajadores_con_sumatoria1} trabajadores")
        print(f"2. TAST/SLN4 + TAST/SLN3: {total_sumatoria2} veces en {trabajadores_con_sumatoria2} trabajadores")
        print(f"3. TANT/NANT + TLPT/NLPT: {total_sumatoria3} veces en {trabajadores_con_sumatoria3} trabajadores")
        print(f"4. Resto de parejas ({', '.join(parejas_resto)}): {total_sumatoria4} veces en {trabajadores_con_sumatoria4} trabajadores")
        print(f"5. Turno BANT: {total_bant} veces en {trabajadores_con_bant} trabajadores")
        print(f"6. Turno BLPT: {total_blpt} veces en {trabajadores_con_blpt} trabajadores")
        
        return datos_trabajadores
        
    except Exception as e:
        print(f"Error al generar el reporte: {e}")
        return None

def crear_excel_formateado_con_turnos_individuales(datos_trabajadores, parejas_ordenadas, archivo_salida, 
                                                  parejas_sumatoria1, parejas_sumatoria2, parejas_sumatoria3, parejas_resto):
    """Crea un archivo Excel con formato profesional incluyendo cuatro columnas de sumatoria y turnos individuales"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    if wb.active:
        wb.remove(wb.active)
    
    # Crear hoja principal
    ws = wb.create_sheet("Reporte Parejas Turnos")
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sumatoria1_header_font = Font(bold=True, color="FFFFFF", size=12)
    sumatoria1_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    sumatoria2_header_font = Font(bold=True, color="FFFFFF", size=12)
    sumatoria2_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    sumatoria3_header_font = Font(bold=True, color="FFFFFF", size=12)
    sumatoria3_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    sumatoria4_header_font = Font(bold=True, color="FFFFFF", size=12)
    sumatoria4_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    turno_individual_font = Font(bold=True, color="FFFFFF", size=12)
    turno_individual_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir título
    ws['A1'] = "REPORTE DE PAREJAS DE TURNOS CON CUATRO SUMATORIAS Y TURNOS INDIVIDUALES"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:N1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Escribir encabezados principales
    headers = ['No.', 'Código Trabajador'] + parejas_ordenadas + [
        'SUMATORIA 1\nMAST/NANR +\nMLPR/NLPR +\nBLPT/NLPR',
        'SUMATORIA 2\nTAST/SLN4 +\nTAST/SLN3',
        'SUMATORIA 3\nTANT/NANT +\nTLPT/NLPT',
        'SUMATORIA 4\nRESTO DE\nPAREJAS',
        'TURNO\nBANT',
        'TURNO\nBLPT'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        
        # Aplicar estilo especial para las columnas de sumatoria y turnos individuales
        if col == len(headers) - 5:  # Sumatoria 1
            cell.font = sumatoria1_header_font
            cell.fill = sumatoria1_header_fill
        elif col == len(headers) - 4:  # Sumatoria 2
            cell.font = sumatoria2_header_font
            cell.fill = sumatoria2_header_fill
        elif col == len(headers) - 3:  # Sumatoria 3
            cell.font = sumatoria3_header_font
            cell.fill = sumatoria3_header_fill
        elif col == len(headers) - 2:  # Sumatoria 4
            cell.font = sumatoria4_header_font
            cell.fill = sumatoria4_header_fill
        elif col == len(headers) - 1:  # Turno BANT
            cell.font = turno_individual_font
            cell.fill = turno_individual_fill
        elif col == len(headers):  # Turno BLPT
            cell.font = turno_individual_font
            cell.fill = turno_individual_fill
        else:
            cell.font = header_font
            cell.fill = header_fill
        
        cell.border = border
        cell.alignment = center_alignment
    
    # Escribir datos
    for row_idx, trabajador in enumerate(datos_trabajadores, 4):
        # Escribir No. y Código
        ws.cell(row=row_idx, column=1, value=trabajador['No.']).border = border
        ws.cell(row=row_idx, column=2, value=trabajador['Codigo']).border = border
        
        # Escribir conteos de parejas
        for col_idx, pareja in enumerate(parejas_ordenadas, 3):
            conteo = trabajador['Parejas'].get(pareja, 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=conteo)
            cell.border = border
            cell.alignment = center_alignment
            
            # Resaltar valores mayores a 0
            if conteo > 0:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Calcular y escribir sumatorias
        sumatoria1 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria1)
        sumatoria2 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria2)
        sumatoria3 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria3)
        sumatoria4 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_resto)
        
        # Conteo de turnos individuales
        bant_count = trabajador['Turnos_Individuales'].get('BANT', 0)
        blpt_count = trabajador['Turnos_Individuales'].get('BLPT', 0)
        
        # Escribir sumatorias y turnos individuales
        col_sumatoria1 = len(headers) - 5
        col_sumatoria2 = len(headers) - 4
        col_sumatoria3 = len(headers) - 3
        col_sumatoria4 = len(headers) - 2
        col_bant = len(headers) - 1
        col_blpt = len(headers)
        
        ws.cell(row=row_idx, column=col_sumatoria1, value=sumatoria1).border = border
        ws.cell(row=row_idx, column=col_sumatoria2, value=sumatoria2).border = border
        ws.cell(row=row_idx, column=col_sumatoria3, value=sumatoria3).border = border
        ws.cell(row=row_idx, column=col_sumatoria4, value=sumatoria4).border = border
        ws.cell(row=row_idx, column=col_bant, value=bant_count).border = border
        ws.cell(row=row_idx, column=col_blpt, value=blpt_count).border = border
        
        # Centrar y resaltar sumatorias y turnos individuales
        for col in [col_sumatoria1, col_sumatoria2, col_sumatoria3, col_sumatoria4, col_bant, col_blpt]:
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = center_alignment
            if cell.value > 0:
                if col == col_sumatoria1:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif col == col_sumatoria2:
                    cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                elif col == col_sumatoria3:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif col == col_sumatoria4:
                    cell.fill = PatternFill(start_color="E6E0EC", end_color="E6E0EC", fill_type="solid")
                elif col == col_bant or col == col_blpt:
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        if col >= len(headers) - 5:  # Columnas de sumatoria y turnos individuales
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Crear hoja de resumen
    ws_resumen = wb.create_sheet("Resumen")
    
    # Título del resumen
    ws_resumen['A1'] = "RESUMEN ESTADÍSTICO CON CUATRO SUMATORIAS Y TURNOS INDIVIDUALES"
    ws_resumen['A1'].font = Font(bold=True, size=16, color="366092")
    ws_resumen.merge_cells('A1:C1')
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados del resumen
    resumen_headers = ['Pareja de Turnos', 'Total Ocurrencias', 'Trabajadores con Pareja']
    for col, header in enumerate(resumen_headers, 1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos del resumen
    for row_idx, pareja in enumerate(parejas_ordenadas, 4):
        total = sum(t['Parejas'].get(pareja, 0) for t in datos_trabajadores)
        trabajadores_con_pareja = sum(1 for t in datos_trabajadores if t['Parejas'].get(pareja, 0) > 0)
        
        ws_resumen.cell(row=row_idx, column=1, value=pareja).border = border
        ws_resumen.cell(row=row_idx, column=2, value=total).border = border
        ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_pareja).border = border
        
        # Centrar valores
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).alignment = center_alignment
    
    # Agregar filas de sumatorias en el resumen
    row_idx = len(parejas_ordenadas) + 4
    
    # Calcular totales de sumatorias
    total_sumatoria1 = sum(sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria1) for t in datos_trabajadores)
    total_sumatoria2 = sum(sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria2) for t in datos_trabajadores)
    total_sumatoria3 = sum(sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria3) for t in datos_trabajadores)
    total_sumatoria4 = sum(sum(t['Parejas'].get(pareja, 0) for pareja in parejas_resto) for t in datos_trabajadores)
    
    trabajadores_con_sumatoria1 = sum(1 for t in datos_trabajadores if sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria1) > 0)
    trabajadores_con_sumatoria2 = sum(1 for t in datos_trabajadores if sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria2) > 0)
    trabajadores_con_sumatoria3 = sum(1 for t in datos_trabajadores if sum(t['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria3) > 0)
    trabajadores_con_sumatoria4 = sum(1 for t in datos_trabajadores if sum(t['Parejas'].get(pareja, 0) for pareja in parejas_resto) > 0)
    
    # Calcular totales de turnos individuales
    total_bant = sum(t['Turnos_Individuales'].get('BANT', 0) for t in datos_trabajadores)
    total_blpt = sum(t['Turnos_Individuales'].get('BLPT', 0) for t in datos_trabajadores)
    trabajadores_con_bant = sum(1 for t in datos_trabajadores if t['Turnos_Individuales'].get('BANT', 0) > 0)
    trabajadores_con_blpt = sum(1 for t in datos_trabajadores if t['Turnos_Individuales'].get('BLPT', 0) > 0)
    
    # Sumatoria 1
    ws_resumen.cell(row=row_idx, column=1, value="SUMATORIA 1 (MAST/NANR + MLPR/NLPR + BLPT/NLPR)").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_sumatoria1).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_sumatoria1).border = border
    
    # Aplicar estilo especial a la fila de sumatoria 1
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = sumatoria1_header_font
        cell.fill = sumatoria1_header_fill
        cell.alignment = center_alignment
    
    # Sumatoria 2
    row_idx += 1
    ws_resumen.cell(row=row_idx, column=1, value="SUMATORIA 2 (TAST/SLN4 + TAST/SLN3)").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_sumatoria2).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_sumatoria2).border = border
    
    # Aplicar estilo especial a la fila de sumatoria 2
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = sumatoria2_header_font
        cell.fill = sumatoria2_header_fill
        cell.alignment = center_alignment
    
    # Sumatoria 3
    row_idx += 1
    ws_resumen.cell(row=row_idx, column=1, value="SUMATORIA 3 (TANT/NANT + TLPT/NLPT)").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_sumatoria3).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_sumatoria3).border = border
    
    # Aplicar estilo especial a la fila de sumatoria 3
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = sumatoria3_header_font
        cell.fill = sumatoria3_header_fill
        cell.alignment = center_alignment
    
    # Sumatoria 4
    row_idx += 1
    ws_resumen.cell(row=row_idx, column=1, value=f"SUMATORIA 4 (RESTO: {', '.join(parejas_resto)})").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_sumatoria4).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_sumatoria4).border = border
    
    # Aplicar estilo especial a la fila de sumatoria 4
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = sumatoria4_header_font
        cell.fill = sumatoria4_header_fill
        cell.alignment = center_alignment
    
    # Turno BANT
    row_idx += 1
    ws_resumen.cell(row=row_idx, column=1, value="TURNO BANT").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_bant).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_bant).border = border
    
    # Aplicar estilo especial a la fila de turno BANT
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = turno_individual_font
        cell.fill = turno_individual_fill
        cell.alignment = center_alignment
    
    # Turno BLPT
    row_idx += 1
    ws_resumen.cell(row=row_idx, column=1, value="TURNO BLPT").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_blpt).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_blpt).border = border
    
    # Aplicar estilo especial a la fila de turno BLPT
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = turno_individual_font
        cell.fill = turno_individual_fill
        cell.alignment = center_alignment
    
    # Ajustar ancho de columnas del resumen
    ws_resumen.column_dimensions['A'].width = 45
    ws_resumen.column_dimensions['B'].width = 18
    ws_resumen.column_dimensions['C'].width = 25
    
    # Crear hoja de detalles por trabajador
    ws_detalles = wb.create_sheet("Detalles por Trabajador")
    
    # Título
    ws_detalles['A1'] = "DETALLES POR TRABAJADOR CON CUATRO SUMATORIAS Y TURNOS INDIVIDUALES"
    ws_detalles['A1'].font = Font(bold=True, size=16, color="366092")
    ws_detalles.merge_cells('A1:I1')
    ws_detalles['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    detalle_headers = ['Trabajador', 'Parejas Encontradas', 'Total Ocurrencias', 
                      'Sumatoria 1\nMAST/MLPR/BLPT', 'Sumatoria 2\nTAST/SLN', 
                      'Sumatoria 3\nTANT/TLPT', 'Sumatoria 4\nRESTO',
                      'Turno BANT', 'Turno BLPT']
    for col, header in enumerate(detalle_headers, 1):
        cell = ws_detalles.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos de detalles
    row_idx = 4
    for trabajador in datos_trabajadores:
        codigo = trabajador['Codigo']
        parejas_trabajador = []
        total_ocurrencias = 0
        
        for pareja in parejas_ordenadas:
            conteo = trabajador['Parejas'].get(pareja, 0)
            if conteo > 0:
                parejas_trabajador.append(f"{pareja} ({conteo} veces)")
                total_ocurrencias += conteo
        
        sumatoria1 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria1)
        sumatoria2 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria2)
        sumatoria3 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_sumatoria3)
        sumatoria4 = sum(trabajador['Parejas'].get(pareja, 0) for pareja in parejas_resto)
        bant_count = trabajador['Turnos_Individuales'].get('BANT', 0)
        blpt_count = trabajador['Turnos_Individuales'].get('BLPT', 0)
        
        if parejas_trabajador:  # Solo mostrar trabajadores con parejas
            ws_detalles.cell(row=row_idx, column=1, value=codigo).border = border
            ws_detalles.cell(row=row_idx, column=2, value=", ".join(parejas_trabajador)).border = border
            ws_detalles.cell(row=row_idx, column=3, value=total_ocurrencias).border = border
            ws_detalles.cell(row=row_idx, column=4, value=sumatoria1).border = border
            ws_detalles.cell(row=row_idx, column=5, value=sumatoria2).border = border
            ws_detalles.cell(row=row_idx, column=6, value=sumatoria3).border = border
            ws_detalles.cell(row=row_idx, column=7, value=sumatoria4).border = border
            ws_detalles.cell(row=row_idx, column=8, value=bant_count).border = border
            ws_detalles.cell(row=row_idx, column=9, value=blpt_count).border = border
            
            # Centrar valores
            for col in range(1, 10):
                ws_detalles.cell(row=row_idx, column=col).alignment = center_alignment
            
            # Resaltar sumatorias si son mayores a 0
            if sumatoria1 > 0:
                ws_detalles.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            if sumatoria2 > 0:
                ws_detalles.cell(row=row_idx, column=5).fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
            if sumatoria3 > 0:
                ws_detalles.cell(row=row_idx, column=6).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            if sumatoria4 > 0:
                ws_detalles.cell(row=row_idx, column=7).fill = PatternFill(start_color="E6E0EC", end_color="E6E0EC", fill_type="solid")
            if bant_count > 0:
                ws_detalles.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            if blpt_count > 0:
                ws_detalles.cell(row=row_idx, column=9).fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            row_idx += 1
    
    # Ajustar ancho de columnas de detalles
    ws_detalles.column_dimensions['A'].width = 15
    ws_detalles.column_dimensions['B'].width = 50
    ws_detalles.column_dimensions['C'].width = 18
    ws_detalles.column_dimensions['D'].width = 20
    ws_detalles.column_dimensions['E'].width = 20
    ws_detalles.column_dimensions['F'].width = 20
    ws_detalles.column_dimensions['G'].width = 20
    ws_detalles.column_dimensions['H'].width = 15
    ws_detalles.column_dimensions['I'].width = 15
    
    # Guardar archivo
    wb.save(archivo_salida)

def main():
    archivo = "conteoTurnosTrabajador.xlsm"
    
    print("=== GENERADOR DE REPORTE EXCEL CON CUATRO SUMATORIAS Y TURNOS INDIVIDUALES ===\n")
    
    # Generar reporte
    datos_trabajadores = generar_reporte_excel_con_turnos_individuales(archivo)
    
    if datos_trabajadores is not None:
        print(f"\nEl archivo Excel se ha guardado como 'reporte_parejas_turnos_con_individuales.xlsx'")
        print("El archivo contiene 3 hojas:")
        print("1. 'Reporte Parejas Turnos' - Tabla completa con cuatro columnas de sumatoria + turnos individuales")
        print("2. 'Resumen' - Estadísticas generales incluyendo las cuatro sumatorias y turnos individuales")
        print("3. 'Detalles por Trabajador' - Lista de parejas por trabajador con las cuatro sumatorias y turnos individuales")

if __name__ == "__main__":
    main() 