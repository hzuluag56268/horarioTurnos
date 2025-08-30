import pandas as pd
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def generar_reporte_excel_con_sumatoria(archivo_excel, archivo_salida="reporte_parejas_turnos_con_sumatoria.xlsx"):
    """
    Genera un reporte Excel con formato profesional del conteo de parejas de turnos,
    incluyendo una columna de sumatoria para MAST/NANR, MLPR/NLPR y BLPT/NLPR.
    
    Args:
        archivo_excel (str): Ruta del archivo Excel
        archivo_salida (str): Nombre del archivo Excel de salida
    """
    try:
        # Leer el archivo Excel
        df = pd.read_excel(archivo_excel, header=None)
        
        # Encontrar la fila de encabezados (días)
        fila_encabezados = None
        for i, row in df.iterrows():
            if any(str(cell).startswith(('FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED', 'THU')) for cell in row if pd.notna(cell)):
                fila_encabezados = i
                break
        
        if fila_encabezados is None:
            print("No se encontraron encabezados de días")
            return
        
        # Encontrar las filas de trabajadores
        filas_trabajadores = []
        for i in range(fila_encabezados + 1, len(df)):
            row = df.iloc[i]
            if pd.notna(row[0]) and pd.notna(row[1]) and str(row[0]).isdigit():
                filas_trabajadores.append(i)
        
        # Recolectar todas las parejas únicas
        todas_parejas = set()
        datos_trabajadores = []
        
        for fila_idx in filas_trabajadores:
            trabajador = df.iloc[fila_idx]
            codigo_trabajador = trabajador[1] if pd.notna(trabajador[1]) else f"Trabajador_{trabajador[0]}"
            numero_trabajador = trabajador[0]
            
            # Contar parejas de turnos
            parejas_encontradas = defaultdict(int)
            
            # Recorrer las columnas de días (cada día tiene 2 columnas)
            for col in range(2, len(trabajador), 2):
                if col + 1 < len(trabajador):
                    turno1 = str(trabajador[col]) if pd.notna(trabajador[col]) else ""
                    turno2 = str(trabajador[col + 1]) if pd.notna(trabajador[col + 1]) else ""
                    
                    # Crear la pareja de turnos
                    pareja = f"{turno1}/{turno2}" if turno1 and turno2 else ""
                    
                    if pareja:
                        parejas_encontradas[pareja] += 1
                        todas_parejas.add(pareja)
            
            # Guardar datos del trabajador
            datos_trabajadores.append({
                'No.': numero_trabajador,
                'Codigo': codigo_trabajador,
                'Parejas': dict(parejas_encontradas)
            })
        
        # Crear DataFrame para el reporte
        parejas_ordenadas = sorted(todas_parejas)
        
        # Crear columnas para cada pareja
        columnas = ['No.', 'Codigo'] + parejas_ordenadas
        
        # Crear filas de datos
        filas_datos = []
        for trabajador in datos_trabajadores:
            fila = [trabajador['No.'], trabajador['Codigo']]
            for pareja in parejas_ordenadas:
                conteo = trabajador['Parejas'].get(pareja, 0)
                fila.append(conteo)
            filas_datos.append(fila)
        
        # Crear DataFrame
        df_reporte = pd.DataFrame(filas_datos, columns=columnas)
        
        # Agregar columna de sumatoria
        parejas_sumatoria = ['MAST/NANR', 'MLPR/NLPR', 'BLPT/NLPR']
        df_reporte['SUMATORIA_MAST_MLPR_BLPT'] = 0
        
        for pareja in parejas_sumatoria:
            if pareja in df_reporte.columns:
                df_reporte['SUMATORIA_MAST_MLPR_BLPT'] += df_reporte[pareja]
        
        # Crear archivo Excel con formato
        crear_excel_formateado_con_sumatoria(df_reporte, archivo_salida, parejas_ordenadas, parejas_sumatoria)
        
        print(f"Reporte Excel generado exitosamente: {archivo_salida}")
        print(f"Total de trabajadores: {len(datos_trabajadores)}")
        print(f"Total de parejas diferentes: {len(parejas_ordenadas)}")
        
        # Mostrar resumen de la sumatoria
        print(f"\nResumen de la sumatoria ({', '.join(parejas_sumatoria)}):")
        total_sumatoria = df_reporte['SUMATORIA_MAST_MLPR_BLPT'].sum()
        trabajadores_con_sumatoria = (df_reporte['SUMATORIA_MAST_MLPR_BLPT'] > 0).sum()
        print(f"  Total de ocurrencias: {total_sumatoria}")
        print(f"  Trabajadores con al menos una ocurrencia: {trabajadores_con_sumatoria}")
        
        # Mostrar trabajadores con mayor sumatoria
        print(f"\nTop 5 trabajadores con mayor sumatoria:")
        top_5 = df_reporte.nlargest(5, 'SUMATORIA_MAST_MLPR_BLPT')[['Codigo', 'SUMATORIA_MAST_MLPR_BLPT']]
        for _, row in top_5.iterrows():
            print(f"  {row['Codigo']}: {row['SUMATORIA_MAST_MLPR_BLPT']} veces")
        
        return df_reporte
        
    except Exception as e:
        print(f"Error al generar el reporte: {e}")
        return None

def crear_excel_formateado_con_sumatoria(df_reporte, archivo_salida, parejas_ordenadas, parejas_sumatoria):
    """Crea un archivo Excel con formato profesional incluyendo la columna de sumatoria"""
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # Crear hoja principal
    ws = wb.create_sheet("Reporte Parejas Turnos")
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sumatoria_header_font = Font(bold=True, color="FFFFFF", size=12)
    sumatoria_header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir título
    ws['A1'] = "REPORTE DE PAREJAS DE TURNOS CON SUMATORIA"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Escribir encabezados principales
    headers = ['No.', 'Código Trabajador'] + parejas_ordenadas + ['SUMATORIA\nMAST/NANR +\nMLPR/NLPR +\nBLPT/NLPR']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        
        # Aplicar estilo especial para la columna de sumatoria
        if col == len(headers):  # Última columna (sumatoria)
            cell.font = sumatoria_header_font
            cell.fill = sumatoria_header_fill
        else:
            cell.font = header_font
            cell.fill = header_fill
        
        cell.border = border
        cell.alignment = center_alignment
    
    # Escribir datos
    for row_idx, (_, row) in enumerate(df_reporte.iterrows(), 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_alignment
            
            # Resaltar valores mayores a 0
            if col_idx > 2 and value > 0:  # Columnas de parejas
                if col_idx == len(headers):  # Columna de sumatoria
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        if col == len(headers):  # Columna de sumatoria
            ws.column_dimensions[chr(64 + col)].width = 20
        else:
            ws.column_dimensions[chr(64 + col)].width = 15
    
    # Crear hoja de resumen
    ws_resumen = wb.create_sheet("Resumen")
    
    # Título del resumen
    ws_resumen['A1'] = "RESUMEN ESTADÍSTICO"
    ws_resumen['A1'].font = Font(bold=True, size=16, color="366092")
    ws_resumen.merge_cells('A1:C1')
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados del resumen
    resumen_headers = ['Pareja de Turnos', 'Total Ocurrencias', 'Trabajadores con Pareja']
    for col, header in enumerate(resumen_headers, 1):
        cell = ws_resumen.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos del resumen
    for row_idx, pareja in enumerate(parejas_ordenadas, 4):
        total = df_reporte[pareja].sum()
        trabajadores_con_pareja = (df_reporte[pareja] > 0).sum()
        
        ws_resumen.cell(row=row_idx, column=1, value=pareja).border = border
        ws_resumen.cell(row=row_idx, column=2, value=total).border = border
        ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_pareja).border = border
        
        # Centrar valores
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).alignment = center_alignment
    
    # Agregar fila de sumatoria en el resumen
    row_idx = len(parejas_ordenadas) + 4
    total_sumatoria = df_reporte['SUMATORIA_MAST_MLPR_BLPT'].sum()
    trabajadores_con_sumatoria = (df_reporte['SUMATORIA_MAST_MLPR_BLPT'] > 0).sum()
    
    ws_resumen.cell(row=row_idx, column=1, value="SUMATORIA (MAST/NANR + MLPR/NLPR + BLPT/NLPR)").border = border
    ws_resumen.cell(row=row_idx, column=2, value=total_sumatoria).border = border
    ws_resumen.cell(row=row_idx, column=3, value=trabajadores_con_sumatoria).border = border
    
    # Aplicar estilo especial a la fila de sumatoria
    for col in range(1, 4):
        cell = ws_resumen.cell(row=row_idx, column=col)
        cell.font = sumatoria_header_font
        cell.fill = sumatoria_header_fill
        cell.alignment = center_alignment
    
    # Ajustar ancho de columnas del resumen
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 18
    ws_resumen.column_dimensions['C'].width = 25
    
    # Crear hoja de detalles por trabajador
    ws_detalles = wb.create_sheet("Detalles por Trabajador")
    
    # Título
    ws_detalles['A1'] = "DETALLES POR TRABAJADOR"
    ws_detalles['A1'].font = Font(bold=True, size=16, color="366092")
    ws_detalles.merge_cells('A1:D1')
    ws_detalles['A1'].alignment = Alignment(horizontal='center')
    
    # Encabezados
    detalle_headers = ['Trabajador', 'Parejas Encontradas', 'Total Ocurrencias', 'Sumatoria MAST/MLPR/BLPT']
    for col, header in enumerate(detalle_headers, 1):
        cell = ws_detalles.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos de detalles
    row_idx = 4
    for _, trabajador in df_reporte.iterrows():
        codigo = trabajador['Codigo']
        parejas_trabajador = []
        total_ocurrencias = 0
        
        for pareja in parejas_ordenadas:
            if trabajador[pareja] > 0:
                parejas_trabajador.append(f"{pareja} ({trabajador[pareja]} veces)")
                total_ocurrencias += trabajador[pareja]
        
        sumatoria_especifica = trabajador['SUMATORIA_MAST_MLPR_BLPT']
        
        if parejas_trabajador:  # Solo mostrar trabajadores con parejas
            ws_detalles.cell(row=row_idx, column=1, value=codigo).border = border
            ws_detalles.cell(row=row_idx, column=2, value=", ".join(parejas_trabajador)).border = border
            ws_detalles.cell(row=row_idx, column=3, value=total_ocurrencias).border = border
            ws_detalles.cell(row=row_idx, column=4, value=sumatoria_especifica).border = border
            
            # Centrar valores
            for col in range(1, 5):
                ws_detalles.cell(row=row_idx, column=col).alignment = center_alignment
            
            # Resaltar sumatoria si es mayor a 0
            if sumatoria_especifica > 0:
                ws_detalles.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            row_idx += 1
    
    # Ajustar ancho de columnas de detalles
    ws_detalles.column_dimensions['A'].width = 15
    ws_detalles.column_dimensions['B'].width = 50
    ws_detalles.column_dimensions['C'].width = 18
    ws_detalles.column_dimensions['D'].width = 25
    
    # Guardar archivo
    wb.save(archivo_salida)

def main():
    archivo = "conteoTurnosTrabajador.xlsm"
    
    print("=== GENERADOR DE REPORTE EXCEL CON SUMATORIA ===\n")
    
    # Generar reporte
    df_reporte = generar_reporte_excel_con_sumatoria(archivo)
    
    if df_reporte is not None:
        print(f"\nEl archivo Excel se ha guardado como 'reporte_parejas_turnos_con_sumatoria.xlsx'")
        print("El archivo contiene 3 hojas:")
        print("1. 'Reporte Parejas Turnos' - Tabla completa con conteos y sumatoria")
        print("2. 'Resumen' - Estadísticas generales incluyendo sumatoria")
        print("3. 'Detalles por Trabajador' - Lista de parejas por trabajador con sumatoria")

if __name__ == "__main__":
    main() 